import zipfile, re, os
from tools.noi_tools import WORK_DIR, _read_shared_strings, _NOI_RCSD_XML, _SHARED_STRINGS

cdg_path = os.path.join(WORK_DIR, '2603 Control De Gestión Renta Comercial vAgente.xlsx')
with zipfile.ZipFile(cdg_path) as z:
    sheet_xml = z.read(_NOI_RCSD_XML).decode('utf-8')
    ss_xml = z.read(_SHARED_STRINGS).decode('utf-8')

ss_dict = _read_shared_strings(ss_xml)

print("=== CDG NOI rows 287-295 labels (col C) ===")
for row_num in range(287, 296):
    row_m = re.search(r'<row r="' + str(row_num) + r'"[^>]*>(.*?)</row>', sheet_xml, re.DOTALL)
    if not row_m:
        print(f'Row {row_num}: not found')
        continue
    row_xml = row_m.group(0)
    c_m = re.search(r'<c r="C' + str(row_num) + r'"[^>]*>.*?<v>(\d+)</v>', row_xml, re.DOTALL)
    if c_m and 't="s"' in row_xml:
        label = ss_dict.get(int(c_m.group(1)), '')
        print(f'Row {row_num}: {label!r}')
    else:
        print(f'Row {row_num}: no string label')

print("\n=== INMOSA ER data (col B labels) ===")
import openpyxl
wb = openpyxl.load_workbook(
    r'C:\Users\raimundo.opazo\OneDrive - Toesca\Inmobiliario Toesca - Documentos\Fondos\Rentas TRI\Activos\INMOSA\Flujos\2026\EEFF y FC Senior Assist Mar.26.xlsx',
    data_only=True)
ws = wb['Activo Pasivo EERR']
for r in range(1, 100):
    b = ws.cell(r, 2).value
    c = ws.cell(r, 3).value
    if b and c is not None and isinstance(c, (int, float)):
        print(f'Row {r}: {b!r} = {c}')
wb.close()
