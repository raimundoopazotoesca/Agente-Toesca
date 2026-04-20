"""
Herramientas para actualizar la hoja NOI-RCSD del CDG Rentas Comerciales.

Activos y fuentes:
  INMOSA (filas 287-295)      : ER-FC INMOSA (SharePoint: Fondo Rentas/Flujos INMOSA)
  Parque Titanium (335-379)   : hoja 'NOI PT' del RR JLL (en WORK_DIR)
  Viña Centro                 : INFORME EEFF VINA CENTRO → crea col en ER Viña → NOI auto
  Fondo Apoquindo (426-456)   : hoja 'NOI PT' del RR JLL (en WORK_DIR)
  Apoquindo 3001 (468-476)    : hoja 'NOI PT' del RR JLL (en WORK_DIR)
  Mall Curicó                 : INFORME EEFF CURICO → crea col en ER Curico → NOI auto

Estructura ER Viña (CDG):
  Fila 5 = UF del mes, fila 6 = fecha del mes
  Col C = código de cuenta (coincide con ESTADO DE RESULTADO del EEFF Viña)
  Valores almacenados en UF = CLP_eeff / UF_mes

Estructura ER Curico (CDG):
  Fila 3 = UF del mes (fórmula), fila 4 = fecha del mes
  Col C = código de cuenta (coincide con ESTADO DE RESULTADO del EEFF Curico)
  Valores almacenados en CLP raw

Implementación: zipfile + XML directo para máxima velocidad con archivo 14MB/87 hojas.
"""

import glob
import os
import re
import shutil
import tempfile
import zipfile
from datetime import date, datetime, timedelta
from typing import Optional

import openpyxl

from config import SHAREPOINT_DIR, WORK_DIR

# ── Rutas base ─────────────────────────────────────────────────────────────────
_TRES_A_BASE = os.path.join(SHAREPOINT_DIR, "EEFF Proveedores", "Informes TresA")
_TRES_A_DIRS = {
    "vina":   os.path.join(_TRES_A_BASE, "Viña Centro"),
    "curico": os.path.join(_TRES_A_BASE, "Curico"),
}
_RR_JLL_BASE = os.path.join(SHAREPOINT_DIR, "Rent Rolls", "JLL")
_INMOSA_BASE = os.path.join(SHAREPOINT_DIR, "EEFF Proveedores", "Flujos INMOSA (Residencias Adulto Mayor)")

# Sheet XML paths within the CDG xlsx (derived from workbook.xml.rels)
_ER_VINA_XML   = "xl/worksheets/sheet54.xml"
_ER_CURICO_XML = "xl/worksheets/sheet56.xml"
_NOI_RCSD_XML  = "xl/worksheets/sheet40.xml"
_SHARED_STRINGS = "xl/sharedStrings.xml"

_NOI_SHEET_NAME = "NOI- RCSD"

# ── Mapeo NOI-RCSD row → ER row (derivado de fórmulas columna BZ) ──────────────
# Curico: NOI filas 260-276 → ER Curico filas de la Section 2 (113+)
_NOI_CURICO_MAP = {
    260: 115, 261: 117, 262: 118, 263: 119, 264: 120, 265: 121,
    266: 122, 267: 123, 268: 124, 269: 125, 270: 126, 271: 127,
    272: 128, 273: 129, 274: 130, 275: 131, 276: 132,
}
# Viña: NOI filas 196-214 → ER Viña filas de la Section 2 (97+)
_NOI_VINA_MAP = {
    196: 98, 197: 99, 198: 100, 199: 101, 200: 102, 201: 103,
    202: 104, 203: 105, 204: 107, 205: 108, 206: 109, 207: 110,
    208: 111, 209: 112, 210: 113, 211: 114, 212: 115, 213: 116, 214: 117,
}
_ER_SHEET_NAME = {
    "vina":   "ER Vi\u00f1a",   # 'ER Viña'
    "curico": "ER Curico",
}
_NOI_ROW_MAP   = {"vina": _NOI_VINA_MAP, "curico": _NOI_CURICO_MAP}
_NOI_ROW7      = 7   # fila de fechas globales en NOI-RCSD


# ── Utilidades de fecha ────────────────────────────────────────────────────────

def _ultimo_dia_mes(año: int, mes: int) -> date:
    if mes == 12:
        return date(año + 1, 1, 1) - timedelta(days=1)
    return date(año, mes + 1, 1) - timedelta(days=1)


def _excel_serial(d: date) -> int:
    return (d - date(1899, 12, 30)).days


# ── Utilidades XML (portadas de gestion_renta_tools) ──────────────────────────

def _col_num(letter: str) -> int:
    n = 0
    for c in letter.upper():
        n = n * 26 + ord(c) - ord("A") + 1
    return n


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _find_cell_bounds(row_xml: str, ref: str) -> tuple:
    """Retorna (start, end) del span completo de la celda en row_xml."""
    tag = f'<c r="{ref}"'
    start = row_xml.find(tag)
    if start == -1:
        return -1, -1
    i = start + len(tag)
    while i < len(row_xml):
        if row_xml[i:i+2] == '/>':
            return start, i + 2
        if row_xml[i] == '>':
            end = row_xml.find('</c>', i)
            return start, (end + 4) if end != -1 else i + 1
        i += 1
    return start, len(row_xml)


def _replace_or_insert_cell(row_xml: str, ref: str, new_cell: str) -> str:
    """Reemplaza la celda si existe, o la inserta en orden de columna."""
    start, end = _find_cell_bounds(row_xml, ref)
    if start != -1:
        return row_xml[:start] + new_cell + row_xml[end:]
    col = re.sub(r"\d", "", ref)
    col_n = _col_num(col)
    for m in re.finditer(r'<c r="([A-Z]+)\d+"', row_xml):
        if _col_num(m.group(1)) > col_n:
            return row_xml[: m.start()] + new_cell + row_xml[m.start():]
    return row_xml + new_cell


def _get_cell_style(row_xml: str, ref: str, default: str = "0") -> str:
    """Extrae el atributo s= de una celda existente en el XML."""
    tag = f'<c r="{ref}"'
    pos = row_xml.find(tag)
    if pos == -1:
        return default
    snippet = row_xml[pos: pos + 200]
    m = re.search(r'\bs="(\d+)"', snippet)
    return m.group(1) if m else default


def _update_row_spans(row_xml: str, target_col: int) -> str:
    """Extiende el atributo spans si target_col queda fuera del rango actual."""
    m = re.search(r'\bspans="(\d+):(\d+)"', row_xml)
    if not m:
        return row_xml
    low, high = int(m.group(1)), int(m.group(2))
    if target_col <= high:
        return row_xml
    new_spans = f'spans="{low}:{target_col}"'
    return row_xml[:m.start()] + new_spans + row_xml[m.end():]


# ── Leer shared strings ────────────────────────────────────────────────────────

def _read_shared_strings(ss_xml: str) -> dict:
    """Retorna {index: string_value} desde sharedStrings.xml."""
    result = {}
    for i, m in enumerate(re.finditer(r'<si>(.*?)</si>', ss_xml, re.DOTALL)):
        # Puede tener <t> o <r><t> (rich text)
        texts = re.findall(r'<t[^>]*>([^<]*)</t>', m.group(1))
        result[i] = "".join(texts)
    return result


# ── Encontrar columna de fecha en una fila ────────────────────────────────────

def _find_col_for_serial(row_xml: str, target_serial: int) -> Optional[int]:
    """
    Busca en el XML de una fila la celda cuyo <v> sea igual al serial objetivo.
    Retorna el número de columna (1-based) o None si no existe.
    """
    for m in re.finditer(r'<c r="([A-Z]+)\d+"[^>]*>(?:<f[^>]*/?>|<f[^>]*>.*?</f>)?<v>(\d+)</v>', row_xml, re.DOTALL):
        if int(m.group(2)) == target_serial:
            return _col_num(m.group(1))
    return None


def _last_date_col(row_xml: str) -> tuple:
    """
    Retorna (col_num, serial) de la última celda con valor numérico en la fila.
    Considera solo celdas con serials en rango plausible de fechas (40000-60000).
    """
    last_col = None
    last_serial = None
    for m in re.finditer(r'<c r="([A-Z]+)\d+"[^>]*>(?:<f[^>]*/?>|<f[^>]*>.*?</f>)?<v>(\d+)</v>', row_xml, re.DOTALL):
        serial = int(m.group(2))
        if 40000 <= serial <= 60000:  # rango de fechas 2009-2064
            col = _col_num(m.group(1))
            if last_col is None or col > last_col:
                last_col = col
                last_serial = serial
    return last_col, last_serial


# ── Encontrar filas con código de cuenta en col C ─────────────────────────────

def _build_account_row_map(sheet_xml: str, ss_dict: dict) -> dict:
    """
    Escanea el XML de la hoja y retorna {account_code: row_number}
    para filas donde col C es un shared string con código de cuenta (ej: '4-1-01-100 ...').
    """
    result = {}
    for row_m in re.finditer(r'<row r="(\d+)"[^>]*>(.*?)</row>', sheet_xml, re.DOTALL):
        row_num = int(row_m.group(1))
        row_xml = row_m.group(2)
        # Buscar celda en col C con t="s"
        c_m = re.search(r'<c r="C' + str(row_num) + r'"[^>]*t="s"[^>]*>.*?<v>(\d+)</v>', row_xml, re.DOTALL)
        if not c_m:
            # También puede ser <c r="C3" s="..." t="s">
            c_m = re.search(r'<c r="C' + str(row_num) + r'"[^>]*>.*?<v>(\d+)</v>', row_xml, re.DOTALL)
            if not c_m:
                continue
        ss_idx = int(c_m.group(1))
        code = ss_dict.get(ss_idx, "")
        # Filtrar: código de cuenta comienza con dígito-guión-dígito
        if re.match(r'^\d[-\d]', code.strip()):
            result[code.strip()] = row_num
    return result


# ── Actualizar/agregar celda de valor numérico en una fila ────────────────────

def _set_numeric_cell(sheet_xml: str, row_num: int, col_num: int,
                      value: float, style: str = "0") -> str:
    """
    En el XML completo de la hoja, actualiza o agrega la celda (row_num, col_num)
    con el valor numérico indicado. Respeta fórmulas existentes (no las sobreescribe).
    """
    ref = f"{_col_letter(col_num)}{row_num}"
    row_tag = f'<row r="{row_num}"'
    row_start = sheet_xml.find(row_tag)
    if row_start == -1:
        # La fila no existe — no se puede agregar fácilmente, saltar
        return sheet_xml

    # Encontrar fin de la fila
    row_end_tag = '</row>'
    row_end = sheet_xml.find(row_end_tag, row_start)
    if row_end == -1:
        return sheet_xml
    row_end += len(row_end_tag)

    row_xml = sheet_xml[row_start:row_end]

    # Verificar si la celda tiene fórmula — si sí, solo actualizar <v>
    cell_start, cell_end = _find_cell_bounds(row_xml, ref)
    if cell_start != -1:
        cell_snippet = row_xml[cell_start:cell_end]
        has_formula = '<f' in cell_snippet
        if has_formula:
            # Solo actualizar el valor cacheado <v>
            val_m = re.search(r'<v>[^<]*</v>', cell_snippet)
            if val_m:
                new_snippet = (cell_snippet[:val_m.start()]
                               + f'<v>{value}</v>'
                               + cell_snippet[val_m.end():])
            else:
                new_snippet = cell_snippet.replace('</c>', f'<v>{value}</v></c>')
            new_row = row_xml[:cell_start] + new_snippet + row_xml[cell_end:]
        else:
            # Reemplazar celda completa
            new_cell = f'<c r="{ref}" s="{style}"><v>{value}</v></c>'
            new_row = _replace_or_insert_cell(row_xml, ref, new_cell)
    else:
        # Celda no existe — crear con el estilo indicado
        new_cell = f'<c r="{ref}" s="{style}"><v>{value}</v></c>'
        new_row = _replace_or_insert_cell(row_xml, ref, new_cell)

    # Extender spans si es necesario
    new_row = _update_row_spans(new_row, col_num)

    return sheet_xml[:row_start] + new_row + sheet_xml[row_end:]


# ── Leer EEFF ESTADO DE RESULTADO ─────────────────────────────────────────────

def _leer_eeff_estado_resultado(eeff_path: str) -> tuple:
    """
    Lee la hoja 'ESTADO DE RESULTADO' del INFORME EEFF (Tres Asociados).
    Retorna (fecha_cierre: date, {codigo_cuenta: valor_clp}).
    Los valores son del mes actual (columna E = col 5).
    """
    try:
        wb = openpyxl.load_workbook(eeff_path, read_only=True, data_only=True)
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), os.path.basename(eeff_path))
        shutil.copy2(eeff_path, tmp)
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)

    # Buscar hoja "ESTADO DE RESULTADO" o "ESTADO DE RESULTADO XXXX"
    er_sheet = next(
        (s for s in wb.sheetnames if s.upper().startswith("ESTADO DE RESULTADO")),
        None,
    )
    if er_sheet is None:
        wb.close()
        return None, {}

    ws = wb[er_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # ── Extraer fecha del título ──────────────────────────────────────────────
    _MESES_NUM = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
        "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
        "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    }
    fecha_cierre = None
    for row in rows[:10]:
        for cell in row:
            if cell and isinstance(cell, str):
                cell_str = str(cell)
                # Formato numérico: AL 31-01-2026 o AL 31/01/2026
                m = re.search(r"AL\s+(\d{1,2})[-/](\d{1,2})[-/](\d{4})", cell_str, re.IGNORECASE)
                if m:
                    try:
                        fecha_cierre = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                    except ValueError:
                        pass
                # Formato textual: AL 31 DE ENERO 2026
                if not fecha_cierre:
                    m2 = re.search(
                        r"AL\s+(\d{1,2})\s+DE\s+(\w+)\s+(\d{4})", cell_str, re.IGNORECASE
                    )
                    if m2:
                        mes_num = _MESES_NUM.get(m2.group(2).lower())
                        if mes_num:
                            try:
                                fecha_cierre = date(int(m2.group(3)), mes_num, int(m2.group(1)))
                            except ValueError:
                                pass
                if fecha_cierre:
                    break
        if fecha_cierre:
            break

    # ── Leer valores (col B = código, col E = mes actual) ─────────────────────
    account_values: dict = {}
    for row in rows[10:]:
        code_raw = row[1] if len(row) > 1 else None
        val_raw  = row[4] if len(row) > 4 else None
        if code_raw is None or val_raw is None:
            continue
        code = str(code_raw).strip()
        if not re.match(r"^\d[-\d]", code):
            continue
        try:
            account_values[code] = float(val_raw)
        except (TypeError, ValueError):
            pass

    return fecha_cierre, account_values


# ── Descubrir archivo EEFF ────────────────────────────────────────────────────

def _find_eeff_file(mall: str, año: int, mes: int) -> Optional[str]:
    """Busca INFORME EEFF en SharePoint TresA/{año}/ y luego en WORK_DIR."""
    mes_str = f"{mes:02d}-{año}"
    keywords = {
        "vina":   ["vi", "centro"],
        "curico": ["curic"],
    }
    kws = keywords.get(mall, [])

    # Primero busca en SharePoint con subcarpeta del año
    sp_año_dir = os.path.join(_TRES_A_DIRS[mall], str(año)) if mall in _TRES_A_DIRS else None
    search_dirs = []
    if sp_año_dir and os.path.isdir(sp_año_dir):
        search_dirs.append(sp_año_dir)
    search_dirs.append(WORK_DIR)

    for d in search_dirs:
        # Buscar con fecha exacta MM-YYYY en nombre
        for f in glob.glob(os.path.join(d, "*.xlsx")):
            bn = os.path.basename(f).lower()
            if mes_str.lower() in bn and all(k in bn for k in kws):
                return f
        # Fallback: más reciente con keywords
        candidates = [
            f for f in glob.glob(os.path.join(d, "*.xlsx"))
            if all(k in os.path.basename(f).lower() for k in kws)
        ]
        if candidates:
            return max(candidates, key=os.path.getmtime)
    return None


# ── Función principal: actualizar ER Viña / ER Curico ─────────────────────────

def _actualizar_er_mall(
    nombre_cdg: str,
    eeff_path: str,
    mall: str,   # 'vina' | 'curico'
) -> str:
    """
    Agrega/actualiza la columna del mes correspondiente en ER Viña o ER Curico del CDG.
    Usa zipfile+XML para máxima velocidad.
    """
    cdg_path = os.path.join(WORK_DIR, nombre_cdg)
    if not os.path.exists(cdg_path):
        return f"Error: '{nombre_cdg}' no encontrado en WORK_DIR."
    if not os.path.exists(eeff_path):
        return f"Error: EEFF no encontrado: {eeff_path}"

    # ── 1. Leer EEFF ──────────────────────────────────────────────────────────
    fecha_cierre, eeff_values = _leer_eeff_estado_resultado(eeff_path)
    if not eeff_values:
        return f"Error: no se pudo leer 'ESTADO DE RESULTADO' en {os.path.basename(eeff_path)}"
    if fecha_cierre is None:
        return "Error: no se pudo determinar la fecha del EEFF."

    año, mes = fecha_cierre.year, fecha_cierre.month
    fecha_fin = _ultimo_dia_mes(año, mes)
    target_serial = _excel_serial(fecha_fin)

    cfg = {
        "vina":   {"xml": _ER_VINA_XML,   "date_row": 6, "uf_row": 5, "in_uf": True},
        "curico": {"xml": _ER_CURICO_XML, "date_row": 4, "uf_row": 3, "in_uf": False},
    }[mall]

    sheet_xml_path = cfg["xml"]
    date_row       = cfg["date_row"]
    uf_row         = cfg["uf_row"]
    in_uf          = cfg["in_uf"]

    # ── 2. Leer XMLs del CDG ──────────────────────────────────────────────────
    with zipfile.ZipFile(cdg_path, "r") as z:
        sheet_xml = z.read(sheet_xml_path).decode("utf-8")
        ss_xml    = z.read(_SHARED_STRINGS).decode("utf-8")

    ss_dict = _read_shared_strings(ss_xml)

    # ── 3. Encontrar columna de fecha ─────────────────────────────────────────
    # Extraer la fila de fecha como texto
    date_row_m = re.search(r'<row r="' + str(date_row) + r'"[^>]*>(.*?)</row>',
                            sheet_xml, re.DOTALL)
    if not date_row_m:
        return f"Error: no se encontró fila {date_row} en {sheet_xml_path}"

    date_row_xml = date_row_m.group(0)
    target_col = _find_col_for_serial(date_row_xml, target_serial)

    if target_col is None:
        # La columna no existe: agregar después de la última columna de fecha
        last_col, _ = _last_date_col(date_row_xml)
        target_col = (last_col or 4) + 1
        # Agregar celda de fecha en la fila de fecha
        style_date = _get_cell_style(date_row_xml,
                                     f"{_col_letter(target_col - 1)}{date_row}", "1239")
        sheet_xml = _set_numeric_cell(sheet_xml, date_row, target_col,
                                       target_serial, style_date)

    col_letter = _col_letter(target_col)

    # ── 4. Obtener UF si necesario (ER Viña almacena en UF) ───────────────────
    uf_mes = None
    if in_uf:
        # Buscar UF en la hoja UF del CDG (openpyxl read_only es rápido)
        try:
            wb_uf = openpyxl.load_workbook(cdg_path, read_only=True, data_only=True)
            if "UF" in wb_uf.sheetnames:
                ws_uf = wb_uf["UF"]
                target_date = fecha_fin
                for row in ws_uf.iter_rows(min_row=5, values_only=True):
                    fecha_cell = row[0]
                    valor_cell = row[1]
                    if fecha_cell is None:
                        continue
                    if isinstance(fecha_cell, datetime):
                        fecha_cell = fecha_cell.date()
                    if fecha_cell == target_date and valor_cell is not None:
                        try:
                            uf_mes = float(valor_cell)
                            break
                        except (TypeError, ValueError):
                            pass
            wb_uf.close()
        except Exception:
            pass

        if uf_mes is None:
            return (f"Error: no se encontró UF para {año}-{mes:02d} en la hoja 'UF' del CDG. "
                    "Actualice la hoja UF primero.")

    # ── 5. Construir mapa código de cuenta → fila ─────────────────────────────
    account_row_map = _build_account_row_map(sheet_xml, ss_dict)

    # ── 6. Obtener estilo de celdas de datos de la columna anterior ───────────
    # Para usar el mismo estilo en celdas nuevas
    prev_col_letter = _col_letter(target_col - 1)
    default_style = "0"
    for code, row_num in list(account_row_map.items())[:3]:
        row_m2 = re.search(r'<row r="' + str(row_num) + r'"[^>]*>(.*?)</row>',
                           sheet_xml, re.DOTALL)
        if row_m2:
            s = _get_cell_style(row_m2.group(0),
                                f"{prev_col_letter}{row_num}", default_style)
            if s != "0":
                default_style = s
                break

    # ── 7. Escribir valores ────────────────────────────────────────────────────
    written = 0
    not_found_in_cdg = []

    for eeff_code, clp_val in eeff_values.items():
        # Matching: normalizar espacios
        eeff_norm = " ".join(eeff_code.strip().split())
        cdg_row = None
        for cdg_code, rn in account_row_map.items():
            if " ".join(cdg_code.strip().split()) == eeff_norm:
                cdg_row = rn
                break

        if cdg_row is None:
            not_found_in_cdg.append(eeff_code)
            continue

        val = clp_val / uf_mes if in_uf else clp_val
        sheet_xml = _set_numeric_cell(sheet_xml, cdg_row, target_col,
                                       val, default_style)
        written += 1

    # ── 8. Actualizar NOI-RCSD: agregar fórmulas para el mes nuevo ───────────
    with zipfile.ZipFile(cdg_path, "r") as z:
        noi_xml = z.read(_NOI_RCSD_XML).decode("utf-8")

    # Encontrar columna NOI para el mes (fila 7 = fecha global)
    noi_row7_m = re.search(r'<row r="' + str(_NOI_ROW7) + r'"[^>]*>(.*?)</row>',
                           noi_xml, re.DOTALL)
    noi_formulae_written = 0
    noi_col_letter = None
    if noi_row7_m:
        noi_col_n = _find_col_for_serial(noi_row7_m.group(0), target_serial)
        if noi_col_n is not None:
            noi_col_letter = _col_letter(noi_col_n)
            er_name = _ER_SHEET_NAME[mall]
            row_map = _NOI_ROW_MAP[mall]
            for noi_row, er_row in row_map.items():
                row_m = re.search(r'<row r="' + str(noi_row) + r'"[^>]*>(.*?)</row>',
                                  noi_xml, re.DOTALL)
                if not row_m:
                    continue
                row_txt = row_m.group(0)
                cell_ref = f"{noi_col_letter}{noi_row}"
                cs, ce = _find_cell_bounds(row_txt, cell_ref)
                if cs != -1 and '<f' in row_txt[cs:ce]:
                    continue  # ya tiene formula
                formula = f"+'{er_name}'!{col_letter}{er_row}"
                new_cell = f'<c r="{cell_ref}" s="61"><f>{formula}</f><v>0</v></c>'
                new_row = _replace_or_insert_cell(row_txt, cell_ref, new_cell)
                new_row = _update_row_spans(new_row, noi_col_n)
                # Reemplazar la fila en noi_xml
                row_start = noi_xml.find(row_m.group(0))
                noi_xml = (noi_xml[:row_start] + new_row
                           + noi_xml[row_start + len(row_m.group(0)):])
                noi_formulae_written += 1

    # ── 9. Guardar ambos sheets en un solo zip ─────────────────────────────────
    tmp_path = cdg_path + ".tmp"
    with zipfile.ZipFile(cdg_path, "r") as zin, \
         zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == sheet_xml_path:
                zout.writestr(item, sheet_xml.encode("utf-8"))
            elif item.filename == _NOI_RCSD_XML:
                zout.writestr(item, noi_xml.encode("utf-8"))
            else:
                zout.writestr(item, zin.read(item.filename))

    # Windows: si el archivo está abierto en Excel, guardar con sufijo
    saved_as = cdg_path
    try:
        os.replace(tmp_path, cdg_path)
    except (PermissionError, OSError):
        alt_path = cdg_path.replace(".xlsx", "_noi.xlsx")
        os.replace(tmp_path, alt_path)
        saved_as = alt_path

    mes_nombre = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
    }[mes]
    unidad = "UF" if in_uf else "CLP"
    lines = [
        f"ER {mall.capitalize()} actualizado — {mes_nombre} {año}",
        f"  Columna ER: {col_letter} (#{target_col})  |  Cuentas escritas: {written} (en {unidad})",
    ]
    if noi_col_letter:
        lines.append(f"  NOI-RCSD col {noi_col_letter}: {noi_formulae_written} formulas agregadas.")
    else:
        lines.append("  [!] No se encontro col {año}-{mes:02d} en NOI-RCSD fila 7.")
    lines.append(f"  Guardado en: {os.path.basename(saved_as)}"
                 + (" [!](archivo en uso)" if saved_as != cdg_path else ""))
    if not_found_in_cdg:
        lines.append(f"  [!]{len(not_found_in_cdg)} cuenta(s) del EEFF sin match en ER CDG:")
        for c in not_found_in_cdg[:10]:
            lines.append(f"    - {c}")
    if mall == "vina":
        lines.append("  Nota: ER Vina Section 2 se llena via formula — verificar NOI en Excel.")
    return "\n".join(lines)


# ── Herramientas públicas: ER Viña y ER Curicó ────────────────────────────────

def actualizar_er_vina(nombre_cdg: str, año: int, mes: int,
                       nombre_eeff: Optional[str] = None) -> str:
    """
    Lee el INFORME EEFF de Viña Centro y agrega la columna del mes indicado
    en la hoja 'ER Viña' del CDG. Valores en UF. NOI-RCSD se actualiza solo.
    Busca el EEFF en WORK_DIR y luego en SharePoint TresA/Viña Centro.
    """
    if nombre_eeff:
        path = (nombre_eeff if os.path.isabs(nombre_eeff)
                else os.path.join(WORK_DIR, nombre_eeff))
    else:
        path = _find_eeff_file("vina", año, mes)
    if not path:
        return (f"Error: no se encontró INFORME EEFF Viña Centro para {mes:02d}-{año}.\n"
                f"Cópielo a WORK_DIR ({WORK_DIR}) o a: {_TRES_A_DIRS['vina']}")
    return _actualizar_er_mall(nombre_cdg, path, "vina")


def actualizar_er_curico(nombre_cdg: str, año: int, mes: int,
                         nombre_eeff: Optional[str] = None) -> str:
    """
    Lee el INFORME EEFF de Curicó y agrega la columna del mes indicado
    en la hoja 'ER Curico' del CDG. Valores en CLP. NOI-RCSD se actualiza solo.
    Busca el EEFF en WORK_DIR y luego en SharePoint TresA/Curico.
    """
    if nombre_eeff:
        path = (nombre_eeff if os.path.isabs(nombre_eeff)
                else os.path.join(WORK_DIR, nombre_eeff))
    else:
        path = _find_eeff_file("curico", año, mes)
    if not path:
        return (f"Error: no se encontró INFORME EEFF Curicó para {mes:02d}-{año}.\n"
                f"Cópielo a WORK_DIR ({WORK_DIR}) o a: {_TRES_A_DIRS['curico']}")
    return _actualizar_er_mall(nombre_cdg, path, "curico")


# ── PT / Apoquindo / Apoquindo 3001 (desde hoja NOI PT del RR JLL) ───────────

def buscar_rr_jll(año: int, mes: int) -> str:
    """
    Busca el archivo Rent Roll JLL del mes indicado en SharePoint.
    Retorna la ruta absoluta o un mensaje de error.
    """
    aamm = f"{str(año)[2:]}{mes:02d}"
    sp_año_dir = os.path.join(_RR_JLL_BASE, str(año))
    search_dirs = []
    if os.path.isdir(sp_año_dir):
        search_dirs.append(sp_año_dir)
    search_dirs.append(WORK_DIR)

    for d in search_dirs:
        for pat in [f"{aamm} Rent Roll y NOI*.xlsx", f"{aamm}*Rent Roll*.xlsx"]:
            matches = glob.glob(os.path.join(d, pat))
            if matches:
                return matches[0]
    return f"Error: no se encontró RR JLL {aamm} en {sp_año_dir} ni en WORK_DIR."


def _actualizar_noi_desde_jll(
    nombre_cdg: str,
    nombre_rr_jll: str,
    año: int,
    mes: int,
    activo: str,
    fila_inicio: int,
    fila_fin: int,
) -> str:
    """
    Copia datos de la hoja 'NOI PT' del RR JLL a las filas indicadas del NOI-RCSD.
    nombre_rr_jll puede ser nombre de archivo (busca en WORK_DIR) o ruta absoluta.
    """
    cdg_path = os.path.join(WORK_DIR, nombre_cdg)
    jll_path = (nombre_rr_jll if os.path.isabs(nombre_rr_jll)
                else os.path.join(WORK_DIR, nombre_rr_jll))

    if not os.path.exists(cdg_path):
        return f"Error: '{nombre_cdg}' no encontrado en WORK_DIR."
    if not os.path.exists(jll_path):
        return f"Error: RR JLL no encontrado en {jll_path}."

    # ── 1. Leer RR JLL (openpyxl read_only, rápido) ───────────────────────────
    wb_jll = openpyxl.load_workbook(jll_path, read_only=True, data_only=True)
    if "NOI PT" not in wb_jll.sheetnames:
        wb_jll.close()
        return f"Error: no se encontró la hoja 'NOI PT' en {nombre_rr_jll}."

    ws_jll = wb_jll["NOI PT"]
    jll_rows = list(ws_jll.iter_rows(values_only=True))
    wb_jll.close()

    fecha_fin_mes = _ultimo_dia_mes(año, mes)
    target_col_jll = None
    header_row_jll = None

    for i, row in enumerate(jll_rows):
        for ci, v in enumerate(row):
            if isinstance(v, (datetime, date)):
                d = v.date() if isinstance(v, datetime) else v
                if d.year == año and d.month == mes:
                    target_col_jll = ci
                    header_row_jll = i
                    break
        if target_col_jll is not None:
            break

    if target_col_jll is None:
        return (f"Error: no se encontró la columna para {mes:02d}-{año} en 'NOI PT'. "
                "Verificar que el archivo RR JLL sea del mes correcto.")

    jll_data: dict = {}
    for row in jll_rows[header_row_jll + 1:]:
        label_raw = row[0] if row[0] is not None else (row[1] if len(row) > 1 else None)
        if label_raw is None:
            continue
        label = " ".join(str(label_raw).strip().split())
        if len(row) > target_col_jll and row[target_col_jll] is not None:
            try:
                jll_data[label] = float(row[target_col_jll])
            except (TypeError, ValueError):
                pass

    # ── 2. Leer XML del NOI-RCSD ──────────────────────────────────────────────
    with zipfile.ZipFile(cdg_path, "r") as z:
        sheet_xml = z.read(_NOI_RCSD_XML).decode("utf-8")
        ss_xml    = z.read(_SHARED_STRINGS).decode("utf-8")

    ss_dict = _read_shared_strings(ss_xml)

    # ── 3. Encontrar columna del mes en el encabezado del activo ─────────────
    enc_row = fila_inicio - 1
    target_serial = _excel_serial(fecha_fin_mes)

    enc_row_m = re.search(r'<row r="' + str(enc_row) + r'"[^>]*>(.*?)</row>',
                           sheet_xml, re.DOTALL)
    target_col_noi = None
    if enc_row_m:
        target_col_noi = _find_col_for_serial(enc_row_m.group(0), target_serial)
        if target_col_noi is None:
            last_col, _ = _last_date_col(enc_row_m.group(0))
            target_col_noi = (last_col or 6) + 1
            style_enc = _get_cell_style(enc_row_m.group(0),
                                        f"{_col_letter(target_col_noi-1)}{enc_row}", "0")
            sheet_xml = _set_numeric_cell(sheet_xml, enc_row, target_col_noi,
                                           target_serial, style_enc)
    else:
        target_col_noi = 7  # fallback

    col_letter = _col_letter(target_col_noi)

    # ── 4. Leer etiquetas (col C) de las filas del activo en NOI-RCSD ────────
    written = skipped = 0
    not_found = []

    for row_num in range(fila_inicio, fila_fin + 1):
        row_m = re.search(r'<row r="' + str(row_num) + r'"[^>]*>(.*?)</row>',
                          sheet_xml, re.DOTALL)
        if not row_m:
            continue
        row_xml = row_m.group(0)

        # Obtener etiqueta de col C (shared string o texto inline)
        label = None
        c_m = re.search(r'<c r="C' + str(row_num) + r'"[^>]*>.*?<v>(\d+)</v>', row_xml, re.DOTALL)
        if c_m:
            # Puede ser shared string (t="s") o numérico
            if 't="s"' in row_xml[row_m.start(): row_m.start() + c_m.end()]:
                label = ss_dict.get(int(c_m.group(1)), "")
            # Si no tiene t="s", es un número, no es etiqueta de texto

        if not label:
            continue
        label_norm = " ".join(label.strip().split())

        # Ver si la celda destino ya tiene fórmula
        cell_ref = f"{col_letter}{row_num}"
        cell_start, _ = _find_cell_bounds(row_xml, cell_ref)
        if cell_start != -1:
            snippet = row_xml[cell_start: cell_start + 300]
            if '<f' in snippet and '>' in snippet:
                skipped += 1
                continue

        jll_val = jll_data.get(label_norm)
        if jll_val is not None:
            default_style = _get_cell_style(row_xml,
                                            f"{_col_letter(target_col_noi-1)}{row_num}", "0")
            sheet_xml = _set_numeric_cell(sheet_xml, row_num, target_col_noi,
                                           jll_val, default_style)
            written += 1
        else:
            not_found.append(f"fila {row_num}: {label_norm!r}")

    # ── 5. Guardar ────────────────────────────────────────────────────────────
    tmp_path = cdg_path + ".tmp"
    with zipfile.ZipFile(cdg_path, "r") as zin, \
         zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == _NOI_RCSD_XML:
                zout.writestr(item, sheet_xml.encode("utf-8"))
            else:
                zout.writestr(item, zin.read(item.filename))
    saved_as = cdg_path
    try:
        os.replace(tmp_path, cdg_path)
    except (PermissionError, OSError):
        alt_path = cdg_path.replace(".xlsx", "_noi.xlsx")
        os.replace(tmp_path, alt_path)
        saved_as = alt_path

    lines = [
        f"NOI-RCSD actualizado — {activo} — {mes:02d}-{año}",
        f"  Filas {fila_inicio}-{fila_fin}, columna {col_letter}",
        f"  Valores escritos: {written}  |  Fórmulas respetadas: {skipped}",
        f"  Guardado en: {os.path.basename(saved_as)}"
        + ("  [!](archivo en uso — cierre Excel)" if saved_as != cdg_path else ""),
    ]
    if not_found:
        lines.append(f"  [!]{len(not_found)} etiqueta(s) sin match en JLL 'NOI PT':")
        for nf in not_found[:10]:
            lines.append(f"    - {nf}")
    return "\n".join(lines)


def actualizar_noi_pt(nombre_cdg: str, nombre_rr_jll: str,
                      año: int, mes: int) -> str:
    """Copia datos de 'NOI PT' del RR JLL a filas 335-379 del NOI-RCSD (Parque Titanium)."""
    return _actualizar_noi_desde_jll(
        nombre_cdg, nombre_rr_jll, año, mes,
        activo="Parque Titanium", fila_inicio=335, fila_fin=379,
    )


def actualizar_noi_apoquindo(nombre_cdg: str, nombre_rr_jll: str,
                             año: int, mes: int) -> str:
    """Copia datos de 'NOI PT' del RR JLL a filas 426-456 del NOI-RCSD (Fondo Apoquindo)."""
    return _actualizar_noi_desde_jll(
        nombre_cdg, nombre_rr_jll, año, mes,
        activo="Fondo Apoquindo", fila_inicio=426, fila_fin=456,
    )


def actualizar_noi_apo3001(nombre_cdg: str, nombre_rr_jll: str,
                           año: int, mes: int) -> str:
    """Copia datos de 'NOI PT' del RR JLL a filas 468-476 del NOI-RCSD (Apoquindo 3001)."""
    return _actualizar_noi_desde_jll(
        nombre_cdg, nombre_rr_jll, año, mes,
        activo="Apoquindo 3001", fila_inicio=468, fila_fin=476,
    )


# ── INMOSA (desde ER-FC INMOSA en SharePoint) ─────────────────────────────────

def buscar_er_inmosa(año: int, mes: int) -> str:
    """
    Busca el archivo ER-FC INMOSA más reciente del año en SharePoint.
    Los archivos están en EEFF Proveedores/Flujos INMOSA/{año}/ y se nombran
    'ER-FC INMOSA {año} {meses}.xlsx'. Cada mes se sube uno nuevo — el más
    reciente contiene los datos del mes indicado.
    Retorna la ruta absoluta o mensaje de error.
    """
    sp_año_dir = os.path.join(_INMOSA_BASE, str(año))
    search_dirs = [sp_año_dir, WORK_DIR] if os.path.isdir(sp_año_dir) else [WORK_DIR]

    for d in search_dirs:
        if not os.path.isdir(d):
            continue
        archivos = [f for f in os.listdir(d) if f.lower().endswith(".xlsx")
                    and "inmosa" in f.lower()]
        if archivos:
            # El más reciente por fecha de modificación
            return max((os.path.join(d, f) for f in archivos), key=os.path.getmtime)

    return f"Error: no se encontró ER-FC INMOSA para {año} en {sp_año_dir} ni WORK_DIR."


def actualizar_noi_inmosa(nombre_cdg: str, nombre_er_inmosa: str,
                          año: int, mes: int) -> str:
    """
    Copia los valores de INMOSA desde la planilla ER-FC INMOSA
    a las filas 287-295 del NOI-RCSD. Fórmulas se respetan.

    El archivo ER-FC INMOSA debe estar en WORK_DIR o indicar ruta completa.
    """
    cdg_path = os.path.join(WORK_DIR, nombre_cdg)
    er_path  = (nombre_er_inmosa if os.path.isabs(nombre_er_inmosa)
                else os.path.join(WORK_DIR, nombre_er_inmosa))

    if not os.path.exists(cdg_path):
        return f"Error: '{nombre_cdg}' no encontrado en WORK_DIR."
    if not os.path.exists(er_path):
        return f"Error: '{nombre_er_inmosa}' no encontrado. Cópielo a WORK_DIR primero."

    # ── 1. Leer ER-FC INMOSA ──────────────────────────────────────────────────
    wb_er = openpyxl.load_workbook(er_path, read_only=True, data_only=True)
    sheet_names = wb_er.sheetnames

    # Buscar hoja con datos: preferir NOI/ER/Estado/Result
    target_sheet = None
    for sheet in sheet_names:
        if any(k in sheet.upper() for k in ("NOI", "ESTADO", "RESULT")):
            target_sheet = sheet
            break
    if target_sheet is None:
        for sheet in sheet_names:
            if any(k in sheet.upper() for k in ("ER", "FLUJO")):
                target_sheet = sheet
                break
    if target_sheet is None and sheet_names:
        target_sheet = sheet_names[0]

    ws_er = wb_er[target_sheet]
    er_rows = list(ws_er.iter_rows(values_only=True))
    wb_er.close()

    fecha_fin_mes = _ultimo_dia_mes(año, mes)
    target_col_er = None
    for row in er_rows:
        for ci, v in enumerate(row):
            if isinstance(v, (datetime, date)):
                d = v.date() if isinstance(v, datetime) else v
                if d.year == año and d.month == mes:
                    target_col_er = ci
                    break
        if target_col_er is not None:
            break

    if target_col_er is None:
        return (f"Error: no se encontró columna para {mes:02d}-{año} en '{target_sheet}'. "
                f"Hojas disponibles: {sheet_names}")

    er_data: dict = {}
    for row in er_rows:
        label_raw = row[0] if row[0] is not None else (row[1] if len(row) > 1 else None)
        if label_raw is None:
            continue
        label = " ".join(str(label_raw).strip().split())
        if len(row) > target_col_er and row[target_col_er] is not None:
            try:
                er_data[label] = float(row[target_col_er])
            except (TypeError, ValueError):
                pass

    # ── 2. Leer XML del NOI-RCSD ──────────────────────────────────────────────
    with zipfile.ZipFile(cdg_path, "r") as z:
        sheet_xml = z.read(_NOI_RCSD_XML).decode("utf-8")
        ss_xml    = z.read(_SHARED_STRINGS).decode("utf-8")

    ss_dict = _read_shared_strings(ss_xml)
    target_serial = _excel_serial(fecha_fin_mes)

    # Encabezado INMOSA en fila 284
    enc_row = 284
    enc_row_m = re.search(r'<row r="' + str(enc_row) + r'"[^>]*>(.*?)</row>',
                           sheet_xml, re.DOTALL)
    target_col_noi = None
    if enc_row_m:
        target_col_noi = _find_col_for_serial(enc_row_m.group(0), target_serial)
        if target_col_noi is None:
            last_col, _ = _last_date_col(enc_row_m.group(0))
            target_col_noi = (last_col or 6) + 1
            sheet_xml = _set_numeric_cell(sheet_xml, enc_row, target_col_noi,
                                           target_serial, "0")
    else:
        target_col_noi = 7

    col_letter = _col_letter(target_col_noi)
    written = skipped = 0
    not_found = []

    for row_num in range(287, 296):
        row_m = re.search(r'<row r="' + str(row_num) + r'"[^>]*>(.*?)</row>',
                          sheet_xml, re.DOTALL)
        if not row_m:
            continue
        row_xml = row_m.group(0)

        label = None
        c_m = re.search(r'<c r="C' + str(row_num) + r'"[^>]*>.*?<v>(\d+)</v>', row_xml, re.DOTALL)
        if c_m and 't="s"' in row_xml:
            label = ss_dict.get(int(c_m.group(1)), "")
        if not label:
            continue
        label_norm = " ".join(label.strip().split())

        cell_ref = f"{col_letter}{row_num}"
        cell_start, _ = _find_cell_bounds(row_xml, cell_ref)
        if cell_start != -1 and '<f' in row_xml[cell_start: cell_start + 300]:
            skipped += 1
            continue

        val = er_data.get(label_norm)
        if val is not None:
            default_style = _get_cell_style(row_xml, f"{_col_letter(target_col_noi-1)}{row_num}", "0")
            sheet_xml = _set_numeric_cell(sheet_xml, row_num, target_col_noi, val, default_style)
            written += 1
        else:
            not_found.append(f"fila {row_num}: {label_norm!r}")

    tmp_path = cdg_path + ".tmp"
    with zipfile.ZipFile(cdg_path, "r") as zin, \
         zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == _NOI_RCSD_XML:
                zout.writestr(item, sheet_xml.encode("utf-8"))
            else:
                zout.writestr(item, zin.read(item.filename))
    saved_as = cdg_path
    try:
        os.replace(tmp_path, cdg_path)
    except (PermissionError, OSError):
        alt_path = cdg_path.replace(".xlsx", "_noi.xlsx")
        os.replace(tmp_path, alt_path)
        saved_as = alt_path

    lines = [
        f"NOI-RCSD INMOSA actualizado — {mes:02d}-{año}",
        f"  Columna {col_letter}, filas 287-295",
        f"  Valores escritos: {written}  |  Fórmulas respetadas: {skipped}",
        f"  Guardado en: {os.path.basename(saved_as)}"
        + ("  [!](archivo en uso — cierre Excel)" if saved_as != cdg_path else ""),
    ]
    if not_found:
        lines.append(f"  [!]{len(not_found)} etiqueta(s) sin match en ER-FC INMOSA:")
        for nf in not_found[:10]:
            lines.append(f"    - {nf}")
    return "\n".join(lines)


# ── Inspección de estructura (diagnóstico) ────────────────────────────────────

def inspeccionar_noi_rcsd(nombre_cdg: str, activo: str) -> str:
    """
    Muestra etiquetas y último valor para un activo del NOI-RCSD.
    activo: 'inmosa' | 'pt' | 'apoquindo' | 'apo3001'
    """
    rangos = {
        "inmosa":    (284, 300),
        "pt":        (334, 382),
        "apoquindo": (425, 458),
        "apo3001":   (467, 480),
    }
    cdg_path = os.path.join(WORK_DIR, nombre_cdg)
    if not os.path.exists(cdg_path):
        return f"Error: '{nombre_cdg}' no encontrado en WORK_DIR."

    fila_ini, fila_fin = rangos.get(activo.lower(), (280, 300))

    with zipfile.ZipFile(cdg_path, "r") as z:
        sheet_xml = z.read(_NOI_RCSD_XML).decode("utf-8")
        ss_xml    = z.read(_SHARED_STRINGS).decode("utf-8")

    ss_dict = _read_shared_strings(ss_xml)

    # Última columna con fecha en la fila de encabezado
    enc_row_m = re.search(r'<row r="' + str(fila_ini) + r'"[^>]*>(.*?)</row>',
                           sheet_xml, re.DOTALL)
    last_col, last_serial = (None, None)
    if enc_row_m:
        last_col, last_serial = _last_date_col(enc_row_m.group(0))

    from datetime import timedelta
    last_date = (date(1899, 12, 30) + timedelta(days=last_serial)
                 if last_serial else None)

    lines = [
        f"NOI-RCSD — {activo.upper()} (filas {fila_ini}-{fila_fin})",
        f"  Última fecha en encabezado: col {last_col} ({last_date})",
        "",
        "  Fila | Etiqueta (col C)                           | Último valor",
        "  " + "-" * 65,
    ]

    for row_num in range(fila_ini + 1, fila_fin + 1):
        row_m = re.search(r'<row r="' + str(row_num) + r'"[^>]*>(.*?)</row>',
                          sheet_xml, re.DOTALL)
        if not row_m:
            continue
        row_xml = row_m.group(0)

        c_m = re.search(r'<c r="C' + str(row_num) + r'"[^>]*>.*?<v>(\d+)</v>', row_xml, re.DOTALL)
        if not c_m:
            continue
        label = ss_dict.get(int(c_m.group(1)), "")
        if not label:
            continue

        val = None
        if last_col:
            cell_ref = f"{_col_letter(last_col)}{row_num}"
            v_m = re.search(r'<c r="' + cell_ref + r'"[^>]*>.*?<v>([^<]+)</v>', row_xml, re.DOTALL)
            if v_m:
                val = v_m.group(1)

        lines.append(f"  {row_num:4d} | {label[:45]:<45} | {val}")

    return "\n".join(lines)
