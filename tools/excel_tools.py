"""Herramientas para leer, validar y actualizar planillas Excel."""
import os
import openpyxl
import pandas as pd
from config import WORK_DIR


def _resolve(filepath: str) -> str:
    """Si la ruta no es absoluta, busca en el directorio de trabajo."""
    if not os.path.isabs(filepath):
        return os.path.join(WORK_DIR, filepath)
    return filepath


def read_excel_file(filepath: str, sheet_name: str = None) -> str:
    """Lee el contenido de un archivo Excel y lo muestra como tabla."""
    try:
        path = _resolve(filepath)
        if not os.path.exists(path):
            return f"Error: No se encontró '{path}'."

        if sheet_name:
            df = pd.read_excel(path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(path)

        # Lista de hojas disponibles
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()

        result = f"Archivo: {os.path.basename(path)}\n"
        result += f"Hojas disponibles: {', '.join(sheets)}\n"
        result += f"Leyendo hoja: {sheet_name or sheets[0]}\n"
        result += f"Filas: {len(df)}  |  Columnas: {len(df.columns)}\n"
        result += f"Columnas: {', '.join(str(c) for c in df.columns)}\n\n"

        preview = df.head(20)
        result += preview.to_string(index=True)

        if len(df) > 20:
            result += f"\n\n... y {len(df) - 20} fila(s) más."
        return result

    except Exception as e:
        return f"Error al leer Excel: {e}"


def validate_excel_file(filepath: str, required_columns: str = None) -> str:
    """Valida un archivo Excel detectando errores, celdas vacías y filas duplicadas."""
    try:
        path = _resolve(filepath)
        if not os.path.exists(path):
            return f"Error: No se encontró '{path}'."

        df = pd.read_excel(path)
        issues = []

        if required_columns:
            required = [c.strip() for c in required_columns.split(",")]
            missing = [c for c in required if c not in df.columns]
            if missing:
                issues.append(f"Columnas obligatorias faltantes: {', '.join(missing)}")

        null_counts = df.isnull().sum()
        for col, count in null_counts[null_counts > 0].items():
            pct = round(count / len(df) * 100, 1)
            issues.append(f"'{col}': {count} celda(s) vacía(s) ({pct}%)")

        duplicates = df.duplicated().sum()
        if duplicates > 0:
            issues.append(f"{duplicates} fila(s) completamente duplicada(s)")

        result = f"=== VALIDACIÓN: {os.path.basename(path)} ===\n"
        result += f"Filas: {len(df)}  |  Columnas: {len(df.columns)}\n"
        result += f"Columnas: {', '.join(str(c) for c in df.columns)}\n\n"

        if issues:
            result += f"⚠️  Se encontraron {len(issues)} problema(s):\n"
            for i, issue in enumerate(issues, 1):
                result += f"  {i}. {issue}\n"
        else:
            result += "✅ El archivo no presenta problemas detectados.\n"

        return result

    except Exception as e:
        return f"Error al validar Excel: {e}"


def update_excel_cell(filepath: str, sheet: str, cell: str, value: str) -> str:
    """Actualiza el valor de una celda en un archivo Excel."""
    try:
        path = _resolve(filepath)
        if not os.path.exists(path):
            return f"Error: No se encontró '{path}'."

        wb = openpyxl.load_workbook(path)
        if sheet not in wb.sheetnames:
            return f"Error: Hoja '{sheet}' no encontrada. Disponibles: {', '.join(wb.sheetnames)}"

        ws = wb[sheet]
        try:
            typed_value: int | float | str = int(value) if "." not in value else float(value)
        except (ValueError, TypeError):
            typed_value = value

        ws[cell] = typed_value
        wb.save(path)
        return f"✅ Celda {cell} (hoja '{sheet}') actualizada a: {value}"

    except Exception as e:
        return f"Error al actualizar celda: {e}"


def list_work_files() -> str:
    """Lista los archivos Excel disponibles en el directorio de trabajo actual."""
    try:
        os.makedirs(WORK_DIR, exist_ok=True)
        files = [
            f for f in os.listdir(WORK_DIR)
            if f.lower().endswith((".xlsx", ".xls"))
        ]

        if not files:
            return f"No hay archivos Excel en el directorio de trabajo ({WORK_DIR})."

        result = f"Archivos Excel en directorio de trabajo ({WORK_DIR}):\n\n"
        for fname in files:
            fpath = os.path.join(WORK_DIR, fname)
            size = os.path.getsize(fpath)
            result += f"  📊 {fname}  ({size:,} bytes)\n"
        return result

    except Exception as e:
        return f"Error al listar archivos de trabajo: {e}"
