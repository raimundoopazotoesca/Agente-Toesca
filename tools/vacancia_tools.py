"""
Herramientas para actualizar la hoja Vacancia y refrescar Tabla Rentas 2 del CDG.

Mapping m2 vacantes Vacancia (filas 47-58) → Resumen:
  47 INMOSA              G34  (Residencias Adulto Mayor)
  48 Machalí             0 fijo (activo sin portfolio)
  49 SUCDEN              H34  (Bodegas Sucden)
  50 PT Oficinas         C96  (Torre A S.A. Oficinas)
  51 PT Locales          H96  (Inmob. Boulevard PT Locales)
  52 PT Bodegas          I96  (Inmob. Boulevard PT Bodegas)
  53 Viña Centro         C34
  54 Apoquindo 4700      J50  (Apo 4700 Total)
  55 Apoquindo 4501      E50  (Apo 4501 Total)
  56 Fondo Apoquindo     E50 + J50  (suma de los dos anteriores)
  57 Curicó              E1   (real, no ponderado — no usar E34)
  58 Apoquindo 3001      I34
"""

import os
import sys
from datetime import datetime

import openpyxl

from config import WORK_DIR


def _leer_valor(ws, row: int, col: int) -> float:
    """Lee celda; retorna 0 si es None o '-'."""
    v = ws.cell(row=row, column=col).value
    if v is None or v == "-":
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def actualizar_vacancia(nombre_cdg: str, año: int, mes: int) -> str:
    """
    Lee m2 vacantes desde la hoja Resumen del CDG y los escribe en la columna
    correspondiente al período año/mes en la hoja Vacancia (filas 47-58).

    Los dates están en la fila 46 de Vacancia. Se busca la columna cuya fecha
    coincida con el primer día del mes indicado.

    Parámetros:
        nombre_cdg : nombre del archivo CDG en WORK_DIR (ej: '2603 CDG.xlsx')
        año        : año del período (ej: 2026)
        mes        : mes del período (ej: 3)
    """
    cdg_path = os.path.join(WORK_DIR, nombre_cdg)
    if not os.path.exists(cdg_path):
        return f"Error: no se encontró '{nombre_cdg}' en WORK_DIR ({WORK_DIR})"

    wb = openpyxl.load_workbook(cdg_path)

    if "Resumen" not in wb.sheetnames:
        wb.close()
        return "Error: no se encontró la hoja 'Resumen' en el CDG."
    if "Vacancia" not in wb.sheetnames:
        wb.close()
        return "Error: no se encontró la hoja 'Vacancia' en el CDG."

    ws_res = wb["Resumen"]
    ws_vac = wb["Vacancia"]

    # ── 1. Leer valores desde Resumen ────────────────────────────────────────
    curico    = _leer_valor(ws_res, 1, 5)    # E1  — Curicó real (no ponderado)
    viña      = _leer_valor(ws_res, 34, 3)   # C34 — Viña Centro
    inmosa    = _leer_valor(ws_res, 34, 7)   # G34 — INMOSA = Residencias AM
    sucden    = _leer_valor(ws_res, 34, 8)   # H34 — Bodegas Sucden
    apo3001   = _leer_valor(ws_res, 34, 9)   # I34 — Apoquindo 3001
    apo4501_t = _leer_valor(ws_res, 50, 5)   # E50 — Apoquindo 4501 Total
    apo4700_t = _leer_valor(ws_res, 50, 10)  # J50 — Apoquindo 4700 Total
    fondo_apo = apo4501_t + apo4700_t        # Fondo Apoquindo = suma
    pt_ofic   = _leer_valor(ws_res, 96, 3)   # C96 — Torre A Oficinas (PT Ofic)
    pt_locs   = _leer_valor(ws_res, 96, 8)   # H96 — Boulevard PT Locales
    pt_bod    = _leer_valor(ws_res, 96, 9)   # I96 — Boulevard PT Bodegas

    # ── 2. Encontrar columna destino en Vacancia (fila 46 = header de fechas) ─
    target_date = datetime(año, mes, 1)
    target_col = None

    for col in range(1, ws_vac.max_column + 2):
        v = ws_vac.cell(row=46, column=col).value
        if v is None:
            continue
        if isinstance(v, datetime) and v.year == año and v.month == mes:
            target_col = col
            break
        if hasattr(v, "year") and v.year == año and v.month == mes:
            target_col = col
            break

    if target_col is None:
        wb.close()
        return (
            f"Error: no se encontró la columna para {año}-{mes:02d} en la fila 46 de Vacancia. "
            f"Verificar que la fecha {año}-{mes:02d}-01 esté en el encabezado."
        )

    # ── 3. Escribir valores en filas 47-58 ───────────────────────────────────
    asset_values = {
        47: ("INMOSA",            inmosa),
        48: ("Machalí",           0),         # activo sin portfolio → siempre 0
        49: ("SUCDEN",            sucden),
        50: ("PT Oficinas",       pt_ofic),
        51: ("PT Locales",        pt_locs),
        52: ("PT Bodegas",        pt_bod),
        53: ("Viña Centro",       viña),
        54: ("Apoquindo 4700",    apo4700_t),
        55: ("Apoquindo 4501",    apo4501_t),
        56: ("Fondo Apoquindo",   fondo_apo),
        57: ("Curicó",            curico),
        58: ("Apoquindo 3001",    apo3001),
    }

    for row, (name, val) in asset_values.items():
        ws_vac.cell(row=row, column=target_col).value = val

    # ── 4. Guardar ────────────────────────────────────────────────────────────
    wb.save(cdg_path)
    wb.close()

    lines = [
        f"Vacancia actualizada — {año}-{mes:02d}",
        f"  Columna destino: {target_col}",
        "",
        "  Activo                  Valor",
        "  " + "-" * 35,
    ]
    for row, (name, val) in asset_values.items():
        lines.append(f"  {name:<22}  {val}")

    lines.append("")
    lines.append("Recuerda actualizar la tabla dinámica 'Tabla Rentas 2' con 'refrescar_tabla_rentas_2'.")
    return "\n".join(lines)


def refrescar_tabla_rentas_2(nombre_cdg: str) -> str:
    """
    Refresca la tabla dinámica en la hoja 'Tabla Rentas 2' del CDG via COM (solo Windows).
    Esto actualiza los datos que usa la hoja Facts Sheet.

    Parámetros:
        nombre_cdg : nombre del archivo CDG en WORK_DIR
    """
    if sys.platform != "win32":
        return "Esta función solo está disponible en Windows (requiere Excel Desktop via COM)."

    cdg_path = os.path.join(WORK_DIR, nombre_cdg)
    if not os.path.exists(cdg_path):
        return f"Error: no se encontró '{nombre_cdg}' en WORK_DIR ({WORK_DIR})"

    try:
        import win32com.client

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(os.path.abspath(cdg_path))
        try:
            ws = wb.Sheets("Tabla Rentas 2")
            count = 0
            for pt in ws.PivotTables():
                pt.RefreshTable()
                count += 1
            wb.Save()
        finally:
            wb.Close(SaveChanges=False)
            excel.Quit()

        return f"Tabla Rentas 2 refrescada ({count} tabla(s) dinámica(s) actualizadas)."

    except Exception as e:
        return f"Error al refrescar Tabla Rentas 2: {e}"
