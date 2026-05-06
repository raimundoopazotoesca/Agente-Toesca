"""
Herramientas para el flujo mensual del Control de Gestión Renta Comercial.
Manipulación directa del XML del xlsx (sin cargar el workbook completo en openpyxl)
para máxima velocidad con archivos de 14MB+ y 87 hojas.
"""
import glob
import os
import re
import shutil
import tempfile
import zipfile
from calendar import monthrange
from datetime import date, datetime, timedelta
from config import SHAREPOINT_DIR, WORK_DIR
import openpyxl


RUTA_COMERCIAL = os.path.join(
    SHAREPOINT_DIR or "",
    "Controles de Gestión", "Renta Comercial", "Controles de Gestión",
)


# ─── Config fija por hoja (derivada del análisis del xlsx 2603) ───────────────
# Verificar en CONTEXT.txt si se agregan/renombran sheets en versiones futuras.
SHEET_CFG = {
    "A&R Apoquindo": {
        "sheet_file": "xl/worksheets/sheet15.xml",
        "table_file": "xl/tables/table2.xml",
        "tabla":      "Tabla133",
        "date_col":   "Fecha",
        "series":     [None],
        "cuotas":     {None: 1585000},
        "has_bursatil": False,
    },
    "A&R PT": {
        "sheet_file": "xl/worksheets/sheet16.xml",
        "table_file": "xl/tables/table3.xml",
        "tabla":      "Tabla13",
        "date_col":   "SF",
        "series":     [None],
        "cuotas":     {None: 1640000},
        "has_bursatil": True,
        "nemotecnico": "CFITRIPT-E",
    },
    "A&R Rentas": {
        "sheet_file": "xl/worksheets/sheet17.xml",
        "table_file": "xl/tables/table4.xml",
        "tabla":      "Tabla1",
        "date_col":   "Fecha",
        "series":     ["A", "C", "I"],
        "cuotas":     {"A": 475667, "C": 1252928, "I": 1091101},
        "has_bursatil": True,
        "nemotecnicos": {"A": "CFITOERI1A", "C": "CFITOERI1C", "I": "CFITOERI1I"},
    },
}

# Hoja Pendientes
PENDIENTES_SHEET = "xl/worksheets/sheet3.xml"


# ─── Utilidades ───────────────────────────────────────────────────────────────

def _excel_date(d) -> int:
    if isinstance(d, datetime):
        d = d.date()
    return (d - date(1899, 12, 30)).days


def _from_excel_date(serial: int) -> date:
    return date(1899, 12, 30) + timedelta(days=serial)


def _last_day(year: int, month: int) -> date:
    return date(year, month, monthrange(year, month)[1])


def _xml_escape(s: str) -> str:
    return (s.replace("&", "&amp;")
             .replace("<", "&lt;")
             .replace(">", "&gt;")
             .replace("'", "&apos;"))


def _col_num(letter: str) -> int:
    n = 0
    for c in letter.upper():
        n = n * 26 + ord(c) - ord("A") + 1
    return n


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _find_or_add_shared_string(ss_xml: str, text: str) -> tuple:
    """Returns (index, updated_ss_xml)."""
    sis = list(re.finditer(r"<si>(.*?)</si>", ss_xml, re.DOTALL))
    for i, m in enumerate(sis):
        t = re.search(r"<t[^>]*>([^<]*)</t>", m.group(1))
        if t and t.group(1) == text:
            return i, ss_xml
    # Add new entry
    idx = len(sis)
    new_si = f'<si><t xml:space="preserve">{_xml_escape(text)}</t></si>'
    ss_xml = ss_xml.replace("</sst>", new_si + "</sst>")
    # Increment count and uniqueCount
    def _inc(m_):
        return re.sub(
            r'(count|uniqueCount)="(\d+)"',
            lambda x: f'{x.group(1)}="{int(x.group(2)) + 1}"',
            m_.group(0),
        )
    ss_xml = re.sub(r"<sst\b[^>]*>", _inc, ss_xml, count=1)
    return idx, ss_xml


def _get_table_ref(table_xml: str) -> str:
    m = re.search(r'\bref="([^"]+)"', table_xml)
    return m.group(1) if m else ""


def _update_table_ref(table_xml: str, new_ref: str) -> str:
    return re.sub(r'\bref="[^"]+"', f'ref="{new_ref}"', table_xml, count=1)


def _table_ref_last_row(table_ref: str) -> int:
    m = re.search(r"[A-Z]+(\d+):[A-Z]+(\d+)", table_ref)
    return int(m.group(2)) if m else 0


def _table_ref_first_data_row(table_ref: str) -> int:
    m = re.search(r"[A-Z]+(\d+)", table_ref)
    return int(m.group(1)) + 1 if m else 0  # +1 para saltar header


def _cell_has_value(sheet_xml: str, ref: str):
    """
    Verifica si una celda tiene valor.
    Retorna: True (tiene <v>), False (self-closing o sin <v>), None (no existe en XML).
    """
    start = sheet_xml.find(f'<c r="{ref}"')
    if start == -1:
        return None
    # Escanear hacia adelante para encontrar /> o >
    i = start + len(f'<c r="{ref}"')
    limit = min(i + 300, len(sheet_xml))
    while i < limit:
        if sheet_xml[i:i+2] == '/>':
            return False  # Self-closing = sin valor
        if sheet_xml[i] == '>':
            # Tiene apertura → buscar </c>
            end = sheet_xml.find('</c>', i)
            if end == -1:
                return False
            return '<v>' in sheet_xml[i:end]
        i += 1
    return None


def _find_first_empty_date_row(sheet_xml: str, table_ref: str) -> tuple:
    """
    Encuentra la primera fila dentro de la tabla donde la celda D no tiene valor.
    Returns (first_empty_row, last_data_row, last_c_value)
    """
    first_data = _table_ref_first_data_row(table_ref)
    last_row = _table_ref_last_row(table_ref)

    last_data_row = first_data - 1
    last_c_value = 0

    for row_num in range(first_data, last_row + 1):
        status = _cell_has_value(sheet_xml, f"D{row_num}")
        if status is True:
            last_data_row = row_num
            c_m = re.search(rf'<c r="C{row_num}"[^>]*>.*?<v>(\d+)</v>', sheet_xml, re.DOTALL)
            if c_m:
                last_c_value = int(c_m.group(1))
        else:
            # False (self-closing) o None (ausente) = fila vacía
            return row_num, last_data_row, last_c_value

    return -1, last_data_row, last_c_value  # tabla llena


def _find_cell_bounds(row_xml: str, ref: str) -> tuple:
    """
    Retorna (start, end) del span completo de la celda en row_xml.
    Maneja correctamente tanto self-closing (<c .../>) como con contenido (<c ...>...</c>).
    Retorna (-1, -1) si no se encuentra.
    """
    tag = f'<c r="{ref}"'
    start = row_xml.find(tag)
    if start == -1:
        return -1, -1
    i = start + len(tag)
    while i < len(row_xml):
        if row_xml[i:i+2] == '/>':
            return start, i + 2
        if row_xml[i] == '>':
            end = row_xml.find('</c>', i)
            return start, (end + 4) if end != -1 else i + 1
        i += 1
    return start, len(row_xml)


def _get_cell_style(row_xml: str, ref: str, default: str) -> str:
    """Extrae el atributo s= de una celda."""
    tag = f'<c r="{ref}"'
    start = row_xml.find(tag)
    if start == -1:
        return default
    end = min(start + 200, len(row_xml))
    snippet = row_xml[start:end]
    m = re.search(r'\bs="(\d+)"', snippet)
    return m.group(1) if m else default


def _replace_or_insert_cell(row_xml: str, ref: str, new_cell: str) -> str:
    """Reemplaza la celda si existe (vacía o con contenido), o la inserta en orden."""
    start, end = _find_cell_bounds(row_xml, ref)
    if start != -1:
        return row_xml[:start] + new_cell + row_xml[end:]
    # Insertar en orden de columna
    col = re.sub(r"\d", "", ref)
    col_n = _col_num(col)
    for m in re.finditer(r'<c r="([A-Z]+)\d+"', row_xml):
        if _col_num(m.group(1)) > col_n:
            return row_xml[: m.start()] + new_cell + row_xml[m.start():]
    return row_xml + new_cell


def _fill_row(sheet_xml: str, ss_xml: str, row_num: int,
              tabla: str, date_col: str,
              date_serial: int, detalle: str, serie,
              precio: float, cuotas: int, prev_c_value: int) -> tuple:
    """
    Rellena una fila pre-asignada con datos.
    Retorna (sheet_xml_actualizado, ss_xml_actualizado).
    """
    d = _from_excel_date(date_serial)
    year, month = d.year, d.month

    # Obtener/agregar strings compartidos
    detalle_idx, ss_xml = _find_or_add_shared_string(ss_xml, detalle)

    # Buscar la fila en el XML
    row_pat = rf'(<row\b[^>]*r="{row_num}"[^>]*>)(.*?)(</row>)'
    row_m = re.search(row_pat, sheet_xml, re.DOTALL)

    if row_m:
        row_open = row_m.group(1)
        row_content = row_m.group(2)
        row_close = row_m.group(3)
    else:
        # Fila no existe, crearla nueva
        row_open = f'<row r="{row_num}" spans="1:25">'
        row_content = ""
        row_close = "</row>"

    rc = row_content  # alias

    # ── A: YEAR ───────────────────────────────────────────────────────────────
    a_s = _get_cell_style(rc, f"A{row_num}", "106")
    a = f'<c r="A{row_num}" s="{a_s}"><f>+YEAR({tabla}[[#This Row],[{date_col}]])</f><v>{year}</v></c>'
    rc = _replace_or_insert_cell(rc, f"A{row_num}", a)

    # ── B: MONTH ──────────────────────────────────────────────────────────────
    b_s = _get_cell_style(rc, f"B{row_num}", "106")
    b = f'<c r="B{row_num}" s="{b_s}"><f>+MONTH({tabla}[[#This Row],[{date_col}]])</f><v>{month}</v></c>'
    rc = _replace_or_insert_cell(rc, f"B{row_num}", b)

    # ── C: ID (solo si no tiene cached value ya correcto) ────────────────────
    c_existing = _cell_has_value(rc, f"C{row_num}")
    if not c_existing:
        c_s = _get_cell_style(rc, f"C{row_num}", "113")
        new_id = prev_c_value + 1
        c = f'<c r="C{row_num}" s="{c_s}"><f>+C{row_num - 1}+1</f><v>{new_id}</v></c>'
        rc = _replace_or_insert_cell(rc, f"C{row_num}", c)

    # ── D: Fecha ──────────────────────────────────────────────────────────────
    d_s = _get_cell_style(rc, f"D{row_num}", "1622")
    date_cell = f'<c r="D{row_num}" s="{d_s}"><v>{date_serial}</v></c>'
    rc = _replace_or_insert_cell(rc, f"D{row_num}", date_cell)

    # ── E: Detalle ────────────────────────────────────────────────────────────
    e_s = _get_cell_style(rc, f"E{row_num}", "133")
    e = f'<c r="E{row_num}" s="{e_s}" t="s"><v>{detalle_idx}</v></c>'
    rc = _replace_or_insert_cell(rc, f"E{row_num}", e)

    # ── F: Serie ──────────────────────────────────────────────────────────────
    if serie is not None:
        serie_idx, ss_xml = _find_or_add_shared_string(ss_xml, str(serie))
        f_s = _get_cell_style(rc, f"F{row_num}", "133")
        f_cell = f'<c r="F{row_num}" s="{f_s}" t="s"><v>{serie_idx}</v></c>'
        rc = _replace_or_insert_cell(rc, f"F{row_num}", f_cell)

    # ── G: Tipo (ya tiene fórmula en filas pre-asignadas, asegurar) ───────────
    g_existing = re.search(rf'<c r="G{row_num}"[^>]*>.*?<f>', rc, re.DOTALL)
    if not g_existing:
        g_s = _get_cell_style(rc, f"G{row_num}", "134")
        g = (f'<c r="G{row_num}" s="{g_s}" t="str">'
             f'<f>+IF(E{row_num}="Aporte",E{row_num},"Reparto")</f>'
             f'<v>Reparto</v></c>')
        rc = _replace_or_insert_cell(rc, f"G{row_num}", g)

    # ── H: Monto$ = precio * cuotas ──────────────────────────────────────────
    h_start, h_end = _find_cell_bounds(rc, f"H{row_num}")
    h_existing = h_start != -1 and '<f>' in rc[h_start:h_end]
    if not h_existing:
        h_s = _get_cell_style(rc, f"H{row_num}", "2551")
        h = (f'<c r="H{row_num}" s="{h_s}">'
             f'<f>+{tabla}[[#This Row],[Monto $ / cuota]]*{tabla}[[#This Row],[Cuotas]]</f>'
             f'<v>0</v></c>')
        rc = _replace_or_insert_cell(rc, f"H{row_num}", h)

    # ── I: Precio por cuota ───────────────────────────────────────────────────
    i_s = _get_cell_style(rc, f"I{row_num}", "1624")
    i = f'<c r="I{row_num}" s="{i_s}"><v>{precio}</v></c>'
    rc = _replace_or_insert_cell(rc, f"I{row_num}", i)

    # ── J: Cuotas ─────────────────────────────────────────────────────────────
    j_s = _get_cell_style(rc, f"J{row_num}", "1625")
    j = f'<c r="J{row_num}" s="{j_s}"><v>{cuotas}</v></c>'
    rc = _replace_or_insert_cell(rc, f"J{row_num}", j)

    # ── K: UF (VLOOKUP) ───────────────────────────────────────────────────────
    k_s0, k_e0 = _find_cell_bounds(rc, f"K{row_num}")
    if k_s0 == -1 or '<f>' not in rc[k_s0:k_e0]:
        k_s = _get_cell_style(rc, f"K{row_num}", "134")
        k = (f'<c r="K{row_num}" s="{k_s}">'
             f"<f>+VLOOKUP({tabla}[[#This Row],[{date_col}]],'UF DIARIA'!$A$2:$E$1048576,5,FALSE)</f>"
             f'<v>0</v></c>')
        rc = _replace_or_insert_cell(rc, f"K{row_num}", k)

    # ── L: Monto UF = H/K ─────────────────────────────────────────────────────
    l_s0, l_e0 = _find_cell_bounds(rc, f"L{row_num}")
    if l_s0 == -1 or '<f>' not in rc[l_s0:l_e0]:
        l_s = _get_cell_style(rc, f"L{row_num}", "115")
        l = f'<c r="L{row_num}" s="{l_s}"><f>+H{row_num}/K{row_num}</f><v>0</v></c>'
        rc = _replace_or_insert_cell(rc, f"L{row_num}", l)

    # ── M: Monto UF/cuota ─────────────────────────────────────────────────────
    m_s0, m_e0 = _find_cell_bounds(rc, f"M{row_num}")
    if m_s0 == -1 or '<f>' not in rc[m_s0:m_e0]:
        m_s = _get_cell_style(rc, f"M{row_num}", "116")
        m_c = (f'<c r="M{row_num}" s="{m_s}">'
               f'<f>+{tabla}[[#This Row],[Monto UF]]/{tabla}[[#This Row],[Cuotas]]</f>'
               f'<v>0</v></c>')
        rc = _replace_or_insert_cell(rc, f"M{row_num}", m_c)

    new_row = row_open + rc + row_close

    if row_m:
        sheet_xml = sheet_xml[: row_m.start()] + new_row + sheet_xml[row_m.end():]
    else:
        # Insertar en el lugar correcto
        next_row_m = re.search(
            rf'<row\b[^>]*r="({row_num + 1}|{row_num + 2}|{row_num + 3})"', sheet_xml
        )
        if next_row_m:
            sheet_xml = sheet_xml[: next_row_m.start()] + new_row + sheet_xml[next_row_m.start():]
        else:
            sheet_xml = sheet_xml.replace("</sheetData>", new_row + "</sheetData>")

    return sheet_xml, ss_xml


def _expand_table_if_needed(table_xml: str, row_num: int) -> str:
    """Expande el ref de la tabla si row_num está fuera del rango actual."""
    current_ref = _get_table_ref(table_xml)
    last = _table_ref_last_row(current_ref)
    if row_num > last:
        col_start = re.search(r"([A-Z]+)\d+:[A-Z]+\d+", current_ref).group(1)
        col_end_m = re.search(r"[A-Z]+\d+:([A-Z]+)\d+", current_ref)
        col_end = col_end_m.group(1) if col_end_m else "Y"
        first = re.search(r"[A-Z]+(\d+)", current_ref).group(1)
        new_ref = f"{col_start}{first}:{col_end}{row_num}"
        table_xml = _update_table_ref(table_xml, new_ref)
    return table_xml


def _apply_to_xlsx(filepath: str, modifications: dict) -> None:
    """Aplica modificaciones al xlsx (edita solo los archivos especificados)."""
    tmp = filepath + ".tmp"
    with zipfile.ZipFile(filepath, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = modifications.get(item.filename)
                if data is not None:
                    zout.writestr(item, data if isinstance(data, bytes) else data.encode("utf-8"))
                else:
                    zout.writestr(item, zin.read(item.filename))
    os.replace(tmp, filepath)


def _resolve_path(nombre_archivo: str) -> str:
    if os.path.isabs(nombre_archivo):
        return nombre_archivo
    return os.path.join(WORK_DIR, nombre_archivo)


# ─── Herramientas públicas ────────────────────────────────────────────────────

def crear_planilla_mes(mes_code_nuevo: str) -> str:
    """
    Crea la planilla del mes siguiente copiando la del mes anterior.
    mes_code_nuevo: código AAMM como '2604'.
    Busca el archivo más reciente en la carpeta del año correspondiente.
    Retorna la ruta del archivo creado.
    """
    try:
        año = "20" + mes_code_nuevo[:2]
        carpeta_año = os.path.join(RUTA_COMERCIAL, año)

        if not os.path.exists(carpeta_año):
            return f"Error: carpeta {carpeta_año} no existe."

        archivos = [f for f in os.listdir(carpeta_año) if f.endswith(".xlsx")]
        if not archivos:
            # Intentar año anterior
            año_ant = str(int(año) - 1)
            carpeta_año_ant = os.path.join(RUTA_COMERCIAL, año_ant)
            archivos_ant = [f for f in os.listdir(carpeta_año_ant) if f.endswith(".xlsx")]
            if not archivos_ant:
                return f"Error: no hay archivos .xlsx en {carpeta_año} ni {carpeta_año_ant}."
            archivos_ant.sort()
            origen = os.path.join(carpeta_año_ant, archivos_ant[-1])
        else:
            archivos.sort()
            # Tomar el más reciente anterior al mes nuevo
            candidatos = [f for f in archivos if f[:4] < mes_code_nuevo]
            if not candidatos:
                return f"Error: no hay archivo anterior a {mes_code_nuevo} en {carpeta_año}."
            origen = os.path.join(carpeta_año, candidatos[-1])

        nombre_nuevo = f"{mes_code_nuevo} Control De Gestión Renta Comercial vAgente.xlsx"
        destino = os.path.join(carpeta_año, nombre_nuevo)

        if os.path.exists(destino):
            return f"Ya existe: {destino}"

        shutil.copy2(origen, destino)
        return f"Planilla creada: {destino}\n(Copia de: {os.path.basename(origen)})"

    except Exception as e:
        return f"Error al crear planilla: {e}"


def guardar_cdg(nombre_archivo: str) -> str:
    """
    Guarda el CDG editado (debe ser un archivo vAgente) de vuelta en SharePoint,
    en la misma carpeta donde van los CDG del año (SharePoint/Controles de Gestión/Renta Comercial/Controles de Gestión/{YYYY}/).

    PROHIBIDO editar/guardar archivos vF o vActualizar — solo vAgente.
    """
    src = _resolve_path(nombre_archivo)
    base = os.path.basename(src)

    if "vagente" not in base.lower():
        return (
            f"Error: solo puedes guardar archivos 'vAgente'. "
            f"'{base}' parece ser una versión que no debes modificar."
        )
    if not os.path.exists(src):
        return f"Error: '{src}' no encontrado."

    # Determinar año desde el código AAMM al inicio del nombre
    import re as _re
    m = _re.match(r"(\d{2})(\d{2})", base)
    año = ("20" + m.group(1)) if m else str(date.today().year)

    carpeta_destino = os.path.join(RUTA_COMERCIAL, año)
    if not os.path.isdir(carpeta_destino):
        return f"Error: carpeta destino no existe: {carpeta_destino}"

    destino = os.path.join(carpeta_destino, base)
    shutil.copy2(src, destino)
    return f"OK: '{base}' guardado en {carpeta_destino}"


def actualizar_fecha_pendientes(nombre_archivo: str, año: int, mes: int) -> str:
    """
    Actualiza la fecha en B2 de la hoja Pendientes al primer día del mes indicado.
    Trabaja sobre el archivo en WORK_DIR (o ruta absoluta).
    """
    filepath = _resolve_path(nombre_archivo)
    if not os.path.exists(filepath):
        return f"Error: no se encontró '{filepath}'."
    try:
        nueva_fecha = _excel_date(date(año, mes, 1))

        with zipfile.ZipFile(filepath, "r") as z:
            sheet_xml = z.read(PENDIENTES_SHEET).decode("utf-8")

        # Reemplazar el valor en B2
        b2_pat = r'(<c r="B2"[^>]*>).*?(<v>)(\d+)(</v>)'
        if re.search(b2_pat, sheet_xml, re.DOTALL):
            sheet_xml = re.sub(
                b2_pat,
                lambda m: m.group(1) + m.group(2) + str(nueva_fecha) + m.group(4),
                sheet_xml,
                flags=re.DOTALL,
            )
        else:
            # Buscar celda self-closing y reemplazar con valor
            b2_empty = re.search(r'<c r="B2"([^>]*)/?>', sheet_xml)
            if b2_empty:
                style = b2_empty.group(1)
                new_cell = f'<c r="B2"{style}><v>{nueva_fecha}</v></c>'
                sheet_xml = sheet_xml[: b2_empty.start()] + new_cell + sheet_xml[b2_empty.end():]
            else:
                return "Error: no se encontró celda B2 en hoja Pendientes."

        _apply_to_xlsx(filepath, {PENDIENTES_SHEET: sheet_xml})
        return f"Fecha Pendientes actualizada a {date(año, mes, 1).strftime('%d/%m/%Y')} en '{nombre_archivo}'."
    except Exception as e:
        return f"Error al actualizar fecha Pendientes: {e}"


def _agregar_vr(nombre_archivo: str, sheet_key: str, año: int, mes: int,
                detalle: str, precios: dict) -> str:
    """
    Función genérica interna para agregar filas VR Bursátil o VR Contable.
    precios: {serie_o_None: precio_cuota}
    """
    filepath = _resolve_path(nombre_archivo)
    if not os.path.exists(filepath):
        return f"Error: no se encontró '{filepath}'."

    cfg = SHEET_CFG[sheet_key]
    sheet_file = cfg["sheet_file"]
    table_file = cfg["table_file"]
    tabla = cfg["tabla"]
    date_col = cfg["date_col"]
    series = cfg["series"]

    fecha = _last_day(año, mes)
    date_serial = _excel_date(fecha)

    try:
        with zipfile.ZipFile(filepath, "r") as z:
            sheet_xml = z.read(sheet_file).decode("utf-8")
            table_xml = z.read(table_file).decode("utf-8")
            ss_xml = z.read("xl/sharedStrings.xml").decode("utf-8")

        table_ref = _get_table_ref(table_xml)
        mods = {}
        resultados = []

        for serie in series:
            precio = precios.get(serie)
            if precio is None:
                resultados.append(f"  [!] Precio no proporcionado para serie {serie}, fila omitida.")
                continue

            cuotas = cfg["cuotas"].get(serie) or cfg["cuotas"].get(None, 0)

            row_num, last_data_row, last_c = _find_first_empty_date_row(sheet_xml, table_ref)
            if row_num == -1:
                resultados.append(f"  [!] Tabla llena en {sheet_key}, expandiendo...")
                last_row = _table_ref_last_row(table_ref)
                row_num = last_row + 1
                # Actualizar ref de tabla
                table_xml = _expand_table_if_needed(table_xml, row_num)
                table_ref = _get_table_ref(table_xml)

            sheet_xml, ss_xml = _fill_row(
                sheet_xml, ss_xml, row_num,
                tabla, date_col,
                date_serial, detalle, serie,
                precio, cuotas, last_c,
            )
            label = f"Serie {serie}" if serie else sheet_key
            resultados.append(f"  OK {label}: fila {row_num}, {detalle}, precio={precio:,.4f}")

        mods[sheet_file] = sheet_xml
        mods[table_file] = table_xml
        mods["xl/sharedStrings.xml"] = ss_xml
        _apply_to_xlsx(filepath, mods)

        return (f"{sheet_key} — {detalle} {fecha.strftime('%d/%m/%Y')} agregado a '{nombre_archivo}':\n"
                + "\n".join(resultados))

    except Exception as e:
        return f"Error al agregar {detalle} en {sheet_key}: {e}"


def agregar_vr_bursatil_pt(nombre_archivo: str, año: int, mes: int,
                            precio_cuota: float) -> str:
    """Agrega fila VR Bursátil mensual en hoja A&R PT."""
    return _agregar_vr(nombre_archivo, "A&R PT", año, mes, "VR Bursátil", {None: precio_cuota})


def agregar_vr_bursatil_rentas(nombre_archivo: str, año: int, mes: int,
                                precio_a: float, precio_c: float, precio_i: float) -> str:
    """Agrega 3 filas VR Bursátil mensuales en hoja A&R Rentas (series A, C, I)."""
    return _agregar_vr(nombre_archivo, "A&R Rentas", año, mes, "VR Bursátil",
                       {"A": precio_a, "C": precio_c, "I": precio_i})


def agregar_vr_contable_pt(nombre_archivo: str, año: int, mes: int,
                            precio_cuota: float) -> str:
    """Agrega fila VR Contable trimestral en hoja A&R PT."""
    return _agregar_vr(nombre_archivo, "A&R PT", año, mes, "VR Contable", {None: precio_cuota})


def agregar_vr_contable_rentas(nombre_archivo: str, año: int, mes: int,
                                precio_a: float, precio_c: float, precio_i: float) -> str:
    """Agrega 3 filas VR Contable trimestrales en hoja A&R Rentas (series A, C, I)."""
    return _agregar_vr(nombre_archivo, "A&R Rentas", año, mes, "VR Contable",
                       {"A": precio_a, "C": precio_c, "I": precio_i})


def agregar_vr_contable_apoquindo(nombre_archivo: str, año: int, mes: int,
                                   precio_cuota: float) -> str:
    """Agrega fila VR Contable trimestral en hoja A&R Apoquindo."""
    return _agregar_vr(nombre_archivo, "A&R Apoquindo", año, mes, "VR Contable",
                       {None: precio_cuota})


def agregar_dividendo_pt(nombre_archivo: str, año: int, mes: int,
                         monto_por_cuota: float) -> str:
    """Agrega fila Dividendo en hoja A&R PT."""
    return _agregar_vr(nombre_archivo, "A&R PT", año, mes, "Dividendo", {None: monto_por_cuota})


def agregar_dividendo_rentas(nombre_archivo: str, año: int, mes: int,
                              monto_a: float, monto_c: float, monto_i: float) -> str:
    """Agrega 3 filas Dividendo en A&R Rentas (series A, C, I)."""
    return _agregar_vr(nombre_archivo, "A&R Rentas", año, mes, "Dividendo",
                       {"A": monto_a, "C": monto_c, "I": monto_i})


def agregar_dividendo_apoquindo(nombre_archivo: str, año: int, mes: int,
                                 monto_por_cuota: float) -> str:
    """Agrega fila Dividendo en hoja A&R Apoquindo."""
    return _agregar_vr(nombre_archivo, "A&R Apoquindo", año, mes, "Dividendo",
                       {None: monto_por_cuota})


def agregar_aporte_pt(nombre_archivo: str, año: int, mes: int,
                      monto_por_cuota: float) -> str:
    """Agrega fila Aporte en hoja A&R PT."""
    return _agregar_vr(nombre_archivo, "A&R PT", año, mes, "Aporte", {None: monto_por_cuota})


def agregar_aporte_rentas(nombre_archivo: str, año: int, mes: int,
                           monto_a: float, monto_c: float, monto_i: float) -> str:
    """Agrega 3 filas Aporte en A&R Rentas (series A, C, I)."""
    return _agregar_vr(nombre_archivo, "A&R Rentas", año, mes, "Aporte",
                       {"A": monto_a, "C": monto_c, "I": monto_i})


def agregar_aporte_apoquindo(nombre_archivo: str, año: int, mes: int,
                              monto_por_cuota: float) -> str:
    """Agrega fila Aporte en hoja A&R Apoquindo."""
    return _agregar_vr(nombre_archivo, "A&R Apoquindo", año, mes, "Aporte",
                       {None: monto_por_cuota})


def info_siguiente_accion(nombre_archivo: str) -> str:
    """
    Lee el estado actual de las hojas A&R y reporta qué falta actualizar.
    Muestra la última fecha registrada en cada hoja y la siguiente fila disponible.
    """
    filepath = _resolve_path(nombre_archivo)
    if not os.path.exists(filepath):
        return f"Error: no se encontró '{filepath}'."
    try:
        report = []
        with zipfile.ZipFile(filepath, "r") as z:
            ss_xml = z.read("xl/sharedStrings.xml").decode("utf-8")
            sis = list(re.finditer(r"<si>(.*?)</si>", ss_xml, re.DOTALL))
            strings = []
            for si in sis:
                t = re.search(r"<t[^>]*>([^<]*)</t>", si.group(1))
                strings.append(t.group(1) if t else "")

            for sheet_key, cfg in SHEET_CFG.items():
                sheet_xml = z.read(cfg["sheet_file"]).decode("utf-8")
                table_xml = z.read(cfg["table_file"]).decode("utf-8")
                table_ref = _get_table_ref(table_xml)
                row_num, last_data_row, last_c = _find_first_empty_date_row(sheet_xml, table_ref)

                # Leer la última fecha
                if last_data_row >= _table_ref_first_data_row(table_ref):
                    d_m = re.search(
                        rf'<c r="D{last_data_row}"[^>]*><v>(\d+)</v>', sheet_xml
                    )
                    last_date = _from_excel_date(int(d_m.group(1))).strftime("%d/%m/%Y") if d_m else "?"

                    e_m = re.search(
                        rf'<c r="E{last_data_row}"[^>]*t="s"[^>]*><v>(\d+)</v>', sheet_xml
                    )
                    last_det = strings[int(e_m.group(1))] if e_m and int(e_m.group(1)) < len(strings) else "?"
                else:
                    last_date = "sin datos"
                    last_det = ""

                next_info = f"fila {row_num}" if row_num != -1 else "tabla llena (necesita expansión)"
                report.append(
                    f"{sheet_key}:\n"
                    f"  Último dato: fila {last_data_row} | {last_date} | {last_det}\n"
                    f"  Próxima fila disponible: {next_info}\n"
                    f"  Tabla: {table_ref}"
                )

        return "\n\n".join(report)
    except Exception as e:
        return f"Error al leer estado: {e}"


# ─── TIR y verificación de archivos ──────────────────────────────────────────

def buscar_tir() -> str:
    """
    Busca el archivo Cálculo TIR Fondo Rentas más reciente en SharePoint
    (busca recursivamente bajo Controles de Gestión/).
    Retorna la ruta absoluta o mensaje de error.
    """
    carpeta = os.path.join(SHAREPOINT_DIR, "Controles de Gestión")
    archivos = glob.glob(os.path.join(carpeta, "**", "*TIR*Rentas*.xlsx"), recursive=True)
    if not archivos:
        archivos = glob.glob(os.path.join(carpeta, "**", "*TIR*.xlsx"), recursive=True)
    if not archivos:
        return f"Error: no se encontró archivo TIR en {carpeta}"
    return max(archivos, key=os.path.getmtime)


def verificar_archivos_cdg(año: int, mes: int) -> str:
    """
    Verifica qué archivos necesarios para actualizar el CDG del mes indicado
    están disponibles y cuáles faltan. Responde a '¿tienes todo para el CDG de [mes]?'

    Archivos requeridos siempre:
      - CDG mes anterior (para copiar)
      - Saldo Caja (más reciente del año)
      - RR JLL (Nicole)
      - RR Tres A Viña (Sebastián)
      - RR Tres A Curicó (Sebastián)
      - EEFF Viña Centro
      - EEFF Curicó
      - EEFF INMOSA

    Adicionales en fin de trimestre (mar/jun/sep/dic):
      - EEFF PT (Toesca Rentas Inmobiliarias PT)
      - EEFF Rentas/TRI (Toesca Rentas Inmobiliarias)
      - EEFF Apoquindo (Fondo Toesca Rentas Apoquindo)
      - TIR Fondo Rentas
    """
    from tools.caja_tools import buscar_saldo_caja
    from tools.noi_tools import buscar_rr_jll, buscar_er_inmosa, _find_eeff_file, _RR_JLL_BASE, _INMOSA_BASE, _TRES_A_DIRS
    from tools.rentroll_tools import _find_file, _RR_JLL_DIR, _RR_TRESA_DIRS
    from tools.eeff_tools import buscar_pdf_eeff, FONDO_RUTAS

    aamm = f"{str(año)[2:]}{mes:02d}"
    es_trimestre = mes in (3, 6, 9, 12)
    meses_es = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
                7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

    checklist = []  # lista de (nombre, ok, ruta_encontrada, donde_subir)

    def _sp(ruta: str) -> str:
        """Convierte ruta absoluta a 'SP: ruta/relativa' para legibilidad."""
        if SHAREPOINT_DIR and ruta and ruta.startswith(SHAREPOINT_DIR):
            rel = ruta[len(SHAREPOINT_DIR):].lstrip(os.sep)
            return f"SP: {rel}"
        return ruta

    def chk(nombre, ruta_o_error, donde_subir: str = ""):
        ok = bool(ruta_o_error and os.path.isfile(ruta_o_error))
        checklist.append((nombre, ok, ruta_o_error if ok else None, donde_subir))

    _MESES_ABR = {"ene":1,"feb":2,"mar":3,"abr":4,"may":5,"jun":6,
                  "jul":7,"ago":8,"sep":9,"oct":10,"nov":11,"dic":12}

    # ── CDG mes anterior — solo versión vF ───────────────────────────────────
    mes_ant = mes - 1 if mes > 1 else 12
    año_ant = año if mes > 1 else año - 1
    aamm_ant = f"{str(año_ant)[2:]}{mes_ant:02d}"
    cdg_prev = None
    sp_cdg_root = os.path.join(SHAREPOINT_DIR, "Controles de Gestión")
    patron = os.path.join(sp_cdg_root, "**", f"{aamm_ant}*vF*.xlsx")
    candidatos_sp = glob.glob(patron, recursive=True)
    if candidatos_sp:
        cdg_prev = max(candidatos_sp, key=os.path.getmtime)
    chk(f"CDG {aamm_ant} vF (mes anterior)", cdg_prev,
        os.path.join(sp_cdg_root, "Renta Comercial", "Controles de Gestión", str(año_ant)))

    # ── Saldo Caja — necesita archivo del mes SIGUIENTE al CDG ───────────────
    mes_sig = mes + 1 if mes < 12 else 1
    año_sig = año if mes < 12 else año + 1
    aamm_sig = f"{str(año_sig)[2:]}{mes_sig:02d}"
    sc = buscar_saldo_caja(año, mes)
    if sc.startswith("Error") and mes == 12:
        sc = buscar_saldo_caja(año + 1, 1)
    sc_valido = None
    if not sc.startswith("Error"):
        m_sc = re.match(r"(\d{2})(\d{2})\d{2}", os.path.basename(sc))
        if m_sc and (m_sc.group(1) + m_sc.group(2)) >= aamm_sig:
            sc_valido = sc
    chk(f"Saldo Caja (necesita {meses_es[mes_sig]} {año_sig} o posterior)", sc_valido,
        os.path.join(SHAREPOINT_DIR, "Controles de Gestión", "Saldo Caja", str(año_sig)))

    # ── RR JLL ────────────────────────────────────────────────────────────────
    rr_jll = buscar_rr_jll(año, mes)
    chk(f"RR JLL {aamm}", rr_jll if not rr_jll.startswith("Error") else None,
        os.path.join(_RR_JLL_DIR, str(año)))

    # ── RR Tres A ─────────────────────────────────────────────────────────────
    chk(f"RR Tres A Viña {meses_es[mes]}", _find_file(año, mes, "vina"),
        os.path.join(_RR_TRESA_DIRS["vina"], str(año)))
    chk(f"RR Tres A Curicó {meses_es[mes]}", _find_file(año, mes, "curico"),
        os.path.join(_RR_TRESA_DIRS["curico"], str(año)))

    # ── EEFF Viña y Curicó — validar que el archivo sea del mes correcto ──────
    mes_str = f"{mes:02d}-{año}"
    for label, mall in [(f"EEFF Viña Centro {mes_str}", "vina"),
                        (f"EEFF Curicó {mes_str}", "curico")]:
        path = _find_eeff_file(mall, año, mes)
        valido = path if (path and mes_str in os.path.basename(path)) else None
        chk(label, valido, os.path.join(_TRES_A_DIRS[mall], str(año)))

    # ── EEFF INMOSA — validar que el archivo cubra el mes del CDG ─────────────
    inmosa = buscar_er_inmosa(año, mes)
    inmosa_valido = None
    if not inmosa.startswith("Error"):
        nombre_lower = os.path.basename(inmosa).lower()
        menciones = [(nombre_lower.find(abr), m)
                     for abr, m in _MESES_ABR.items() if abr in nombre_lower]
        if menciones:
            ultimo_mes = max(menciones, key=lambda x: x[0])[1]
            inmosa_valido = inmosa if ultimo_mes >= mes else None
        else:
            inmosa_valido = inmosa
    chk(f"EEFF INMOSA (necesita cubrir {meses_es[mes]})", inmosa_valido,
        os.path.join(_INMOSA_BASE, str(año)))

    # ── Fin de trimestre ──────────────────────────────────────────────────────
    # Los EEFF son del trimestre ANTERIOR (ej: CDG marzo → EEFF diciembre año-1)
    _FONDO_DISPLAY = {
        "A&R PT":        "EEFF PT (Toesca Rentas Inmobiliarias PT)",
        "A&R Rentas":    "EEFF Rentas/TRI (Toesca Rentas Inmobiliarias)",
        "A&R Apoquindo": "EEFF Apoquindo (Fondo Toesca Rentas Apoquindo)",
    }
    if es_trimestre:
        trim_map = {3: (12, año - 1), 6: (3, año), 9: (6, año), 12: (9, año)}
        mes_eeff, año_eeff = trim_map[mes]
        for fondo_key, label_base in _FONDO_DISPLAY.items():
            ruta_pdf = buscar_pdf_eeff(fondo_key, año_eeff, mes_eeff)
            ok = os.path.isfile(ruta_pdf) if ruta_pdf else False
            chk(f"{label_base} {mes_eeff:02d}-{año_eeff}",
                ruta_pdf if ok else None,
                os.path.join(FONDO_RUTAS[fondo_key], str(año_eeff)))

        tir = buscar_tir()
        chk("TIR Fondo Rentas", tir if not tir.startswith("Error") else None,
            os.path.join(SHAREPOINT_DIR, "Controles de Gestión"))

    # ── Resumen ───────────────────────────────────────────────────────────────
    sufijo = " (fin de trimestre)" if es_trimestre else ""
    encontrados = [(n, r)    for n, ok, r, _  in checklist if ok]
    faltantes   = [(n, dest) for n, ok, _, dest in checklist if not ok]
    n_total  = len(checklist)
    n_faltan = len(faltantes)

    lines = [f"Verificación CDG {meses_es[mes]} {año}{sufijo}", ""]

    lines.append(f"Archivos encontrados ({len(encontrados)}/{n_total}):")
    for nombre, ruta in encontrados:
        lines.append(f"  ✓  {nombre}")
        lines.append(f"     {_sp(ruta)}")
        lines.append("")

    lines.append(f"Archivos faltantes ({n_faltan}/{n_total}):")
    if faltantes:
        for nombre, dest in faltantes:
            lines.append(f"  ✗  {nombre}")
            if dest:
                lines.append(f"     → Subir a: {_sp(dest)}")
            lines.append("")
    else:
        lines.append("  (ninguno)")
        lines.append("")

    if n_faltan:
        lines.append(f"=> Faltan {n_faltan} archivo(s). No se puede iniciar hasta tenerlos todos.")
    else:
        lines.append("=> Todo listo para actualizar el CDG.")
    return "\n".join(lines)


# ── Consulta histórica de datos del CDG ──────────────────────────────────────

def leer_cdg_historico(mes: int, año: int, hoja: str, filtro: str = None) -> str:
    """
    Lee una hoja de cualquier CDG histórico directamente desde el servidor
    sin necesidad de copiar el archivo al WORK_DIR.
    Útil para responder preguntas históricas: vacancia, NOI, precios cuota, etc.

    Parámetros:
        mes, año : período del CDG
        hoja     : nombre exacto de la hoja ("Vacancia", "NOI-RCSD", "Input AP", ...)
        filtro   : keyword para filtrar filas (busca en todas las celdas)
    """
    aamm = f"{str(año)[2:]}{mes:02d}"
    nombre = f"{aamm} Control De Gestión Renta Comercial.xlsx"
    ruta = os.path.join(RUTA_COMERCIAL, nombre)

    if not os.path.exists(ruta):
        return f"No se encontró el CDG '{nombre}' en {RUTA_COMERCIAL}"

    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), nombre)
        shutil.copy2(ruta, tmp)
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)

    if hoja not in wb.sheetnames:
        sugeridas = [s for s in wb.sheetnames if hoja.lower() in s.lower()]
        wb.close()
        if sugeridas:
            return f"Hoja '{hoja}' no encontrada. Similares: {', '.join(sugeridas[:5])}"
        primeras = wb.sheetnames[:20]
        return f"Hoja '{hoja}' no encontrada en {nombre}. Primeras hojas: {', '.join(primeras)}"

    ws = wb[hoja]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return f"Hoja '{hoja}' vacía en {nombre}"

    if filtro:
        fl = filtro.lower()
        matched = [r for r in rows if any(fl in str(v).lower() for v in r if v is not None)]
        if not matched:
            return f"No se encontraron filas con '{filtro}' en hoja '{hoja}' de {nombre}"
        rows_out = matched[:50]
        header = f"CDG {nombre} — Hoja: {hoja} | Filtro: '{filtro}' ({len(matched)} fila(s))"
        trailer = f"\n... y {len(matched) - 50} fila(s) más" if len(matched) > 50 else ""
    else:
        rows_out = rows[:80]
        header = f"CDG {nombre} — Hoja: {hoja}"
        trailer = f"\n... y {len(rows) - 80} fila(s) más. Usa 'filtro' para buscar datos específicos." if len(rows) > 80 else ""

    lines = [header, ""]
    for row in rows_out:
        cells = [str(v).strip() for v in row if v is not None and str(v).strip() not in ("", "None")]
        if cells:
            lines.append("  |  ".join(cells[:12]))

    if trailer:
        lines.append(trailer)

    return "\n".join(lines)
