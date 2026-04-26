"""
Herramientas para la sección 'RENTABILIDAD DEL FONDO (en UF)' de DATOS FS.

Flujo típico por fondo:
  PT / Apoquindo:
    1. actualizar_fecha_ar(archivo, fondo, fecha_serial)  → cambia D2 a fecha contable
    2. [Abrir Excel para recalcular XIRR]
    3. leer_rentabilidades_ar(archivo, fondo)              → lee N10/O10/P10 cacheados
    4. pegar_rentabilidades_datos_fs(archivo, fondo, rent) → escribe en DATOS FS

  Rentas (3 series, inicio viene del archivo TIR Fondo):
    1. actualizar_fecha_ar(archivo, "A&R Rentas", fecha_serial)
    2. [Abrir Excel para recalcular]
    3. leer_rentabilidades_ar(archivo, "A&R Rentas")       → YTD/12M por serie
    4. copiar_datos_tir_rentas(archivo_cg, archivo_tir)    → A&R Rentas C:M → TIR Fondo B:L
    5. leer_tir_rentas_resumen(archivo_tir)                → desde-inicio por serie
    6. pegar_rentabilidades_datos_fs(archivo, "A&R Rentas", rent_por_serie)
"""
import os
import re
import zipfile
import shutil
from config import WORK_DIR

# ─── Mapping de sheet files A&R (sincronizado con gestion_renta_tools.SHEET_CFG) ─
AR_SHEET_FILES = {
    "A&R PT":        "xl/worksheets/sheet16.xml",
    "A&R Apoquindo": "xl/worksheets/sheet15.xml",
    "A&R Rentas":    "xl/worksheets/sheet17.xml",
}

# ─── Celdas XIRR en hojas A&R ──────────────────────────────────────────────────
# PT y Apoquindo: N10=inicio, O10=YTD, P10=12M (fórmulas XIRR, requieren Excel abierto)
# Rentas: row 12 por serie (también XIRR)
AR_RENT_CELLS = {
    "A&R PT": {
        None: {"inicio": "N10", "ytd": "O10", "12m": "P10"},
    },
    "A&R Apoquindo": {
        None: {"inicio": "N10", "ytd": "O10", "12m": "P10"},
    },
    "A&R Rentas": {
        # Rentas no tiene 'inicio' aquí; ese dato viene del archivo TIR Fondo Resumen
        "A": {"ytd": "P12", "12m": "Q12"},
        "C": {"ytd": "Y12", "12m": "Z12"},
        "I": {"ytd": "AH12", "12m": "AI12"},
    },
}

# ─── Celdas hardcoded (sin fórmula) en DATOS FS ────────────────────────────────
# Columnas Libro: solo estas se escriben; las columnas Bursátil son fórmulas.
DATOS_FS_CELLS = {
    "A&R Rentas": {
        "A": {"inicio": "H10", "ytd": "H11", "12m": "H12"},
        "C": {"inicio": "J10", "ytd": "J11", "12m": "J12"},
        "I": {"inicio": "L10", "ytd": "L11", "12m": "L12"},
    },
    "A&R PT": {
        None: {"inicio": "H98", "ytd": "H99", "12m": "H100"},
    },
    "A&R Apoquindo": {
        None: {"inicio": "H136", "ytd": "H137", "12m": "H138"},
    },
}

# ─── Nombre de la hoja TIR Fondo dentro del archivo TIR ────────────────────────
TIR_FONDO_SHEET  = "TIR Fondo"
TIR_RESUMEN_SHEET = "Resumen"


# ─── Helpers internos ─────────────────────────────────────────────────────────

def _resolve_path(nombre_archivo: str) -> str:
    if os.path.isabs(nombre_archivo):
        return nombre_archivo
    return os.path.join(WORK_DIR, nombre_archivo)


def _find_cell_bounds(xml: str, ref: str) -> tuple:
    """Retorna (start, end) del span completo de la celda. (-1,-1) si no existe."""
    tag = f'<c r="{ref}"'
    start = xml.find(tag)
    if start == -1:
        return -1, -1
    i = start + len(tag)
    while i < len(xml):
        if xml[i:i+2] == '/>':
            return start, i + 2
        if xml[i] == '>':
            end = xml.find('</c>', i)
            return start, (end + 4) if end != -1 else i + 1
        i += 1
    return start, len(xml)


def _get_cell_attr(xml: str, ref: str, attr: str, default: str) -> str:
    """Extrae el valor de un atributo de la celda (p.ej. s='14')."""
    tag = f'<c r="{ref}"'
    start = xml.find(tag)
    if start == -1:
        return default
    snippet = xml[start: min(start + 300, len(xml))]
    m = re.search(rf'\b{attr}="([^"]+)"', snippet)
    return m.group(1) if m else default


def _read_cell_numeric(xml: str, ref: str) -> float | None:
    """Lee el valor numérico cacheado de una celda (<v>). None si vacío o ausente."""
    start, end = _find_cell_bounds(xml, ref)
    if start == -1:
        return None
    cell_xml = xml[start:end]
    # Ignorar celdas con tipo string (t="s") o boolean
    if 't="s"' in cell_xml or 't="b"' in cell_xml:
        return None
    v_m = re.search(r'<v>([^<]+)</v>', cell_xml)
    if not v_m:
        return None
    try:
        return float(v_m.group(1))
    except ValueError:
        return None


def _replace_or_insert_cell(row_xml: str, ref: str, new_cell: str) -> str:
    """Reemplaza la celda si existe, o la inserta en orden de columna."""
    start, end = _find_cell_bounds(row_xml, ref)
    if start != -1:
        return row_xml[:start] + new_cell + row_xml[end:]
    # Insertar en orden
    col = re.sub(r"\d", "", ref)
    col_n = _col_num(col)
    for m in re.finditer(r'<c r="([A-Z]+)\d+"', row_xml):
        if _col_num(m.group(1)) > col_n:
            return row_xml[:m.start()] + new_cell + row_xml[m.start():]
    return row_xml + new_cell


def _col_num(letter: str) -> int:
    n = 0
    for c in letter.upper():
        n = n * 26 + ord(c) - ord("A") + 1
    return n


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _write_cell_in_sheet(sheet_xml: str, cell_ref: str, value: float) -> str:
    """
    Escribe un valor numérico en una celda, preservando el estilo existente.
    Si la celda no existe, la crea sin estilo. Elimina fórmulas existentes.
    """
    row_num = int(re.sub(r"[A-Z]", "", cell_ref))
    row_pat = rf'(<row\b[^>]*\br="{row_num}"[^>]*>)(.*?)(</row>)'
    row_m = re.search(row_pat, sheet_xml, re.DOTALL)

    if row_m:
        row_open = row_m.group(1)
        row_content = row_m.group(2)
        row_close = "</row>"
    else:
        row_open = f'<row r="{row_num}" spans="1:20">'
        row_content = ""
        row_close = "</row>"

    style = _get_cell_attr(row_content, cell_ref, "s", "0")
    # Formatear el valor: porcentajes se guardan como decimales (ej 0.0523)
    new_cell = f'<c r="{cell_ref}" s="{style}"><v>{value}</v></c>'
    row_content = _replace_or_insert_cell(row_content, cell_ref, new_cell)
    new_row = row_open + row_content + row_close

    if row_m:
        return sheet_xml[:row_m.start()] + new_row + sheet_xml[row_m.end():]
    else:
        # Insertar la fila en orden
        next_row_m = re.search(
            rf'<row\b[^>]*\br="({row_num + 1}|{row_num + 2}|{row_num + 3})"',
            sheet_xml,
        )
        if next_row_m:
            return sheet_xml[:next_row_m.start()] + new_row + sheet_xml[next_row_m.start():]
        return sheet_xml.replace("</sheetData>", new_row + "</sheetData>")


def _apply_to_xlsx(filepath: str, modifications: dict) -> None:
    """Aplica modificaciones al xlsx editando solo los archivos especificados."""
    tmp = filepath + ".tmp"
    with zipfile.ZipFile(filepath, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = modifications.get(item.filename)
                if data is not None:
                    zout.writestr(item, data if isinstance(data, bytes) else data.encode("utf-8"))
                else:
                    zout.writestr(item, zin.read(item.filename))
    os.replace(tmp, filepath)


def _find_sheet_xml_path(xlsx_path: str, sheet_name: str) -> str | None:
    """
    Busca la ruta interna del XML de una hoja por nombre.
    Ej: 'DATOS FS' → 'xl/worksheets/sheet5.xml'
    """
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        wb_xml = zf.read("xl/workbook.xml").decode("utf-8")

        # Buscar el rId del sheet
        m = re.search(
            r'<sheet\b(?=[^>]*\bname="' + re.escape(sheet_name) + r'")[^>]*\br:id="([^"]+)"',
            wb_xml,
        )
        if not m:
            # Intentar con orden inverso de atributos
            m = re.search(
                r'<sheet\b(?=[^>]*\br:id="([^"]+)")[^>]*\bname="' + re.escape(sheet_name) + r'"',
                wb_xml,
            )
            if not m:
                return None
            rid = m.group(1)
        else:
            rid = m.group(1)

        rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        m2 = re.search(
            r'<Relationship\b(?=[^>]*\bId="' + re.escape(rid) + r'")[^>]*\bTarget="([^"]+)"',
            rels_xml,
        )
        if not m2:
            return None
        target = m2.group(1)
        # El Target puede ser relativo, ej: "worksheets/sheet5.xml"
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        return target


def _read_shared_strings(xlsx_path: str) -> list:
    """Lee el array de strings compartidos."""
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        if "xl/sharedStrings.xml" not in zf.namelist():
            return []
        ss_xml = zf.read("xl/sharedStrings.xml").decode("utf-8")
    result = []
    for si_m in re.finditer(r"<si>(.*?)</si>", ss_xml, re.DOTALL):
        t = re.search(r"<t[^>]*>([^<]*)</t>", si_m.group(1))
        result.append(t.group(1) if t else "")
    return result


# ─── Herramientas públicas ─────────────────────────────────────────────────────

def actualizar_fecha_ar(nombre_archivo: str, fondo_key: str, fecha_serial: int) -> str:
    """
    Actualiza la celda D2 en la hoja A&R del fondo indicado con el serial de fecha.
    Usar para cambiar entre fecha contable y fecha bursátil antes de leer XIRR.

    IMPORTANTE: Después de ejecutar esta función, abrir el archivo en Excel y guardar
    para que las fórmulas XIRR (N10/O10/P10) se recalculen con la nueva fecha.

    Args:
        nombre_archivo: Archivo en WORK_DIR (o ruta absoluta).
        fondo_key: 'A&R PT', 'A&R Apoquindo' o 'A&R Rentas'.
        fecha_serial: Serial Excel de la fecha (ej: 46112 = 31/03/2026).
    """
    if fondo_key not in AR_SHEET_FILES:
        return f"Error: fondo '{fondo_key}' no reconocido. Disponibles: {list(AR_SHEET_FILES)}"

    filepath = _resolve_path(nombre_archivo)
    if not os.path.exists(filepath):
        return f"Error: no se encontró '{filepath}'."

    sheet_file = AR_SHEET_FILES[fondo_key]
    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            sheet_xml = zf.read(sheet_file).decode("utf-8")

        sheet_xml = _write_cell_in_sheet(sheet_xml, "D2", fecha_serial)
        _apply_to_xlsx(filepath, {sheet_file: sheet_xml})

        from datetime import date, timedelta
        d = date(1899, 12, 30) + timedelta(days=fecha_serial)
        return (
            f"OK: D2 en '{fondo_key}' actualizado a {d.strftime('%d/%m/%Y')} (serial {fecha_serial}).\n"
            f"IMPORTANTE: Abrir el archivo en Excel y guardar para recalcular XIRR."
        )
    except Exception as e:
        return f"Error al actualizar fecha: {e}"


def leer_rentabilidades_ar(nombre_archivo: str, fondo_key: str) -> str:
    """
    Lee los valores cacheados de rentabilidad contable desde la hoja A&R.

    - PT / Apoquindo: N10 (inicio), O10 (YTD), P10 (12M) — fórmulas XIRR.
    - Rentas: P12/Q12 (Serie A), Y12/Z12 (Serie C), AH12/AI12 (Serie I).

    Si las celdas están vacías, el archivo debe abrirse en Excel para recalcular.
    Los valores se retornan como porcentajes decimales (0.0523 = 5.23%).
    """
    if fondo_key not in AR_SHEET_FILES:
        return f"Error: fondo '{fondo_key}' no reconocido."

    filepath = _resolve_path(nombre_archivo)
    if not os.path.exists(filepath):
        return f"Error: no se encontró '{filepath}'."

    sheet_file = AR_SHEET_FILES[fondo_key]
    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            sheet_xml = zf.read(sheet_file).decode("utf-8")

        cell_map = AR_RENT_CELLS[fondo_key]
        lines = [f"Rentabilidades contables — {fondo_key}:"]
        found_any = False

        for serie, cells in cell_map.items():
            label = f"Serie {serie}" if serie else "Fondo"
            for metric, ref in cells.items():
                val = _read_cell_numeric(sheet_xml, ref)
                if val is not None:
                    pct = val * 100
                    lines.append(f"  {label} | {metric.upper():6s} ({ref}): {pct:.4f}%  [{val:.8f}]")
                    found_any = True
                else:
                    lines.append(f"  {label} | {metric.upper():6s} ({ref}): [vacío — Excel debe recalcular]")

        if not found_any:
            lines.append(
                "\nNinguna celda tiene valor. Pasos:\n"
                "  1. Actualizar D2 con la fecha correcta (usar actualizar_fecha_ar)\n"
                "  2. Abrir el archivo en Excel, esperar que calcule y guardarlo\n"
                "  3. Volver a ejecutar leer_rentabilidades_ar"
            )
        return "\n".join(lines)
    except Exception as e:
        return f"Error al leer rentabilidades: {e}"


# ─── Celdas DATOS FS para todas las métricas del FS (bursátil + libro) ────────
# Estructura: {serie_o_None: {métrica: {bursatil: ref, libro: ref}}}
# Apoquindo no tiene bursátil (sin ticker) → solo 'libro'
DATOS_FS_RENT_CELLS = {
    "A&R PT": {
        None: {
            "inicio":   {"bursatil": "G98",  "libro": "H98"},
            "ytd":      {"bursatil": "G99",  "libro": "H99"},
            "12m":      {"bursatil": "G100", "libro": "H100"},
            "dy":       {"bursatil": "G101", "libro": "H101"},
            "dy_amort": {"bursatil": "G102", "libro": "H102"},
        }
    },
    "A&R Apoquindo": {
        None: {
            "inicio":   {"libro": "H136"},
            "ytd":      {"libro": "H137"},
            "12m":      {"libro": "H138"},
            "dy":       {"libro": "H139"},
            "dy_amort": {"libro": "H140"},
        }
    },
    "A&R Rentas": {
        "A": {
            "inicio":   {"bursatil": "G10", "libro": "H10"},
            "ytd":      {"bursatil": "G11", "libro": "H11"},
            "12m":      {"bursatil": "G12", "libro": "H12"},
            "dy":       {"bursatil": "G13", "libro": "H13"},
            "dy_amort": {"bursatil": "G14", "libro": "H14"},
        },
        "C": {
            "inicio":   {"bursatil": "I10", "libro": "J10"},
            "ytd":      {"bursatil": "I11", "libro": "J11"},
            "12m":      {"bursatil": "I12", "libro": "J12"},
            "dy":       {"bursatil": "I13", "libro": "J13"},
            "dy_amort": {"bursatil": "I14", "libro": "J14"},
        },
        "I": {
            "inicio":   {"bursatil": "K10", "libro": "L10"},
            "ytd":      {"bursatil": "K11", "libro": "L11"},
            "12m":      {"bursatil": "K12", "libro": "L12"},
            "dy":       {"bursatil": "K13", "libro": "L13"},
            "dy_amort": {"bursatil": "K14", "libro": "L14"},
        },
    },
}

# Hoja DATOS FS
DATOS_FS_SHEET = "xl/worksheets/sheet9.xml"


def _fmt_pct(val: float) -> str:
    """0.0860 → '8,6%'  |  -0.0638 → '-6,4%'"""
    return f"{val * 100:.1f}%".replace(".", ",")


def leer_rentabilidades_completas_fs(nombre_archivo: str, fondo_key: str) -> str:
    """
    Lee todas las métricas de rentabilidad del FS desde la hoja DATOS FS del CDG:
    inicio, YTD, 12M, Dividend Yield, DY+Amortización — para bursátil y libro.

    Para A&R Apoquindo solo retorna valores libro (sin ticker bursátil).
    Para A&R Rentas retorna las 3 series (A, C, I).

    Retorna texto con los valores y JSON listo para datos_json['rentabilidad']
    en actualizar_fs_pt / actualizar_fs_apoquindo / actualizar_fs_tri.

    IMPORTANTE: el CDG debe corresponder al mes del FS y estar recalculado en Excel.
    """
    if fondo_key not in DATOS_FS_RENT_CELLS:
        return f"Error: fondo '{fondo_key}' no reconocido."

    filepath = _resolve_path(nombre_archivo)
    if not os.path.exists(filepath):
        return f"Error: no se encontró '{filepath}'."

    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            sheet_xml = zf.read(DATOS_FS_SHEET).decode("utf-8")
    except Exception as e:
        return f"Error al leer DATOS FS: {e}"

    cell_map = DATOS_FS_RENT_CELLS[fondo_key]
    metricas = ["inicio", "ytd", "12m", "dy", "dy_amort"]
    labels = {
        "inicio":   "Desde inicio (anualizada)",
        "ytd":      "YTD (anualizada)",
        "12m":      "Últimos 12 meses",
        "dy":       "Dividend Yield",
        "dy_amort": "DY + Amortización",
    }

    resultado = {}
    lines = [f"Rentabilidades FS — {fondo_key}:"]

    for serie, metr_map in cell_map.items():
        label_serie = f"Serie {serie}" if serie else "Fondo"
        lines.append(f"\n  {label_serie}:")
        serie_data = {}

        for met in metricas:
            refs = metr_map.get(met, {})
            row_data = {}
            for tipo, ref in refs.items():
                val = _read_cell_numeric(sheet_xml, ref)
                if val is not None:
                    row_data[tipo] = _fmt_pct(val)
                else:
                    row_data[tipo] = "[vacío]"
            # Formatear línea
            if "bursatil" in row_data and "libro" in row_data:
                lines.append(f"    {labels[met]:35s} B:{row_data['bursatil']:8s} L:{row_data['libro']}")
            elif "libro" in row_data:
                lines.append(f"    {labels[met]:35s} L:{row_data['libro']}")
            # Guardar para JSON — claves según fondo
            if serie:
                for tipo, val in row_data.items():
                    resultado[f"{met}_{tipo}_{serie.lower()}"] = val
            else:
                for tipo, val in row_data.items():
                    resultado[f"{met}_{tipo}"] = val

        serie_data = resultado

    lines.append(f"\nJSON listo para datos_json['rentabilidad']:\n{__import__('json').dumps(resultado, ensure_ascii=False)}")
    return "\n".join(lines)


def pegar_rentabilidades_datos_fs(
    nombre_archivo: str,
    fondo_key: str,
    rentabilidades: dict,
) -> str:
    """
    Escribe los valores de rentabilidad libro en las celdas hardcoded de DATOS FS.

    Args:
        nombre_archivo: Archivo en WORK_DIR.
        fondo_key: 'A&R PT', 'A&R Apoquindo' o 'A&R Rentas'.
        rentabilidades: dict con los valores a escribir.
          - Para PT/Apoquindo:
              {None: {"inicio": 0.0523, "ytd": 0.0312, "12m": 0.0489}}
          - Para Rentas:
              {
                "A": {"inicio": 0.04, "ytd": 0.02, "12m": 0.038},
                "C": {"inicio": 0.041, "ytd": 0.021, "12m": 0.039},
                "I": {"inicio": 0.042, "ytd": 0.022, "12m": 0.040},
              }

    Los valores deben ser decimales (0.05 = 5%). Las celdas ya tienen formato %.
    """
    if fondo_key not in DATOS_FS_CELLS:
        return f"Error: fondo '{fondo_key}' no reconocido."

    filepath = _resolve_path(nombre_archivo)
    if not os.path.exists(filepath):
        return f"Error: no se encontró '{filepath}'."

    # Encontrar sheet DATOS FS dinámicamente
    datos_fs_path = _find_sheet_xml_path(filepath, "DATOS FS")
    if not datos_fs_path:
        return "Error: no se encontró la hoja 'DATOS FS' en el archivo."

    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            sheet_xml = zf.read(datos_fs_path).decode("utf-8")

        cell_cfg = DATOS_FS_CELLS[fondo_key]
        written = []
        skipped = []

        for serie, metrics in cell_cfg.items():
            rent_serie = rentabilidades.get(serie, {})
            label = f"Serie {serie}" if serie else "Fondo"
            for metric, cell_ref in metrics.items():
                if metric in rent_serie and rent_serie[metric] is not None:
                    val = float(rent_serie[metric])
                    sheet_xml = _write_cell_in_sheet(sheet_xml, cell_ref, val)
                    pct = val * 100
                    written.append(f"  {label} {metric.upper():6s} → {cell_ref} = {pct:.4f}%")
                else:
                    skipped.append(f"  {label} {metric.upper():6s} → {cell_ref} [no proporcionado]")

        _apply_to_xlsx(filepath, {datos_fs_path: sheet_xml})

        lines = [f"DATOS FS actualizado — {fondo_key}:"]
        if written:
            lines.append("Escritos:")
            lines.extend(written)
        if skipped:
            lines.append("Omitidos (no se proporcionó valor):")
            lines.extend(skipped)
        return "\n".join(lines)
    except Exception as e:
        return f"Error al escribir en DATOS FS: {e}"


def copiar_datos_tir_rentas(archivo_cg: str, archivo_tir: str) -> str:
    """
    Copia los datos de la hoja 'A&R Rentas' (columnas C:M) del archivo CG al
    archivo TIR Fondo (columnas B:L de hoja 'TIR Fondo').

    Solo copia los valores numéricos de las filas de datos (omite la fila de headers).
    Requiere openpyxl.

    Args:
        archivo_cg:  Archivo Control de Gestión en WORK_DIR.
        archivo_tir: Archivo TIR Fondo Rentas en WORK_DIR (o ruta absoluta).
    """
    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl no instalado. Ejecutar: pip install openpyxl"

    cg_path  = _resolve_path(archivo_cg)
    tir_path = _resolve_path(archivo_tir)

    if not os.path.exists(cg_path):
        return f"Error: no se encontró '{cg_path}'."
    if not os.path.exists(tir_path):
        return f"Error: no se encontró '{tir_path}'."

    try:
        # Leer datos de A&R Rentas columnas C:M (índices 3..13 en 1-based)
        wb_cg = openpyxl.load_workbook(cg_path, read_only=True, data_only=True)
        if "A&R Rentas" not in wb_cg.sheetnames:
            wb_cg.close()
            return "Error: hoja 'A&R Rentas' no encontrada en archivo CG."

        ws_rentas = wb_cg["A&R Rentas"]
        # Leer a partir de fila 2 (fila 1 = headers)
        data_rows = []
        for row in ws_rentas.iter_rows(min_row=2, min_col=3, max_col=13):
            vals = [cell.value for cell in row]
            # Detener si la fila de fecha (primera celda del rango) está vacía
            if vals[0] is None:
                break
            data_rows.append(vals)
        wb_cg.close()

        if not data_rows:
            return "No hay filas de datos en A&R Rentas (columnas C:M)."

        # Escribir en TIR Fondo, columnas B:L, a partir de fila 2
        wb_tir = openpyxl.load_workbook(tir_path)
        if TIR_FONDO_SHEET not in wb_tir.sheetnames:
            wb_tir.close()
            return f"Error: hoja '{TIR_FONDO_SHEET}' no encontrada en '{archivo_tir}'."

        ws_tir = wb_tir[TIR_FONDO_SHEET]
        # Limpiar rango destino antes de escribir
        for r in range(2, ws_tir.max_row + 2):
            for c in range(2, 13):  # columnas B:L (2..12)
                ws_tir.cell(row=r, column=c).value = None

        for i, row_vals in enumerate(data_rows):
            row_num = i + 2
            for j, val in enumerate(row_vals):
                ws_tir.cell(row=row_num, column=j + 2).value = val

        wb_tir.save(tir_path)
        wb_tir.close()

        return (
            f"OK: {len(data_rows)} filas copiadas de 'A&R Rentas' (C:M) "
            f"→ '{TIR_FONDO_SHEET}' (B:L) en '{archivo_tir}'."
        )
    except Exception as e:
        return f"Error al copiar datos TIR Rentas: {e}"


def leer_tir_rentas_resumen(archivo_tir: str) -> str:
    """
    Lee las rentabilidades 'desde inicio' por serie desde la hoja 'Resumen'
    del archivo TIR Fondo Rentas.

    Retorna los valores encontrados y un dict JSON listo para pasar a
    pegar_rentabilidades_datos_fs como campo 'inicio' de cada serie.

    Busca en la hoja 'Resumen' valores por serie (A, C, I) usando
    openpyxl (data_only=True para obtener valores calculados).
    """
    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl no instalado."

    tir_path = _resolve_path(archivo_tir)
    if not os.path.exists(tir_path):
        return f"Error: no se encontró '{tir_path}'."

    try:
        wb = openpyxl.load_workbook(tir_path, read_only=True, data_only=True)
        if TIR_RESUMEN_SHEET not in wb.sheetnames:
            hojas = ", ".join(wb.sheetnames)
            wb.close()
            return (
                f"Error: hoja '{TIR_RESUMEN_SHEET}' no encontrada en '{archivo_tir}'.\n"
                f"Hojas disponibles: {hojas}"
            )

        ws = wb[TIR_RESUMEN_SHEET]

        # Leer todo el contenido para mostrarlo al agente
        lines = [f"Hoja '{TIR_RESUMEN_SHEET}' de '{archivo_tir}':"]
        rows_data = []
        for row in ws.iter_rows(max_row=50, values_only=True):
            if any(v is not None for v in row):
                rows_data.append(row)
                line = "  " + " | ".join(
                    (f"{v:.6f}" if isinstance(v, float) else str(v) if v is not None else "")
                    for v in row
                )
                lines.append(line)
        wb.close()

        lines.append(
            "\nNOTA: Identificar manualmente qué fila/columna contiene la TIR "
            "desde inicio anualizada para cada serie (A, C, I) y pasar esos "
            "valores como 'inicio' en pegar_rentabilidades_datos_fs."
        )
        return "\n".join(lines)
    except Exception as e:
        return f"Error al leer TIR Fondo Resumen: {e}"
