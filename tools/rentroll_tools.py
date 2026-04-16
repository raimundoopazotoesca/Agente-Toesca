"""
Herramientas para revisar planillas de Rent Roll (RR).

Proveedores:
  - JLL (Nicole Carvajal): archivo "[AAMM] Rent Roll y NOI.xlsx"
  - Tres Asociados (Sebastián Bravo): "Excel Tres A Viña [Mes] [Año].xlsx"
                                       "Excel Tres A Curicó [Mes] [Año].xlsx"

Validaciones:
  1. Coherencia de Vacantes
  2. Consistencia en Absorción (requiere archivo del mes anterior)
  3. Renta escalonada (comparando a 2 decimales)
  4. Fechas de término de contrato vencidas
"""

import glob
import json
import os
import shutil
import tempfile
from datetime import date, datetime
from typing import Optional

import openpyxl

from config import WORK_DIR
from tools.email_tools import send_email

# ── Constantes de proveedores ────────────────────────────────────────────────
NICOLE_EMAIL = "nicole.carvajal@jll.com"
SEBASTIAN_EMAIL = "sebastian.bravo@tresasociados.cl"

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

# Almacena el resultado de la última revisión para uso en enviar_emails
_ultimo_resultado: dict = {}


# ── Utilidades internas ──────────────────────────────────────────────────────

def _cierre_mes(año: int, mes: int) -> date:
    """Último día del mes."""
    if mes == 12:
        return date(año + 1, 1, 1).replace(day=1) - __import__("datetime").timedelta(days=1)
    return date(año, mes + 1, 1) - __import__("datetime").timedelta(days=1)


def _find_file(año: int, mes: int, proveedor: str) -> Optional[str]:
    """
    Busca el archivo de RR en WORK_DIR según proveedor.
    proveedor: 'jll', 'vina', 'curico'
    """
    aamm = f"{str(año)[2:]}{mes:02d}"
    mes_nombre = MESES_ES[mes]

    patterns = {
        "jll":    [f"{aamm} Rent Roll y NOI*.xlsx", f"{aamm}*Rent Roll*.xlsx"],
        "vina":   [f"*Vi*a*{mes_nombre}*{año}*.xlsx", f"*Vina*{mes_nombre}*{año}*.xlsx",
                   f"*Vi*a*{aamm}*.xlsx"],
        "curico": [f"*Curic*{mes_nombre}*{año}*.xlsx", f"*Curico*{mes_nombre}*{año}*.xlsx",
                   f"*Curic*{aamm}*.xlsx"],
    }

    for pat in patterns.get(proveedor, []):
        matches = glob.glob(os.path.join(WORK_DIR, pat))
        if matches:
            return matches[0]
    return None


def _load_ws_rows(filepath: str, sheetname: str):
    """Carga una hoja en modo read-only y retorna las filas como lista."""
    # Copiar a temp si el archivo está bloqueado
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), os.path.basename(filepath))
        shutil.copy2(filepath, tmp)
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)

    if sheetname not in wb.sheetnames:
        wb.close()
        return None, None
    ws = wb[sheetname]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows, sheetname


def _find_header_row(rows, keyword="Arrendatario"):
    for i, row in enumerate(rows):
        if any(keyword in str(v) for v in row if v):
            return i
    return None


def _get_col_map(rows, h_idx):
    header = rows[h_idx]
    col_map = {}
    fecha_positions = []
    for i, h in enumerate(header):
        if h is None:
            continue
        name = str(h).strip()
        if name not in col_map:
            col_map[name] = i
        if name == "Fecha":
            fecha_positions.append(i)
    return col_map, fecha_positions


# ── Validación 1: Coherencia de Vacantes ────────────────────────────────────

def _val1_vacantes(rows, col_map, data_start):
    errors = []
    keys = ["Tipo Activo 1", "Tipo Activo 3", "Arrendatario", "Tipo Arrendatario"]
    cols = {k: col_map.get(k) for k in keys}
    if any(v is None for v in cols.values()):
        return []
    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        if all(row[cols[k]] is None for k in keys):
            continue
        vals = {k: str(row[cols[k]] or "").strip() for k in keys}
        has_v = {k: "vacante" in vals[k].lower() for k in keys}
        if any(has_v.values()) and not all(has_v.values()):
            errors.append({
                "fila": row_idx + 1,
                "valores": vals,
            })
    return errors


# ── Validación 2: Consistencia en Absorción ─────────────────────────────────

def _read_rr_locals(rows, col_map, data_start):
    """Retorna dict {(activo, local): arrendatario}."""
    data = {}
    activo_col = col_map.get("Activo2") or col_map.get("Activo1")
    local_col = col_map.get("Local")
    arr_col = col_map.get("Arrendatario")
    if arr_col is None or local_col is None:
        return data
    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        local = str(row[local_col] or "").strip()
        arr = str(row[arr_col] or "").strip()
        activo = str(row[activo_col] or "").strip() if activo_col is not None else ""
        if not local or local in ("None", ""):
            continue
        data[(activo, local)] = arr
    return data


def _read_absorcion(rows, col_map, data_start):
    absorcion = []
    status_col = col_map.get("Status")
    arr_col = col_map.get("Arrendatario")
    new_arr_col = col_map.get("Nuevo Arrendatario")
    activo_col = col_map.get("Activo")
    if status_col is None:
        return absorcion
    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        status = str(row[status_col] or "").strip()
        if not status or status in ("None", "Status"):
            continue
        absorcion.append({
            "activo": str(row[activo_col] or "").strip() if activo_col else "",
            "status": status,
            "arr": str(row[arr_col] or "").strip() if arr_col else "",
            "new_arr": str(row[new_arr_col] or "").strip() if new_arr_col else "",
        })
    return absorcion


def _val2_absorcion(rr_prev, rr_curr, absorcion):
    errors = []
    all_locals = set(rr_prev.keys()) | set(rr_curr.keys())
    for key in sorted(all_locals):
        arr_prev = rr_prev.get(key, "(no existía)")
        arr_curr = rr_curr.get(key, "(desapareció)")
        if arr_prev == arr_curr:
            continue
        activo, local = key
        vac_prev = "vacante" in arr_prev.lower()
        vac_curr = "vacante" in arr_curr.lower()

        if vac_prev and not vac_curr:
            expected_status, expected_new = "Nuevo Contrato", arr_curr
            found = any(
                "nuevo" in m["status"].lower() and expected_new.lower() in m["new_arr"].lower()
                for m in absorcion
            )
        elif not vac_prev and vac_curr:
            expected_status, expected_new = "Término", arr_prev
            found = any(
                ("término" in m["status"].lower() or "termino" in m["status"].lower())
                and arr_prev.lower() in m["arr"].lower()
                for m in absorcion
            )
        else:
            # Cambio de arrendatario — verificación básica
            found = True

        if not found:
            errors.append({
                "local": local,
                "activo": activo,
                "anterior": arr_prev,
                "actual": arr_curr,
                "movimiento_esperado": expected_status,
            })
    return errors


# ── Validación 3: Renta Escalonada ──────────────────────────────────────────

def _val3_escalonada(rows, col_map, fecha_positions, data_start, cierre_date):
    errors = []
    renta_fija_key = next((k for k in col_map if "Renta Fija" in k and "UF/m2" in k), None)
    if not renta_fija_key:
        return errors

    escalon_pairs = []
    for i in range(1, 6):
        val_key = next((k for k in col_map if k.startswith(str(i)) and "UF/m2/mes" in k), None)
        if val_key:
            val_col = col_map[val_key]
            fecha_col = next((f for f in sorted(fecha_positions) if f > val_col), None)
            escalon_pairs.append((i, val_col, fecha_col))
    if not escalon_pairs:
        return errors

    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        arr_col = col_map.get("Arrendatario")
        arr = str(row[arr_col] or "").strip() if arr_col is not None else ""
        if "vacante" in arr.lower():
            continue
        renta_raw = row[col_map[renta_fija_key]]
        if renta_raw is None:
            continue
        try:
            renta_fija = round(float(renta_raw), 2)
        except (TypeError, ValueError):
            continue

        escalones = []
        for (num, val_col, fecha_col) in escalon_pairs:
            val_raw = row[val_col] if val_col < len(row) else None
            fecha_raw = row[fecha_col] if fecha_col is not None and fecha_col < len(row) else None
            if val_raw is None or fecha_raw is None:
                continue
            try:
                val = round(float(val_raw), 2)
            except (TypeError, ValueError):
                continue
            if isinstance(fecha_raw, datetime):
                fecha_raw = fecha_raw.date()
            if isinstance(fecha_raw, date):
                escalones.append((num, val, fecha_raw))
        if not escalones:
            continue

        activos = [(num, val, f) for (num, val, f) in escalones if f <= cierre_date]
        if not activos:
            continue  # ningún escalón ha entrado en vigor
        active = max(activos, key=lambda x: x[0])

        if renta_fija != active[1]:
            errors.append({
                "fila": row_idx + 1,
                "arrendatario": arr,
                "renta_fija": renta_fija,
                "escalon_num": active[0],
                "escalon_val": active[1],
                "escalon_desde": str(active[2]),
            })
    return errors


# ── Validación 4: Fechas de término vencidas ────────────────────────────────

def _val4_terminos(rows, col_map, data_start, cierre_date):
    errors = []
    termino_col = col_map.get("Término del Contrato")
    arr_col = col_map.get("Arrendatario")
    if termino_col is None:
        return errors

    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        arr = str(row[arr_col] or "").strip() if arr_col is not None else ""
        if "vacante" in arr.lower():
            continue
        fecha_raw = row[termino_col]
        if fecha_raw is None:
            continue
        if isinstance(fecha_raw, datetime):
            fecha_raw = fecha_raw.date()
        if isinstance(fecha_raw, date) and fecha_raw < cierre_date:
            errors.append({
                "fila": row_idx + 1,
                "arrendatario": arr,
                "fecha_termino": str(fecha_raw),
            })
    return errors


# ── Función principal: validar un archivo ───────────────────────────────────

def _validar_archivo(filepath: str, cierre_date: date,
                     prev_filepath: Optional[str] = None) -> dict:
    """Corre las 4 validaciones sobre un archivo de RR."""
    result = {"archivo": os.path.basename(filepath), "errores": {}}

    rows, _ = _load_ws_rows(filepath, "Rent Roll")
    if rows is None:
        result["errores"]["lectura"] = "No se encontró la hoja 'Rent Roll'"
        return result

    h = _find_header_row(rows)
    if h is None:
        result["errores"]["lectura"] = "No se encontró fila de encabezados"
        return result

    col_map, fecha_pos = _get_col_map(rows, h)
    data_start = h + 1

    # Val 1
    v1 = _val1_vacantes(rows, col_map, data_start)
    if v1:
        result["errores"]["val1_vacantes"] = v1

    # Val 2 (solo si hay archivo previo)
    if prev_filepath:
        prev_rows, _ = _load_ws_rows(prev_filepath, "Rent Roll")
        if prev_rows:
            prev_h = _find_header_row(prev_rows)
            if prev_h is not None:
                prev_col_map, _ = _get_col_map(prev_rows, prev_h)
                rr_prev = _read_rr_locals(prev_rows, prev_col_map, prev_h + 1)
                rr_curr = _read_rr_locals(rows, col_map, data_start)

                abs_rows, _ = _load_ws_rows(filepath, "Absorción")
                absorcion = []
                if abs_rows:
                    abs_h = _find_header_row(abs_rows, "Status")
                    if abs_h is not None:
                        abs_col_map, _ = _get_col_map(abs_rows, abs_h)
                        absorcion = _read_absorcion(abs_rows, abs_col_map, abs_h + 1)

                v2 = _val2_absorcion(rr_prev, rr_curr, absorcion)
                if v2:
                    result["errores"]["val2_absorcion"] = v2

    # Val 3
    v3 = _val3_escalonada(rows, col_map, fecha_pos, data_start, cierre_date)
    if v3:
        result["errores"]["val3_escalonada"] = v3

    # Val 4
    v4 = _val4_terminos(rows, col_map, data_start, cierre_date)
    if v4:
        result["errores"]["val4_terminos"] = v4

    return result


# ── Herramienta pública: revisar_rent_rolls ──────────────────────────────────

def revisar_rent_rolls(año: int, mes: int) -> str:
    """
    Busca los archivos de Rent Roll del mes indicado en WORK_DIR,
    ejecuta las 4 validaciones y retorna el resumen de errores.
    También busca el archivo del mes anterior para la validación de absorción.
    """
    global _ultimo_resultado

    cierre = _cierre_mes(año, mes)
    mes_prev = mes - 1 if mes > 1 else 12
    año_prev = año if mes > 1 else año - 1

    resultado = {"año": año, "mes": mes, "cierre": str(cierre), "proveedores": {}}
    encontrados = []
    no_encontrados = []

    # ── JLL ──────────────────────────────────────────────────────────────────
    jll_path = _find_file(año, mes, "jll")
    if jll_path:
        encontrados.append(f"JLL: {os.path.basename(jll_path)}")
        resultado["proveedores"]["jll"] = _validar_archivo(jll_path, cierre)
    else:
        no_encontrados.append(f"JLL: {str(año)[2:]}{mes:02d} Rent Roll y NOI.xlsx")

    # ── Tres A Viña ───────────────────────────────────────────────────────────
    vina_path = _find_file(año, mes, "vina")
    vina_prev = _find_file(año_prev, mes_prev, "vina")
    if vina_path:
        encontrados.append(f"Viña: {os.path.basename(vina_path)}")
        resultado["proveedores"]["vina"] = _validar_archivo(vina_path, cierre, vina_prev)
    else:
        no_encontrados.append(f"Viña: Excel Tres A Viña {MESES_ES[mes]} {año}.xlsx")

    # ── Tres A Curicó ─────────────────────────────────────────────────────────
    curico_path = _find_file(año, mes, "curico")
    curico_prev = _find_file(año_prev, mes_prev, "curico")
    if curico_path:
        encontrados.append(f"Curicó: {os.path.basename(curico_path)}")
        resultado["proveedores"]["curico"] = _validar_archivo(curico_path, cierre, curico_prev)
    else:
        no_encontrados.append(f"Curicó: Excel Tres A Curicó {MESES_ES[mes]} {año}.xlsx")

    _ultimo_resultado = resultado

    # ── Formatear salida ──────────────────────────────────────────────────────
    lines = [f"Revisión Rent Roll — {MESES_ES[mes]} {año} (cierre {cierre})", ""]

    if no_encontrados:
        lines.append("⚠ Archivos NO encontrados en WORK_DIR:")
        for nf in no_encontrados:
            lines.append(f"  - {nf}")
        lines.append("")

    if encontrados:
        lines.append("Archivos revisados:")
        for e in encontrados:
            lines.append(f"  - {e}")
        lines.append("")

    for proveedor, res in resultado["proveedores"].items():
        label = {"jll": "JLL (Nicole)", "vina": "Tres A Viña (Sebastián)",
                 "curico": "Tres A Curicó (Sebastián)"}.get(proveedor, proveedor)
        lines.append(f"── {label} ──────────────────────")
        errores = res.get("errores", {})
        if not errores:
            lines.append("  Sin errores.")
        else:
            if "lectura" in errores:
                lines.append(f"  ERROR lectura: {errores['lectura']}")

            if "val1_vacantes" in errores:
                lines.append(f"  VAL1 — Coherencia vacantes: {len(errores['val1_vacantes'])} error(es)")
                for e in errores["val1_vacantes"]:
                    vals = e["valores"]
                    lines.append(f"    Fila {e['fila']}: TA1={vals['Tipo Activo 1']} | TA3={vals['Tipo Activo 3']} | Arr={vals['Arrendatario']} | TipoArr={vals['Tipo Arrendatario']}")

            if "val2_absorcion" in errores:
                lines.append(f"  VAL2 — Absorción: {len(errores['val2_absorcion'])} movimiento(s) sin registro")
                for e in errores["val2_absorcion"]:
                    lines.append(f"    Local {e['local']} ({e['activo']}): [{e['anterior']}] → [{e['actual']}] — falta {e['movimiento_esperado']}")

            if "val3_escalonada" in errores:
                lines.append(f"  VAL3 — Renta escalonada: {len(errores['val3_escalonada'])} error(es)")
                for e in errores["val3_escalonada"]:
                    lines.append(f"    Fila {e['fila']} [{e['arrendatario']}]: Renta={e['renta_fija']} | Escalón {e['escalon_num']}={e['escalon_val']} (desde {e['escalon_desde']})")

            if "val4_terminos" in errores:
                lines.append(f"  VAL4 — Contratos vencidos: {len(errores['val4_terminos'])} error(es)")
                for e in errores["val4_terminos"]:
                    lines.append(f"    Fila {e['fila']} [{e['arrendatario']}]: venció el {e['fecha_termino']}")
        lines.append("")

    lines.append("Cuando confirmes los errores usa 'enviar_emails_rent_roll' para enviar los correos.")
    return "\n".join(lines)


# ── Consolidación de Absorción en CDG ────────────────────────────────────────

# Columnas a copiar de la Absorción del proveedor al CDG (por nombre)
_ABS_COLS = [
    "Activo", "Tipo Activo", "Status",
    "Arrendatario", "Nuevo Arrendatario",
    "Antes (UF)", "Hoy (UF)",            # solo JLL; None si no existe
    "Antes (UF/m2)", "Hoy (UF/m2)",
    "M2", "%",
    "Vencimiento", "Inicio Nuevo Contrato", "Nuevo Vencimiento",
]


def _abs_key(rec: dict):
    """Clave única para deduplicar entradas de Absorción."""
    def _d(v):
        if isinstance(v, datetime):
            return v.date()
        return v
    return (
        str(rec.get("Activo") or "").strip(),
        str(rec.get("Tipo Activo") or "").strip(),
        str(rec.get("Status") or "").strip(),
        str(rec.get("Arrendatario") or "").strip(),
        str(rec.get("Nuevo Arrendatario") or "").strip(),
        _d(rec.get("Vencimiento")),
        _d(rec.get("Inicio Nuevo Contrato")),
    )


def _read_abs_source(filepath: str, sheetname: str = "Absorción") -> list:
    """Lee hoja Absorción de un proveedor → lista de dicts."""
    rows, _ = _load_ws_rows(filepath, sheetname)
    if rows is None:
        return []
    h = _find_header_row(rows, "Status")
    if h is None:
        return []
    col_map, _ = _get_col_map(rows, h)
    records = []
    for row in rows[h + 1:]:
        status_idx = col_map.get("Status")
        if status_idx is None:
            continue
        status = str(row[status_idx] or "").strip()
        if not status or status == "Status":
            continue
        rec = {}
        for col_name in _ABS_COLS:
            idx = col_map.get(col_name)
            rec[col_name] = row[idx] if idx is not None and idx < len(row) else None
        records.append(rec)
    return records


def consolidar_absorcion(año: int, mes: int, nombre_cdg: str) -> str:
    """
    Sincroniza la hoja 'Absorcion' del CDG con las hojas Absorción de los
    proveedores (JLL y Tres A) del período indicado.

    Solo agrega entradas que aún no existen en el CDG (deduplicación por clave
    Activo+TipoActivo+Status+Arrendatario+NuevoArrendatario+Vencimiento+Inicio).
    Las nuevas filas se insertan al final del bloque del activo correspondiente.
    """
    from config import WORK_DIR

    cdg_path = os.path.join(WORK_DIR, nombre_cdg)
    if not os.path.exists(cdg_path):
        return f"Error: no se encontró '{nombre_cdg}' en WORK_DIR ({WORK_DIR})"

    # ── 1. Leer Absorción de proveedores ─────────────────────────────────────
    nuevas: list = []
    for proveedor in ("jll", "vina", "curico"):
        path = _find_file(año, mes, proveedor)
        if path:
            nuevas.extend(_read_abs_source(path))

    if not nuevas:
        return "No se encontraron archivos de RR con hoja Absorción."

    # ── 2. Abrir CDG y leer Absorción existente ───────────────────────────────
    wb_cdg = openpyxl.load_workbook(cdg_path)
    ws_abs = wb_cdg["Absorcion"]

    cdg_rows = list(ws_abs.iter_rows(values_only=True))
    h_cdg = _find_header_row(cdg_rows, "Status")
    if h_cdg is None:
        wb_cdg.close()
        return "Error: no se encontró encabezado en hoja Absorcion del CDG."

    cdg_col_map, _ = _get_col_map(cdg_rows, h_cdg)

    # Construir set de claves existentes
    existing_keys: set = set()
    status_col_cdg = cdg_col_map.get("Status")
    for row in cdg_rows[h_cdg + 1:]:
        if status_col_cdg is None:
            break
        status = str(row[status_col_cdg] or "").strip()
        if not status or status == "Status":
            continue
        rec = {col: (row[idx] if idx is not None and idx < len(row) else None)
               for col, idx in cdg_col_map.items()}
        existing_keys.add(_abs_key(rec))

    # ── 3. Filtrar entradas realmente nuevas ──────────────────────────────────
    to_add: dict = {}  # activo → list of records
    for rec in nuevas:
        k = _abs_key(rec)
        if k in existing_keys:
            continue
        activo = str(rec.get("Activo") or "").strip()
        if activo not in to_add:
            to_add[activo] = []
        to_add[activo].append(rec)
        existing_keys.add(k)

    if not to_add:
        wb_cdg.close()
        return "No hay entradas nuevas que agregar a Absorcion."

    # ── 4. Encontrar última fila de cada activo en CDG ────────────────────────
    activo_last_row: dict = {}
    header_excel_row = h_cdg + 1
    data_start = header_excel_row + 1
    activo_col_cdg = cdg_col_map.get("Activo")

    for excel_row in range(data_start, ws_abs.max_row + 1):
        if activo_col_cdg is None or status_col_cdg is None:
            break
        activo_val = str(ws_abs.cell(row=excel_row, column=activo_col_cdg + 1).value or "").strip()
        status_val = str(ws_abs.cell(row=excel_row, column=status_col_cdg + 1).value or "").strip()
        if activo_val and status_val and status_val != "Status":
            activo_last_row[activo_val] = excel_row

    # ── 5. Insertar nuevas filas al final del bloque de cada activo ───────────
    # Procesar en orden inverso de posición para no desplazar índices
    activos_ordenados = sorted(
        to_add.keys(),
        key=lambda a: activo_last_row.get(a, 0),
        reverse=True,
    )

    total_agregadas = 0
    no_mapeados = []
    resumen = []

    for activo in activos_ordenados:
        records = to_add[activo]
        last_row = activo_last_row.get(activo)
        if last_row is None:
            no_mapeados.append(f"{activo} ({len(records)} fila(s)) — activo no encontrado en CDG")
            continue

        insert_at = last_row + 1
        n = len(records)
        ws_abs.insert_rows(insert_at, n)

        for offset, rec in enumerate(records):
            excel_row = insert_at + offset
            for col_name in _ABS_COLS:
                cdg_col_idx = cdg_col_map.get(col_name)
                if cdg_col_idx is None:
                    continue
                val = rec.get(col_name)
                ws_abs.cell(row=excel_row, column=cdg_col_idx + 1).value = val

        total_agregadas += n
        resumen.append(f"  {activo}: +{n} fila(s)")

    # ── 6. Guardar ────────────────────────────────────────────────────────────
    wb_cdg.save(cdg_path)
    wb_cdg.close()

    lines = [f"Consolidación Absorción — {MESES_ES[mes]} {año}",
             f"  Filas nuevas agregadas: {total_agregadas}"]
    lines.extend(resumen)
    if no_mapeados:
        lines.append("  ⚠ Sin sección en CDG (no agregados):")
        for nm in no_mapeados:
            lines.append(f"    - {nm}")
    return "\n".join(lines)


# ── Herramienta pública: enviar_emails_rent_roll ─────────────────────────────

def _tabla_val3(errores):
    """Genera texto con tabla de errores de renta escalonada."""
    lines = []
    for e in errores:
        lines.append(
            f"    Fila {e['fila']} [{e['arrendatario']}]: "
            f"Renta Fija registrada = {e['renta_fija']} UF/m2/mes | "
            f"Escalón {e['escalon_num']} esperado = {e['escalon_val']} (activo desde {e['escalon_desde']})"
        )
    return "\n".join(lines)


def enviar_emails_rent_roll() -> str:
    """
    Envía los correos a Nicole (JLL) y Sebastián (Tres A) con los errores
    encontrados en la última revisión de Rent Rolls.
    """
    global _ultimo_resultado

    if not _ultimo_resultado:
        return "No hay resultados de revisión. Ejecuta primero 'revisar_rent_rolls'."

    año = _ultimo_resultado["año"]
    mes = _ultimo_resultado["mes"]
    mes_nombre = MESES_ES[mes]
    proveedores = _ultimo_resultado.get("proveedores", {})
    enviados = []

    # ── Correo a Nicole (JLL) ────────────────────────────────────────────────
    jll = proveedores.get("jll", {})
    jll_errores = jll.get("errores", {})
    if jll_errores and "lectura" not in jll_errores:
        aamm = f"{str(año)[2:]}{mes:02d}"
        asunto = f"Revisión Rent Roll y NOI — {aamm} ({mes_nombre} {año})"

        cuerpo_parts = [f"Nicole,\n\nRevisando la planilla {aamm} me encontré con unas cosas que te quería consultar:\n"]

        if "val1_vacantes" in jll_errores:
            cuerpo_parts.append("1. Inconsistencia en columna Vacante:")
            for e in jll_errores["val1_vacantes"]:
                v = e["valores"]
                cuerpo_parts.append(
                    f"   Fila {e['fila']}: Tipo Activo 1={v['Tipo Activo 1']} | "
                    f"Tipo Activo 3={v['Tipo Activo 3']} | "
                    f"Arrendatario={v['Arrendatario']} | "
                    f"Tipo Arrendatario={v['Tipo Arrendatario']}\n"
                    f"   Si el espacio está vacante, debería marcarse como 'Vacante' en todas esas columnas. ¿Podrías revisar?"
                )
            cuerpo_parts.append("")

        if "val2_absorcion" in jll_errores:
            cuerpo_parts.append("2. Movimientos sin registro en Absorción:")
            for e in jll_errores["val2_absorcion"]:
                cuerpo_parts.append(
                    f"   Local {e['local']} ({e['activo']}): "
                    f"pasó de [{e['anterior']}] a [{e['actual']}] "
                    f"pero no aparece el '{e['movimiento_esperado']}' en la hoja Absorción."
                )
            cuerpo_parts.append("")

        if "val3_escalonada" in jll_errores:
            n = len(jll_errores["val3_escalonada"])
            cuerpo_parts.append(f"3. Renta escalonada ({n} caso(s)):")
            cuerpo_parts.append("   Los siguientes arrendatarios tienen escalones que ya deberían estar vigentes, pero la Renta Fija registrada no coincide:\n")
            cuerpo_parts.append(_tabla_val3(jll_errores["val3_escalonada"]))
            cuerpo_parts.append("\n   ¿Podrías revisar si hay una actualización pendiente?")
            cuerpo_parts.append("")

        if "val4_terminos" in jll_errores:
            cuerpo_parts.append("4. Contratos con fecha de término vencida:")
            for e in jll_errores["val4_terminos"]:
                cuerpo_parts.append(
                    f"   Fila {e['fila']} [{e['arrendatario']}]: venció el {e['fecha_termino']}."
                )
            cuerpo_parts.append("\n   ¿Estos contratos están renovados y pendiente de actualizar la fecha, o ya terminaron?")
            cuerpo_parts.append("")

        cuerpo_parts.append("Gracias,\nRaimundo")
        cuerpo = "\n".join(cuerpo_parts)

        resultado = send_email(NICOLE_EMAIL, asunto, cuerpo)
        enviados.append(f"Nicole ({NICOLE_EMAIL}): {resultado}")

    # ── Correo a Sebastián (Tres A) ──────────────────────────────────────────
    # Junta errores de Viña y Curicó
    seb_errores_vina = proveedores.get("vina", {}).get("errores", {})
    seb_errores_curico = proveedores.get("curico", {}).get("errores", {})

    hay_errores_seb = any(
        k != "lectura" for k in list(seb_errores_vina.keys()) + list(seb_errores_curico.keys())
    )

    if hay_errores_seb:
        asunto = f"Revisión Rent Roll — {mes_nombre} {año}"
        cuerpo_parts = [f"Sebastián,\n\nRevisando las planillas de {mes_nombre} me encontré con unas inconsistencias que te quería comentar:\n"]

        for activo, errores in [("Viña Centro", seb_errores_vina), ("Curicó", seb_errores_curico)]:
            if not errores or "lectura" in errores:
                continue
            cuerpo_parts.append(f"── {activo} ──")

            if "val1_vacantes" in errores:
                cuerpo_parts.append("  Inconsistencia en columna Vacante:")
                for e in errores["val1_vacantes"]:
                    v = e["valores"]
                    cuerpo_parts.append(
                        f"    Fila {e['fila']}: TA1={v['Tipo Activo 1']} | TA3={v['Tipo Activo 3']} | "
                        f"Arr={v['Arrendatario']} | TipoArr={v['Tipo Arrendatario']}"
                    )
                cuerpo_parts.append("")

            if "val2_absorcion" in errores:
                cuerpo_parts.append("  Movimientos sin registro en Absorción:")
                for e in errores["val2_absorcion"]:
                    cuerpo_parts.append(
                        f"    Local {e['local']}: [{e['anterior']}] → [{e['actual']}] — falta '{e['movimiento_esperado']}'"
                    )
                cuerpo_parts.append("")

            if "val3_escalonada" in errores:
                n = len(errores["val3_escalonada"])
                cuerpo_parts.append(f"  Renta escalonada ({n} caso(s)):")
                cuerpo_parts.append(_tabla_val3(errores["val3_escalonada"]))
                cuerpo_parts.append("")

            if "val4_terminos" in errores:
                cuerpo_parts.append("  Contratos con fecha de término vencida:")
                for e in errores["val4_terminos"]:
                    cuerpo_parts.append(f"    Fila {e['fila']} [{e['arrendatario']}]: venció el {e['fecha_termino']}")
                cuerpo_parts.append("")

        cuerpo_parts.append("¿Podrías revisar y enviarnos las planillas actualizadas?\n\nGracias,\nRaimundo")
        cuerpo = "\n".join(cuerpo_parts)

        resultado = send_email(SEBASTIAN_EMAIL, asunto, cuerpo)
        enviados.append(f"Sebastián ({SEBASTIAN_EMAIL}): {resultado}")

    if not enviados:
        return "No se encontraron errores que reportar. No se enviaron correos."

    return "Correos enviados:\n" + "\n".join(enviados)


# ── Consolidación de Rent Roll en CDG ────────────────────────────────────────

# Columnas a copiar (por nombre). El orden no importa; se buscan por header.
_COLS_COPY = [
    "Activo1", "Activo2",
    "Tipo Activo 1", "Tipo Activo 2 (no va vacante)", "Tipo Activo 3",
    "Arrendatario", "Tipo Arrendatario", "Rol", "Ex Arrendatario",
    "Detalle Activo", "Area Arrendable (m2)",
    "Renta Fija (UF/m2 /mes)",
    "IVA",                              # opcional; se omite si no existe
    "Inicio Pago Renta",
    "Fecha Inicio", "Término del Contrato",
]
# Escalones: pares (valor_col_name, siguiente_Fecha). Se manejan aparte.
_ESCALON_NAMES = [
    "1\n(UF/m2/mes)", "2\n(UF/m2/mes)", "3\n(UF/m2/mes)",
    "4\n(UF/m2/mes)", "5\n(UF/m2/mes)",
]


def _build_escalon_pairs(col_map: dict, fecha_positions: list) -> list:
    """
    Retorna lista de (nombre_escalon, idx_valor, idx_fecha) en orden.
    idx_fecha es la primera posición de 'Fecha' > idx_valor.
    """
    pairs = []
    for name in _ESCALON_NAMES:
        val_idx = col_map.get(name)
        if val_idx is None:
            continue
        fecha_idx = next((f for f in sorted(fecha_positions) if f > val_idx), None)
        pairs.append((name, val_idx, fecha_idx))
    return pairs


def _read_source_data(filepath: str) -> dict:
    """
    Lee la hoja Rent Roll de un archivo proveedor y devuelve
    {(activo2, detalle): {col_name: value, ...}} con todas las columnas a copiar.
    """
    rows, _ = _load_ws_rows(filepath, "Rent Roll")
    if rows is None:
        return {}

    h = _find_header_row(rows)
    if h is None:
        return {}

    col_map, fecha_positions = _get_col_map(rows, h)
    escalon_pairs = _build_escalon_pairs(col_map, fecha_positions)

    # Columnas escalón: necesito los índices de Fecha por escalón
    # Ya tenidos en escalon_pairs como (name, val_idx, fecha_idx)

    source = {}
    activo2_col = col_map.get("Activo2")
    detalle_col = col_map.get("Detalle Activo")
    if activo2_col is None or detalle_col is None:
        return {}

    for row in rows[h + 1:]:
        activo2 = str(row[activo2_col] or "").strip()
        detalle = str(row[detalle_col] or "").strip()
        if not activo2 or not detalle:
            continue

        key = (activo2, detalle)
        record = {}

        # Columnas simples
        for col_name in _COLS_COPY:
            idx = col_map.get(col_name)
            if idx is not None and idx < len(row):
                record[col_name] = row[idx]

        # Escalones (pares valor + fecha)
        for (name, val_idx, fecha_idx) in escalon_pairs:
            record[name] = row[val_idx] if val_idx < len(row) else None
            if fecha_idx is not None:
                record[f"__fecha_{name}"] = row[fecha_idx] if fecha_idx < len(row) else None

        source[key] = record

    return source


def consolidar_rent_rolls(año: int, mes: int, nombre_cdg: str) -> str:
    """
    Consolida los datos de los Rent Rolls de proveedores (JLL y Tres A) en la
    hoja 'Rent Roll' del CDG.

    Matching por (Activo2, Detalle Activo). Solo actualiza las celdas de las
    columnas especificadas; no mueve filas ni toca columnas calculadas.

    Parámetros:
        año, mes : período de los RR a usar
        nombre_cdg : nombre del archivo CDG en WORK_DIR (ej: '2601 CDG.xlsx')
    """
    from config import WORK_DIR

    cdg_path = os.path.join(WORK_DIR, nombre_cdg)
    if not os.path.exists(cdg_path):
        return f"Error: no se encontró '{nombre_cdg}' en WORK_DIR ({WORK_DIR})"

    # ── 1. Leer datos de proveedores ─────────────────────────────────────────
    source_data: dict = {}   # (activo2, detalle) → record

    for proveedor in ("jll", "vina", "curico"):
        path = _find_file(año, mes, proveedor)
        if path:
            data = _read_source_data(path)
            conflicts = set(source_data.keys()) & set(data.keys())
            if conflicts:
                return (
                    f"Error: clave duplicada entre proveedores para {conflicts}. "
                    "Revisar archivos."
                )
            source_data.update(data)

    if not source_data:
        return "No se encontraron archivos de RR en WORK_DIR para el período indicado."

    # ── 2. Abrir CDG y mapear headers ────────────────────────────────────────
    wb_cdg = openpyxl.load_workbook(cdg_path)
    ws_cdg = wb_cdg["Rent Roll"]

    # Encontrar fila de header en CDG
    cdg_rows_iter = ws_cdg.iter_rows(values_only=True)
    cdg_rows = list(cdg_rows_iter)
    h_cdg = _find_header_row(cdg_rows)
    if h_cdg is None:
        wb_cdg.close()
        return "Error: no se encontró la fila de encabezados en la hoja Rent Roll del CDG."

    cdg_col_map, cdg_fecha_pos = _get_col_map(cdg_rows, h_cdg)
    cdg_escalon_pairs = _build_escalon_pairs(cdg_col_map, cdg_fecha_pos)

    # Construir mapa escalon_name → (idx_valor_cdg, idx_fecha_cdg)
    cdg_escalon_map = {name: (vi, fi) for name, vi, fi in cdg_escalon_pairs}

    activo2_cdg = cdg_col_map.get("Activo2")
    detalle_cdg = cdg_col_map.get("Detalle Activo")
    if activo2_cdg is None or detalle_cdg is None:
        wb_cdg.close()
        return "Error: no se encontraron columnas Activo2 o Detalle Activo en CDG."

    # ── 3. Actualizar celdas ─────────────────────────────────────────────────
    actualizados = 0
    no_encontrados = []   # claves en proveedor que no están en CDG
    sin_proveedor = []    # claves en CDG sin dato de proveedor (solo para info)

    # Iterar filas del CDG (empezando después del header)
    header_excel_row = h_cdg + 1  # 1-based row en Excel
    data_start_excel = header_excel_row + 1

    for excel_row in range(data_start_excel, ws_cdg.max_row + 1):
        activo2_val = ws_cdg.cell(row=excel_row, column=activo2_cdg + 1).value
        detalle_val = ws_cdg.cell(row=excel_row, column=detalle_cdg + 1).value

        activo2_val = str(activo2_val or "").strip()
        detalle_val = str(detalle_val or "").strip()

        if not activo2_val or not detalle_val:
            continue

        key = (activo2_val, detalle_val)
        if key not in source_data:
            continue  # fila del CDG no cubierta por ningún proveedor (ej: otro activo)

        record = source_data[key]

        # Columnas simples
        for col_name in _COLS_COPY:
            if col_name not in record:
                continue  # columna no existe en fuente (ej: IVA en Tres A)
            cdg_col_idx = cdg_col_map.get(col_name)
            if cdg_col_idx is None:
                continue
            new_val = record[col_name]
            ws_cdg.cell(row=excel_row, column=cdg_col_idx + 1).value = new_val

        # Escalones
        for esc_name in _ESCALON_NAMES:
            if esc_name not in record:
                continue
            if esc_name not in cdg_escalon_map:
                continue
            val_cdg_idx, fecha_cdg_idx = cdg_escalon_map[esc_name]
            ws_cdg.cell(row=excel_row, column=val_cdg_idx + 1).value = record[esc_name]
            if fecha_cdg_idx is not None:
                fecha_key = f"__fecha_{esc_name}"
                ws_cdg.cell(row=excel_row, column=fecha_cdg_idx + 1).value = record.get(fecha_key)

        actualizados += 1

    # Verificar que todos los registros del proveedor se mapearon
    cdg_keys = set()
    for excel_row in range(data_start_excel, ws_cdg.max_row + 1):
        a2 = str(ws_cdg.cell(row=excel_row, column=activo2_cdg + 1).value or "").strip()
        dt = str(ws_cdg.cell(row=excel_row, column=detalle_cdg + 1).value or "").strip()
        if a2 and dt:
            cdg_keys.add((a2, dt))

    for key in source_data:
        if key not in cdg_keys:
            no_encontrados.append(f"({key[0]}, detalle={key[1]})")

    # ── 4. Guardar ────────────────────────────────────────────────────────────
    wb_cdg.save(cdg_path)
    wb_cdg.close()

    lines = [
        f"Consolidación Rent Roll — {MESES_ES[mes]} {año}",
        f"  Filas actualizadas en CDG: {actualizados}",
    ]
    if no_encontrados:
        lines.append(
            f"  ⚠ {len(no_encontrados)} fila(s) del proveedor NO encontradas en CDG "
            "(podrían ser locales nuevos):"
        )
        for nf in no_encontrados[:20]:
            lines.append(f"    - {nf}")
    else:
        lines.append("  Todas las filas del proveedor encontradas en CDG.")

    return "\n".join(lines)
