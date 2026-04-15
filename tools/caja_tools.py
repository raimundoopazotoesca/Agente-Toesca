"""
Herramientas para actualizar la hoja 'Caja' del Control de Gestión.

Fuente de datos: "Saldo Caja + FFMM Inmobiliario" (enviado por María José Castro
todos los lunes por correo).

Flujo:
  1. listar_hojas_saldo_caja(archivo_saldo)          → ver hojas disponibles con fechas
  2. copiar_datos_saldo_caja(archivo_cg, archivo_saldo, nombre_hoja)
                                                     → pega A:I en hoja Caja del CDG
  3. [Abrir CDG en Excel y guardar para que R5/R22/R26 recalculen]
  4. leer_celdas_caja(archivo_cg)                    → lee R5, R22, R26 (valores cacheados)
  5. inspeccionar_caja_historica(archivo_cg)         → muestra estructura de tabla histórica
  6. agregar_fila_caja_historica(archivo_cg, año, mes, col_fecha, col_r5, col_r22, col_r26)
                                                     → añade fila al histórico

Sobre limpieza de números:
  El archivo Saldo Caja a veces trae valores como texto "1.234.567" (puntos como
  separadores de miles). Se detectan y convierten automáticamente.
"""
import os
import re
import zipfile
from calendar import monthrange
from datetime import date, timedelta
from config import WORK_DIR


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _resolve_path(nombre_archivo: str) -> str:
    if os.path.isabs(nombre_archivo):
        return nombre_archivo
    return os.path.join(WORK_DIR, nombre_archivo)


def _excel_date(d: date) -> int:
    return (d - date(1899, 12, 30)).days


def _last_day(year: int, month: int) -> date:
    return date(year, month, monthrange(year, month)[1])


def _col_num(letter: str) -> int:
    n = 0
    for c in letter.upper():
        n = n * 26 + ord(c) - ord("A") + 1
    return n


def _xml_escape(s: str) -> str:
    return (s.replace("&", "&amp;")
             .replace("<", "&lt;")
             .replace(">", "&gt;"))


def _find_cell_bounds(xml: str, ref: str) -> tuple:
    """Retorna (start, end) del span completo de la celda."""
    tag = f'<c r="{ref}"'
    start = xml.find(tag)
    if start == -1:
        return -1, -1
    i = start + len(tag)
    while i < len(xml):
        if xml[i:i+2] == '/>':
            return start, i + 2
        if xml[i] == '>':
            end = xml.find('</c>', i)
            return start, (end + 4) if end != -1 else i + 1
        i += 1
    return start, len(xml)


def _get_cell_attr(xml: str, ref: str, attr: str, default: str) -> str:
    tag = f'<c r="{ref}"'
    start = xml.find(tag)
    if start == -1:
        return default
    snippet = xml[start: min(start + 300, len(xml))]
    m = re.search(rf'\b{attr}="([^"]+)"', snippet)
    return m.group(1) if m else default


def _read_cell_cached(xml: str, ref: str, shared_strings: list) -> tuple:
    """
    Lee el valor cacheado de una celda.
    Retorna (valor_python, tipo) donde tipo es 'num', 'str', 'empty'.
    """
    start, end = _find_cell_bounds(xml, ref)
    if start == -1:
        return None, 'empty'
    cell_xml = xml[start:end]
    v_m = re.search(r'<v>([^<]+)</v>', cell_xml)
    if not v_m:
        return None, 'empty'
    raw = v_m.group(1)
    if 't="s"' in cell_xml:
        idx = int(raw)
        val = shared_strings[idx] if idx < len(shared_strings) else raw
        return val, 'str'
    try:
        return float(raw), 'num'
    except ValueError:
        return raw, 'str'


def _replace_or_insert_cell(row_xml: str, ref: str, new_cell: str) -> str:
    start, end = _find_cell_bounds(row_xml, ref)
    if start != -1:
        return row_xml[:start] + new_cell + row_xml[end:]
    col = re.sub(r"\d", "", ref)
    col_n = _col_num(col)
    for m in re.finditer(r'<c r="([A-Z]+)\d+"', row_xml):
        if _col_num(m.group(1)) > col_n:
            return row_xml[:m.start()] + new_cell + row_xml[m.start():]
    return row_xml + new_cell


def _write_cells_in_row(sheet_xml: str, ss_xml: str, row_num: int,
                        cells: dict) -> tuple:
    """
    Escribe varias celdas en una fila.
    cells: {col_letter: value} donde value puede ser float/int o str.
    Retorna (sheet_xml_actualizado, ss_xml_actualizado).
    """
    row_pat = rf'(<row\b[^>]*\br="{row_num}"[^>]*>)(.*?)(</row>)'
    row_m = re.search(row_pat, sheet_xml, re.DOTALL)

    if row_m:
        row_open = row_m.group(1)
        row_content = row_m.group(2)
    else:
        row_open = f'<row r="{row_num}" spans="1:18">'
        row_content = ""

    rc = row_content
    for col, val in cells.items():
        ref = f"{col}{row_num}"
        style = _get_cell_attr(rc, ref, "s", "0")
        if val is None:
            # Dejar celda vacía (self-closing)
            new_cell = f'<c r="{ref}" s="{style}"/>'
        elif isinstance(val, str):
            # Buscar/agregar shared string
            sis = list(re.finditer(r"<si>(.*?)</si>", ss_xml, re.DOTALL))
            idx = None
            for i_si, si_m in enumerate(sis):
                t = re.search(r"<t[^>]*>([^<]*)</t>", si_m.group(1))
                if t and t.group(1) == val:
                    idx = i_si
                    break
            if idx is None:
                idx = len(sis)
                new_si = f'<si><t xml:space="preserve">{_xml_escape(val)}</t></si>'
                ss_xml = ss_xml.replace("</sst>", new_si + "</sst>")
                ss_xml = re.sub(
                    r'(count|uniqueCount)="(\d+)"',
                    lambda x: f'{x.group(1)}="{int(x.group(2)) + 1}"',
                    ss_xml, count=2,
                )
            new_cell = f'<c r="{ref}" s="{style}" t="s"><v>{idx}</v></c>'
        else:
            # Numérico
            new_cell = f'<c r="{ref}" s="{style}"><v>{val}</v></c>'
        rc = _replace_or_insert_cell(rc, ref, new_cell)

    new_row = row_open + rc + "</row>"
    if row_m:
        return sheet_xml[:row_m.start()] + new_row + sheet_xml[row_m.end():], ss_xml
    else:
        next_row_m = re.search(
            rf'<row\b[^>]*\br="({row_num + 1}|{row_num + 2}|{row_num + 3})"',
            sheet_xml,
        )
        if next_row_m:
            return sheet_xml[:next_row_m.start()] + new_row + sheet_xml[next_row_m.start():], ss_xml
        return sheet_xml.replace("</sheetData>", new_row + "</sheetData>"), ss_xml


def _apply_to_xlsx(filepath: str, modifications: dict) -> None:
    tmp = filepath + ".tmp"
    with zipfile.ZipFile(filepath, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = modifications.get(item.filename)
                if data is not None:
                    zout.writestr(item, data if isinstance(data, bytes) else data.encode("utf-8"))
                else:
                    zout.writestr(item, zin.read(item.filename))
    os.replace(tmp, filepath)


def _find_sheet_xml_path(xlsx_path: str, sheet_name: str) -> str | None:
    """Busca la ruta XML interna de una hoja por nombre."""
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
        m = re.search(
            r'<sheet\b(?=[^>]*\bname="' + re.escape(sheet_name) + r'")[^>]*\br:id="([^"]+)"',
            wb_xml,
        )
        if not m:
            m = re.search(
                r'<sheet\b[^>]*\bname="' + re.escape(sheet_name) + r'"[^>]*\br:id="([^"]+)"',
                wb_xml,
            )
        if not m:
            return None
        rid = m.group(1)
        rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        m2 = re.search(
            r'<Relationship\b(?=[^>]*\bId="' + re.escape(rid) + r'")[^>]*\bTarget="([^"]+)"',
            rels_xml,
        )
        if not m2:
            return None
        target = m2.group(1)
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        return target


def _leer_shared_strings(zf) -> list:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    ss_xml = zf.read("xl/sharedStrings.xml").decode("utf-8")
    result = []
    for si_m in re.finditer(r"<si>(.*?)</si>", ss_xml, re.DOTALL):
        t = re.search(r"<t[^>]*>([^<]*)</t>", si_m.group(1))
        result.append(t.group(1) if t else "")
    return result


def _limpiar_numero(val) -> float | None:
    """
    Convierte un valor a float limpio.
    Maneja:
      - int/float:    retorna directamente
      - "1.234.567"  (puntos = separadores de miles) → 1234567.0
      - "1.234,56"   (punto = miles, coma = decimal) → 1234.56
      - "1,234,567"  (comas = miles) → 1234567.0
      - "1234567"    → 1234567.0
    Si no es numérico retorna None.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if not isinstance(val, str):
        return None
    s = val.strip().replace(" ", "").replace("$", "").replace("%", "")
    if not s:
        return None
    # Si tiene coma: puede ser "1.234,56" o "1,234,567"
    if "," in s and "." in s:
        # Determinar cuál es el separador decimal
        last_comma = s.rfind(",")
        last_dot = s.rfind(".")
        if last_comma > last_dot:
            # Coma es decimal: "1.234,56"
            s = s.replace(".", "").replace(",", ".")
        else:
            # Punto es decimal: "1,234.56"
            s = s.replace(",", "")
    elif "," in s:
        # Solo comas: puede ser "1,234,567" (miles) o "1,56" (decimal)
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            # Probablemente decimal: "1,56"
            s = s.replace(",", ".")
        else:
            # Comas como miles: "1,234,567"
            s = s.replace(",", "")
    elif "." in s:
        # Solo puntos
        parts = s.split(".")
        if len(parts) == 2 and len(parts[1]) <= 2:
            # Decimal único: "1234.56" o "1234.5"
            pass  # ya está bien
        else:
            # Múltiples puntos o 3 decimales → miles: "1.234.567" o "1.234"
            s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


# ─── Herramientas públicas ────────────────────────────────────────────────────

def listar_hojas_saldo_caja(archivo_saldo_caja: str) -> str:
    """
    Lista todas las hojas del archivo Saldo Caja + FFMM Inmobiliario,
    mostrando sus nombres (que son fechas) para elegir la más apropiada.
    """
    filepath = _resolve_path(archivo_saldo_caja)
    if not os.path.exists(filepath):
        return f"Error: no se encontró '{filepath}'."
    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
        hojas = re.findall(r'<sheet\b[^>]*\bname="([^"]+)"', wb_xml)
        if not hojas:
            return "No se encontraron hojas en el archivo."
        lines = [f"Hojas disponibles en '{archivo_saldo_caja}':"]
        for h in hojas:
            lines.append(f"  - {h}")
        lines.append(
            "\nTip: elegir la hoja cuya fecha sea la más cercana al mes del CDG. "
            "Ej: para CDG 2601 (ene 2026), usar hoja con fecha ≈ 02/02/2026."
        )
        return "\n".join(lines)
    except Exception as e:
        return f"Error al leer hojas: {e}"


def copiar_datos_saldo_caja(
    archivo_cg: str,
    archivo_saldo_caja: str,
    nombre_hoja: str,
) -> str:
    """
    Copia las columnas A:I de la hoja indicada del archivo Saldo Caja
    a las columnas A:I de la hoja 'Caja' en el CDG.

    Solo se copian filas que tienen algún valor en las primeras 9 columnas.
    Los valores numéricos almacenados como texto (ej: "1.234.567") se limpian
    automáticamente.

    IMPORTANTE: Después de copiar, abrir el CDG en Excel y guardar para que
    las fórmulas en R5, R22 y R26 recalculen con los nuevos datos.
    """
    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl no instalado."

    cg_path    = _resolve_path(archivo_cg)
    saldo_path = _resolve_path(archivo_saldo_caja)

    if not os.path.exists(cg_path):
        return f"Error: no se encontró '{cg_path}'."
    if not os.path.exists(saldo_path):
        return f"Error: no se encontró '{saldo_path}'."

    # ── 1. Leer datos de Saldo Caja ──
    try:
        wb_saldo = openpyxl.load_workbook(saldo_path, read_only=True, data_only=True)
        if nombre_hoja not in wb_saldo.sheetnames:
            hojas = ", ".join(wb_saldo.sheetnames)
            wb_saldo.close()
            return f"Error: hoja '{nombre_hoja}' no existe. Disponibles: {hojas}"
        ws = wb_saldo[nombre_hoja]

        columnas = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
        filas_data = []  # list of (row_num, {col: cleaned_val})

        for row in ws.iter_rows(max_col=9):
            row_vals = [cell.value for cell in row]
            if all(v is None for v in row_vals):
                continue  # Fila completamente vacía, omitir

            row_idx = row[0].row
            cells = {}
            for i, col in enumerate(columnas):
                raw = row_vals[i]
                # Intentar limpiar como número primero
                num = _limpiar_numero(raw)
                if num is not None and isinstance(raw, str) and num != raw:
                    cells[col] = num  # era texto con número, limpiar
                elif isinstance(raw, str):
                    cells[col] = raw  # texto genuino
                elif raw is not None:
                    cells[col] = float(raw) if isinstance(raw, (int, float)) else raw
                else:
                    cells[col] = None
            filas_data.append((row_idx, cells))

        wb_saldo.close()
    except Exception as e:
        return f"Error al leer '{archivo_saldo_caja}': {e}"

    if not filas_data:
        return f"No hay datos en columnas A:I de la hoja '{nombre_hoja}'."

    # ── 2. Encontrar hoja Caja en CDG ──
    caja_path = _find_sheet_xml_path(cg_path, "Caja")
    if not caja_path:
        return "Error: no se encontró la hoja 'Caja' en el CDG."

    # ── 3. Escribir en CDG vía XML ──
    try:
        with zipfile.ZipFile(cg_path, "r") as zf:
            sheet_xml = zf.read(caja_path).decode("utf-8")
            ss_xml = zf.read("xl/sharedStrings.xml").decode("utf-8") \
                if "xl/sharedStrings.xml" in zf.namelist() else "<sst></sst>"

        advertencias = []
        for row_num, cells in filas_data:
            sheet_xml, ss_xml = _write_cells_in_row(sheet_xml, ss_xml, row_num, cells)

        mods = {caja_path: sheet_xml, "xl/sharedStrings.xml": ss_xml}
        _apply_to_xlsx(cg_path, mods)

        # Reportar si hubo valores de texto que se limpiaron
        convertidos = [
            f"  Fila {r}: {sum(1 for v in c.values() if v is not None)} celdas"
            for r, c in filas_data
        ]
        lines = [
            f"OK: {len(filas_data)} filas copiadas de '{nombre_hoja}' → hoja 'Caja' del CDG.",
            f"Rango escrito: A1:I{filas_data[-1][0]}",
            "",
            "SIGUIENTE PASO: Abrir el CDG en Excel y guardar (Ctrl+S) para que",
            "las celdas R5, R22 y R26 recalculen sus formulas con los nuevos datos.",
        ]
        if advertencias:
            lines.append("\nAdvertencias:")
            lines.extend(advertencias)
        return "\n".join(lines)
    except Exception as e:
        return f"Error al escribir en CDG: {e}"


def leer_celdas_caja(archivo_cg: str) -> str:
    """
    Lee los valores cacheados de las celdas R5, R22 y R26 de la hoja 'Caja'.
    Estos valores son correctos solo después de haber abierto y guardado el
    CDG en Excel (para que recalculen las fórmulas).

    Retorna los valores y una línea lista para copiar a la tabla histórica.
    """
    cg_path = _resolve_path(archivo_cg)
    if not os.path.exists(cg_path):
        return f"Error: no se encontró '{cg_path}'."

    caja_path = _find_sheet_xml_path(cg_path, "Caja")
    if not caja_path:
        return "Error: no se encontró la hoja 'Caja' en el CDG."

    try:
        with zipfile.ZipFile(cg_path, "r") as zf:
            sheet_xml = zf.read(caja_path).decode("utf-8")
            strings = _leer_shared_strings(zf)

        resultados = {}
        for celda in ("R5", "R22", "R26"):
            val, tipo = _read_cell_cached(sheet_xml, celda, strings)
            resultados[celda] = (val, tipo)

        lines = ["Valores en hoja 'Caja':"]
        for ref, (val, tipo) in resultados.items():
            if val is None:
                lines.append(f"  {ref}: [vacío — Excel debe recalcular]")
            elif tipo == 'num':
                lines.append(f"  {ref}: {val:,.2f}")
            else:
                lines.append(f"  {ref}: {val}")

        all_values = [v for v, t in resultados.values() if v is not None]
        if len(all_values) == 3:
            lines.append(
                f"\nValores para tabla histórica (R5 | R22 | R26):\n"
                f"  {resultados['R5'][0]:,.2f} | {resultados['R22'][0]:,.2f} | {resultados['R26'][0]:,.2f}"
            )
        return "\n".join(lines)
    except Exception as e:
        return f"Error al leer celdas Caja: {e}"


def inspeccionar_caja_historica(archivo_cg: str) -> str:
    """
    Muestra el contenido de las filas 28–36 de la hoja 'Caja' para identificar
    la estructura de la tabla histórica (cabeceras, columnas, última fila con datos).
    Ejecutar antes de agregar_fila_caja_historica para saber qué columnas usar.
    """
    cg_path = _resolve_path(archivo_cg)
    if not os.path.exists(cg_path):
        return f"Error: no se encontró '{cg_path}'."

    caja_path = _find_sheet_xml_path(cg_path, "Caja")
    if not caja_path:
        return "Error: no se encontró la hoja 'Caja' en el CDG."

    try:
        with zipfile.ZipFile(cg_path, "r") as zf:
            sheet_xml = zf.read(caja_path).decode("utf-8")
            strings = _leer_shared_strings(zf)

        lines = ["Contenido de filas 28-40 (hoja 'Caja'):"]
        cols = [chr(c) for c in range(ord('A'), ord('V') + 1)]  # A..U

        for row_num in range(28, 41):
            row_cells = {}
            for col in cols:
                ref = f"{col}{row_num}"
                val, tipo = _read_cell_cached(sheet_xml, ref, strings)
                if val is not None:
                    row_cells[col] = val

            if row_cells:
                cell_strs = [f"{col}={v!r}" for col, v in sorted(row_cells.items(),
                                                                    key=lambda x: _col_num(x[0]))]
                lines.append(f"  Fila {row_num}: {', '.join(cell_strs)}")
            else:
                lines.append(f"  Fila {row_num}: [vacía]")

        lines.append(
            "\nIdentificar:\n"
            "  - La fila de cabecera (ej: fila 31) con etiquetas como 'Fecha', 'Saldo', etc.\n"
            "  - Las columnas donde van la fecha y los valores de R5, R22, R26.\n"
            "  - La última fila con datos para saber dónde insertar la nueva."
        )
        return "\n".join(lines)
    except Exception as e:
        return f"Error al inspeccionar Caja histórica: {e}"


def agregar_fila_caja_historica(
    archivo_cg: str,
    año: int,
    mes: int,
    col_fecha: str,
    col_r5: str,
    col_r22: str,
    col_r26: str,
    valor_r5: float,
    valor_r22: float,
    valor_r26: float,
    fila_inicio_datos: int = 32,
) -> str:
    """
    Añade una nueva fila a la tabla Caja Histórica.

    La fecha se calcula como el último día del mes indicado.

    Args:
        archivo_cg:        Archivo CDG en WORK_DIR.
        año, mes:          Mes del CDG (ej: 2026, 1 para enero 2026).
        col_fecha:         Columna donde va la fecha (ej: 'A').
        col_r5:            Columna del valor R5 (ej: 'B').
        col_r22:           Columna del valor R22 (ej: 'C').
        col_r26:           Columna del valor R26 (ej: 'D').
        valor_r5:          Valor numérico de celda R5.
        valor_r22:         Valor numérico de celda R22.
        valor_r26:         Valor numérico de celda R26.
        fila_inicio_datos: Primera fila de datos del histórico (default: 32,
                           asumiendo que fila 31 es cabecera).
    """
    cg_path = _resolve_path(archivo_cg)
    if not os.path.exists(cg_path):
        return f"Error: no se encontró '{cg_path}'."

    caja_path = _find_sheet_xml_path(cg_path, "Caja")
    if not caja_path:
        return "Error: no se encontró la hoja 'Caja' en el CDG."

    try:
        with zipfile.ZipFile(cg_path, "r") as zf:
            sheet_xml = zf.read(caja_path).decode("utf-8")
            ss_xml = zf.read("xl/sharedStrings.xml").decode("utf-8") \
                if "xl/sharedStrings.xml" in zf.namelist() else "<sst></sst>"
            strings = _leer_shared_strings(zf)

        # ── Encontrar la primera fila vacía desde fila_inicio_datos ──
        # Se considera vacía si la celda de fecha no tiene valor
        target_row = fila_inicio_datos
        for r in range(fila_inicio_datos, fila_inicio_datos + 200):
            ref_fecha = f"{col_fecha}{r}"
            val, _ = _read_cell_cached(sheet_xml, ref_fecha, strings)
            if val is None:
                target_row = r
                break
        else:
            return (
                f"Error: no se encontró fila vacía en columna {col_fecha} "
                f"entre filas {fila_inicio_datos} y {fila_inicio_datos + 199}."
            )

        # ── Calcular fecha: último día del mes ──
        fecha_serial = _excel_date(_last_day(año, mes))
        fecha_str = _last_day(año, mes).strftime("%d/%m/%Y")

        # ── Escribir la nueva fila ──
        cells = {
            col_fecha: fecha_serial,
            col_r5:    float(valor_r5),
            col_r22:   float(valor_r22),
            col_r26:   float(valor_r26),
        }
        sheet_xml, ss_xml = _write_cells_in_row(sheet_xml, ss_xml, target_row, cells)
        _apply_to_xlsx(cg_path, {caja_path: sheet_xml, "xl/sharedStrings.xml": ss_xml})

        return (
            f"OK: nueva fila añadida en fila {target_row} de Caja Histórica.\n"
            f"  {col_fecha}{target_row} = {fecha_str} (serial {fecha_serial})\n"
            f"  {col_r5}{target_row}  = {valor_r5:,.2f}\n"
            f"  {col_r22}{target_row} = {valor_r22:,.2f}\n"
            f"  {col_r26}{target_row} = {valor_r26:,.2f}"
        )
    except Exception as e:
        return f"Error al agregar fila histórica: {e}"
