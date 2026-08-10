"""Ingesta de histórico trimestral vacancia/renta Oficinas Las Condes (Clase A/B)
desde SharePoint RAW/"mercado oficinas DB.xlsx" -> raw_mercado_oficinas_evolucion.

Formato fuente (hoja única "Hoja1"): fila 3 = años (col C+, uno cada 4 cols),
fila 4 = etiqueta de trimestre (1Q..4Q), filas 5-8 = "Vacancia A", "Vacancia B",
"Renta A", "Renta B" con un valor por columna de trimestre.

No expone CLI de producción (no hay flujo recurrente todavía, el archivo es
histórico manual) — se corre una vez y se re-corre si el archivo se actualiza
(idempotente vía UNIQUE(periodo, submercado, clase, file_hash) + superseded_at).
"""
from __future__ import annotations

import hashlib
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from tools.db.connection import apply_migrations, get_conn_for, DEFAULT_DB_PATH  # noqa: E402

SUBMERCADO = "Las Condes (CBD)"
ROW_LABELS = {
    "Vacancia A": ("A", "vacancia_pct"),
    "Vacancia B": ("B", "vacancia_pct"),
    "Renta A": ("A", "renta_uf_m2"),
    "Renta B": ("B", "renta_uf_m2"),
}


def _quarter_end_month(trimestre: int) -> str:
    return {1: "03", 2: "06", 3: "09", 4: "12"}[trimestre]


def ingest(xlsx_path: str, db_path: str = DEFAULT_DB_PATH, sheet: str = "Hoja1") -> int:
    import openpyxl

    apply_migrations(db_path)
    file_hash = hashlib.sha256(Path(xlsx_path).read_bytes()).hexdigest()

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    years_row = rows[2]
    quarters_row = rows[3]

    # Construir lista (col_idx, anio, trimestre) para cada columna con dato.
    cols = []
    current_year = None
    for idx in range(2, len(quarters_row)):
        if years_row[idx] is not None:
            current_year = int(years_row[idx])
        q = quarters_row[idx]
        if q is None or current_year is None:
            continue
        trimestre = int(str(q)[0])
        cols.append((idx, current_year, trimestre))

    conn = get_conn_for(db_path)
    n_inserted = 0
    try:
        conn.execute(
            "UPDATE raw_mercado_oficinas_evolucion SET superseded_at = datetime('now') "
            "WHERE submercado = ? AND superseded_at IS NULL AND file_hash != ?",
            (SUBMERCADO, file_hash),
        )
        for row in rows[4:]:
            label = row[1]
            if label not in ROW_LABELS:
                continue
            clase, campo = ROW_LABELS[label]
            for idx, anio, trimestre in cols:
                val = row[idx]
                if val is None:
                    continue
                periodo = f"{anio}-{_quarter_end_month(trimestre)}"
                cur = conn.execute(
                    """
                    SELECT id FROM raw_mercado_oficinas_evolucion
                    WHERE periodo = ? AND submercado = ? AND clase = ? AND superseded_at IS NULL
                    """,
                    (periodo, SUBMERCADO, clase),
                )
                existing = cur.fetchone()
                if existing:
                    conn.execute(
                        f"UPDATE raw_mercado_oficinas_evolucion SET {campo} = ?, "
                        f"file_hash = ?, source_file = ?, source_sheet = ? WHERE id = ?",
                        (val, file_hash, str(xlsx_path), sheet, existing[0]),
                    )
                else:
                    conn.execute(
                        """
                        INSERT INTO raw_mercado_oficinas_evolucion
                            (periodo, anio, trimestre, submercado, clase, {campo},
                             source_file, source_sheet, file_hash)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """.format(campo=campo),
                        (periodo, anio, trimestre, SUBMERCADO, clase, val,
                         str(xlsx_path), sheet, file_hash),
                    )
                    n_inserted += 1
        conn.commit()
    finally:
        conn.close()
    return n_inserted


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else (
        r"C:\Users\raimundo.opazo\OneDrive - Toesca\Inmobiliario Toesca - Documentos"
        r"\RAW\mercado oficinas DB.xlsx"
    )
    n = ingest(path)
    print(f"OK -> {n} filas nuevas insertadas en raw_mercado_oficinas_evolucion")
