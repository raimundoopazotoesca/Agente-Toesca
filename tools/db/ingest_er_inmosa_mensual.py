"""Ingesta mensual de ingresos/NOI de INMOSA desde la planilla fuente
('NUEVO ORDEN/Renta Comercial/Ingresos Activos/inmosa/<año>/Flujo Caja
MM.YY Senior Assist.xlsx'), hoja única 'FC MM.YYYY'.

A diferencia de ingest_er_mensual.py (hoja 'ESTADO DE RESULTADO' con
cuenta_codigo tipo '4-1-01-100'), esta planilla es un Flujo de Caja: filas
sin código de cuenta, solo etiqueta, en UF (columna 'Real UF'). Cada archivo
trae el histórico acumulado del año (enero..mes de cierre) en columnas; el
usuario elige un solo mes por corrida en la UI.

Layout confirmado 2026-07-27 sobre 'Flujo Caja Jun.26 Senior Assist.xlsx':
  - hoja única (ej. 'FC 06.2026')
  - columna A vacía, columna B trae las etiquetas de fila
  - fila de fechas (datetime, día 25 de cada mes) unas filas arriba de la
    fila ancla 'Ingresos operacionales'
  - fila ancla 'Ingresos operacionales' → INGRESOS_OPERACION
  - 7 filas de gasto operacional debajo (todas GASTOS_OPERACION):
    Remuneraciones y Administración, Aseo/mantención/Postventa, Seguros,
    Contribuciones, Oficina Central, Otros gastos operacionales, IVA
  - subtotal de control 'Costos Operacionales' (= suma de las 7 filas de
    gasto) y 'Flujo operativo (antes de Leasing e Hipotecas)' (= Ingresos +
    Costos Operacionales) — ninguno de los dos se persiste, solo validan
    integridad
  - fila 'Leasing' marca el fin del bloque operacional relevante para NOI;
    todo lo que sigue (flujo financiero, deuda, etc.) se ignora acá

monto_uf = pesos reales del mes elegido, en UF (la planilla no trae CLP).
Por convención con ingest_er_inmosa.py (fuente histórica previa, misma
activo_key 'INMOSA'), el valor UF se guarda en monto_clp y monto_uf queda
en None.
"""
from __future__ import annotations

import hashlib
import re
import sqlite3
from typing import Optional

import openpyxl

_ACTIVO_KEY = "INMOSA"

_ANCLA = re.compile(r"^ingresos operacionales$", re.I)
_TERMINATOR = re.compile(r"^leasing$", re.I)
_SUBTOTAL_COSTOS = re.compile(r"^costos operacionales$", re.I)
_SUBTOTAL_FLUJO_OPERATIVO = re.compile(r"^flujo operativo \(antes de leasing e hipotecas\)$", re.I)

_CATEGORIAS: dict[str, str] = {
    "remuneraciones y administraci�n": "INMOSA_ADM",
    "remuneraciones y administracion": "INMOSA_ADM",
    "remuneraciones y administración": "INMOSA_ADM",
    "aseo, mantenci�n, postventa": "INMOSA_ASEO",
    "aseo, mantencion, postventa": "INMOSA_ASEO",
    "aseo, mantención, postventa": "INMOSA_ASEO",
    "seguros": "INMOSA_SEG",
    "contribuciones": "INMOSA_CONTRIB",
    "oficina central (asesor�as araucana)": "INMOSA_OFICINA_CENTRAL",
    "oficina central (asesorias araucana)": "INMOSA_OFICINA_CENTRAL",
    "oficina central (asesorías araucana)": "INMOSA_OFICINA_CENTRAL",
    "otros gastos operacionales": "INMOSA_OTROS_GASTOS",
    "iva": "INMOSA_IVA",
}

_TOLERANCIA_UF = 0.5


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip())


def _file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def _pick_sheet(wb) -> "openpyxl.worksheet.worksheet.Worksheet":
    return wb.worksheets[0]


def _find_ancla(all_rows) -> int:
    for i, row in enumerate(all_rows):
        val = _norm(row[1].value if len(row) > 1 else None)
        if _ANCLA.match(val):
            return i
    raise ValueError("No se encontró la fila ancla 'Ingresos operacionales'.")


def _find_header_row(all_rows, ancla_idx: int) -> dict[int, str]:
    """Busca hacia arriba desde la ancla la primera fila con >=3 celdas fecha.
    Devuelve {col_idx (0-based): periodo 'YYYY-MM'}."""
    for i in range(ancla_idx - 1, -1, -1):
        row = all_rows[i]
        cells: dict[int, str] = {}
        for col_idx, cell in enumerate(row):
            v = cell.value
            if hasattr(v, "month") and hasattr(v, "year"):
                cells[col_idx] = f"{v.year:04d}-{v.month:02d}"
        if len(cells) >= 3:
            return cells
    raise ValueError("No se encontró la fila de fechas antes de la fila ancla.")


def parse_mes(xlsx_path: str, activo_key: str, periodo: str,
              conn: "sqlite3.Connection | None" = None) -> list[dict]:
    """Extrae las filas operacionales de un único mes del Flujo de Caja INMOSA.

    Lanza ValueError si el activo no es INMOSA, si el periodo no existe en
    el archivo, o si la validación de integridad falla.
    """
    if activo_key != _ACTIVO_KEY:
        raise ValueError(f"Activo no soportado por ingest_er_inmosa_mensual: {activo_key!r}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = _pick_sheet(wb)
    sheet_name = ws.title
    all_rows = list(ws.iter_rows(values_only=False))
    wb.close()

    ancla_idx = _find_ancla(all_rows)
    period_by_col = _find_header_row(all_rows, ancla_idx)

    col_for_periodo = next((c for c, p in period_by_col.items() if p == periodo), None)
    if col_for_periodo is None:
        disponibles = ", ".join(sorted(period_by_col.values()))
        raise ValueError(f"El archivo no tiene columna para el periodo {periodo}. Disponibles: {disponibles}")

    out: list[dict] = []
    suma_gastos = 0.0
    monto_ingreso: Optional[float] = None
    subtotal_costos: Optional[float] = None
    subtotal_flujo_operativo: Optional[float] = None
    terminador_encontrado = False

    for i in range(ancla_idx, len(all_rows)):
        row = all_rows[i]
        raw_label = row[1].value if len(row) > 1 else None
        label = _norm(raw_label)
        if not label:
            continue

        if _TERMINATOR.match(label):
            terminador_encontrado = True
            break

        cell = row[col_for_periodo] if col_for_periodo < len(row) else None
        monto = float(cell.value) if cell is not None and cell.value is not None else 0.0

        if _ANCLA.match(label):
            monto_ingreso = monto
            out.append({
                "activo_key":     activo_key,
                "periodo":        periodo,
                "cuenta_codigo":  "INMOSA_ING_OP",
                "cuenta_nombre":  str(raw_label).strip(),
                "monto_clp":      monto,
                "monto_uf":       None,
                "seccion":        "INGRESOS_OPERACION",
                "es_operacional": 1,
                "source_file":    xlsx_path,
                "source_sheet":   sheet_name,
                "source_row":     i + 1,
            })
            continue

        if _SUBTOTAL_COSTOS.match(label):
            subtotal_costos = monto
            continue
        if _SUBTOTAL_FLUJO_OPERATIVO.match(label):
            subtotal_flujo_operativo = monto
            continue

        codigo = _CATEGORIAS.get(label.lower())
        if codigo is None:
            raise ValueError(f"Categoría no reconocida en {xlsx_path}, fila {i + 1}: {raw_label!r}")

        suma_gastos += monto
        out.append({
            "activo_key":     activo_key,
            "periodo":        periodo,
            "cuenta_codigo":  codigo,
            "cuenta_nombre":  str(raw_label).strip(),
            "monto_clp":      monto,
            "monto_uf":       None,
            "seccion":        "GASTOS_OPERACION",
            "es_operacional": 1,
            "source_file":    xlsx_path,
            "source_sheet":   sheet_name,
            "source_row":     i + 1,
        })

    if not terminador_encontrado:
        raise ValueError("No se encontró la fila 'Leasing'.")
    if not out:
        raise ValueError(f"No se encontraron cuentas con datos para el periodo {periodo}.")

    if subtotal_costos is not None and abs(suma_gastos - subtotal_costos) >= _TOLERANCIA_UF:
        raise ValueError(
            f"Validación de integridad falló en {periodo}: suma gastos operacionales={suma_gastos!r} "
            f"!= Costos Operacionales={subtotal_costos!r}"
        )
    if (subtotal_flujo_operativo is not None and monto_ingreso is not None
            and subtotal_costos is not None):
        esperado = monto_ingreso + subtotal_costos
        if abs(esperado - subtotal_flujo_operativo) >= _TOLERANCIA_UF:
            raise ValueError(
                f"Validación de integridad falló en {periodo}: Ingresos + Costos Operacionales={esperado!r} "
                f"!= Flujo operativo (antes de Leasing e Hipotecas)={subtotal_flujo_operativo!r}"
            )

    return out


def ultimo_periodo_disponible(xlsx_path: str, activo_key: str) -> Optional[str]:
    """Best-effort: último periodo del archivo con datos en la fila ancla,
    para avisar en la UI si el usuario eligió un mes distinto."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = _pick_sheet(wb)
    all_rows = list(ws.iter_rows(values_only=False))
    wb.close()

    ancla_idx = _find_ancla(all_rows)
    period_by_col = _find_header_row(all_rows, ancla_idx)
    row = all_rows[ancla_idx]
    filled = [
        p for c, p in period_by_col.items()
        if c < len(row) and row[c].value is not None
    ]
    return max(filled) if filled else None


# ── Persistencia ─────────────────────────────────────────────────────────

def persist_mes(xlsx_path: str, activo_key: str, periodo: str,
                 conn: "sqlite3.Connection | None" = None) -> dict:
    """Ingesta idempotente de UN mes en raw_er_activo_line.

    Supersede por (activo_key, periodo), igual que ingest_er_mensual.py:
    subir el mismo archivo el próximo mes con una columna más no debe tocar
    los meses ya cerrados de otros periodos.
    """
    from tools.db import repo_audit, repo_er_activo

    owns_conn = conn is None
    if owns_conn:
        from tools.db.connection import get_conn
        conn = get_conn()

    try:
        with open(xlsx_path, "rb") as f:
            file_hash = _file_hash(f.read())

        prev = conn.execute(
            """SELECT 1 FROM raw_er_activo_line
                WHERE file_hash=? AND activo_key=? AND periodo=? AND superseded_at IS NULL
                LIMIT 1""",
            (file_hash, activo_key, periodo),
        ).fetchone()
        if prev is not None:
            return {"status": "skipped_idempotent", "rows": 0, "file_hash": file_hash, "ingest_run_id": None}

        lines = parse_mes(xlsx_path, activo_key, periodo, conn=conn)
        for line in lines:
            line["file_hash"] = file_hash

        n_previas = conn.execute(
            """SELECT COUNT(*) FROM raw_er_activo_line
                WHERE activo_key=? AND periodo=? AND superseded_at IS NULL""",
            (activo_key, periodo),
        ).fetchone()[0]
        status = "superseded_and_reinserted" if n_previas else "inserted"
        conn.execute(
            """UPDATE raw_er_activo_line SET superseded_at=datetime('now')
                WHERE activo_key=? AND periodo=? AND superseded_at IS NULL""",
            (activo_key, periodo),
        )

        run_id = repo_audit.start_ingest_run(
            conn, tool="ingest_er_inmosa_mensual", source_file=xlsx_path, file_hash=file_hash,
        )
        inserted = repo_er_activo.insert_lines(conn, lines, run_id)
        repo_audit.finish_ingest_run(
            conn, run_id, rows_in=len(lines), rows_loaded=inserted, status="ok",
        )
        return {"status": status, "rows": inserted, "file_hash": file_hash, "ingest_run_id": run_id}
    finally:
        if owns_conn:
            conn.close()
