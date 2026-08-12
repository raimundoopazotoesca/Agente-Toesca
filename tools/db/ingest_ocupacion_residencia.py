"""Ingesta de ocupación mensual por residencia de INMOSA (Living Acalis).

Fuente: SharePoint RAW/"Ocupacion Acalis - <mes> <año> - Toesca.xlsx".
Una hoja por residencia. Dos formatos de hoja observados:

- Formato A (Candil, Colombia, Errazuriz, Medina):
    Mes/Año | Cantidad de residentes | Ocupación %
- Formato B (Coventry, Montahue, Montemar, VA - Cordillera, VA - Valle):
    Mes/Año | Ingresos | Egresos | Fallecimiento | Cantidad de residentes | Ocupación %
    (+ fila final "Total" con promedios y una nota al pie — ambas se ignoran)

Layout fijo en cada hoja: fila 2 = nombre residencia, fila 3 col C = camas
(capacidad), fila 5 = header de columnas, datos desde fila 6 hasta la
primera fila cuyo Mes/Año no matchea MM-YYYY (fin de la serie real).

ocupacion_pct se recalcula siempre como cantidad_residentes/camas (decisión
del usuario 2026-08-12: el % de la fuente puede venir inconsistente, ej.
Colombia 03-2026 trae 102% con camas=106). El valor crudo de la fuente se
preserva en ocupacion_pct_fuente para auditoría, nunca se usa en cálculos.
"""
from __future__ import annotations

import hashlib
import re
import sqlite3
from pathlib import Path
from typing import Optional

import openpyxl

_MES_ANO_RE = re.compile(r"^(\d{2})-(\d{4})$")

# Nombre de hoja -> (residencia_key, activo_key en dim_activo)
_HOJA_A_RESIDENCIA = {
    "Candil":         "Candil",
    "Colombia":       "Colombia",
    "Coventry":       "Coventry",
    "Errazuriz":      "Errazuriz",
    "Medina":         "Medina",
    "Montahue":       "Montahue",
    "Montemar":       "Montemar",
    "VA - Cordillera": "VA - Cordillera",
    "VA - Valle":     "VA - Valle",
}
_ACTIVO_KEY = "INMOSA"

# Vendidas / fuera del portafolio (confirmado por el usuario 2026-08-12).
# Se sigue ingestando su histórico para gráficos, pero se marcan no-vigentes
# vía dim_residencia.vigente_hasta = último período con datos en la fuente.
_RESIDENCIAS_VENDIDAS = {"Montahue", "Montemar", "VA - Cordillera", "VA - Valle"}


def _file_hash(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _to_float(v) -> Optional[float]:
    """Convierte a float; tolera coma decimal chilena (ej. '98,4')."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().replace(",", ".")
    return float(v)


def _periodo(mes_ano: str) -> str:
    m = _MES_ANO_RE.match(mes_ano.strip())
    if not m:
        raise ValueError(f"Mes/Año inválido: {mes_ano!r}")
    mm, yyyy = m.groups()
    return f"{yyyy}-{mm}"


def parse_hoja(ws, sheet_name: str, xlsx_path: str) -> list[dict]:
    residencia_key = _HOJA_A_RESIDENCIA.get(sheet_name)
    if residencia_key is None:
        raise ValueError(f"Hoja sin mapear a residencia: {sheet_name!r}")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError(f"Hoja {sheet_name!r} demasiado corta ({len(rows)} filas)")

    camas = rows[2][2]  # fila 3, columna C
    header = [str(v).strip() if v else "" for v in rows[4]]  # fila 5
    tiene_movimientos = "Ingresos" in header

    if tiene_movimientos:
        # B=Mes/Año C=Ingresos D=Egresos E=Fallecimiento F=Residentes G=Ocupación%
        col_mes, col_ing, col_egr, col_fall, col_res, col_ocup = 1, 2, 3, 4, 5, 6
    else:
        # B=Mes/Año C=Residentes D=Ocupación%
        col_mes, col_res, col_ocup = 1, 2, 3
        col_ing = col_egr = col_fall = None

    out: list[dict] = []
    for i in range(5, len(rows)):  # fila 6 en adelante (índice 5)
        row = rows[i]
        raw_mes = row[col_mes] if col_mes < len(row) else None
        if raw_mes is None or not _MES_ANO_RE.match(str(raw_mes).strip()):
            continue  # fin de serie real, o fila "Total"/nota al pie

        periodo = _periodo(str(raw_mes))
        residentes = row[col_res] if col_res < len(row) else None
        ocup_fuente = row[col_ocup] if col_ocup < len(row) else None

        residentes_f = _to_float(residentes)
        ocup_calc = (residentes_f / camas) if (residentes_f is not None and camas) else None

        out.append({
            "residencia_key":       residencia_key,
            "periodo":              periodo,
            "camas":                camas,
            "cantidad_residentes":  residentes_f,
            "ocupacion_pct":        ocup_calc,
            "ocupacion_pct_fuente": _to_float(ocup_fuente),
            "ingresos":       _to_float(row[col_ing]) if col_ing is not None else None,
            "egresos":        _to_float(row[col_egr]) if col_egr is not None else None,
            "fallecimientos": _to_float(row[col_fall]) if col_fall is not None else None,
            "source_file":  xlsx_path,
            "source_sheet": sheet_name,
            "source_row":   i + 1,
        })
    return out


def parse_planilla(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        out: list[dict] = []
        for sheet_name in wb.sheetnames:
            out.extend(parse_hoja(wb[sheet_name], sheet_name, xlsx_path))
        return out
    finally:
        wb.close()


# ── Persistencia ─────────────────────────────────────────────────────────

_INSERT_COLS = [
    "residencia_key", "periodo", "camas", "cantidad_residentes",
    "ocupacion_pct", "ocupacion_pct_fuente",
    "ingresos", "egresos", "fallecimientos",
    "source_file", "source_sheet", "source_row", "file_hash", "ingest_run_id",
]


def _ensure_dim_residencia(conn: sqlite3.Connection) -> None:
    for hoja, residencia_key in _HOJA_A_RESIDENCIA.items():
        conn.execute(
            """INSERT INTO dim_residencia (residencia_key, nombre, activo_key)
               VALUES (?, ?, ?)
               ON CONFLICT(residencia_key) DO NOTHING""",
            (residencia_key, hoja, _ACTIVO_KEY),
        )


def persist(xlsx_path: str, conn: "sqlite3.Connection | None" = None) -> dict:
    """Ingesta idempotente. Reingesta (hash distinto) supersede lo anterior
    por residencia_key (una planilla nueva trae todas las residencias)."""
    from tools.db import repo_audit

    owns_conn = conn is None
    if owns_conn:
        from tools.db.connection import get_conn
        conn = get_conn()

    try:
        file_hash = _file_hash(xlsx_path)

        prev = conn.execute(
            """SELECT 1 FROM raw_ocupacion_residencia_line
                WHERE file_hash = ? AND superseded_at IS NULL LIMIT 1""",
            (file_hash,),
        ).fetchone()
        if prev is not None:
            return {"status": "skipped_idempotent", "rows": 0,
                    "file_hash": file_hash, "ingest_run_id": None}

        lines = parse_planilla(xlsx_path)
        for line in lines:
            line["file_hash"] = file_hash

        _ensure_dim_residencia(conn)
        # actualiza camas si cambió (última planilla manda)
        for hoja, residencia_key in _HOJA_A_RESIDENCIA.items():
            camas_vistas = {l["camas"] for l in lines if l["residencia_key"] == residencia_key}
            if camas_vistas:
                conn.execute(
                    "UPDATE dim_residencia SET camas = ? WHERE residencia_key = ?",
                    (max(camas_vistas), residencia_key),
                )

        prev_hashes = conn.execute(
            """SELECT DISTINCT file_hash FROM raw_ocupacion_residencia_line
                WHERE file_hash != ? AND superseded_at IS NULL""",
            (file_hash,),
        ).fetchall()

        run_id = repo_audit.start_ingest_run(
            conn, tool="ingest_ocupacion_residencia",
            source_file=xlsx_path, file_hash=file_hash,
        )

        if prev_hashes:
            for row in prev_hashes:
                conn.execute(
                    """UPDATE raw_ocupacion_residencia_line
                          SET superseded_at = datetime('now')
                        WHERE file_hash = ? AND superseded_at IS NULL""",
                    (row[0],),
                )
            status = "superseded_and_reinserted"
        else:
            status = "inserted"

        cols_sql = ", ".join(_INSERT_COLS)
        placeholders = ", ".join(["?"] * len(_INSERT_COLS))
        sql = f"INSERT OR IGNORE INTO raw_ocupacion_residencia_line ({cols_sql}) VALUES ({placeholders})"
        inserted = 0
        for line in lines:
            values = tuple(run_id if c == "ingest_run_id" else line.get(c) for c in _INSERT_COLS)
            cur = conn.execute(sql, values)
            inserted += cur.rowcount if cur.rowcount > 0 else 0

        for residencia_key in _RESIDENCIAS_VENDIDAS:
            conn.execute(
                """UPDATE dim_residencia
                      SET vigente_hasta = (
                          SELECT MAX(periodo) FROM raw_ocupacion_residencia_line
                           WHERE residencia_key = ? AND superseded_at IS NULL)
                    WHERE residencia_key = ?""",
                (residencia_key, residencia_key),
            )
        conn.commit()

        repo_audit.finish_ingest_run(
            conn, run_id, rows_in=len(lines), rows_loaded=inserted, status="ok",
        )

        return {"status": status, "rows": inserted,
                "file_hash": file_hash, "ingest_run_id": run_id}
    finally:
        if owns_conn:
            conn.close()


# ── CLI ───────────────────────────────────────────────────────────────────

def main(argv: Optional[list[str]] = None) -> int:
    import argparse
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="Path a la planilla xlsx")
    ap.add_argument("--dry-run", action="store_true",
                     help="Parsea e imprime resumen, no escribe DB")
    args = ap.parse_args(argv)

    if args.dry_run:
        rows = parse_planilla(args.xlsx)
        print(f"Parsed {len(rows)} filas de {args.xlsx}")
        residencias = sorted({r["residencia_key"] for r in rows})
        print(f"  residencias: {residencias}")
        for r in residencias:
            sub = [x for x in rows if x["residencia_key"] == r]
            periodos = sorted(x["periodo"] for x in sub)
            print(f"  {r}: {len(sub)} filas, {periodos[0]}..{periodos[-1]}")
        anomalias = [
            r for r in rows
            if r["ocupacion_pct_fuente"] is not None
            and (r["ocupacion_pct_fuente"] > 1.2 or r["ocupacion_pct_fuente"] < 0)
        ]
        if anomalias:
            print(f"  ANOMALÍAS ({len(anomalias)}):")
            for a in anomalias:
                print(f"    {a['residencia_key']} {a['periodo']}: "
                      f"fuente={a['ocupacion_pct_fuente']!r} vs recalculado={a['ocupacion_pct']!r}")
        return 0

    res = persist(args.xlsx)
    print(res)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
