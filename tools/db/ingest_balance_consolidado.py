"""Valida y persiste balances consolidados trimestrales desde una planilla.

Formato principal:
  - una hoja por periodo, con nombre MM-YYYY (ej. 03-2026)
  - dentro de la hoja, tres bloques: TRI, Apoquindo y PT

Como fallback tambien acepta una hoja por fondo. El parser busca las 10
cuentas ESF que consume el factsheet, por codigo canonico (`ESF.total_activo`)
o por descripcion, y persiste un snapshot por (fondo_key, periodo) en
raw_balance_consolidado_line.
"""
from __future__ import annotations

import hashlib
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl

from tools.db.connection import get_conn_for

DB_PATH = Path(__file__).resolve().parents[2] / "memory" / "agente_toesca_v2.db"

FONDOS_REQUERIDOS = {
    "TRI": {
        "fondo_key": "TRI",
        "labels": ("tri", "rentas", "rentas tri", "toesca rentas inmobiliarias"),
    },
    "PT": {
        "fondo_key": "PT",
        "labels": ("pt", "rentas pt", "parque titanium", "titanium"),
    },
    "Apo": {
        "fondo_key": "Apo",
        "labels": ("apo", "apoquindo", "rentas apoquindo"),
    },
}

UNIDAD_MULTIPLIER = {
    "CLP": 1,
    "M$": 1_000,
    "MM$": 1_000_000,
}

CUENTAS_REQUERIDAS = [
    "ESF.efectivo",
    "ESF.otros_activos_corrientes",
    "ESF.propiedades_inversion",
    "ESF.otros_activos_no_corrientes",
    "ESF.total_activo",
    "ESF.prestamos",
    "ESF.pasivos_impuestos_diferidos",
    "ESF.otros_pasivos",
    "ESF.patrimonio_neto",
    "ESF.total_pasivo_patrimonio",
]

ACTIVO_COMPONENTES = [
    "ESF.efectivo",
    "ESF.otros_activos_corrientes",
    "ESF.propiedades_inversion",
    "ESF.otros_activos_no_corrientes",
]

PASIVO_PATRIMONIO_COMPONENTES = [
    "ESF.prestamos",
    "ESF.pasivos_impuestos_diferidos",
    "ESF.otros_pasivos",
    "ESF.patrimonio_neto",
]

CUENTAS_CERO_SI_FALTAN = {
    "ESF.pasivos_impuestos_diferidos",
}

CUENTA_LABEL = {
    "ESF.efectivo": "Efectivo y equivalentes",
    "ESF.otros_activos_corrientes": "Otros activos corrientes",
    "ESF.propiedades_inversion": "Propiedades de inversion",
    "ESF.otros_activos_no_corrientes": "Otros activos no corrientes",
    "ESF.total_activo": "Total activo",
    "ESF.prestamos": "Prestamos bancarios",
    "ESF.pasivos_impuestos_diferidos": "Pasivos por impuestos diferidos",
    "ESF.otros_pasivos": "Otros pasivos",
    "ESF.patrimonio_neto": "Patrimonio neto",
    "ESF.total_pasivo_patrimonio": "Total pasivo y patrimonio",
}

BALANCE_TOLERANCE_CLP = 1_000.0

CUENTA_ALIASES = {
    "ESF.efectivo": (
        "efectivo y equivalentes al efectivo",
        "efectivo y efectivo equivalente",
        "efectivo y equivalentes",
        "caja",
    ),
    "ESF.otros_activos_corrientes": (
        "otros activos corrientes",
        "otros activos corriente",
    ),
    "ESF.propiedades_inversion": (
        "propiedades de inversion",
        "propiedades de inversiones",
        "inversiones inmobiliarias",
        "inversiones en propiedades",
        "propiedad de inversion",
    ),
    "ESF.otros_activos_no_corrientes": (
        "otros activos no corrientes",
        "otros activos no corriente",
        "activo por impuestos diferidos",
        "activos por impuestos diferidos",
    ),
    "ESF.total_activo": (
        "total activo",
        "total activos",
        "total de activos",
    ),
    "ESF.prestamos": (
        "prestamos",
        "prestamos bancarios",
        "obligaciones con bancos",
        "deuda bancaria",
        "deuda financiera",
    ),
    "ESF.pasivos_impuestos_diferidos": (
        "pasivos por impuestos diferidos",
        "pasivo por impuestos diferidos",
        "impuestos diferidos pasivo",
    ),
    "ESF.otros_pasivos": (
        "otros pasivos",
        "otros pasivos corrientes y no corrientes",
    ),
    "ESF.patrimonio_neto": (
        "patrimonio neto",
        "total patrimonio neto",
        "total patrimonio",
        "patrimonio",
    ),
    "ESF.total_pasivo_patrimonio": (
        "total pasivo y patrimonio",
        "total pasivos y patrimonio",
        "total pasivo y patrimonio neto",
        "total pasivos y patrimonio neto",
        "total pasivo + patrimonio",
        "total pasivos + patrimonio",
    ),
}


class ValidationResult:
    def __init__(self):
        self.ok = True
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.data: dict[str, Any] = {}

    def add_error(self, msg: str) -> None:
        self.errors.append(msg)
        self.ok = False

    def add_warning(self, msg: str) -> None:
        self.warnings.append(msg)

    def to_dict(self) -> dict[str, Any]:
        return {"ok": self.ok, "errors": self.errors, "warnings": self.warnings, **self.data}


@dataclass
class ParsedFund:
    fondo_key: str
    sheet_name: str
    rows: list[dict[str, Any]]
    warnings: list[str]
    balance_check: dict[str, Any] | None = None


def _norm(text: Any) -> str:
    s = str(text or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^a-z0-9$]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


_ALIASES_NORM = {
    alias: cuenta
    for cuenta, aliases in CUENTA_ALIASES.items()
    for alias in [_norm(cuenta), *(_norm(a) for a in aliases)]
}


def _file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def _validar_periodo(periodo: str, result: ValidationResult) -> bool:
    if not periodo or len(periodo) != 7 or periodo[4] != "-":
        result.add_error("Periodo debe tener formato YYYY-MM.")
        return False
    try:
        mes = int(periodo[5:7])
    except ValueError:
        result.add_error("Periodo debe tener formato YYYY-MM.")
        return False
    if mes not in (3, 6, 9, 12):
        result.add_error("El balance consolidado debe ser trimestral: meses 03, 06, 09 o 12.")
        return False
    return True


def _periodo_sheet_name(periodo: str) -> str:
    return f"{periodo[5:7]}-{periodo[:4]}"


def _load_workbook(file_bytes: bytes):
    return openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)


def _match_sheet(sheet_name: str) -> str | None:
    n = _norm(sheet_name)
    if "apoquindo" in n or re.search(r"\bapo\b", n):
        return "Apo"
    if "parque titanium" in n or "titanium" in n or re.search(r"\bpt\b", n):
        return "PT"
    if "tri" in n or "rentas inmobiliarias" in n:
        return "TRI"

    exact_matches: list[tuple[int, str]] = []
    for fondo, cfg in FONDOS_REQUERIDOS.items():
        for label in cfg["labels"]:
            label_norm = _norm(label)
            if n == label_norm:
                exact_matches.append((len(label_norm), fondo))
    if exact_matches:
        return sorted(exact_matches, reverse=True)[0][1]

    contains_matches: list[tuple[int, str]] = []
    for fondo, cfg in FONDOS_REQUERIDOS.items():
        for label in cfg["labels"]:
            label_norm = _norm(label)
            if label_norm and label_norm in n:
                contains_matches.append((len(label_norm), fondo))
    if contains_matches:
        return sorted(contains_matches, reverse=True)[0][1]
    return None


def _find_period_sheet(wb, periodo: str) -> str | None:
    expected_names = {
        _norm(_periodo_sheet_name(periodo)),
        _norm(periodo),
        _norm(periodo.replace("-", ".")),
        _norm(periodo.replace("-", "/")),
    }
    for sheet_name in wb.sheetnames:
        if _norm(sheet_name) in expected_names:
            return sheet_name
    return None


def _find_sheets(wb) -> tuple[dict[str, str], list[str]]:
    matched: dict[str, str] = {}
    warnings: list[str] = []
    for sheet_name in wb.sheetnames:
        fondo = _match_sheet(sheet_name)
        if not fondo:
            continue
        if fondo in matched:
            warnings.append(
                f"Hay mas de una hoja que parece ser {fondo}: '{matched[fondo]}' y '{sheet_name}'. "
                f"Se usara '{matched[fondo]}'."
            )
            continue
        matched[fondo] = sheet_name
    return matched, warnings


def _period_value_matches(value: Any, periodo: str) -> bool:
    year = int(periodo[:4])
    month = int(periodo[5:7])
    if isinstance(value, (date, datetime)):
        return value.year == year and value.month == month
    if value is None:
        return False
    s = _norm(value)
    if not s:
        return False
    candidates = {
        periodo,
        f"{periodo[5:7]} {periodo[:4]}",
        f"{periodo[5:7]}-{periodo[:4]}",
        f"{periodo[5:7]}.{periodo[:4]}",
        f"{periodo[5:7]}/{periodo[:4]}",
    }
    raw = str(value).strip().lower()
    if any(c in raw for c in candidates):
        return True
    return str(year) in s and f"{month:02d}" in s


def _find_period_col(ws, periodo: str) -> int | None:
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > 30:
            break
        for col_idx, value in enumerate(row):
            if _period_value_matches(value, periodo):
                return col_idx
    return None


def _to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    s = value.strip()
    if not s:
        return None
    lowered = s.lower()
    for token in ("clp", "m$", "mm$", "$"):
        lowered = lowered.replace(token, "")
    if re.search(r"[a-zA-Z]", lowered):
        return None
    negative = lowered.startswith("(") and lowered.endswith(")")
    lowered = lowered.strip("() ")
    lowered = re.sub(r"[^0-9,.\-]", "", lowered)
    if not lowered or lowered in {"-", ",", "."}:
        return None
    if "," in lowered and "." in lowered:
        lowered = lowered.replace(".", "").replace(",", ".")
    elif "," in lowered:
        lowered = lowered.replace(".", "").replace(",", ".")
    elif lowered.count(".") > 1:
        lowered = lowered.replace(".", "")
    try:
        n = float(lowered)
    except ValueError:
        return None
    return -n if negative else n


def _account_from_value(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    exact_code = value.strip()
    if exact_code in CUENTAS_REQUERIDAS:
        return exact_code
    code_match = re.search(r"ESF\.[A-Za-z0-9_]+", exact_code)
    if code_match and code_match.group(0) in CUENTAS_REQUERIDAS:
        return code_match.group(0)
    norm = _norm(value)
    if norm in _ALIASES_NORM:
        return _ALIASES_NORM[norm]
    return None


def _find_account(row: tuple[Any, ...]) -> str | None:
    for value in row:
        cuenta = _account_from_value(value)
        if cuenta:
            return cuenta
    return None


def _find_amount(row: tuple[Any, ...], period_col: int | None) -> tuple[float | None, bool]:
    if period_col is not None and period_col < len(row):
        amount = _to_number(row[period_col])
        if amount is not None:
            return amount, False
    nums = [_to_number(value) for value in row]
    nums = [n for n in nums if n is not None]
    if not nums:
        return None, False
    return nums[-1], len(nums) > 1 and period_col is None


def _infer_multiplier_from_rows(rows: list[tuple[Any, ...]], unidad: str) -> int:
    if unidad != "AUTO":
        return UNIDAD_MULTIPLIER[unidad]
    for row in rows[:20]:
        text = " ".join(str(v) for v in row if isinstance(v, str))
        n = _norm(text)
        if "miles de pesos" in n or "m$" in text.lower():
            return UNIDAD_MULTIPLIER["M$"]
        if "millones de pesos" in n or "mm$" in text.lower():
            return UNIDAD_MULTIPLIER["MM$"]
    return UNIDAD_MULTIPLIER["CLP"]


def _parse_account_pairs(row: tuple[Any, ...]) -> list[tuple[str, float]]:
    pairs: list[tuple[str, float]] = []
    for idx, value in enumerate(row):
        cuenta = _account_from_value(value)
        if not cuenta:
            continue
        amount = None
        for value_right in row[idx + 1:]:
            if _account_from_value(value_right):
                break
            amount = _to_number(value_right)
            if amount is not None:
                break
        if amount is not None:
            pairs.append((cuenta, amount))
    return pairs


def _parse_fund_block(
    rows: list[tuple[Any, ...]],
    start_idx: int,
    end_idx: int,
    fondo_key: str,
    periodo: str,
    multiplier: int,
) -> ParsedFund:
    values_by_account: dict[str, float] = {}
    warnings: list[str] = []
    for row in rows[start_idx + 1:end_idx]:
        for cuenta, amount in _parse_account_pairs(row):
            if cuenta in values_by_account:
                warnings.append(
                    f"{fondo_key}: cuenta duplicada '{CUENTA_LABEL[cuenta]}'; se uso el ultimo valor encontrado."
                )
            values_by_account[cuenta] = amount * multiplier

    for cuenta in CUENTAS_CERO_SI_FALTAN:
        if cuenta not in values_by_account:
            values_by_account[cuenta] = 0.0
            warnings.append(f"{fondo_key}: '{CUENTA_LABEL[cuenta]}' no viene en la planilla; se guardara 0.")

    parsed_rows = [
        {
            "fondo_key": fondo_key,
            "periodo": periodo,
            "cuenta_codigo": cuenta,
            "cuenta_label": CUENTA_LABEL[cuenta],
            "monto_clp": values_by_account[cuenta],
        }
        for cuenta in CUENTAS_REQUERIDAS
        if cuenta in values_by_account
    ]
    return ParsedFund(fondo_key=fondo_key, sheet_name="", rows=parsed_rows, warnings=warnings)


def _parse_period_sheet(ws, periodo: str, multiplier: int) -> list[ParsedFund]:
    rows = list(ws.iter_rows(values_only=True))
    block_starts: list[tuple[int, str]] = []
    for idx, row in enumerate(rows):
        row_text = " ".join(str(v) for v in row if isinstance(v, str))
        if "fondo" not in _norm(row_text):
            continue
        fondo = _match_sheet(row_text)
        if fondo:
            block_starts.append((idx, FONDOS_REQUERIDOS[fondo]["fondo_key"]))

    parsed: list[ParsedFund] = []
    for i, (start_idx, fondo_key) in enumerate(block_starts):
        end_idx = block_starts[i + 1][0] if i + 1 < len(block_starts) else len(rows)
        fund = _parse_fund_block(rows, start_idx, end_idx, fondo_key, periodo, multiplier)
        fund.sheet_name = ws.title
        parsed.append(fund)
    return parsed


def _parse_sheet(ws, fondo_key: str, periodo: str, multiplier: int) -> ParsedFund:
    period_col = _find_period_col(ws, periodo)
    values_by_account: dict[str, float] = {}
    warnings: list[str] = []
    warned_multi_numeric = False

    for row in ws.iter_rows(values_only=True):
        cuenta = _find_account(row)
        if not cuenta:
            continue
        amount, multi_numeric = _find_amount(row, period_col)
        if amount is None:
            warnings.append(f"{fondo_key}: '{CUENTA_LABEL[cuenta]}' no tiene monto legible.")
            continue
        if multi_numeric and not warned_multi_numeric:
            warnings.append(
                f"{fondo_key}: se detectaron varias columnas numericas y no una columna explicita "
                f"para {periodo}; se uso el ultimo numero de cada fila."
            )
            warned_multi_numeric = True
        values_by_account[cuenta] = amount * multiplier

    rows = [
        {
            "fondo_key": fondo_key,
            "periodo": periodo,
            "cuenta_codigo": cuenta,
            "cuenta_label": CUENTA_LABEL[cuenta],
            "monto_clp": values_by_account[cuenta],
        }
        for cuenta in CUENTAS_REQUERIDAS
        if cuenta in values_by_account
    ]
    return ParsedFund(fondo_key=fondo_key, sheet_name=ws.title, rows=rows, warnings=warnings)


def _existentes(periodo: str) -> dict[str, int]:
    con = get_conn_for(str(DB_PATH))
    try:
        rows = con.execute(
            "SELECT fondo_key, COUNT(*) FROM raw_balance_consolidado_line "
            "WHERE periodo=? AND superseded_at IS NULL GROUP BY fondo_key",
            (periodo,),
        ).fetchall()
        return {row[0]: row[1] for row in rows}
    finally:
        con.close()


def _previous_snapshot(fondo_key: str, periodo: str) -> tuple[str | None, dict[str, float]]:
    con = get_conn_for(str(DB_PATH))
    try:
        row = con.execute(
            "SELECT MAX(periodo) FROM raw_balance_consolidado_line "
            "WHERE fondo_key=? AND periodo<? AND superseded_at IS NULL",
            (fondo_key, periodo),
        ).fetchone()
        prev_periodo = row[0] if row else None
        if not prev_periodo:
            return None, {}
        rows = con.execute(
            "SELECT cuenta_codigo, SUM(monto_clp) AS monto_clp "
            "FROM raw_balance_consolidado_line "
            "WHERE fondo_key=? AND periodo=? AND superseded_at IS NULL "
            "GROUP BY cuenta_codigo",
            (fondo_key, prev_periodo),
        ).fetchall()
        return prev_periodo, {row[0]: float(row[1]) for row in rows if row[1] is not None}
    finally:
        con.close()


def _add_deltas(parsed: list[ParsedFund], periodo: str) -> dict[str, str | None]:
    prev_periodos: dict[str, str | None] = {}
    for fund in parsed:
        prev_periodo, prev_values = _previous_snapshot(fund.fondo_key, periodo)
        prev_periodos[fund.fondo_key] = prev_periodo
        for row in fund.rows:
            prev = prev_values.get(row["cuenta_codigo"])
            row["monto_clp_anterior"] = prev
            if prev is None or prev == 0:
                row["delta_pct"] = None
            else:
                row["delta_pct"] = (row["monto_clp"] / prev - 1) * 100
    return prev_periodos


def _validate_parsed(parsed: list[ParsedFund], result: ValidationResult) -> None:
    for fund in parsed:
        present = {row["cuenta_codigo"] for row in fund.rows}
        missing = [CUENTA_LABEL[c] for c in CUENTAS_REQUERIDAS if c not in present]
        if missing:
            result.add_error(
                f"{fund.fondo_key} ({fund.sheet_name}): faltan cuentas requeridas: "
                + ", ".join(missing)
            )
        values = {row["cuenta_codigo"]: row["monto_clp"] for row in fund.rows}
        total_activo = values.get("ESF.total_activo")
        total_pp = values.get("ESF.total_pasivo_patrimonio")
        if total_activo is not None and total_pp is not None:
            activo_componentes = sum(values.get(c, 0.0) for c in ACTIVO_COMPONENTES)
            pp_componentes = sum(values.get(c, 0.0) for c in PASIVO_PATRIMONIO_COMPONENTES)
            diff = total_activo - total_pp
            diff_activo_componentes = activo_componentes - total_activo
            diff_pp_componentes = pp_componentes - total_pp
            ok = (
                abs(diff) <= BALANCE_TOLERANCE_CLP
                and abs(diff_activo_componentes) <= BALANCE_TOLERANCE_CLP
                and abs(diff_pp_componentes) <= BALANCE_TOLERANCE_CLP
            )
            fund.balance_check = {
                "total_activo": total_activo,
                "total_pasivo_patrimonio": total_pp,
                "activo_componentes": activo_componentes,
                "pasivo_patrimonio_componentes": pp_componentes,
                "diff_clp": diff,
                "diff_activo_componentes_clp": diff_activo_componentes,
                "diff_pasivo_patrimonio_componentes_clp": diff_pp_componentes,
                "ok": ok,
                "tolerance_clp": BALANCE_TOLERANCE_CLP,
            }
            if not ok:
                result.add_error(
                    f"{fund.fondo_key}: balance descuadrado. "
                    f"Activo - Pasivo+Patrimonio = {diff:,.0f} CLP; "
                    f"Componentes de activo - Total activo = {diff_activo_componentes:,.0f} CLP; "
                    f"Componentes de pasivo+patrimonio - Total pasivo+patrimonio = {diff_pp_componentes:,.0f} CLP. "
                    "No se puede ingestar un balance descuadrado."
                )


def parse(file_bytes: bytes, periodo: str, unidad: str = "M$") -> tuple[ValidationResult, list[ParsedFund]]:
    result = ValidationResult()
    if not _validar_periodo(periodo, result):
        return result, []
    if unidad not in {*UNIDAD_MULTIPLIER, "AUTO"}:
        result.add_error("Unidad invalida. Usa AUTO, CLP, M$ o MM$.")
        return result, []

    try:
        wb = _load_workbook(file_bytes)
    except Exception as exc:  # noqa: BLE001
        result.add_error(f"No se pudo abrir la planilla: {exc}")
        return result, []

    period_sheet_name = _find_period_sheet(wb, periodo)
    parsed: list[ParsedFund] = []
    if period_sheet_name:
        ws = wb[period_sheet_name]
        rows_for_unit = list(ws.iter_rows(values_only=True))
        multiplier = _infer_multiplier_from_rows(rows_for_unit, unidad)
        # Reusar las filas ya leidas evitando una segunda pasada read-only.
        block_starts: list[tuple[int, str]] = []
        for idx, row in enumerate(rows_for_unit):
            row_text = " ".join(str(v) for v in row if isinstance(v, str))
            if "fondo" not in _norm(row_text):
                continue
            fondo = _match_sheet(row_text)
            if fondo:
                block_starts.append((idx, FONDOS_REQUERIDOS[fondo]["fondo_key"]))
        for i, (start_idx, fondo_key) in enumerate(block_starts):
            end_idx = block_starts[i + 1][0] if i + 1 < len(block_starts) else len(rows_for_unit)
            parsed_fund = _parse_fund_block(rows_for_unit, start_idx, end_idx, fondo_key, periodo, multiplier)
            parsed_fund.sheet_name = period_sheet_name
            parsed.append(parsed_fund)
    else:
        multiplier = _infer_multiplier_from_rows([], unidad)
        sheets, sheet_warnings = _find_sheets(wb)
        for warning in sheet_warnings:
            result.add_warning(warning)
        missing_sheets = [f for f in FONDOS_REQUERIDOS if f not in sheets]
        if missing_sheets:
            result.add_error(
                f"No existe hoja '{_periodo_sheet_name(periodo)}' en la planilla."
            )
            return result, []
        for fondo, sheet_name in sheets.items():
            fondo_key = FONDOS_REQUERIDOS[fondo]["fondo_key"]
            parsed_fund = _parse_sheet(wb[sheet_name], fondo_key, periodo, multiplier)
            parsed.append(parsed_fund)

    present_funds = {fund.fondo_key for fund in parsed}
    required_funds = {cfg["fondo_key"] for cfg in FONDOS_REQUERIDOS.values()}
    if not required_funds.issubset(present_funds):
        missing = sorted(required_funds - present_funds)
        result.add_error("La hoja del periodo no contiene bloques para todos los fondos. Faltan: " + ", ".join(missing) + ".")
        return result, parsed

    for parsed_fund in parsed:
        for warning in parsed_fund.warnings:
            result.add_warning(warning)

    _validate_parsed(parsed, result)
    prev_periodos = _add_deltas(parsed, periodo)
    existentes = _existentes(periodo)
    for fund in parsed:
        if fund.fondo_key in existentes:
            result.add_warning(
                f"{fund.fondo_key} {periodo} ya tiene {existentes[fund.fondo_key]} fila(s) vigentes; "
                "confirmar reemplazara ese snapshot."
            )

    result.data = {
        "periodo": periodo,
        "unidad": unidad,
        "multiplicador": multiplier,
        "fondos": [
            {
                "fondo_key": fund.fondo_key,
                "sheet_name": fund.sheet_name,
                "n_lineas": len(fund.rows),
                "balance_check": fund.balance_check,
                "periodo_anterior": prev_periodos.get(fund.fondo_key),
                "rows": fund.rows,
            }
            for fund in parsed
        ],
        "n_lineas": sum(len(fund.rows) for fund in parsed),
        "file_hash": _file_hash(file_bytes),
        "existentes": existentes,
    }
    return result, parsed


def validate(file_bytes: bytes, filename: str, periodo: str, unidad: str = "M$") -> ValidationResult:
    result, _ = parse(file_bytes, periodo, unidad)
    result.data["filename"] = filename
    return result


def commit(file_bytes: bytes, filename: str, periodo: str, unidad: str = "M$") -> dict[str, Any]:
    result, parsed = parse(file_bytes, periodo, unidad)
    if not result.ok:
        raise ValueError("; ".join(result.errors) or "Error de validacion.")

    file_hash = result.data["file_hash"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    con = get_conn_for(str(DB_PATH))
    cur = con.cursor()
    cur.execute(
        """INSERT INTO ingest_run
           (tool, source_file, file_hash, started_at, status, periodo_declarado)
           VALUES (?,?,?,?,?,?)""",
        ("ingest_balance_consolidado", filename, file_hash, now, "running", periodo),
    )
    run_id = cur.lastrowid

    inserted = 0
    superseded: dict[str, int] = {}
    try:
        for fund in parsed:
            cur.execute(
                "UPDATE raw_balance_consolidado_line SET superseded_at=? "
                "WHERE fondo_key=? AND periodo=? AND superseded_at IS NULL",
                (now, fund.fondo_key, periodo),
            )
            superseded[fund.fondo_key] = cur.rowcount
            for row in fund.rows:
                cur.execute(
                    """INSERT INTO raw_balance_consolidado_line
                       (fondo_key, periodo, cuenta_codigo, monto_clp, source_file,
                        ingest_run_id, loaded_at)
                       VALUES (?,?,?,?,?,?,?)""",
                    (
                        fund.fondo_key,
                        periodo,
                        row["cuenta_codigo"],
                        row["monto_clp"],
                        filename,
                        run_id,
                        now,
                    ),
                )
                inserted += 1

        cur.execute(
            "UPDATE ingest_run SET ended_at=?, status='ok', rows_in=?, rows_loaded=? WHERE id=?",
            (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), result.data["n_lineas"], inserted, run_id),
        )
        con.commit()
    except Exception:
        con.rollback()
        cur.execute(
            "UPDATE ingest_run SET ended_at=?, status='error' WHERE id=?",
            (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), run_id),
        )
        con.commit()
        raise
    finally:
        con.close()

    return {
        "status": "ok",
        "periodo": periodo,
        "unidad": unidad,
        "run_id": run_id,
        "filas_insertadas": inserted,
        "filas_superseded": sum(superseded.values()),
        "superseded": superseded,
        "fondos": [
            {"fondo_key": fund.fondo_key, "sheet_name": fund.sheet_name, "n_lineas": len(fund.rows)}
            for fund in parsed
        ],
    }
