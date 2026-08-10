"""Backfill de una sola corrida: histórico semestral vacancia/UF de Bodegas
desde SharePoint RAW/"mercado bodegas db.xlsx" -> raw_mercado_bodegas_evolucion.

Formato fuente (hoja única "Hoja1"): fila 3 = headers ('UF/m2' en col D,
'Vacancia' en col E), filas 4+ = semestre (col C, ej. '2S-2015'), UF/m² (col
D), vacancia como fracción (col E).

No expone CLI de producción — el archivo es histórico manual que no se
vuelve a actualizar; se corre una vez para poblar la DB y el gráfico del
fact sheet lee siempre de raw_mercado_bodegas_evolucion, nunca del xlsx.
Re-ejecutable de forma idempotente (UNIQUE(semestre, file_hash) +
superseded_at) por si el archivo se corrige a mano en el futuro.
"""
from __future__ import annotations

import hashlib
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from tools.db.connection import apply_migrations, get_conn_for, DEFAULT_DB_PATH  # noqa: E402


def _anio_periodo_num(semestre: str) -> tuple[int, int]:
    # '2S-2015' -> (2015, 2), '1S-2026' -> (2026, 1)
    num, anio = semestre.split("-")
    return int(anio), int(num[0])


def ingest(xlsx_path: str, db_path: str = DEFAULT_DB_PATH, sheet: str = "Hoja1") -> int:
    import openpyxl

    apply_migrations(db_path)
    file_hash = hashlib.sha256(Path(xlsx_path).read_bytes()).hexdigest()

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    conn = get_conn_for(db_path)
    n_inserted = 0
    try:
        conn.execute(
            "UPDATE raw_mercado_bodegas_evolucion SET superseded_at = datetime('now') "
            "WHERE superseded_at IS NULL AND file_hash != ?",
            (file_hash,),
        )
        for row in rows[3:]:
            semestre = row[2]
            if semestre is None:
                continue
            uf_m2 = row[3]
            vacancia_pct = row[4]
            anio, periodo_num = _anio_periodo_num(str(semestre))
            existing = conn.execute(
                "SELECT id FROM raw_mercado_bodegas_evolucion "
                "WHERE semestre = ? AND superseded_at IS NULL",
                (semestre,),
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE raw_mercado_bodegas_evolucion "
                    "SET uf_m2=?, vacancia_pct=?, file_hash=?, source_file=? WHERE id=?",
                    (uf_m2, vacancia_pct, file_hash, str(xlsx_path), existing[0]),
                )
            else:
                conn.execute(
                    """INSERT INTO raw_mercado_bodegas_evolucion
                        (semestre, anio, periodo_num, uf_m2, vacancia_pct, source_file, file_hash)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (semestre, anio, periodo_num, uf_m2, vacancia_pct, str(xlsx_path), file_hash),
                )
                n_inserted += 1
        conn.commit()
    finally:
        conn.close()
    return n_inserted


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else (
        r"C:\Users\raimundo.opazo\OneDrive - Toesca\Inmobiliario Toesca - Documentos"
        r"\RAW\mercado bodegas db.xlsx"
    )
    n = ingest(path)
    print(f"OK -> {n} filas nuevas insertadas en raw_mercado_bodegas_evolucion")
