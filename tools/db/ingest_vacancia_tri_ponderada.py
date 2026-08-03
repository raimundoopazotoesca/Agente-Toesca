"""Ingesta de la vacancia consolidada de TRI ya calculada por el usuario →
derived_kpi (entidad_tipo='fondo', entidad_key='TRI', kpi='vacancia_pct').

Fuente: 'RAW/Vacancia histórica DB.xlsx', hoja 'Hoja1', fila 37 "Vacancia
Ponderada Fondo Rentas" — el usuario agregó esta fila 2026-08-03 con su
propio cálculo ponderado (m² vacantes / m² útiles por activo, ponderado por
participación), cubriendo 2017-06 a 2026-06.

Reemplaza el cálculo desde cero que hacía _fetch_vacancia_tri en
scripts/build_factsheet.py (ver wiki/fondos/tri-consolidacion-ingresos-noi-
vacancia.md): ese cálculo, reconstruido a mano desde v_vacancia_activo /
v_vacancia_activo_tipo / v_vacancia_pt_consolidado_tipo, daba números que no
calzaban contra los del usuario en varios períodos históricos (ej. 2020-01:
6.9% calculado vs 5.84% real) — solo coincidía casi exacto en jun-2026
(5.96% vs 5.945%, el único mes con rent roll completo de los 9 activos).
En vez de seguir depurando la reconstrucción, se usa directamente el número
que el usuario ya valida como correcto.
"""
from __future__ import annotations

import hashlib
import sqlite3
from typing import Optional

import openpyxl

_SHEET_NAME = "Hoja1"
_DATE_ROW = 2
_VACANCIA_PONDERADA_ROW = 37
_KPI = "vacancia_pct"
_FORMULA = "vacancia_ponderada_fondo_rentas_manual (fila 37, Vacancia histórica DB.xlsx)"


def _file_hash(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_planilla(xlsx_path: str) -> dict[str, float]:
    """{periodo: vacancia_pct} desde la fila 37, en % (0-100)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[_SHEET_NAME]

    periodos: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=_DATE_ROW, column=col).value
        if hasattr(v, "year") and hasattr(v, "month"):
            periodos[col] = f"{v.year:04d}-{v.month:02d}"
    wb.close()

    out: dict[str, float] = {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[_SHEET_NAME]
    for col, periodo in periodos.items():
        v = ws.cell(row=_VACANCIA_PONDERADA_ROW, column=col).value
        if v is None:
            continue
        out[periodo] = round(float(v) * 100.0, 4)
    wb.close()
    return out


def persist(xlsx_path: str, conn: "sqlite3.Connection | None" = None) -> dict:
    from tools.db import repo_audit

    owns_conn = conn is None
    if owns_conn:
        from tools.db.connection import get_conn
        conn = get_conn()

    try:
        file_hash = _file_hash(xlsx_path)

        prev = conn.execute(
            "SELECT 1 FROM derived_kpi WHERE entidad_tipo='fondo' AND entidad_key='TRI' "
            "AND kpi=? AND formula=? LIMIT 1",
            (_KPI, _FORMULA),
        ).fetchone()

        serie = parse_planilla(xlsx_path)

        run_id = repo_audit.start_ingest_run(
            conn, tool="ingest_vacancia_tri_ponderada",
            source_file=xlsx_path, file_hash=file_hash,
        )

        conn.execute(
            "DELETE FROM derived_kpi WHERE entidad_tipo='fondo' AND entidad_key='TRI' "
            "AND kpi=? AND formula=?",
            (_KPI, _FORMULA),
        )
        for periodo, valor in serie.items():
            conn.execute(
                "INSERT INTO derived_kpi (entidad_tipo, entidad_key, periodo, kpi, valor, unidad, formula, ingest_run_id) "
                "VALUES ('fondo','TRI',?,?,?, '%', ?, ?)",
                (periodo, _KPI, valor, _FORMULA, run_id),
            )
        conn.commit()

        repo_audit.finish_ingest_run(
            conn, run_id, rows_in=len(serie), rows_loaded=len(serie), status="ok",
        )

        return {"status": "reinserted" if prev else "inserted",
                "rows": len(serie), "file_hash": file_hash, "ingest_run_id": run_id}
    finally:
        if owns_conn:
            conn.close()


def main(argv: Optional[list[str]] = None) -> int:
    import argparse
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="Path a 'Vacancia histórica DB.xlsx'")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)

    if args.dry_run:
        serie = parse_planilla(args.xlsx)
        periodos = sorted(serie)
        print(f"Parsed {len(serie)} periodos de {args.xlsx}")
        print(f"  rango: {periodos[0]}..{periodos[-1]}")
        for p in periodos[:3] + periodos[-3:]:
            print(f"    {p}: {serie[p]:.2f}%")
        return 0

    res = persist(args.xlsx)
    print(res)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
