"""Ingesta de tasaciones y valores de adquisición desde Excel.

Uso:
    python -X utf8 -m tools.db.ingest_tasaciones ruta/al/archivo.xlsx

Hoja de tasaciones: detectada por nombre (contiene "tasacion", ej.
'Consolidado Tasaciones' o 'Tasaciones'). Las columnas se resuelven por
**encabezado** (fila con celda "Tasador"), no por posición fija — soporta
tanto el layout legado (activo_key|periodo|fecha|tasador|valor_uf|...|notas)
como el de `tablaflujos.xlsx` (Fondo|Activo|Fecha/período|Tasador|
Tasación UF|Tasa dcto), y cualquier reordenamiento futuro de columnas.

La columna de activo puede traer el activo_key crudo (legado) o el nombre
de display del activo (ej. "Paseo Curicó") — _resolve_activo_key() normaliza
(minúsculas, sin tildes, espacios→"_") antes de buscar en _KEY_MAP, así que
funciona para ambos.

La columna de período puede traer solo el año (2020) o una fecha completa
('31-12-2024', datetime) — _parse_periodo_fecha() separa ambos casos.

'Adquisiciones': A=activo_key, B=fecha_adquisicion, C=precio_uf_fondo,
D=valor_activo_100%, E=m², F=UF/m², G=%adquirido, ..., P=notas

Machalí está excluido (cualquier variante de "Strip Center Machali").
"""
import re
import sys
import unicodedata
from pathlib import Path
import datetime

import openpyxl

from tools.db.connection import get_conn, apply_migrations, DEFAULT_DB_PATH
from tools.db import repo_tasacion, repo_audit


def _normalize(s: str) -> str:
    """minúsculas, sin tildes, no-alfanumérico → '_', underscores colapsados."""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return re.sub(r"_+", "_", s).strip("_")


_EXCLUIDOS_NORM = {_normalize("Strip Center Machali"), "machali"}

# Mapeo de activo (crudo o display, normalizado) → activo_key en dim_activo
_KEY_MAP_RAW = {
    "torre_a":                                          "Torre A",
    "apoquindo_4501":                                   "Apo4501",
    "apoquindo_4700":                                   "Apo4700",
    "apoquindo_3001":                                   "Apo3001",
    "inmobiliaria_boulevard":                           "Boulevard",
    "paseo_viña_centro":                                "Viña Centro",
    "paseo_curico":                                     "Mall Curicó",
    "paseo_curico_":                                    "Mall Curicó",
    "sucden":                                           "Sucden",
    "residencia_arturo_medina":                         "Residencia Arturo Medina",
    "residencia_candil":                                "Residencia Candil",
    "residencia_colombia":                              "Residencia Colombia",
    "residencia_coventry":                              "Residencia Coventry",
    "residencia_domingo_calderon":                      "Residencia Domingo Calderón",
    "residencia_padre_errazuriz__leonardo_da_vinci":    "Residencia Padre Errázuriz",
    "ed._guardiamarina":                                "Ed. Guardiamarina",
    "ed._placilla":                                     "Ed. Placilla",
}
# Normalizado una vez: acepta tanto claves crudas legado como nombres display
# ("Paseo Curicó", "Torre A", "Apoquindo 4501", ...) sin depender de que
# ambos lados usen la misma convención de tildes/guiones bajos.
_KEY_MAP = {_normalize(k): v for k, v in _KEY_MAP_RAW.items()}


def _resolve_activo_key(raw: str) -> str:
    norm = _normalize(raw)
    return _KEY_MAP.get(norm, raw)


def _is_excluido(raw: str) -> bool:
    return _normalize(raw) in _EXCLUIDOS_NORM


# ── Resolución de columnas por encabezado ───────────────────────────────────

_HEADER_ALIASES = {
    "activo_key":    ["activo", "activo_key"],
    "periodo":       ["periodo"],
    "fecha":         ["fecha"],  # si calza con el mismo header que 'periodo', se usa parseo combinado
    "tasador":       ["tasador"],
    "valor_uf":      ["valor_uf", "tasacion_uf", "valor_tasado", "valor_uf_tasacion"],
    "superficie_m2": ["superficie_m2", "m2", "m"],
    "uf_m2":         ["uf_m2", "uf_m"],
    "variacion_pct": ["variacion_pct", "variacion"],
    "tasa_dcto":     ["tasa_dcto", "tasa_descuento", "tasa"],
    "cap_rate":      ["cap_rate", "caprate"],
    "ltv":           ["ltv"],
    "ltc":           ["ltc"],
    "leverage_fin":  ["leverage_fin", "leverage"],
    "notas":         ["notas"],
}


def _find_header_row(ws, max_scan: int = 10):
    """Busca la fila de encabezado: la primera con una celda normalizada == 'tasador'."""
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        for cell in row:
            if cell and _normalize(str(cell)) == "tasador":
                return row
    return None


def _resolve_columns(header_row) -> dict:
    """Devuelve {campo_logico: indice_columna} resolviendo por alias de encabezado.

    Si 'periodo' y 'fecha' resuelven al mismo índice (o 'fecha' no aparece pero
    el header de 'periodo' contiene la palabra 'fecha'), se marca modo combinado
    seteando cols['periodo_fecha_combinado'] = True.
    """
    normalized = [(_normalize(str(c)) if c else "") for c in header_row]
    cols: dict = {}
    for campo, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                cols[campo] = normalized.index(alias)
                break
        else:
            # match parcial (ej. "fecha_periodo" contiene "fecha" y "periodo")
            for i, h in enumerate(normalized):
                if h and any(a in h for a in aliases):
                    cols[campo] = i
                    break

    if "periodo" in cols and ("fecha" not in cols or cols["fecha"] == cols["periodo"]):
        cols["periodo_fecha_combinado"] = True
        cols["fecha"] = cols["periodo"]

    return cols


def _parse_periodo_fecha(val):
    """Devuelve (periodo:'YYYY' str, fecha:'YYYY-MM-DD' str|None) desde un valor
    que puede ser año puro (int/str), datetime, o fecha 'DD-MM-YYYY' / 'YYYY-MM-DD'.
    """
    if val is None or val == "":
        return None, None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return str(val.year), val.strftime("%Y-%m-%d")

    s = str(val).strip()
    if re.fullmatch(r"\d{4}", s):
        return s, None
    m = re.fullmatch(r"(\d{1,2})-(\d{1,2})-(\d{4})", s)  # DD-MM-YYYY
    if m:
        d, mo, y = m.groups()
        return y, f"{y}-{int(mo):02d}-{int(d):02d}"
    m = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)  # YYYY-MM-DD
    if m:
        y, mo, d = m.groups()
        return y, f"{y}-{int(mo):02d}-{int(d):02d}"
    return None, None


_ADQ_COLS = {
    "activo_key":            0,
    "fecha_adquisicion":     1,
    "precio_uf":             2,
    "valor_activo_uf":       3,
    "superficie_m2":         4,
    "uf_m2":                 5,
    "porcentaje_adquirido":  6,
    "notas":                15,   # col P
}


def _cell(row: tuple, idx: int):
    """Devuelve el valor de la celda o None si está fuera de rango."""
    try:
        val = row[idx]
    except IndexError:
        return None
    if val == "" or val is None:
        return None
    return val


def _float(row: tuple, idx: int):
    val = _cell(row, idx)
    if val is None:
        return None
    try:
        return float(str(val).replace("%", "").replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def _str(row: tuple, idx: int):
    val = _cell(row, idx)
    return str(val).strip() if val is not None else None


def _date_str(row: tuple, idx: int) -> str | None:
    """Convierte datetime o string de fecha a YYYY-MM-DD."""
    val = _cell(row, idx)
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s[:10] if s else None


def ingest_tasaciones(path: str) -> dict:
    """Ingesta un Excel de tasaciones y adquisiciones.

    Devuelve un dict con contadores de filas procesadas.
    """
    apply_migrations(DEFAULT_DB_PATH)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    counts = {"tasaciones_ok": 0, "tasaciones_skip": 0,
              "adquisiciones_ok": 0, "adquisiciones_skip": 0, "errores": []}

    with get_conn() as conn:
        run_id = repo_audit.start_ingest_run(conn, tool="ingest_tasaciones", source_file=path, file_hash=None)

        # ── Hoja de tasaciones (layout resuelto por encabezado) ─────────────
        sheet_tas = next((s for s in wb.sheetnames if "tasacion" in s.lower()), None)
        if sheet_tas:
            ws = wb[sheet_tas]
            header_row = _find_header_row(ws)
            if header_row is None:
                counts["errores"].append(
                    f"Hoja '{sheet_tas}': no se encontró fila de encabezado (celda 'Tasador')."
                )
            else:
                tcols = _resolve_columns(header_row)
                faltantes = {"activo_key", "periodo", "tasador", "valor_uf"} - tcols.keys()
                if faltantes:
                    counts["errores"].append(
                        f"Hoja '{sheet_tas}': faltan columnas requeridas {faltantes} en el encabezado."
                    )
                else:
                    # Ubicar la fila real del encabezado para saber dónde empiezan los datos
                    header_row_num = next(
                        i for i, row in enumerate(
                            ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1
                        ) if row == header_row
                    )
                    for i, row in enumerate(
                        ws.iter_rows(min_row=header_row_num + 1, values_only=True),
                        start=header_row_num + 1,
                    ):
                        raw_activo = _str(row, tcols["activo_key"])
                        tasador = _str(row, tcols["tasador"])
                        valor_uf = _float(row, tcols["valor_uf"])
                        if not raw_activo or not tasador or valor_uf is None:
                            counts["tasaciones_skip"] += 1
                            continue
                        if _is_excluido(raw_activo):
                            counts["tasaciones_skip"] += 1
                            continue

                        periodo, fecha_combinada = _parse_periodo_fecha(_cell(row, tcols["periodo"]))
                        if periodo is None:
                            counts["tasaciones_skip"] += 1
                            counts["errores"].append(f"Tasaciones fila {i}: período no parseable ({_cell(row, tcols['periodo'])!r})")
                            continue
                        fecha = fecha_combinada if tcols.get("periodo_fecha_combinado") else _date_str(row, tcols.get("fecha", tcols["periodo"]))

                        activo_key = _resolve_activo_key(raw_activo)
                        try:
                            repo_tasacion.upsert_tasacion(
                                conn,
                                activo_key=activo_key,
                                periodo=periodo,
                                tasador=tasador,
                                fecha=fecha,
                                valor_uf=valor_uf,
                                superficie_m2=_float(row, tcols["superficie_m2"]) if "superficie_m2" in tcols else None,
                                uf_m2=_float(row, tcols["uf_m2"]) if "uf_m2" in tcols else None,
                                variacion_pct=_float(row, tcols["variacion_pct"]) if "variacion_pct" in tcols else None,
                                tasa_dcto=_float(row, tcols["tasa_dcto"]) if "tasa_dcto" in tcols else None,
                                cap_rate=_float(row, tcols["cap_rate"]) if "cap_rate" in tcols else None,
                                ltv=_float(row, tcols["ltv"]) if "ltv" in tcols else None,
                                ltc=_float(row, tcols["ltc"]) if "ltc" in tcols else None,
                                leverage_fin=_float(row, tcols["leverage_fin"]) if "leverage_fin" in tcols else None,
                                notas=_str(row, tcols["notas"]) if "notas" in tcols else None,
                                ingest_run_id=run_id,
                            )
                            counts["tasaciones_ok"] += 1
                        except Exception as e:
                            counts["errores"].append(f"Tasaciones fila {i}: {e}")
        else:
            counts["errores"].append("Hoja de tasaciones no encontrada en el Excel.")

        # ── Hoja Adquisiciones ───────────────────────────────────────────────
        if "Adquisiciones" in wb.sheetnames:
            ws = wb["Adquisiciones"]
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                raw_key           = _str(row, _ADQ_COLS["activo_key"])
                fecha_adquisicion = _date_str(row, _ADQ_COLS["fecha_adquisicion"])
                if not raw_key:
                    counts["adquisiciones_skip"] += 1
                    continue
                if _is_excluido(raw_key):
                    counts["adquisiciones_skip"] += 1
                    continue
                activo_key = _resolve_activo_key(raw_key)
                # Si no hay fecha, usar solo el año de adquisición si lo hay
                if not fecha_adquisicion:
                    año = _cell(row, 12)  # col M = año_adquisición
                    if año and str(año).strip().isdigit():
                        fecha_adquisicion = f"{str(año).strip()[:4]}-01-01"
                    else:
                        counts["adquisiciones_skip"] += 1
                        continue
                try:
                    repo_tasacion.upsert_adquisicion(
                        conn,
                        activo_key=activo_key,
                        fecha_adquisicion=fecha_adquisicion,
                        precio_uf=_float(row, _ADQ_COLS["precio_uf"]),
                        valor_activo_uf=_float(row, _ADQ_COLS["valor_activo_uf"]),
                        superficie_m2=_float(row, _ADQ_COLS["superficie_m2"]),
                        uf_m2=_float(row, _ADQ_COLS["uf_m2"]),
                        porcentaje_adquirido=_float(row, _ADQ_COLS["porcentaje_adquirido"]),
                        notas=_str(row, _ADQ_COLS["notas"]) or _str(row, 7),
                        ingest_run_id=run_id,
                    )
                    counts["adquisiciones_ok"] += 1
                except Exception as e:
                    counts["errores"].append(f"Adquisiciones fila {i}: {e}")
        else:
            counts["errores"].append("Hoja 'Adquisiciones' no encontrada en el Excel.")

        total = counts["tasaciones_ok"] + counts["adquisiciones_ok"]
        repo_audit.finish_ingest_run(conn, run_id, rows_in=total, rows_loaded=total)

    wb.close()
    return counts


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python -X utf8 -m tools.db.ingest_tasaciones <ruta_excel>")
        sys.exit(1)
    path = sys.argv[1]
    if not Path(path).exists():
        print(f"Archivo no encontrado: {path}")
        sys.exit(1)
    result = ingest_tasaciones(path)
    print(f"Tasaciones: {result['tasaciones_ok']} ok, {result['tasaciones_skip']} skip")
    print(f"Adquisiciones: {result['adquisiciones_ok']} ok, {result['adquisiciones_skip']} skip")
    if result["errores"]:
        print("Errores:")
        for e in result["errores"]:
            print(f"  {e}")
