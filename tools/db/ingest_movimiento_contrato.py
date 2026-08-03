"""Ingesta de SharePoint RAW/"Absorcion Histórica DB.xlsx" -> raw_movimiento_contrato.

Formato fuente: 3 hojas (jll, viña, curico), una fila por movimiento de
contrato (nuevo/renovó/término), columnas desde B. Insumo para derivar
absorción histórica; ver 075_movimiento_contrato.sql.
"""
import argparse
import hashlib
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from tools.db.connection import get_conn

ACTIVO_MAP = {
    "Apoquindo 3001": "Apo3001",
    "Apoquindo 4501": "Apo4501",
    "Apoquindo 4700": "Apo4700",
    "Inm Boulevard PT": "Boulevard",
    "Torre A": "Torre A",
    "Viña Centro": "Viña Centro",
    "Curicó": "Mall Curicó",
}

TIPO_UNIDAD_MAP = {
    "bodega": "Bodega",
    "bodegas": "Bodega",
    "estacionamientos": "Estacionamientos",
    "estacionamiento": "Estacionamientos",
    "local": "Local",
    "oficina": "Oficina",
    "espacio": "Espacio",
    "modulo": "Módulo",
    "módulo": "Módulo",
}

# hoja -> (fila inicio datos, fila fin datos, tiene columnas UF totales)
SHEET_CFG = {
    "jll": (5, 478, True),
    "viña": (3, 361, False),
    "curico": (3, 352, True),
}

# columnas (1-indexed) comunes desde B: Activo, Tipo Activo, Status, Arrendatario,
# Nuevo Arrendatario, [Antes(UF), Hoy(UF),] Antes(UF/m2), Hoy(UF/m2), M2, %,
# Vencimiento, Inicio Nuevo Contrato, Nuevo Vencimiento, [Comentarios]


def _file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _norm_tipo(raw):
    if not raw:
        return None
    return TIPO_UNIDAD_MAP.get(raw.strip().lower(), raw.strip())


def _to_iso(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v)


def _read_sheet(ws, sheet_name, start_row, end_row, has_uf_total):
    rows = []
    for r in range(start_row, end_row + 1):
        activo_raw = ws.cell(row=r, column=2).value
        if not activo_raw:
            continue
        activo_key = ACTIVO_MAP.get(activo_raw.strip())
        if not activo_key:
            print(f"[WARN] {sheet_name} fila {r}: activo sin mapear {activo_raw!r}")
            continue

        if has_uf_total:
            cols = {
                "tipo_unidad": 3, "status": 4, "arrendatario": 5,
                "nuevo_arrendatario": 6, "antes_uf": 7, "hoy_uf": 8,
                "antes_uf_m2": 9, "hoy_uf_m2": 10, "m2": 11, "pct_variacion": 12,
                "vencimiento": 13, "inicio_nuevo_contrato": 14,
                "nuevo_vencimiento": 15, "comentarios": 16,
            }
        else:
            cols = {
                "tipo_unidad": 3, "status": 4, "arrendatario": 5,
                "nuevo_arrendatario": 6, "antes_uf_m2": 7, "hoy_uf_m2": 8,
                "m2": 9, "pct_variacion": 10, "vencimiento": 11,
                "inicio_nuevo_contrato": 12, "nuevo_vencimiento": 13,
                "comentarios": 14,
            }

        def cell(name):
            col = cols.get(name)
            return ws.cell(row=r, column=col).value if col else None

        rows.append({
            "activo_key": activo_key,
            "tipo_unidad": _norm_tipo(cell("tipo_unidad")),
            "status": cell("status"),
            "arrendatario": cell("arrendatario"),
            "nuevo_arrendatario": cell("nuevo_arrendatario"),
            "antes_uf": cell("antes_uf"),
            "hoy_uf": cell("hoy_uf"),
            "antes_uf_m2": cell("antes_uf_m2"),
            "hoy_uf_m2": cell("hoy_uf_m2"),
            "m2": cell("m2"),
            "pct_variacion": cell("pct_variacion"),
            "vencimiento": _to_iso(cell("vencimiento")),
            "inicio_nuevo_contrato": _to_iso(cell("inicio_nuevo_contrato")),
            "nuevo_vencimiento": _to_iso(cell("nuevo_vencimiento")),
            "comentarios": cell("comentarios"),
            "source_row": r,
        })
    return rows


def ingest(path: Path, dry_run: bool = False, record_path: Path | None = None) -> int:
    record_path = record_path or path
    file_hash = _file_hash(path)
    wb = openpyxl.load_workbook(path, data_only=True)

    all_rows = []
    for sheet_name, (start, end, has_uf_total) in SHEET_CFG.items():
        ws = wb[sheet_name]
        rows = _read_sheet(ws, sheet_name, start, end, has_uf_total)
        for row in rows:
            row["source_sheet"] = sheet_name
        all_rows.extend(rows)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        "INSERT INTO ingest_run (tool, source_file, file_hash) VALUES (?, ?, ?)",
        ("ingest_movimiento_contrato", str(record_path), file_hash),
    )
    ingest_run_id = cur.lastrowid

    cur.execute(
        """
        UPDATE raw_movimiento_contrato
           SET superseded_at = datetime('now')
         WHERE source_file = ? AND superseded_at IS NULL
        """,
        (str(record_path),),
    )

    n = 0
    for row in all_rows:
        cur.execute(
            """
            INSERT OR IGNORE INTO raw_movimiento_contrato
                (activo_key, tipo_unidad, status, arrendatario, nuevo_arrendatario,
                 antes_uf, hoy_uf, antes_uf_m2, hoy_uf_m2, m2, pct_variacion,
                 vencimiento, inicio_nuevo_contrato, nuevo_vencimiento, comentarios,
                 source_file, source_sheet, source_row, file_hash, ingest_run_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row["activo_key"], row["tipo_unidad"], row["status"],
                row["arrendatario"], row["nuevo_arrendatario"],
                row["antes_uf"], row["hoy_uf"], row["antes_uf_m2"], row["hoy_uf_m2"],
                row["m2"], row["pct_variacion"],
                row["vencimiento"], row["inicio_nuevo_contrato"], row["nuevo_vencimiento"],
                row["comentarios"],
                str(record_path), row["source_sheet"], row["source_row"],
                file_hash, ingest_run_id,
            ),
        )
        n += cur.rowcount

    if dry_run:
        conn.rollback()
        print(f"[DRY RUN] {n} filas se habrian insertado")
    else:
        conn.commit()
        print(f"{n} filas insertadas (ingest_run_id={ingest_run_id})")

    conn.close()
    return n


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("path", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--record-path", type=Path, default=None)
    args = parser.parse_args()
    ingest(args.path, dry_run=args.dry_run, record_path=args.record_path)
