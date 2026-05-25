"""
Backfill histórico de la DB del agente (Fase 2).

Recorre los archivos de proveedor ya sincronizados en SharePoint y los ingesta
reusando las MISMAS funciones del flujo en vivo (consistencia garantizada).
Idempotente: reejecutar no duplica (UNIQUE file_hash+source_row).

Uso:
    python -m tools.db.backfill                 # backfill de todo lo disponible
    python -m tools.db.backfill rent_roll       # solo rent rolls
"""
import glob
import os
import re
import sys

from tools.sharepoint_paths import (
    RR_JLL_DIR,
    TRI_VINA_RENT_ROLL_DIR,
    TRI_CURICO_RENT_ROLL_DIR,
)

_MES_NUM = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


def _periodo_jll(fname: str) -> str | None:
    """'2509 Rent Roll y NOI.xlsx' → '2025-09'."""
    m = re.match(r"(\d{2})(\d{2})\b", os.path.basename(fname))
    if not m:
        return None
    return f"20{m.group(1)}-{m.group(2)}"


def _periodo_tresa(fname: str) -> str | None:
    """'Excel Tres A Viña Marzo 2026.xlsx' → '2026-03'."""
    low = os.path.basename(fname).lower()
    año = None
    ma = re.search(r"(20\d{2})", low)
    if ma:
        año = ma.group(1)
    mes = None
    for nombre, num in _MES_NUM.items():
        if nombre in low:
            mes = num
            break
    if año and mes:
        return f"{año}-{mes:02d}"
    return None


def _listar_xlsx(base_dir: str) -> list[str]:
    if not os.path.isdir(base_dir):
        return []
    out = []
    # Estructura típica: base/{año}/archivo.xlsx ; también base/archivo.xlsx
    out += glob.glob(os.path.join(base_dir, "*.xlsx"))
    out += glob.glob(os.path.join(base_dir, "*", "*.xlsx"))
    # Excluir temporales de Excel (~$...)
    return sorted(p for p in out if not os.path.basename(p).startswith("~$"))


def backfill_rent_roll(verbose: bool = True) -> dict:
    """Backfill de rent rolls JLL + Tres A (Viña/Curicó)."""
    import tools.rentroll_tools as rr

    fuentes = [
        ("jll", RR_JLL_DIR, _periodo_jll),
        ("vina", TRI_VINA_RENT_ROLL_DIR, _periodo_tresa),
        ("curico", TRI_CURICO_RENT_ROLL_DIR, _periodo_tresa),
    ]
    reporte = {"archivos": 0, "filas": 0, "sin_periodo": [], "sin_datos": [], "detalle": []}

    for proveedor, base, parse_periodo in fuentes:
        for path in _listar_xlsx(base):
            periodo = parse_periodo(path)
            bn = os.path.basename(path)
            if not periodo:
                reporte["sin_periodo"].append(bn)
                continue
            try:
                data = rr._read_source_data(path)
            except Exception as e:
                reporte["sin_datos"].append(f"{bn}: {e}")
                continue
            if not data:
                reporte["sin_datos"].append(f"{bn}: hoja 'Rent Roll' vacía o ausente")
                continue
            n = rr._persist_rent_roll(path, periodo, data, proveedor)
            reporte["archivos"] += 1
            reporte["filas"] += n
            reporte["detalle"].append(f"{proveedor} {periodo}: {n} filas <- {bn}")
            if verbose:
                print(f"  [{proveedor}] {periodo}: {n} filas <- {bn}")

    return reporte


def backfill_er(verbose: bool = True) -> dict:
    """Backfill de ER Viña/Curicó desde los INFORME EEFF (raw_er_activo_line)."""
    import tools.noi_tools as noi
    from tools.sharepoint_paths import TRI_VINA_EEFF_DIR, TRI_CURICO_EEFF_DIR

    fuentes = [("vina", TRI_VINA_EEFF_DIR), ("curico", TRI_CURICO_EEFF_DIR)]
    rep = {"archivos": 0, "filas": 0, "sin_datos": [], "detalle": []}
    for mall, base in fuentes:
        for path in _listar_xlsx(base):
            bn = os.path.basename(path)
            try:
                fecha_cierre, eeff_values = noi._leer_eeff_estado_resultado(path)
            except Exception as e:
                rep["sin_datos"].append(f"{bn}: {e}")
                continue
            if not eeff_values or fecha_cierre is None:
                rep["sin_datos"].append(f"{bn}: sin ESTADO DE RESULTADO o fecha")
                continue
            periodo = f"{fecha_cierre.year}-{fecha_cierre.month:02d}"
            n = noi._persist_er_lines(mall, path, periodo, eeff_values)
            rep["archivos"] += 1
            rep["filas"] += n
            rep["detalle"].append(f"{mall} {periodo}: {n} filas <- {bn}")
            if verbose:
                print(f"  [er-{mall}] {periodo}: {n} filas <- {bn}")
    return rep


def backfill_inmosa(verbose: bool = True) -> dict:
    """Backfill de flujos INMOSA (raw_flujo_line). El archivo tiene meses en columnas."""
    import openpyxl
    from datetime import date, datetime
    import tools.noi_tools as noi
    from tools.sharepoint_paths import TRI_INMOSA_FLUJOS_DIR

    rep = {"archivos": 0, "filas": 0, "sin_datos": [], "detalle": []}
    for path in _listar_xlsx(TRI_INMOSA_FLUJOS_DIR):
        bn = os.path.basename(path)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            rep["sin_datos"].append(f"{bn}: {e}")
            continue
        # Misma selección de hoja que actualizar_noi_inmosa
        target = None
        for s in wb.sheetnames:
            if any(k in s.upper() for k in ("NOI", "ESTADO", "RESULT")):
                target = s
                break
        if target is None:
            for s in wb.sheetnames:
                if any(k in s.upper() for k in ("ER", "FLUJO")):
                    target = s
                    break
        if target is None and wb.sheetnames:
            target = wb.sheetnames[0]
        if target is None:
            wb.close()
            rep["sin_datos"].append(f"{bn}: sin hojas")
            continue
        rows = list(wb[target].iter_rows(values_only=True))
        wb.close()

        # Detectar columna→fecha (primera fecha vista por columna)
        col_dates: dict = {}
        for row in rows:
            for ci, v in enumerate(row):
                if isinstance(v, (datetime, date)):
                    d = v.date() if isinstance(v, datetime) else v
                    col_dates.setdefault(ci, d)
        if not col_dates:
            rep["sin_datos"].append(f"{bn}: sin columnas de fecha en hoja '{target}'")
            continue

        for ci, d in col_dates.items():
            periodo = f"{d.year}-{d.month:02d}"
            er_data: dict = {}
            for row in rows:
                label_raw = row[0] if row[0] is not None else (row[1] if len(row) > 1 else None)
                if label_raw is None:
                    continue
                label = " ".join(str(label_raw).strip().split())
                if len(row) > ci and row[ci] is not None and not isinstance(row[ci], (datetime, date)):
                    try:
                        er_data[label] = float(row[ci])
                    except (TypeError, ValueError):
                        pass
            if not er_data:
                continue
            n = noi._persist_flujo_lines(
                "INMOSA", path, target, periodo, er_data,
                tool="backfill_inmosa", hash_extra=periodo,
            )
            if n:
                rep["archivos"] += 1
                rep["filas"] += n
                rep["detalle"].append(f"INMOSA {periodo}: {n} filas <- {bn}")
                if verbose:
                    print(f"  [inmosa] {periodo}: {n} filas <- {bn}")
    return rep


def backfill_eeff_valor_cuota(verbose: bool = True) -> dict:
    """Backfill de valor cuota libro desde PDFs de EEFF (derived_kpi). Sparse (regex)."""
    import tools.eeff_tools as eeff

    rep = {"intentos": 0, "ok": 0, "sin_dato": [], "detalle": []}
    meses_cierre = [3, 6, 9, 12]
    for fondo_key, base in eeff.FONDO_RUTAS.items():
        if not os.path.isdir(base):
            continue
        for año_dir in sorted(os.listdir(base)):
            ruta_año = os.path.join(base, año_dir)
            if not os.path.isdir(ruta_año) or not año_dir.isdigit():
                continue
            año = int(año_dir)
            for mes in meses_cierre:
                pdf = eeff.buscar_pdf_eeff(fondo_key, año, mes)
                if not os.path.isfile(pdf):
                    continue
                rep["intentos"] += 1
                datos = eeff.extraer_datos_eeff(pdf, fondo_key)
                vc = datos.get("valor_cuota") or {}
                if not vc:
                    rep["sin_dato"].append(f"{fondo_key} {año}-{mes:02d}: {os.path.basename(pdf)}")
                    continue
                eeff._persist_valor_cuota_libro(fondo_key, f"{año}-{mes:02d}", vc)
                rep["ok"] += 1
                rep["detalle"].append(f"{fondo_key} {año}-{mes:02d}: {vc}")
                if verbose:
                    print(f"  [eeff] {fondo_key} {año}-{mes:02d}: {vc}")
    return rep


def backfill_precios(verbose: bool = True) -> dict:
    """Backfill de precios de cuota: una llamada datachart por nemotécnico,
    persiste el último día bursátil de cada mes disponible (fact_precio_cuota)."""
    import tools.web_bursatil_tools as wb
    from tools.db.connection import get_conn
    from tools.db import repo_fact

    rep = {"nemos": 0, "filas": 0, "errores": []}
    for nemo in wb.NEMOTECNICOS:
        try:
            notation_id = wb._get_notation_id(nemo)
            url = (f"https://mercados.larrainvial.com/www/datachart.html"
                   f"?ID_NOTATION={notation_id}&TIME_SPAN=2Y&QUALITY=1&VOLUME=true")
            raw = wb._fetch(url)
            history = wb._parse_datachart(raw)
        except Exception as e:
            rep["errores"].append(f"{nemo}: {e}")
            continue
        if not history:
            rep["errores"].append(f"{nemo}: sin datos")
            continue
        # último día bursátil por (año, mes)
        por_mes: dict = {}
        for e in history:
            key = (e["year"], e["js_month"] + 1)
            por_mes[key] = e  # history en orden → último gana
        n = 0
        with get_conn() as conn:
            for (año, mes), e in por_mes.items():
                fecha = f"{año}-{mes:02d}-{e['day']:02d}"
                repo_fact.upsert_precio(conn, nemo, fecha, e["close"], "LarraínVial")
                n += 1
        rep["nemos"] += 1
        rep["filas"] += n
        if verbose:
            print(f"  [precios] {nemo}: {n} meses")
    return rep


def _find_cdg() -> str | None:
    """Encuentra el CDG más reciente (WORK_DIR o SharePoint Control de Gestión)."""
    from config import WORK_DIR, SHAREPOINT_DIR
    cands = []
    bases = [WORK_DIR, os.path.join(SHAREPOINT_DIR, "Control de Gestión")]
    for base in bases:
        if os.path.isdir(base):
            cands += glob.glob(os.path.join(base, "**", "*Control De Gesti*.xlsx"), recursive=True)
    cands = [c for c in set(cands) if not os.path.basename(c).startswith("~$")]
    return sorted(cands)[-1] if cands else None


def backfill_uf(verbose: bool = True) -> dict:
    """Backfill de UF diaria desde la hoja 'UF' del CDG más reciente (fact_uf)."""
    import openpyxl
    from datetime import date, datetime
    from tools.db.connection import get_conn
    from tools.db import repo_fact

    rep = {"archivos": 0, "filas": 0, "sin_datos": []}
    cdg = _find_cdg()
    if not cdg:
        rep["sin_datos"].append("No se encontró ningún CDG")
        return rep
    try:
        wb = openpyxl.load_workbook(cdg, read_only=True, data_only=True)
    except Exception as e:
        rep["sin_datos"].append(f"{os.path.basename(cdg)}: {e}")
        return rep
    if "UF" not in wb.sheetnames:
        wb.close()
        rep["sin_datos"].append(f"{os.path.basename(cdg)}: sin hoja 'UF'")
        return rep
    ws = wb["UF"]
    pares = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        fecha_cell, valor_cell = (row[0] if len(row) > 0 else None), (row[1] if len(row) > 1 else None)
        if fecha_cell is None or valor_cell is None:
            continue
        if isinstance(fecha_cell, datetime):
            fecha_cell = fecha_cell.date()
        if not isinstance(fecha_cell, date):
            continue
        try:
            valor = float(valor_cell)
        except (TypeError, ValueError):
            continue
        pares.append((fecha_cell.isoformat(), valor))
    wb.close()
    if pares:
        with get_conn() as conn:
            for fecha, valor in pares:
                repo_fact.upsert_uf(conn, fecha, valor)
        rep["archivos"] = 1
        rep["filas"] = len(pares)
        if verbose:
            print(f"  [uf] {len(pares)} días <- {os.path.basename(cdg)} "
                  f"({pares[0][0]}..{pares[-1][0]})")
    return rep


_RENTAS_SERIE_NEMO = {"A": "CFITOERI1A", "C": "CFITOERI1C", "I": "CFITOERI1I"}

# Filas de la hoja 'Vacancia' del CDG → nombre de segmento (m² vacantes).
_VACANCIA_ROWS = {
    47: "INMOSA", 48: "Machalí", 49: "SUCDEN",
    50: "PT Oficinas", 51: "PT Locales", 52: "PT Bodegas",
    53: "Viña Centro", 54: "Apoquindo 4700", 55: "Apoquindo 4501",
    56: "Fondo Apoquindo", 57: "Curicó", 58: "Apoquindo 3001",
}


def backfill_vacancia(verbose: bool = True) -> dict:
    """Espeja la vacancia oficial (m² vacantes) de la hoja 'Vacancia' del CDG a derived_kpi.

    Fila 46 = fechas; filas 47-58 = segmentos. kpi='m2_vacantes', entidad_tipo='activo'.
    """
    import openpyxl
    from datetime import date, datetime
    from tools.db.connection import get_conn
    from tools.db import repo_kpi

    rep = {"archivos": 0, "filas": 0, "sin_datos": []}
    cdg = _find_cdg()
    if not cdg:
        rep["sin_datos"].append("No se encontró ningún CDG")
        return rep
    try:
        wb = openpyxl.load_workbook(cdg, read_only=True, data_only=True)
    except Exception as e:
        rep["sin_datos"].append(f"{os.path.basename(cdg)}: {e}")
        return rep
    if "Vacancia" not in wb.sheetnames:
        wb.close()
        rep["sin_datos"].append(f"{os.path.basename(cdg)}: sin hoja 'Vacancia'")
        return rep
    # En read_only, ws.cell(r,c) es O(n) por llamada → iterar filas UNA vez.
    filas: dict = {}
    for i, row in enumerate(wb["Vacancia"].iter_rows(min_row=46, max_row=58, values_only=True)):
        filas[46 + i] = row
    wb.close()

    row46 = filas.get(46, ())
    # índice de columna → periodo (solo fechas día=1: los headers de vacancia son mensuales)
    col_periodo: dict = {}
    for ci, v in enumerate(row46):
        if isinstance(v, datetime):
            v = v.date()
        if isinstance(v, date) and v.day == 1:
            col_periodo[ci] = f"{v.year}-{v.month:02d}"

    n = 0
    with get_conn() as conn:
        for row, seg in _VACANCIA_ROWS.items():
            rowvals = filas.get(row, ())
            for ci, periodo in col_periodo.items():
                if ci >= len(rowvals):
                    continue
                val = rowvals[ci]
                if val in (None, "-", ""):
                    continue
                try:
                    m2 = float(val)
                except (TypeError, ValueError):
                    continue
                repo_kpi.upsert(conn, "activo", seg, periodo,
                                "m2_vacantes", m2, "m2", "cdg_vacancia_v1")
                n += 1
    rep["archivos"] = 1
    rep["filas"] = n
    if verbose:
        print(f"  [vacancia] {n} valores ({len(col_periodo)} meses × segmentos) <- {os.path.basename(cdg)}")
    return rep


def _find_header_dividendos(rows: list) -> int | None:
    """Fila (0-based) cuyo header tiene 'Detalle' (col E) y 'Tipo' (col G)."""
    for i, r in enumerate(rows):
        e = r[4] if len(r) > 4 else None
        g = r[6] if len(r) > 6 else None
        if isinstance(e, str) and e.strip().lower() == "detalle" and \
           isinstance(g, str) and g.strip().lower() == "tipo":
            return i
    return None


def backfill_dividendos(verbose: bool = True) -> dict:
    """Backfill de dividendos desde el CDG (hojas A&R *).

    PT y Rentas (con nemotécnico) → fact_dividendo.
    Apoquindo (sin nemotécnico) → derived_kpi (kpi='dividendo_por_cuota', entidad fondo).
    Monto = 'Monto $ / cuota' (col I). Fecha = col D.
    """
    from datetime import date, datetime
    import openpyxl
    from tools.db.connection import get_conn
    from tools.db import repo_fact, repo_kpi

    rep = {"archivos": 0, "filas": 0, "sin_datos": [], "detalle": []}
    cdg = _find_cdg()
    if not cdg:
        rep["sin_datos"].append("No se encontró ningún CDG")
        return rep
    try:
        wb = openpyxl.load_workbook(cdg, read_only=True, data_only=True)
    except Exception as e:
        rep["sin_datos"].append(f"{os.path.basename(cdg)}: {e}")
        return rep

    bn = os.path.basename(cdg)
    n_fact = n_kpi = 0
    with get_conn() as conn:
        for hoja, fondo_key in [("A&R PT", "A&R PT"), ("A&R Rentas", "A&R Rentas"),
                                ("A&R Apoquindo", "A&R Apoquindo")]:
            if hoja not in wb.sheetnames:
                continue
            rows = list(wb[hoja].iter_rows(values_only=True))
            h = _find_header_dividendos(rows)
            if h is None:
                rep["sin_datos"].append(f"{hoja}: sin header Detalle/Tipo")
                continue
            for r in rows[h + 1:]:
                detalle = r[4] if len(r) > 4 else None
                if not (isinstance(detalle, str) and detalle.strip().lower() == "dividendo"):
                    continue
                fecha = r[3] if len(r) > 3 else None
                monto = r[8] if len(r) > 8 else None  # Monto $ / cuota
                if fecha is None or monto is None:
                    continue
                if isinstance(fecha, datetime):
                    fecha = fecha.date()
                if not isinstance(fecha, date):
                    continue
                try:
                    monto = float(monto)
                except (TypeError, ValueError):
                    continue
                if hoja == "A&R PT":
                    repo_fact.upsert_dividendo(conn, "CFITRIPT-E", fecha.isoformat(), monto)
                    n_fact += 1
                elif hoja == "A&R Rentas":
                    serie = str(r[5]).strip() if len(r) > 5 and r[5] is not None else None
                    nemo = _RENTAS_SERIE_NEMO.get(serie)
                    if nemo is None:
                        continue
                    repo_fact.upsert_dividendo(conn, nemo, fecha.isoformat(), monto)
                    n_fact += 1
                else:  # Apoquindo: fondo-level, sin nemotécnico
                    repo_kpi.upsert(
                        conn, "fondo", fondo_key, fecha.isoformat(),
                        "dividendo_por_cuota", monto, "CLP", "cdg_dividendo_v1",
                    )
                    n_kpi += 1

    wb.close()
    rep["archivos"] = 1
    rep["filas"] = n_fact + n_kpi
    rep["detalle"].append(f"fact_dividendo: {n_fact} | derived_kpi (Apoquindo): {n_kpi}")
    if verbose:
        print(f"  [dividendos] fact_dividendo: {n_fact} | Apoquindo→kpi: {n_kpi} <- {bn}")
    return rep


def _print_reporte(nombre: str, rep: dict) -> None:
    print(f"\n=== Backfill {nombre} ===")
    print(f"Archivos ingestados: {rep['archivos']}  |  Filas insertadas: {rep['filas']}")
    if rep.get("sin_periodo"):
        print(f"Sin período detectable ({len(rep['sin_periodo'])}): {rep['sin_periodo']}")
    if rep.get("sin_datos"):
        print(f"Sin datos ({len(rep['sin_datos'])}):")
        for s in rep["sin_datos"]:
            print(f"  - {s}")


def main(argv: list[str]) -> None:
    dominios = argv[1:] or ["rent_roll", "er", "inmosa", "uf", "eeff", "precios", "dividendos", "vacancia"]
    if "rent_roll" in dominios:
        _print_reporte("rent_roll", backfill_rent_roll(verbose=True))
    if "er" in dominios:
        _print_reporte("er", backfill_er(verbose=True))
    if "inmosa" in dominios:
        _print_reporte("inmosa", backfill_inmosa(verbose=True))
    if "uf" in dominios:
        _print_reporte("uf", backfill_uf(verbose=True))
    if "dividendos" in dominios:
        _print_reporte("dividendos", backfill_dividendos(verbose=True))
    if "vacancia" in dominios:
        _print_reporte("vacancia", backfill_vacancia(verbose=True))
    if "eeff" in dominios:
        rep = backfill_eeff_valor_cuota(verbose=True)
        print(f"\n=== Backfill eeff (valor cuota) ===")
        print(f"Intentos: {rep['intentos']}  |  Con dato: {rep['ok']}")
        if rep["sin_dato"]:
            print(f"Sin valor cuota auto ({len(rep['sin_dato'])}):")
            for s in rep["sin_dato"]:
                print(f"  - {s}")
    if "precios" in dominios:
        rep = backfill_precios(verbose=True)
        print(f"\n=== Backfill precios ===")
        print(f"Nemotécnicos: {rep['nemos']}  |  Filas: {rep['filas']}")
        if rep["errores"]:
            for e in rep["errores"]:
                print(f"  - {e}")


if __name__ == "__main__":
    main(sys.argv)
