"""Ingesta mensual del parking Parque Titanium desde la planilla
'MM-YYYY Liquidacion Parque Titanium.xlsx' (SABA).

Ubicación operativa:
SharePoint > Inmobiliario Toesca > Renta Comercial > Fondo PT > Parking PT.
El archivo lo envía Marcos Quiroga por mail.
El período a ingestar puede ser anterior o igual al mes del archivo.

Formato de la planilla (distinto al histórico 'Parking PT DB.xlsx'):
- Hoja 'Resumen':
    fila 3 col D-O: fechas por mes (2026-01-01 ... 2026-12-01)
    filas 5-11: líneas de ventas (col A=código, col B=+/-, col C=descripción)
    fila 13 col C: 'Total Ingresos Mensual'
    filas 14-25: líneas de gastos
    fila 26 col C: 'Total Gastos'
    fila 28 (Neto), 29 (IVA), 30 (Bruto): Facturación SABA
    fila 32 (Neto), 33 (IVA), 34 (Bruto): Liquidación factura
    fila 36 col B: 'Pago a Parque Titanium' (montos en la fila desde col D)
- Hoja 'MM-YYYY' (ej. '05-2026'): detalle diario
    fila 4 encabezados, fila 5+ datos hasta fila con día vacío
    col B=día, col C=Total Trans (tickets), col J=Total Bruto

Idempotencia: como el archivo se actualiza mes a mes, NO se rechaza por
file_hash. En cambio, antes de insertar se marcan como superseded las filas
existentes del mismo periodo/activo (y las tickets del mismo mes).

Feriados: se determinan automáticamente con `holidays.CL()`.
"""
from __future__ import annotations

import hashlib
import re
import sqlite3
import tempfile
from datetime import date, datetime
from pathlib import Path

import holidays
import openpyxl

DB_PATH = Path(__file__).resolve().parents[2] / "memory" / "agente_toesca_v2.db"
ACTIVO_KEY = "Parking PT"

PERIOD_HEADER_ROW = 3
FIRST_PERIOD_COL = 4  # col D

VENTAS_ROWS = range(5, 12)      # 5..11
GASTOS_ROWS = range(14, 26)     # 14..25
TOTAL_INGRESOS_ROW = 13
TOTAL_GASTOS_ROW = 26
FACTURACION_ROWS = {
    28: "saba_neto",
    29: "saba_iva",
    30: "saba_bruto",
    32: "liquidacion_neto",
    33: "liquidacion_iva",
    34: "liquidacion_bruto",
    36: "pago_a_pt",
}


class ValidationResult:
    def __init__(self):
        self.ok = True
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.data: dict = {}

    def add_error(self, msg: str):
        self.errors.append(msg)
        self.ok = False

    def add_warning(self, msg: str):
        self.warnings.append(msg)

    def to_dict(self) -> dict:
        return {"ok": self.ok, "errors": self.errors, "warnings": self.warnings, **self.data}


def _file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def _periodo_archivo(filename: str) -> str | None:
    """Extrae YYYY-MM desde nombres tipo 'MM-YYYY Liquidacion Parque Titanium.xlsx'."""
    m = re.search(r"(?<!\d)(\d{2})-(\d{4})(?!\d)", Path(filename).name)
    if not m:
        return None
    month, year = int(m.group(1)), int(m.group(2))
    if not 1 <= month <= 12:
        return None
    return f"{year:04d}-{month:02d}"


def _write_tmp(file_bytes: bytes, filename: str) -> str:
    suffix = Path(filename).suffix or ".xlsx"
    fd, tmp = tempfile.mkstemp(suffix=suffix)
    with open(fd, "wb") as f:
        f.write(file_bytes)
    return tmp


def _find_period_col(ws, periodo: str) -> int | None:
    target_year, target_month = int(periodo[:4]), int(periodo[5:7])
    c = FIRST_PERIOD_COL
    while True:
        v = ws.cell(row=PERIOD_HEADER_ROW, column=c).value
        if v is None:
            return None
        if isinstance(v, datetime) and v.year == target_year and v.month == target_month:
            return c
        c += 1
        if c > 40:
            return None


def _parse_resumen(ws, col: int) -> dict:
    """Devuelve dict con ingresos[], gastos[], facturacion[], totales_planilla."""
    ingresos = []
    gastos = []

    for r in VENTAS_ROWS:
        codigo = ws.cell(r, 1).value
        signo_txt = ws.cell(r, 2).value
        nombre = ws.cell(r, 3).value
        monto = ws.cell(r, col).value
        if nombre is None:
            continue
        signo = 1 if signo_txt == "+" else (-1 if signo_txt == "-" else 1)
        ingresos.append({
            "codigo": str(codigo) if codigo is not None else None,
            "nombre": str(nombre).strip(),
            "signo": signo,
            "monto": float(monto) if monto is not None else 0.0,
        })

    for r in GASTOS_ROWS:
        codigo = ws.cell(r, 1).value
        signo_txt = ws.cell(r, 2).value
        nombre = ws.cell(r, 3).value
        monto = ws.cell(r, col).value
        if nombre is None:
            continue
        signo = 1 if signo_txt == "+" else -1
        gastos.append({
            "codigo": str(codigo) if codigo is not None else None,
            "nombre": str(nombre).strip(),
            "signo": signo,
            "monto": float(monto) if monto is not None else 0.0,
        })

    facturacion = []
    for r, concepto in FACTURACION_ROWS.items():
        v = ws.cell(r, col).value
        if v is None:
            continue
        facturacion.append({"concepto": concepto, "monto": float(v)})

    total_ing = ws.cell(TOTAL_INGRESOS_ROW, col).value
    total_gas = ws.cell(TOTAL_GASTOS_ROW, col).value

    return {
        "ingresos": ingresos,
        "gastos": gastos,
        "facturacion": facturacion,
        "total_ingresos_planilla": float(total_ing) if total_ing is not None else None,
        "total_gastos_planilla": float(total_gas) if total_gas is not None else None,
    }


def _parse_diario(ws, periodo: str) -> list[dict]:
    year, month = int(periodo[:4]), int(periodo[5:7])
    cl_holidays = holidays.CL(years=year)
    out = []
    r = 5
    while True:
        dia = ws.cell(r, 2).value
        if dia is None:
            break
        try:
            dia_int = int(dia)
        except (TypeError, ValueError):
            break
        tickets = ws.cell(r, 3).value
        bruto = ws.cell(r, 10).value  # col J = Total Bruto
        fecha = date(year, month, dia_int)
        out.append({
            "fecha": fecha.strftime("%Y-%m-%d"),
            "tickets": int(tickets) if tickets is not None else 0,
            "monto_bruto_clp": float(bruto) if bruto is not None else None,
            "feriado": 1 if fecha in cl_holidays else 0,
        })
        r += 1
    return out


def _periodo_anterior(periodo: str) -> str:
    y, m = int(periodo[:4]), int(periodo[5:7])
    if m == 1:
        return f"{y - 1:04d}-12"
    return f"{y:04d}-{m - 1:02d}"


def _uf_ult_dia(con, periodo: str) -> float | None:
    row = con.execute(
        "SELECT valor FROM raw_uf_diaria WHERE substr(fecha,1,7)=? "
        "ORDER BY fecha DESC LIMIT 1",
        (periodo,),
    ).fetchone()
    return float(row[0]) if row else None


def _estacionamientos_no_abonados(con) -> float | None:
    row = con.execute(
        "SELECT estacionamientos_no_abonados FROM v_parking_ratio_no_abonados"
    ).fetchone()
    return float(row[0]) if row else None


def _kpis_mes_anterior(con, periodo: str) -> dict:
    prev = _periodo_anterior(periodo)
    r_uf = con.execute(
        "SELECT ingresos_abonados_uf, ingresos_variables_uf, resultado_neto_uf "
        "FROM v_parking_resultado_uf WHERE periodo=?",
        (prev,),
    ).fetchone()
    r_oc = con.execute(
        "SELECT ocupacion_mensual FROM v_parking_ocupacion_mensual "
        "WHERE activo_key='Parking PT' AND periodo=?",
        (prev,),
    ).fetchone()
    return {
        "periodo": prev,
        "ingresos_abonados_uf": float(r_uf[0]) if r_uf and r_uf[0] is not None else None,
        "ingresos_variables_uf": float(r_uf[1]) if r_uf and r_uf[1] is not None else None,
        "resultado_neto_uf": float(r_uf[2]) if r_uf and r_uf[2] is not None else None,
        "ocupacion_mensual": float(r_oc[0]) if r_oc and r_oc[0] is not None else None,
    }


def _pct_delta(curr, prev):
    if curr is None or prev is None or prev == 0:
        return None
    return (curr - prev) / abs(prev)


def _compute_kpis(periodo: str, resumen: dict, diario: list, suma_ing: float, suma_gas: float) -> dict:
    con = sqlite3.connect(str(DB_PATH))
    try:
        uf = _uf_ult_dia(con, periodo)
        est = _estacionamientos_no_abonados(con)
        prev = _kpis_mes_anterior(con, periodo)
    finally:
        con.close()

    # Split ingresos abonados vs variables (mismo criterio que la vista)
    abon_clp = sum(x["monto"] for x in resumen["ingresos"] if x["codigo"] == "70500003-250")
    var_clp = suma_ing - abon_clp
    resultado_clp = suma_ing - suma_gas

    def _div(a, b):
        return a / b if (a is not None and b) else None

    ing_abon_uf = _div(abon_clp, uf)
    ing_var_uf = _div(var_clp, uf)
    resultado_uf = _div(resultado_clp, uf)

    # Ocupación mensual = sum(bruto/40) / (n_dias * 8*60 * est)
    bruto_mes = sum((d["monto_bruto_clp"] or 0.0) for d in diario)
    if est and len(diario) > 0:
        tiempo_total = bruto_mes / 40.0
        tiempo_disp = len(diario) * 8 * 60 * est
        ocupacion = tiempo_total / tiempo_disp if tiempo_disp else None
    else:
        ocupacion = None

    return {
        "uf_valor": uf,
        "estacionamientos_no_abonados": est,
        "ingresos_abonados_uf": ing_abon_uf,
        "ingresos_variables_uf": ing_var_uf,
        "resultado_neto_uf": resultado_uf,
        "ocupacion_mensual": ocupacion,
        "mes_anterior": prev,
        "delta": {
            "ingresos_abonados_uf": _pct_delta(ing_abon_uf, prev["ingresos_abonados_uf"]),
            "ingresos_variables_uf": _pct_delta(ing_var_uf, prev["ingresos_variables_uf"]),
            "resultado_neto_uf": _pct_delta(resultado_uf, prev["resultado_neto_uf"]),
            "ocupacion_mensual": _pct_delta(ocupacion, prev["ocupacion_mensual"]),
        },
    }


def _parse(path: str, periodo: str) -> tuple[ValidationResult, dict | None]:
    result = ValidationResult()
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        result.add_error(f"No se pudo abrir el archivo: {exc}")
        return result, None

    if "Resumen" not in wb.sheetnames:
        result.add_error("La planilla no contiene la hoja 'Resumen'.")
        return result, None

    ws_res = wb["Resumen"]
    col = _find_period_col(ws_res, periodo)
    if col is None:
        result.add_error(
            f"No se encontró la columna del periodo {periodo} en la fila 3 de 'Resumen'."
        )
        return result, None

    resumen = _parse_resumen(ws_res, col)

    # Nombre esperado de la hoja diaria: MM-YYYY
    mm, yyyy = periodo[5:7], periodo[:4]
    sheet_diario = f"{mm}-{yyyy}"
    if sheet_diario not in wb.sheetnames:
        result.add_error(
            f"No se encontró la hoja diaria '{sheet_diario}'. "
            f"Hojas disponibles: {wb.sheetnames}"
        )
        return result, None

    diario = _parse_diario(wb[sheet_diario], periodo)
    if not diario:
        result.add_error(f"La hoja '{sheet_diario}' no contiene días con datos.")
        return result, None

    # Validaciones blandas
    suma_ing = sum(x["monto"] for x in resumen["ingresos"])
    if resumen["total_ingresos_planilla"] is not None:
        diff = abs(suma_ing - resumen["total_ingresos_planilla"])
        if diff > 1.0:
            result.add_warning(
                f"Suma de líneas de ingresos ({suma_ing:,.0f}) difiere de "
                f"'Total Ingresos Mensual' ({resumen['total_ingresos_planilla']:,.0f}) "
                f"por {diff:,.0f} CLP."
            )
    suma_gas = sum(x["monto"] for x in resumen["gastos"])
    if resumen["total_gastos_planilla"] is not None:
        diff = abs(suma_gas - resumen["total_gastos_planilla"])
        if diff > 1.0:
            result.add_warning(
                f"Suma de líneas de gastos ({suma_gas:,.0f}) difiere de "
                f"'Total Gastos' ({resumen['total_gastos_planilla']:,.0f}) "
                f"por {diff:,.0f} CLP."
            )

    n_feriados = sum(1 for d in diario if d["feriado"])
    tickets_total = sum(d["tickets"] for d in diario)

    # KPIs derivados (mismas fórmulas que v_parking_resultado_uf y
    # v_parking_ocupacion_mensual, computados aquí para poder mostrarlos en
    # el preview antes de commit).
    kpis = _compute_kpis(periodo, resumen, diario, suma_ing, suma_gas)

    parsed = {
        "periodo": periodo,
        "resumen": resumen,
        "diario": diario,
    }
    result.data = {
        "periodo": periodo,
        "n_ingresos": len(resumen["ingresos"]),
        "n_gastos": len(resumen["gastos"]),
        "n_facturacion": len(resumen["facturacion"]),
        "n_dias": len(diario),
        "n_feriados": n_feriados,
        "tickets_total": tickets_total,
        "suma_ingresos": suma_ing,
        "suma_gastos": suma_gas,
        "pago_a_pt": next(
            (x["monto"] for x in resumen["facturacion"] if x["concepto"] == "pago_a_pt"),
            None,
        ),
        "kpis": kpis,
    }
    return result, parsed


def validate(file_bytes: bytes, filename: str, periodo: str) -> ValidationResult:
    if not periodo or len(periodo) != 7 or periodo[4] != "-":
        r = ValidationResult()
        r.add_error("Periodo debe tener formato YYYY-MM.")
        return r
    archivo_periodo = _periodo_archivo(filename)
    if archivo_periodo and periodo > archivo_periodo:
        r = ValidationResult()
        r.add_error(
            f"El periodo seleccionado ({periodo}) no puede ser posterior al mes del archivo "
            f"({archivo_periodo})."
        )
        return r
    tmp = _write_tmp(file_bytes, filename)
    try:
        result, _ = _parse(tmp, periodo)
        return result
    finally:
        Path(tmp).unlink(missing_ok=True)


def _get_or_create_concepto(cur, codigo, nombre, tipo, signo) -> int:
    # Deduplicar por (nombre, tipo, signo). El código puede variar entre
    # planillas (histórico usa códigos contables, la mensual SABA usa códigos
    # internos como '988'), pero el concepto de negocio es el mismo.
    row = cur.execute(
        "SELECT id FROM dim_concepto_parking WHERE nombre=? AND tipo=? AND signo=?",
        (nombre, tipo, signo),
    ).fetchone()
    if row:
        return row[0]
    cur.execute(
        "INSERT INTO dim_concepto_parking (codigo, nombre, tipo, signo) VALUES (?,?,?,?)",
        (codigo, nombre, tipo, signo),
    )
    return cur.lastrowid


def commit(file_bytes: bytes, filename: str, periodo: str) -> dict:
    if not periodo or len(periodo) != 7 or periodo[4] != "-":
        raise ValueError("Periodo debe tener formato YYYY-MM.")
    archivo_periodo = _periodo_archivo(filename)
    if archivo_periodo and periodo > archivo_periodo:
        raise ValueError(
            f"El periodo seleccionado ({periodo}) no puede ser posterior al mes del archivo "
            f"({archivo_periodo})."
        )

    tmp = _write_tmp(file_bytes, filename)
    try:
        result, parsed = _parse(tmp, periodo)
        if not result.ok or parsed is None:
            raise ValueError("; ".join(result.errors) or "Error de validación.")
        file_hash = _file_hash(file_bytes)
    finally:
        Path(tmp).unlink(missing_ok=True)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    con = sqlite3.connect(str(DB_PATH))
    con.execute("PRAGMA foreign_keys = ON")
    cur = con.cursor()

    cur.execute(
        """INSERT INTO ingest_run (tool, source_file, file_hash, started_at, status)
           VALUES (?,?,?,?,?)""",
        ("ingest_parking_pt_mensual", filename, file_hash, now, "running"),
    )
    run_id = cur.lastrowid

    try:
        # Supersede filas previas del mismo periodo/activo
        cur.execute(
            "UPDATE raw_parking_ingreso_line SET superseded_at=? "
            "WHERE activo_key=? AND periodo=? AND superseded_at IS NULL",
            (now, ACTIVO_KEY, periodo),
        )
        n_super_ing = cur.rowcount
        cur.execute(
            "UPDATE raw_parking_gasto_line SET superseded_at=? "
            "WHERE activo_key=? AND periodo=? AND superseded_at IS NULL",
            (now, ACTIVO_KEY, periodo),
        )
        n_super_gas = cur.rowcount
        cur.execute(
            "UPDATE raw_parking_facturacion_line SET superseded_at=? "
            "WHERE activo_key=? AND periodo=? AND superseded_at IS NULL",
            (now, ACTIVO_KEY, periodo),
        )
        n_super_fac = cur.rowcount
        cur.execute(
            "UPDATE raw_parking_ticket_line SET superseded_at=? "
            "WHERE activo_key=? AND fecha LIKE ? AND superseded_at IS NULL",
            (now, ACTIVO_KEY, f"{periodo}-%"),
        )
        n_super_tk = cur.rowcount

        # Insertar ingresos
        n_ing = 0
        for x in parsed["resumen"]["ingresos"]:
            cid = _get_or_create_concepto(cur, x["codigo"], x["nombre"], "venta", x["signo"])
            cur.execute(
                """INSERT INTO raw_parking_ingreso_line
                   (activo_key, periodo, concepto_id, monto_clp,
                    source_file, file_hash, ingest_run_id, loaded_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (ACTIVO_KEY, periodo, cid, x["monto"], filename, file_hash, run_id, now),
            )
            n_ing += 1

        # Insertar gastos
        n_gas = 0
        for x in parsed["resumen"]["gastos"]:
            cid = _get_or_create_concepto(cur, x["codigo"], x["nombre"], "gasto", x["signo"])
            cur.execute(
                """INSERT INTO raw_parking_gasto_line
                   (activo_key, periodo, concepto_id, monto_clp,
                    source_file, file_hash, ingest_run_id, loaded_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (ACTIVO_KEY, periodo, cid, x["monto"], filename, file_hash, run_id, now),
            )
            n_gas += 1

        # Facturación
        n_fac = 0
        for x in parsed["resumen"]["facturacion"]:
            cur.execute(
                """INSERT INTO raw_parking_facturacion_line
                   (activo_key, periodo, concepto, monto_clp,
                    source_file, file_hash, ingest_run_id, loaded_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (ACTIVO_KEY, periodo, x["concepto"], x["monto"], filename, file_hash, run_id, now),
            )
            n_fac += 1

        # Tickets diarios
        n_tk = 0
        for d in parsed["diario"]:
            cur.execute(
                """INSERT INTO raw_parking_ticket_line
                   (activo_key, fecha, tickets, feriado, monto_bruto_clp,
                    source_file, file_hash, ingest_run_id, loaded_at)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (ACTIVO_KEY, d["fecha"], d["tickets"], d["feriado"], d["monto_bruto_clp"],
                 filename, file_hash, run_id, now),
            )
            n_tk += 1

        cur.execute(
            "UPDATE ingest_run SET ended_at=?, status='ok', rows_loaded=? WHERE id=?",
            (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), n_ing + n_gas + n_fac + n_tk, run_id),
        )
        con.commit()
    except Exception:
        con.rollback()
        cur.execute(
            "UPDATE ingest_run SET ended_at=?, status='error' WHERE id=?",
            (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), run_id),
        )
        con.commit()
        raise
    finally:
        con.close()

    return {
        "periodo": periodo,
        "run_id": run_id,
        "insertados": {
            "ingresos": n_ing, "gastos": n_gas,
            "facturacion": n_fac, "tickets": n_tk,
        },
        "superseded": {
            "ingresos": n_super_ing, "gastos": n_super_gas,
            "facturacion": n_super_fac, "tickets": n_super_tk,
        },
    }
