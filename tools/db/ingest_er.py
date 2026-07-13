"""
Ingesta de INFORME EEFF Tres Asociados a raw_er_activo_line.

Función principal: read_er_eeff() lee la hoja 'ESTADO DE RESULTADO' del
INFORME EEFF y extrae código de cuenta + monto CLP + metadata (sección, es_operacional).

Función persist_er_lines() realiza dual-write best-effort a la DB.
"""

import hashlib
import os
import re
import shutil
import tempfile
from datetime import date
from typing import Optional

import openpyxl

from tools.db.connection import get_conn as _db_get_conn
from tools.db import repo_audit, repo_er_activo


# ── Mapa: mall slug → activo_key (key para la DB) ────────────────────────────

_ER_ACTIVO_KEY = {"vina": "Viña Centro", "curico": "Mall Curicó"}


# ── Utilidades hash ────────────────────────────────────────────────────────────

def _file_hash(path: str) -> str:
    """Calcula el SHA256 de un archivo."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ── Leer EEFF ESTADO DE RESULTADO ──────────────────────────────────────────────

def read_er_eeff(eeff_path: str) -> tuple:
    """
    Lee la hoja 'ESTADO DE RESULTADO' del INFORME EEFF (Tres Asociados).

    Retorna (fecha_cierre: date, {codigo_cuenta: valor_clp}, {codigo: metadata}).
    Los valores son del mes actual (columna E = col 5 / index 4).

    Metadata es {codigo: {"seccion": str, "es_operacional": int}} que se usa
    en persist_er_lines para enriquecer el registro.
    """
    try:
        wb = openpyxl.load_workbook(eeff_path, read_only=True, data_only=True)
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), os.path.basename(eeff_path))
        shutil.copy2(eeff_path, tmp)
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)

    # Buscar hoja "ESTADO DE RESULTADO" o "ESTADO DE RESULTADO XXXX"
    er_sheet = next(
        (s for s in wb.sheetnames if s.upper().startswith("ESTADO DE RESULTADO")),
        None,
    )
    if er_sheet is None:
        wb.close()
        return None, {}, {}

    ws = wb[er_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # ── Extraer fecha del título ───────────────────────────────────────────────
    _MESES_NUM = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
        "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
        "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    }
    fecha_cierre = None
    for row in rows[:10]:
        for cell in row:
            if cell and isinstance(cell, str):
                cell_str = str(cell)
                # Formato numérico: AL 31-01-2026 o AL 31/01/2026
                m = re.search(r"AL\s+(\d{1,2})[-/](\d{1,2})[-/](\d{4})", cell_str, re.IGNORECASE)
                if m:
                    try:
                        fecha_cierre = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                    except ValueError:
                        pass
                # Formato textual: AL 31 DE ENERO 2026
                if not fecha_cierre:
                    m2 = re.search(
                        r"AL\s+(\d{1,2})\s+DE\s+(\w+)\s+(\d{4})", cell_str, re.IGNORECASE
                    )
                    if m2:
                        mes_num = _MESES_NUM.get(m2.group(2).lower())
                        if mes_num:
                            try:
                                fecha_cierre = date(int(m2.group(3)), mes_num, int(m2.group(1)))
                            except ValueError:
                                pass
                if fecha_cierre:
                    break
        if fecha_cierre:
            break

    # ── Leer valores + secciones (col B = código/encabezado, col E = mes actual) ─
    account_values: dict = {}
    meta_map: dict = {}   # {codigo: {"seccion": str, "es_operacional": int}}
    seccion_actual: str | None = None
    es_operacional: int = 1  # todo lo que precede a TOTAL OPERACIONAL es operacional

    for row in rows[10:]:
        code_raw = row[1] if len(row) > 1 else None
        val_raw  = row[4] if len(row) > 4 else None
        if code_raw is None:
            continue
        code = str(code_raw).strip()
        if not code:
            continue

        if re.match(r"^\d[-\d]", code):
            # Línea de cuenta: guardar valor y metadata de sección
            if val_raw is None:
                continue
            try:
                account_values[code] = float(val_raw)
                meta_map[code] = {"seccion": seccion_actual, "es_operacional": es_operacional}
            except (TypeError, ValueError):
                pass
        else:
            # Encabezado de sección: actualizar contexto
            seccion_actual = code
            # El marcador de fin del bloque operacional
            if "TOTAL OPERACIONAL" in code.upper():
                es_operacional = 0

    return fecha_cierre, account_values, meta_map


# ── Dual-write a DB (Fase 1) ──────────────────────────────────────────────────

def persist_er_lines(
    activo_key: str, source_file: str, periodo: str, eeff_values: dict,
    meta_map: dict | None = None,
) -> int:
    """
    Dual-write best-effort de líneas del ER a raw_er_activo_line.

    Parámetros:
      activo_key: 'Viña Centro' | 'Mall Curicó' (usar _ER_ACTIVO_KEY.get(mall))
      source_file: nombre base del archivo EEFF (para auditoría)
      periodo: 'YYYY-MM' (ej. '2026-01')
      eeff_values: {codigo_cuenta: monto_clp} (del read_er_eeff)
      meta_map: {codigo: {"seccion": str, "es_operacional": int}} (del read_er_eeff)

    Retorna: número de líneas insertadas.
    Nunca propaga errores: si la DB falla, el flujo de Excel debe seguir.
    """
    mall_alias = activo_key if activo_key in _ER_ACTIVO_KEY else None
    activo_key = _ER_ACTIVO_KEY.get(activo_key, activo_key)
    if not activo_key or not eeff_values:
        return 0
    meta_map = meta_map or {}
    try:
        fh = (
            _file_hash(source_file)
            if os.path.isfile(source_file)
            else hashlib.sha256(source_file.encode("utf-8")).hexdigest()
        )
        conn = _db_get_conn()
        try:
            run_id = repo_audit.start_ingest_run(
                conn,
                tool=f"actualizar_er_{mall_alias}" if mall_alias else "ingest_er",
                source_file=os.path.basename(source_file),
                file_hash=fh,
            )
            lines = []
            for i, (codigo, clp_val) in enumerate(eeff_values.items()):
                meta = meta_map.get(codigo, {})
                lines.append({
                    "activo_key": activo_key,
                    "periodo": periodo,
                    "cuenta_codigo": None,
                    "cuenta_nombre": codigo,
                    "monto_clp": clp_val,
                    "monto_uf": None,
                    "seccion": meta.get("seccion"),
                    "es_operacional": meta.get("es_operacional"),
                    "source_file": os.path.basename(source_file),
                    "source_sheet": "ESTADO DE RESULTADO",
                    "source_row": i,
                    "file_hash": fh,
                })
            n = repo_er_activo.insert_lines(conn, lines, run_id)
            repo_audit.finish_ingest_run(
                conn, run_id, rows_in=len(lines), rows_loaded=n, status="ok"
            )
            return n
        finally:
            conn.close()
    except Exception as e:
        print(f"[ingest_er] no se pudo persistir ER en DB: {e}")
        return 0


# ── Versión interna (con file_hash completo) para backfill ────────────────────

def _persist_er_lines_with_hash(
    mall: str, eeff_path: str, periodo: str, eeff_values: dict,
    meta_map: dict | None = None,
) -> int:
    """
    Versión interna que calcula el file_hash desde la ruta.
    Usada por actualizar_er_* y backfill.
    """
    activo_key = _ER_ACTIVO_KEY.get(mall)
    if not activo_key or not eeff_values:
        return 0
    meta_map = meta_map or {}
    try:
        fh = _file_hash(eeff_path)
        conn = _db_get_conn()
        try:
            run_id = repo_audit.start_ingest_run(
                conn,
                tool=f"actualizar_er_{mall}",
                source_file=os.path.basename(eeff_path),
                file_hash=fh,
            )
            lines = []
            for i, (codigo, clp_val) in enumerate(eeff_values.items()):
                meta = meta_map.get(codigo, {})
                lines.append({
                    "activo_key": activo_key,
                    "periodo": periodo,
                    "cuenta_codigo": None,
                    "cuenta_nombre": codigo,
                    "monto_clp": clp_val,
                    "monto_uf": None,
                    "seccion": meta.get("seccion"),
                    "es_operacional": meta.get("es_operacional"),
                    "source_file": os.path.basename(eeff_path),
                    "source_sheet": "ESTADO DE RESULTADO",
                    "source_row": i,
                    "file_hash": fh,
                })
            n = repo_er_activo.insert_lines(conn, lines, run_id)
            repo_audit.finish_ingest_run(
                conn, run_id, rows_in=len(lines), rows_loaded=n, status="ok"
            )
            return n
        finally:
            conn.close()
    except Exception as e:
        print(f"[ingest_er] no se pudo persistir ER {mall} en DB: {e}")
        return 0
