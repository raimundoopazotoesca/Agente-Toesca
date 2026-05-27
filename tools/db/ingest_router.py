"""
Router de ingesta: detecta tipo de archivo de proveedor por nombre y
delega al módulo de ingesta correcto (ingest_er, ingest_flujo, ...).
"""
import os
import re
from datetime import date, datetime
from typing import Optional

import openpyxl


def detect_tipo(path: str) -> Optional[str]:
    """Devuelve uno de: er_vina, er_curico, flujo_inmosa, rent_roll_jll,
    rent_roll_tresa_vina, rent_roll_tresa_curico, eeff_pdf, o None."""
    bn = os.path.basename(path).lower()
    if "informe eeff" in bn and ("viña" in bn or "vina" in bn):
        return "er_vina"
    if "informe eeff" in bn and ("curico" in bn or "curicó" in bn):
        return "er_curico"
    if "inmosa" in bn and ("flujo" in bn or "er-fc" in bn or "er fc" in bn):
        return "flujo_inmosa"
    if re.match(r"\d{4}\s*rent roll y noi", bn):
        return "rent_roll_jll"
    if "tres a" in bn and "viña" in bn:
        return "rent_roll_tresa_vina"
    if "tres a" in bn and ("curico" in bn or "curicó" in bn):
        return "rent_roll_tresa_curico"
    if bn.endswith(".pdf") and "eeff" in bn:
        return "eeff_pdf"
    return None


def ingestar_archivo(path: str, periodo: Optional[str] = None) -> dict:
    """Detecta tipo de archivo y ejecuta ingesta correspondiente.

    Args:
        path: Ruta absoluta al archivo a ingestar.
        periodo: YYYY-MM. Opcional; se infiere del archivo si no se entrega.

    Returns:
        {'tipo': str, 'filas': int, 'periodo': str, 'activo': str} si OK,
        {'error': str} si falla.
    """
    if not os.path.isfile(path):
        return {"error": f"Archivo no existe: {path}"}

    tipo = detect_tipo(path)
    if tipo is None:
        return {"error": f"No se pudo detectar tipo: {os.path.basename(path)}"}

    if tipo in ("er_vina", "er_curico"):
        from tools.db.ingest_er import read_er_eeff, persist_er_lines
        try:
            fecha_cierre, eeff_values, meta_map = read_er_eeff(path)
        except Exception as e:
            return {"error": f"Lectura ER falló: {e}"}
        if not eeff_values or fecha_cierre is None:
            return {"error": "Sin ESTADO DE RESULTADO o fecha"}
        periodo = periodo or f"{fecha_cierre.year}-{fecha_cierre.month:02d}"
        activo = "vina_centro" if tipo == "er_vina" else "power_center_curico"
        n = persist_er_lines(activo, path, periodo, eeff_values, meta_map)
        return {"tipo": tipo, "filas": n, "periodo": periodo, "activo": activo}

    if tipo == "flujo_inmosa":
        from tools.db.ingest_flujo import persist_flujo_lines
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            return {"error": f"Lectura xlsx falló: {e}"}
        target = None
        for s in wb.sheetnames:
            if any(k in s.upper() for k in ("NOI", "ESTADO", "RESULT")):
                target = s
                break
        if target is None and wb.sheetnames:
            target = wb.sheetnames[0]
        if target is None:
            wb.close()
            return {"error": "xlsx sin hojas"}
        rows = list(wb[target].iter_rows(values_only=True))
        wb.close()

        col_dates: dict = {}
        for row in rows:
            for ci, v in enumerate(row):
                if isinstance(v, (datetime, date)):
                    d = v.date() if isinstance(v, datetime) else v
                    col_dates.setdefault(ci, d)
        if not col_dates:
            return {"error": "Sin columnas de fecha"}

        total_filas = 0
        periodos_ingestados = []
        for ci, d in col_dates.items():
            p = f"{d.year}-{d.month:02d}"
            if periodo and p != periodo:
                continue
            er_data: dict = {}
            for row in rows:
                label_raw = row[0] if row[0] is not None else (row[1] if len(row) > 1 else None)
                if label_raw is None:
                    continue
                label = " ".join(str(label_raw).strip().split())
                if len(row) > ci and row[ci] is not None and not isinstance(row[ci], (datetime, date)):
                    try:
                        er_data[label] = float(row[ci])
                    except (TypeError, ValueError):
                        pass
            if not er_data:
                continue
            n = persist_flujo_lines("INMOSA", path, target, p, er_data,
                                     tool="ingestar_archivo", hash_extra=p)
            total_filas += n
            periodos_ingestados.append(p)
        return {"tipo": tipo, "filas": total_filas, "periodos": periodos_ingestados, "activo": "INMOSA"}

    return {"error": f"Tipo {tipo} reconocido pero ingesta no implementada todavía. Usa el script o backfill correspondiente."}
