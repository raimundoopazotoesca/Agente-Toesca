"""Ingesta de saldo caja consolidado por fondo desde la planilla histórica
'NUEVO ORDEN/Saldo Caja/Saldo Caja + FFMM Inmobiliario (final).xlsx'.

La planilla trae una hoja por fecha de corte (nombre 'DD-MM-YYYY'), con un
bloque por sociedad: fila de nombre de sociedad → fila de encabezado
('Banco' | 'N° Cuenta' | 'Saldo' | | 'Fondo' | 'FFMM') → filas de detalle
(cuentas bancarias en col E, fondos mutuos en col H) → fila en blanco.

Para cada mes se usa la hoja cuya fecha está más cerca del cierre de mes
(último día calendario). El saldo caja de cada fondo (TRI/PT/Apo) es una
suma ponderada de sociedades — ver FUND_WEIGHTS, confirmado por el usuario
2026-07-27.

Nombres de sociedad cambiaron en el tiempo (ej. 'Inmobiliaria Boulevard PT
SPA' se llamó 'Inmobiliaria Centro de Convenciones Spa.' en 2018): NO se
mapean alias históricos por ahora. Si una sociedad requerida por un fondo no
aparece en la hoja elegida, ese fondo se omite ese mes (no se inserta un
total parcial silenciosamente) y queda listado en el reporte de la corrida.

Cuentas/FFMM en USD: se convierten a CLP con `fact_dolar` (tabla
raw_dolar_diaria, poblada por `tools/db/ingest_dolar.py` desde
cl.dolarapi.com — solo cotización del día, sin serie histórica). Si no hay
dólar real para la fecha de la hoja, se usa FALLBACK_USD_CLP=900 (pedido
usuario 2026-07-27) y el fondo/mes queda listado en
`usd_convertido_con_fallback` para trazabilidad; el `source_file` de esa
fila también queda marcado con `|usd_fallback=900`.
"""
from __future__ import annotations

import calendar
import os
import re
import sqlite3
import unicodedata
from datetime import date, datetime
from typing import Optional

import openpyxl

WORKBOOK_RELATIVE_PATH = os.path.join("NUEVO ORDEN", "Saldo Caja", "Saldo Caja + FFMM Inmobiliario (final).xlsx")

# claves canónicas de sociedad (normalizadas) → peso por fondo
_TRI_FIP = "toesca rentas inmobiliarias fondo inversion"
_PT_FIP = "toesca rentas inmobiliarias pt fondo inversion"
_VC = "inmobiliaria vc spa"
_CHANARCILLO = "inmobiliaria chanarcillo limitada"
_TORRE_A = "torre a sa"
_BOULEVARD = "inmobiliaria boulevard pt spa"
_APO_FIP = "toesca rentas inmobiliarias apoquindo fondo inversion"
_APO_SA = "inmobiliaria apoquindo sa"
_VINA = "vina centro spa"
_CURICO = "power center curico ltda"

FUND_WEIGHTS: dict[str, dict[str, float]] = {
    "Apo": {
        _APO_FIP: 1.0,
        _APO_SA: 1.0,
    },
    "PT": {
        _PT_FIP: 1.0,
        _TORRE_A: 1.0,
        _BOULEVARD: 1.0,
    },
    "TRI": {
        _TRI_FIP: 1.0,
        _PT_FIP: 1.0 / 3,
        _VC: 1.0,
        _CHANARCILLO: 1.0,
        _TORRE_A: 1.0 / 3,
        _BOULEVARD: 1.0 / 3,
        _APO_FIP: 0.3,
        _APO_SA: 0.3,
        _VINA: 1.0,
        _CURICO: 0.8,
    },
}

_SHEET_NAME_RE = re.compile(r"^(\d{1,2})-(\d{1,2})-(\d{4})\s*$")
_USD_RE = re.compile(r"usd|us\$|u\$", re.I)


def _norm_text(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip())


def _normalize_name(v) -> str:
    """Clave de comparación: sin tildes, minúsculas, sin puntuación, sin ':' final."""
    s = _norm_text(v).rstrip(":").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[.,]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _is_usd(label: str) -> bool:
    return bool(_USD_RE.search(label))


def default_workbook_path() -> str:
    from config import SHAREPOINT_DIR

    if not SHAREPOINT_DIR:
        raise ValueError("SHAREPOINT_DIR no está configurado (.env)")
    return os.path.join(SHAREPOINT_DIR, WORKBOOK_RELATIVE_PATH)


def list_sheet_dates(wb) -> list[tuple[date, str]]:
    """[(fecha, nombre_hoja), ...] para hojas cuyo nombre calza 'DD-MM-YYYY'."""
    out = []
    for name in wb.sheetnames:
        m = _SHEET_NAME_RE.match(name)
        if not m:
            continue
        d, mth, y = (int(x) for x in m.groups())
        try:
            out.append((date(y, mth, d), name))
        except ValueError:
            continue
    return sorted(out)


def pick_month_end_sheets(sheet_dates: list[tuple[date, str]], max_gap_days: int = 45) -> dict[str, tuple[date, str]]:
    """Para cada mes cubierto por el rango de hojas, la hoja más cercana al
    cierre de ese mes (entre TODAS las hojas, no solo las del propio mes).

    Devuelve {periodo YYYY-MM: (fecha_hoja, nombre_hoja)}. Una misma hoja no
    se reutiliza para dos meses; si el mejor candidato ya fue tomado por otro
    mes, o la distancia excede `max_gap_days`, el mes se omite.
    """
    if not sheet_dates:
        return {}
    start, end = sheet_dates[0][0], sheet_dates[-1][0]
    months: list[str] = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m == 13:
            m = 1
            y += 1

    used: set[str] = set()
    out: dict[str, tuple[date, str]] = {}
    for periodo in months:
        py, pm = (int(x) for x in periodo.split("-"))
        month_end = date(py, pm, calendar.monthrange(py, pm)[1])
        candidates = sorted(sheet_dates, key=lambda ds: abs((ds[0] - month_end).days))
        for d, name in candidates:
            if name in used:
                continue
            if abs((d - month_end).days) > max_gap_days:
                break
            used.add(name)
            out[periodo] = (d, name)
            break
    return out


def parse_sheet_companies(ws) -> dict[str, dict]:
    """{clave_normalizada: {'clp': float, 'usd': float, 'raw_name': str}} para una hoja."""
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 2000, 2000), max_col=10, values_only=True))
    n = len(rows)
    result: dict[str, dict] = {}
    last_header_text: Optional[str] = None
    idx = 0
    while idx < n:
        row = rows[idx]
        c2 = row[2] if len(row) > 2 else None
        c2_norm = _norm_text(c2)
        if c2_norm and c2_norm.lower() != "banco":
            last_header_text = c2_norm
            idx += 1
            continue
        if c2_norm.lower() != "banco":
            idx += 1
            continue

        company_name = last_header_text
        j = idx + 1
        clp_total = 0.0
        usd_total = 0.0
        while j < n:
            r = rows[j]
            c2j = _norm_text(r[2] if len(r) > 2 else None)
            if c2j.lower() == "banco":
                break
            block_cells = [r[k] for k in range(2, 8) if k < len(r)]
            if all((v is None or (isinstance(v, str) and not v.strip())) for v in block_cells):
                break

            bank_name = _norm_text(r[2] if len(r) > 2 else None)
            cur_marker = _norm_text(r[5] if len(r) > 5 else None)
            saldo = r[4] if len(r) > 4 else None
            if isinstance(saldo, (int, float)):
                if _is_usd(bank_name) or _is_usd(cur_marker):
                    usd_total += saldo
                else:
                    clp_total += saldo

            ffmm_name = _norm_text(r[6] if len(r) > 6 else None)
            ffmm_val = r[7] if len(r) > 7 else None
            if isinstance(ffmm_val, (int, float)):
                if _is_usd(ffmm_name):
                    usd_total += ffmm_val
                else:
                    clp_total += ffmm_val
            j += 1

        if company_name:
            key = _normalize_name(company_name)
            result[key] = {"clp": clp_total, "usd": usd_total, "raw_name": company_name}
        idx = j if j > idx else idx + 1
        last_header_text = None

    return result


FALLBACK_USD_CLP = 900.0  # tipo de cambio fijo para fechas sin fact_dolar (pedido usuario 2026-07-27)


def compute_fund_totals(
    companies: dict[str, dict],
    fecha_hoja: date,
    conn: "sqlite3.Connection | None",
) -> tuple[dict[str, float], dict[str, list[str]], dict[str, float]]:
    """Aplica FUND_WEIGHTS sobre las sociedades encontradas en una hoja.

    Devuelve (totales_clp_por_fondo, sociedades_faltantes_por_fondo,
    usd_convertido_con_fallback_por_fondo). Un fondo con al menos una
    sociedad faltante no aparece en `totales_clp_por_fondo`. El USD se
    convierte con `fact_dolar` (fecha <= fecha_hoja) si existe; si no,
    con FALLBACK_USD_CLP — el tercer dict lista cuánto USD por fondo se
    convirtió con ese fallback en vez de un dólar real, para trazabilidad.
    """
    dolar = None
    if conn is not None:
        from tools.db import repo_fact
        from tools.db.errors import NotFoundError

        try:
            dolar = repo_fact.get_dolar(conn, fecha_hoja.isoformat())
        except NotFoundError:
            dolar = None
    tasa = dolar or FALLBACK_USD_CLP

    totales: dict[str, float] = {}
    faltantes: dict[str, list[str]] = {}
    usd_fallback: dict[str, float] = {}

    for fondo_key, weights in FUND_WEIGHTS.items():
        missing = [k for k in weights if k not in companies]
        if missing:
            faltantes[fondo_key] = missing
            continue
        total = 0.0
        usd_con_fallback = 0.0
        for company_key, weight in weights.items():
            c = companies[company_key]
            total += c["clp"] * weight
            if c["usd"]:
                total += c["usd"] * tasa * weight
                if dolar is None:
                    usd_con_fallback += c["usd"] * weight
        totales[fondo_key] = total
        if usd_con_fallback:
            usd_fallback[fondo_key] = usd_con_fallback

    return totales, faltantes, usd_fallback


def ingestar(
    path: Optional[str] = None,
    conn: "sqlite3.Connection | None" = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    dry_run: bool = False,
) -> dict:
    """Recorre todas las hojas de la planilla, elige la hoja de cierre de
    cada mes (opcionalmente acotado a [desde, hasta] en 'YYYY-MM') y upsertea
    raw_caja_line para Apo/PT/TRI. Idempotente (ON CONFLICT fondo_key+fecha).

    Con `dry_run=True` no escribe en la DB (solo parsea y calcula) — se usa
    para el paso "Validar y previsualizar" de la ingesta web.
    """
    from tools.db import repo_fact

    xlsx_path = path or default_workbook_path()

    owns_conn = conn is None
    if owns_conn:
        from tools.db.connection import get_conn

        conn = get_conn()

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            sheet_dates = list_sheet_dates(wb)
            if not sheet_dates:
                raise ValueError(
                    "No se encontró ninguna hoja con nombre de fecha 'DD-MM-YYYY' en la planilla."
                )
            month_sheets = pick_month_end_sheets(sheet_dates)
            if desde:
                month_sheets = {p: v for p, v in month_sheets.items() if p >= desde}
            if hasta:
                month_sheets = {p: v for p, v in month_sheets.items() if p <= hasta}

            filas_insertadas = 0
            meses_procesados: list[str] = []
            reporte_faltantes: dict[str, dict[str, list[str]]] = {}
            reporte_usd_fallback: dict[str, dict[str, float]] = {}
            totales_por_periodo: dict[str, dict[str, float]] = {}

            for periodo in sorted(month_sheets):
                fecha_hoja, sheet_name = month_sheets[periodo]
                ws = wb[sheet_name]
                companies = parse_sheet_companies(ws)
                totales, faltantes, usd_fallback = compute_fund_totals(companies, fecha_hoja, conn)

                py, pm = (int(x) for x in periodo.split("-"))
                fecha_cierre = date(py, pm, calendar.monthrange(py, pm)[1]).isoformat()
                base_source = f"{os.path.basename(xlsx_path)}|hoja={sheet_name}"

                for fondo_key, total_clp in totales.items():
                    if not dry_run:
                        source_file = base_source
                        if fondo_key in usd_fallback:
                            source_file += f"|usd_fallback={FALLBACK_USD_CLP:g}"
                        repo_fact.upsert_caja(conn, fondo_key, fecha_cierre, total_clp, source_file)
                    filas_insertadas += 1

                if faltantes:
                    reporte_faltantes[periodo] = faltantes
                if usd_fallback:
                    reporte_usd_fallback[periodo] = usd_fallback
                totales_por_periodo[periodo] = {
                    "fecha_cierre": fecha_cierre,
                    "fecha_hoja": fecha_hoja.isoformat(),
                    **{f"{k}_clp": v for k, v in totales.items()},
                }
                meses_procesados.append(periodo)

            return {
                "status": "ok",
                "meses_procesados": meses_procesados,
                "filas_upsert": filas_insertadas,
                "sociedades_faltantes": reporte_faltantes,
                "usd_convertido_con_fallback": reporte_usd_fallback,
                "totales_por_periodo": totales_por_periodo,
            }
        finally:
            wb.close()
    finally:
        if owns_conn:
            conn.close()


if __name__ == "__main__":
    import argparse
    import json

    parser = argparse.ArgumentParser(description="Ingesta saldo caja consolidado por fondo")
    parser.add_argument("--path", default=None, help="Ruta a la planilla (default: SharePoint NUEVO ORDEN/Saldo Caja)")
    parser.add_argument("--desde", default=None, help="Periodo mínimo YYYY-MM")
    parser.add_argument("--hasta", default=None, help="Periodo máximo YYYY-MM")
    args = parser.parse_args()

    resultado = ingestar(path=args.path, desde=args.desde, hasta=args.hasta)
    print(json.dumps(resultado, ensure_ascii=False, indent=2))
