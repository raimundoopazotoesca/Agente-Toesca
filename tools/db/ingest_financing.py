"""
Ingesta planilla resumen deudas y financiamiento.xlsx → agente_toesca_v2.db

Tablas destino:
  dim_credito              (migracion 020)
  raw_deuda_saldo_line     (migracion 021)
  raw_pagare_intercompania (migracion 022)
"""

import argparse
import sqlite3
import uuid
from datetime import date
from pathlib import Path

import openpyxl

MIGRATIONS_DIR = Path(__file__).parent / "migrations"
FECHA_CONTABLE_ARCHIVO = "2025-12"  # periodo de corte; posterior = proyección

# Mapeo columna Sección 2 → credito_key (Machalí omitido)
COL_TO_CREDITO = {
    "Sucden":           "TRI_SUCDEN_BICE",
    "Curicó":           "TRI_CURICO_METLIFE",
    "Viña Centro":      "TRI_VINA_PRINCIPAL",
    "Inmob. VC":        "TRI_INMOBVC_ARIZONA",
    "Apoquindo 3001":   "TRI_APO3001_SCOTIABANK",
    "Medina":           "TRI_MEDINA_METLIFE",
    "Candil":           "TRI_CANDIL_METLIFE",
    "Padre Errazuriz":  "TRI_PADREERRA_METLIFE",
    "Coventry":         "TRI_COVENTRY_CONFUTURO",
    "Colombia":         "TRI_COLOMBIA_PRINCIPAL",
    "Dom. Calderón":    "TRI_DOMCALDERON_ZURIC",
    "Torre A":          "PT_TORREA_SECURITY",
    "Inmob. Boulevard": "PT_BOULEVARD_SECURITY",
    "Apoquindo Euroam.":"APO_APO_EUROAMERICA",
    "Apoquindo BTG":    "APO_APO_BTG",
}

# Créditos a insertar en dim_credito — todos los de Sección 1 excepto Machalí
# (fondo, sociedad, activo_key, acreedor, tipo_deuda, part_fondo, deuda_ini, tasa, cuota_mensual,
#  fecha_inicio, fecha_vencimiento, estado, encargado, perfil)
DIM_CREDITO_SEED = [
    ("TRI_SUCDEN_BICE",         "TRI", "Inmob. Chañarcillo Ltda.",  "Sucden",       "Bice",        "Leasing",           1.0,  162759.82,  0.0456, 1068.12,  "2026-01-26", "2027-07-26", "VIGENTE", "TOESCA",         "Amortizing 3% anual. Cutón UF 151.592 (97%)"),
    ("TRI_APO3001_SCOTIABANK",   "TRI", "Inmob. Chañarcillo Ltda.",  "Apo3001",      "Scotiabank",  "Leasing",           1.0,  300180.05,  0.0525, 1741.67,  "2024-01-09", "2028-01-10", "VIGENTE", "TOESCA",         "Amortizing 3% anual. Cutón UF 240.769"),
    ("TRI_VINA_PRINCIPAL",       "TRI", "Viña Centro SpA",           "Viña Centro","Principal","Mutuo Hipotecario",1.0, 1000000.0,  0.0507, 4129.89,  "2018-01-09", "2040-01-05", "VIGENTE", "TRES A",         "Bullet hasta dic-28. Luego amortizing 6% anual, sin cutón"),
    ("TRI_INMOBVC_ARIZONA",      "TRI", "Inmobiliaria VC SpA",       "Viña Centro","Arizona - Text Rent","Crédito",1.0,180000.0,  0.089,     0.0,  "2023-01-01", "2026-01-03", "PAGADO",  "TOESCA",         "Bullet. PAGADO."),
    ("TRI_CURICO_METLIFE",       "TRI", "Power Center Curicó SpA",   "Mall Curicó","Metlife",  "Leasing",           0.8,  301307.94,  0.0476, 1636.17,  "2020-01-15", "2042-02-15", "VIGENTE", "TRES A",         "Amortizing 2% anual. Cutón UF 103.285"),
    ("TRI_MEDINA_METLIFE",       "TRI", "Inmosa",                    "INMOSA",       "Metlife",     "Leasing",           0.43,  97554.16,  0.052,   624.91,  "2017-01-06", "2038-01-09", "VIGENTE", "GRUPO ARAUCANA", "Amortizing cuotas iguales, sin cutón"),
    ("TRI_CANDIL_METLIFE",       "TRI", "Inmosa",                    "INMOSA",       "Metlife",     "Leasing",           0.43, 166811.79,  0.0532, 1023.9,   "2017-01-06", "2040-01-11", "VIGENTE", "GRUPO ARAUCANA", "Amortizing cuotas iguales, sin cutón"),
    ("TRI_PADREERRA_METLIFE",    "TRI", "Inmosa",                    "INMOSA",       "Metlife",     "Leasing",           0.43, 145864.35,  0.0517,  860.44,  "2017-01-06", "2042-01-03", "VIGENTE", "GRUPO ARAUCANA", "Amortizing cuotas iguales, sin cutón"),
    ("TRI_COVENTRY_CONFUTURO",   "TRI", "Inmosa",                    "INMOSA",       "Confuturo",   "Leasing",           0.43, 169427.87,  0.05,    993.8,   "2017-01-06", "2042-01-03", "VIGENTE", "GRUPO ARAUCANA", "Amortizing cuotas iguales, sin cutón"),
    ("TRI_COLOMBIA_PRINCIPAL",   "TRI", "Inmosa",                    "INMOSA",       "Principal",   "Leasing",           0.43,  92849.87,  0.051,   557.72,  "2017-01-06", "2040-01-11", "VIGENTE", "GRUPO ARAUCANA", "Amortizing cuotas iguales, sin cutón"),
    ("TRI_DOMCALDERON_ZURIC",    "TRI", "Inmosa",                    "INMOSA",       "Zuric",       "Leasing",           0.43, 247619.0,   0.0537, 1375.76,  "2023-01-01", "2067-01-09", "VIGENTE", "GRUPO ARAUCANA", "Amortizing cuotas iguales, sin cutón. Vence 2067"),
    ("PT_TORREA_SECURITY",       "PT",  "Torre A S.A.",              "Torre A",      "Security",    "Crédito Sindicado", 0.333,1705313.76, 0.0415, 5663.3,   "2017-01-10", "2029-01-11", "VIGENTE", "TOESCA",         "Bullet c/amorts. nov-26/27/28 de UF 14.200 c/u. Cutón UF 1.662.682"),
    ("PT_BOULEVARD_SECURITY",    "PT",  "Inmob. Boulevard PT SpA",   "Boulevard",    "Security",    "Crédito Sindicado", 0.333,1081306.48, 0.0411, 2200.28,  "2017-01-10", "2029-01-11", "VIGENTE", "TOESCA",         "Bullet c/amorts. nov-26/27/28 de UF 5.800 c/u. Cutón UF 645.170"),
    ("APO_APO_EUROAMERICA",      "Apo", "Inmobiliaria Apoquindo I",  "Apo4501",      "Euroamérica", "Leasing",           0.3, 2800000.0,  0.0273,    0.0,   "2019-01-07", "2027-01-01", "PAGADO",  "TOESCA",         "Bullet semestral. PAGADO."),
    ("APO_APO_BTG",              "Apo", "Inmobiliaria Apoquindo II", "Apo4700",      "BTG",         "Crédito",           0.3,  100000.0,  0.0465,    0.0,   "2024-01-10", "2026-01-12", "PAGADO",  "TOESCA",         "Amortizing cuotas iguales. PAGADO."),
]

# Mapeo nombre dim_credito → activo_key real en dim_activo
CREDITO_ACTIVO_KEY = {
    "TRI_SUCDEN_BICE":        "Sucden",
    "TRI_APO3001_SCOTIABANK":  "Apo3001",
    "TRI_VINA_PRINCIPAL":      "Viña Centro",
    "TRI_INMOBVC_ARIZONA":     "Viña Centro",
    "TRI_CURICO_METLIFE":      "Mall Curicó",
    "TRI_MEDINA_METLIFE":      "INMOSA",
    "TRI_CANDIL_METLIFE":      "INMOSA",
    "TRI_PADREERRA_METLIFE":   "INMOSA",
    "TRI_COVENTRY_CONFUTURO":  "INMOSA",
    "TRI_COLOMBIA_PRINCIPAL":  "INMOSA",
    "TRI_DOMCALDERON_ZURIC":   "INMOSA",
    "PT_TORREA_SECURITY":      "Torre A",
    "PT_BOULEVARD_SECURITY":   "Boulevard",
    "APO_APO_EUROAMERICA":     "Apo4501",
    "APO_APO_BTG":             "Apo4700",
}


def _apply_migrations(conn: sqlite3.Connection) -> None:
    for n in (20, 21, 22):
        sql_file = MIGRATIONS_DIR / f"0{n:02d}_{'dim_credito' if n==20 else 'raw_deuda_saldo_line' if n==21 else 'raw_pagare_intercompania'}.sql"
        conn.executescript(sql_file.read_text(encoding="utf-8"))
    conn.commit()


def _parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (date,)):
        return val.isoformat()
    s = str(val)
    if "PAGADO" in s.upper():
        return "PAGADO"
    # Try stripping timestamp
    if " " in s:
        s = s.split(" ")[0]
    return s if s else None


def ingest_dim_credito(conn: sqlite3.Connection) -> int:
    rows = []
    for (ck, fondo, sociedad, activo_label, acreedor, tipo, part, deuda_ini, tasa,
         cuota, f_ini, f_ven, estado, encargado, perfil) in DIM_CREDITO_SEED:
        activo_key = CREDITO_ACTIVO_KEY[ck]
        rows.append((ck, activo_key, fondo, sociedad, acreedor, tipo, part,
                     deuda_ini, tasa, cuota, f_ini, f_ven, estado, encargado, perfil))
    conn.executemany(
        """INSERT OR REPLACE INTO dim_credito
           (credito_key, activo_key, fondo_key, sociedad, acreedor, tipo_deuda,
            part_fondo, deuda_inicial_uf, tasa_anual, cuota_mensual_uf,
            fecha_inicio, fecha_vencimiento, estado, encargado, perfil_amortizacion)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        rows,
    )
    conn.commit()
    return len(rows)


def ingest_saldos_historicos(conn: sqlite3.Connection, wb: openpyxl.Workbook) -> int:
    ws = wb.active
    run_id = str(uuid.uuid4())[:8]

    # Encuentra fila de encabezado de Sección 2 (busca "Año" en col A y activos en col C+)
    header_row = None
    for row in ws.iter_rows():
        if row[0].value and str(row[0].value).strip() in ("Año", "Año", "A?o"):
            header_row = row[0].row
            break
        # También detectar por valor parcial
        if row[0].value and "a" in str(row[0].value).lower() and row[1].value and "fecha" in str(row[1].value).lower():
            header_row = row[0].row
            break

    if header_row is None:
        # Fallback: la Sección 2 comienza en fila 32 (según markitdown)
        header_row = 32

    # Lee encabezados de columnas
    headers = [ws.cell(header_row, c).value for c in range(1, 20)]

    # Identifica columnas de activos (col index → credito_key)
    col_map: dict[int, str] = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).strip()
        # Normaliza caracteres especiales
        for col_name, ckey in COL_TO_CREDITO.items():
            if col_name.lower() in h_str.lower() or h_str.lower() in col_name.lower():
                col_map[i + 1] = ckey  # 1-based
                break

    # Lee filas de datos (hasta que Año sea NaN o cambie de sección)
    rows_to_insert = []
    for r in range(header_row + 1, ws.max_row + 1):
        year_val = ws.cell(r, 1).value
        fecha_val = ws.cell(r, 2).value
        if year_val is None or str(year_val).strip() in ("", "NaN"):
            break
        try:
            int(float(str(year_val)))
        except (ValueError, TypeError):
            break

        # Construye periodo YYYY-MM desde columna Fecha
        if fecha_val is None:
            continue
        if hasattr(fecha_val, "strftime"):
            periodo = fecha_val.strftime("%Y-%m")
        else:
            s = str(fecha_val)
            if len(s) >= 7:
                periodo = s[:7]
            else:
                continue

        is_proy = 1 if periodo > FECHA_CONTABLE_ARCHIVO else 0

        for col_idx, ckey in col_map.items():
            saldo = ws.cell(r, col_idx).value
            if saldo is None:
                saldo = 0.0
            try:
                saldo = float(saldo)
            except (TypeError, ValueError):
                saldo = 0.0
            rows_to_insert.append((run_id, ckey, periodo, saldo, is_proy))

    conn.executemany(
        """INSERT OR REPLACE INTO raw_deuda_saldo_line
           (run_id, credito_key, periodo, saldo_uf, is_proyeccion)
           VALUES (?,?,?,?,?)""",
        rows_to_insert,
    )
    conn.commit()
    return len(rows_to_insert)


def ingest_pagares(conn: sqlite3.Connection, wb: openpyxl.Workbook) -> int:
    ws = wb.active

    # Primero localiza la fila del título de Sección 4
    sec4_title_row = None
    for row in ws.iter_rows():
        val = str(row[0].value or "").strip()
        if "PAGAR" in val.upper() and "INTERCOMPA" in val.upper():
            sec4_title_row = row[0].row
            break

    if sec4_title_row is None:
        print("  [WARN] No se encontro titulo de Seccion 4 (Pagares)")
        return 0

    # Luego busca el encabezado dentro de las 5 filas siguientes
    pagare_header_row = None
    for r in range(sec4_title_row + 1, sec4_title_row + 6):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, 10)]
        joined = " ".join(vals).lower()
        if "acreedor" in joined or "deudor" in joined:
            pagare_header_row = r
            break

    if pagare_header_row is None:
        print("  [WARN] No se encontró encabezado de Sección 4 (Pagarés)")
        return 0

    rows_to_insert = []
    for r in range(pagare_header_row + 1, ws.max_row + 1):
        # Col A=acreedor_fondo (a veces 0), B=deudor_fondo, C=deudor_sociedad, D=tipo, E=fecha_inicio, F=fecha_vencimiento, G=monto, H=tasa, I=saldo
        # Según markitdown: col0=0, col1=Toesca FI, col2=INMOSA, col3=Pagaré, col4=fecha_inicio, col5=fecha_ven, col6=monto, col7=tasa, col8=saldo
        row_vals = [ws.cell(r, c).value for c in range(1, 10)]
        if all(v is None for v in row_vals):
            continue
        # Col B (idx 1) = fondo acreedor; si está vacío es fin de sección
        if row_vals[1] is None:
            break

        # Layout real: A=vacío/0, B=fondo_acreedor, C=deudor_sociedad, D=tipo,
        #              E=fecha_inicio, F=fecha_vencimiento, G=monto, H=tasa, I=saldo
        acreedor_fondo = str(row_vals[1]).strip() if row_vals[1] else ""
        deudor_soc = str(row_vals[2]).strip() if row_vals[2] else ""
        tipo = str(row_vals[3]).strip() if row_vals[3] else ""
        f_ini = _parse_date(row_vals[4])
        f_ven = _parse_date(row_vals[5])

        try:
            monto = float(row_vals[6]) if row_vals[6] is not None else None
        except (TypeError, ValueError):
            monto = None

        # Col H = saldo c/intereses (la columna Tasa no tiene datos en la planilla)
        try:
            saldo = float(row_vals[7]) if row_vals[7] is not None else None
        except (TypeError, ValueError):
            saldo = None
        tasa = None

        rows_to_insert.append((acreedor_fondo, deudor_soc, tipo, f_ini, f_ven, monto, tasa, saldo))

    conn.execute("DELETE FROM raw_pagare_intercompania")
    conn.executemany(
        """INSERT INTO raw_pagare_intercompania
           (acreedor_fondo, deudor_sociedad, tipo, fecha_inicio, fecha_vencimiento,
            monto_uf, tasa, saldo_c_intereses)
           VALUES (?,?,?,?,?,?,?,?)""",
        rows_to_insert,
    )
    conn.commit()
    return len(rows_to_insert)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--db", required=True)
    args = parser.parse_args()

    conn = sqlite3.connect(args.db)
    print("Aplicando migraciones 020-022...")
    _apply_migrations(conn)

    print("Cargando workbook...")
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    print("Insertando dim_credito...")
    n_creditos = ingest_dim_credito(conn)
    print(f"  -> {n_creditos} creditos")

    print("Insertando raw_deuda_saldo_line...")
    n_saldos = ingest_saldos_historicos(conn, wb)
    print(f"  -> {n_saldos} filas de saldo")

    print("Insertando raw_pagare_intercompania...")
    n_pagares = ingest_pagares(conn, wb)
    print(f"  -> {n_pagares} pagares")

    wb.close()
    conn.close()
    print("\nListo.")


if __name__ == "__main__":
    main()
