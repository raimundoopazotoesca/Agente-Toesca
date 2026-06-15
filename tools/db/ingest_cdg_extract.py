"""
Ingesta de CDG Extract (planilla consolidada de TRI con dividendos, valores, capital).
Lee desde Excel y persiste a raw_dividendo_line y raw_cuota_en_circulacion_line.
"""

import hashlib
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import openpyxl

from tools.db.connection import get_conn


def ingest_cdg_extract_tri(excel_path: str) -> Dict:
    """
    Lee cdg_extract.xlsx (hoja 'A&R Rentas' para TRI) y persiste dividendos y cuotas.

    Returns:
        {
            'dividendos_insertados': int,
            'cuotas_insertadas': int,
            'periodos': set,
            'error': str (opcional)
        }
    """
    path = Path(excel_path)
    if not path.exists():
        return {'error': f'Archivo no encontrado: {excel_path}'}

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['A&R Rentas']
    except Exception as e:
        return {'error': f'Error abriendo Excel: {e}'}

    # Mapeo de series a nemotecnicos TRI
    NEMO_MAP = {'A': 'CFITOERI1A', 'C': 'CFITOERI1C', 'I': 'CFITOERI1I'}

    dividendos = {}  # {fecha_str: {nemo: {...}}}
    cuotas = {}      # {fecha_str: {nemo: cuota_count}}

    # Escanear tabla (rows 16+, columns D=fecha, E=tipo, F=serie, J=cuotas, K=monto_uf_cuota)
    for row_num in range(16, 900):
        e_val = ws[f'E{row_num}'].value
        if not e_val or 'Dividendo' not in str(e_val):
            continue

        fecha = ws[f'D{row_num}'].value
        serie = ws[f'F{row_num}'].value
        monto_clp_cuota = ws[f'I{row_num}'].value   # $/cuota (columna I)
        cuota_count = ws[f'J{row_num}'].value
        monto_uf_cuota = ws[f'M{row_num}'].value     # UF/cuota (columna M)

        if not (fecha and serie and monto_clp_cuota):
            continue

        nemo = NEMO_MAP.get(str(serie).strip())
        if not nemo:
            continue

        fecha_str = fecha.strftime('%Y-%m-%d')
        periodo = f"{fecha.year}-{fecha.month:02d}"

        if fecha_str not in dividendos:
            dividendos[fecha_str] = {}
            cuotas[fecha_str] = {}

        dividendos[fecha_str][nemo] = {
            'monto_clp_cuota': monto_clp_cuota,
            'monto_uf_cuota': monto_uf_cuota,
            'periodo': periodo,
            'fecha': fecha_str
        }
        cuotas[fecha_str][nemo] = cuota_count

    # Persistir en DB
    conn = get_conn()

    # Crear tablas si no existen
    conn.execute("""
        CREATE TABLE IF NOT EXISTS raw_dividendo_line (
            id INTEGER PRIMARY KEY,
            fondo_key TEXT NOT NULL,
            nemotecnico TEXT NOT NULL,
            fecha_pago TEXT NOT NULL,
            monto_clp_cuota REAL,
            monto_uf_cuota REAL,
            periodo TEXT,
            source_file TEXT,
            file_hash TEXT,
            loaded_at TEXT DEFAULT CURRENT_TIMESTAMP,
            superseded_at TEXT,
            UNIQUE(nemotecnico, fecha_pago, file_hash)
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS raw_cuota_en_circulacion_line (
            id INTEGER PRIMARY KEY,
            fondo_key TEXT NOT NULL,
            nemotecnico TEXT NOT NULL,
            fecha TEXT NOT NULL,
            cuotas REAL NOT NULL,
            periodo TEXT,
            source_file TEXT,
            file_hash TEXT,
            loaded_at TEXT DEFAULT CURRENT_TIMESTAMP,
            superseded_at TEXT,
            UNIQUE(nemotecnico, fecha, file_hash)
        )
    """)

    # Calcular hash del archivo
    file_hash = _hash_file(excel_path)
    source_file = Path(excel_path).name

    # Insertar dividendos
    dividendos_insertados = 0
    for fecha_str, series_data in dividendos.items():
        for nemo, data in series_data.items():
            try:
                conn.execute("""
                    INSERT OR REPLACE INTO raw_dividendo_line
                    (fondo_key, nemotecnico, fecha_pago, monto_clp_cuota, monto_uf_cuota, periodo, source_file, file_hash)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    'TRI', nemo, fecha_str,
                    data['monto_clp_cuota'], data['monto_uf_cuota'],
                    data['periodo'], source_file, file_hash
                ))
                dividendos_insertados += 1
            except sqlite3.IntegrityError:
                pass

    # Insertar cuotas en circulación
    cuotas_insertadas = 0
    for fecha_str, series_cuotas in cuotas.items():
        for nemo, cuota_count in series_cuotas.items():
            if nemo in dividendos[fecha_str]:
                periodo = dividendos[fecha_str][nemo]['periodo']
                try:
                    conn.execute("""
                        INSERT OR REPLACE INTO raw_cuota_en_circulacion_line
                        (fondo_key, nemotecnico, fecha, cuotas, periodo, source_file, file_hash)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (
                        'TRI', nemo, fecha_str, cuota_count,
                        periodo, source_file, file_hash
                    ))
                    cuotas_insertadas += 1
                except sqlite3.IntegrityError:
                    pass

    conn.commit()

    periodos = set(d['periodo'] for f_data in dividendos.values() for d in f_data.values())

    return {
        'dividendos_insertados': dividendos_insertados,
        'cuotas_insertadas': cuotas_insertadas,
        'periodos': sorted(periodos),
        'fechas': len(dividendos)
    }


def ingest_capital_suscrito(excel_path: str) -> Dict:
    """
    Calcula capital suscrito acumulado por serie desde movimientos históricos.
    Capital Suscrito = Aportes acumulados + Canjes acumulados - Disminuciones acumuladas
    """
    path = Path(excel_path)
    if not path.exists():
        return {'error': f'Archivo no encontrado: {excel_path}'}

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['A&R Rentas']
    except Exception as e:
        return {'error': f'Error abriendo Excel: {e}'}

    NEMO_MAP = {'A': 'CFITOERI1A', 'C': 'CFITOERI1C', 'I': 'CFITOERI1I'}
    TIPOS_CAPITAL = {'Aporte': 1, 'Canje Cuotas': 1, 'Disminución': -1}

    # Acumular movimientos por serie y fecha
    movimientos = {}
    for nemo in NEMO_MAP.values():
        movimientos[nemo] = {}

    for row_num in range(16, 900):
        fecha = ws[f'D{row_num}'].value
        serie = ws[f'F{row_num}'].value
        detalle = ws[f'E{row_num}'].value   # col E = Detalle (no col G = Tipo)
        monto_uf = ws[f'L{row_num}'].value

        if not (fecha and serie and detalle and monto_uf is not None):
            continue

        tipo_str = str(detalle).strip()
        if tipo_str not in TIPOS_CAPITAL:
            continue

        nemo = NEMO_MAP.get(str(serie).strip())
        if not nemo:
            continue

        fecha_str = fecha.strftime('%Y-%m-%d')
        signo = TIPOS_CAPITAL[tipo_str]

        if fecha_str not in movimientos[nemo]:
            movimientos[nemo][fecha_str] = 0
        movimientos[nemo][fecha_str] += monto_uf * signo

    # Calcular acumulado y persistir
    conn = get_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS raw_capital_suscrito_line (
            id INTEGER PRIMARY KEY,
            fondo_key TEXT NOT NULL,
            nemotecnico TEXT NOT NULL,
            fecha_fin_periodo TEXT NOT NULL,
            capital_suscrito_uf REAL NOT NULL,
            periodo TEXT,
            source_file TEXT,
            file_hash TEXT,
            loaded_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(nemotecnico, fecha_fin_periodo, file_hash)
        )
    """)

    file_hash = _hash_file(excel_path)
    source_file = Path(excel_path).name

    insertados = 0
    for nemo in NEMO_MAP.values():
        acum = 0
        for fecha_str in sorted(movimientos[nemo].keys()):
            acum += movimientos[nemo][fecha_str]
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
            periodo = f"{fecha_obj.year}-{fecha_obj.month:02d}"

            try:
                conn.execute("""
                    INSERT OR REPLACE INTO raw_capital_suscrito_line
                    (fondo_key, nemotecnico, fecha_fin_periodo, capital_suscrito_uf, periodo, source_file, file_hash)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, ('TRI', nemo, fecha_str, acum, periodo, source_file, file_hash))
                insertados += 1
            except sqlite3.IntegrityError:
                pass

    conn.commit()

    return {
        'capital_suscrito_insertados': insertados,
        'series': list(NEMO_MAP.values())
    }


def ingest_vr_contable(excel_path: str) -> Dict:
    """
    Lee hoja 'A&R Rentas' y persiste VR Contable (valor libro) por serie/fecha.
    VR Contable = Monto UF/cuota (col M) donde Detalle = 'VR Contable', Serie = X, Fecha = exacta.
    Guarda en raw_valor_cuota_line con tipo='contable' y source_file='cdg_extract.xlsx'.
    Solo inserta si no existe ya una fila para (nemotecnico, fecha, tipo='contable') de EEFF PDF.
    """
    path = Path(excel_path)
    if not path.exists():
        return {'error': f'Archivo no encontrado: {excel_path}'}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['A&R Rentas']
    except Exception as e:
        return {'error': f'Error abriendo Excel: {e}'}

    NEMO_MAP = {'A': 'CFITOERI1A', 'C': 'CFITOERI1C', 'I': 'CFITOERI1I'}

    # Recoger valores: (fecha_str, nemo) → (precio_uf_cuota, cuotas, monto_uf)
    data: dict = {}
    for row_num in range(16, 1200):
        fecha = ws[f'D{row_num}'].value
        detalle = ws[f'E{row_num}'].value
        serie = ws[f'F{row_num}'].value
        cuotas = ws[f'J{row_num}'].value
        monto_uf = ws[f'L{row_num}'].value
        muf_cuota = ws[f'M{row_num}'].value   # Monto UF / cuota = VR Contable

        if fecha is None:
            break
        if not (detalle and serie and muf_cuota is not None):
            continue
        if 'VR Contable' not in str(detalle):
            continue

        nemo = NEMO_MAP.get(str(serie).strip())
        if not nemo:
            continue

        fecha_str = fecha.strftime('%Y-%m-%d')
        data[(fecha_str, nemo)] = (muf_cuota, cuotas, monto_uf)

    conn = get_conn()
    file_hash = _hash_file(excel_path)
    source_file = Path(excel_path).name

    insertados = 0
    omitidos = 0
    discrepancias = []

    for (fecha_str, nemo), (precio_uf, cuotas, monto_uf) in data.items():
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
        periodo = f"{fecha_obj.year}-{fecha_obj.month:02d}"

        # Verificar si ya existe una fila de EEFF PDF para esta fecha/nemo
        existing = conn.execute("""
            SELECT precio_uf, source_file FROM raw_valor_cuota_line
            WHERE fondo_key = 'TRI' AND nemotecnico = ? AND fecha = ? AND tipo = 'contable'
              AND source_file NOT LIKE '%cdg_extract%'
        """, (nemo, fecha_str)).fetchone()

        if existing:
            diff = abs(existing[0] - precio_uf) / precio_uf if precio_uf else 0
            if diff > 0.001:  # más del 0.1% de diferencia
                discrepancias.append((fecha_str, nemo, precio_uf, existing[0], existing[1]))
            omitidos += 1
            continue  # EEFF PDF tiene precedencia

        try:
            conn.execute("""
                INSERT INTO raw_valor_cuota_line
                (fondo_key, nemotecnico, fecha, tipo, precio_uf, cuotas, periodo, source_file, file_hash)
                VALUES (?, ?, ?, 'contable', ?, ?, ?, ?, ?)
            """, ('TRI', nemo, fecha_str, precio_uf, cuotas, periodo, source_file, file_hash))
            insertados += 1
        except Exception:
            omitidos += 1

    conn.commit()
    result: Dict = {'insertados': insertados, 'omitidos_eeff_prioritario': omitidos}
    if discrepancias:
        result['discrepancias'] = [
            {'fecha': d[0], 'nemo': d[1], 'ar_rentas': round(d[2], 9),
             'eeff_pdf': round(d[3], 9), 'pdf_source': d[4]}
            for d in discrepancias
        ]
    return result


def ingest_patrimonio_bursatil(excel_path: str) -> Dict:
    """
    Lee hoja 'A&R Rentas' y persiste el Patrimonio Bursátil por serie/fecha.
    Patrimonio Bursátil = SUM(Monto UF) donde Detalle = 'VR Bursátil', Serie = X, Fecha = exacta.
    Cada fecha de corte genera una fila por serie en raw_patrimonio_bursatil_line.
    """
    path = Path(excel_path)
    if not path.exists():
        return {'error': f'Archivo no encontrado: {excel_path}'}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['A&R Rentas']
    except Exception as e:
        return {'error': f'Error abriendo Excel: {e}'}

    NEMO_MAP = {'A': 'CFITOERI1A', 'C': 'CFITOERI1C', 'I': 'CFITOERI1I'}

    # Acumular por (fecha, serie) — puede haber más de una fila por fecha/serie
    from collections import defaultdict
    acum: dict = defaultdict(float)

    for row_num in range(16, 1200):
        fecha = ws[f'D{row_num}'].value
        detalle = ws[f'E{row_num}'].value
        serie = ws[f'F{row_num}'].value
        monto_uf = ws[f'L{row_num}'].value

        if fecha is None:
            break
        if not (detalle and serie and monto_uf is not None):
            continue
        if 'VR Burs' not in str(detalle):
            continue

        nemo = NEMO_MAP.get(str(serie).strip())
        if not nemo:
            continue

        fecha_str = fecha.strftime('%Y-%m-%d')
        acum[(fecha_str, nemo)] += monto_uf

    conn = get_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS raw_patrimonio_bursatil_line (
            id          INTEGER PRIMARY KEY,
            fondo_key   TEXT NOT NULL,
            nemotecnico TEXT NOT NULL,
            fecha       TEXT NOT NULL,
            patrimonio_uf REAL NOT NULL,
            periodo     TEXT,
            source_file TEXT,
            file_hash   TEXT,
            loaded_at   TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(nemotecnico, fecha, file_hash)
        )
    """)

    file_hash = _hash_file(excel_path)
    source_file = Path(excel_path).name

    insertados = 0
    for (fecha_str, nemo), pat_uf in acum.items():
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
        periodo = f"{fecha_obj.year}-{fecha_obj.month:02d}"
        try:
            conn.execute("""
                INSERT OR REPLACE INTO raw_patrimonio_bursatil_line
                (fondo_key, nemotecnico, fecha, patrimonio_uf, periodo, source_file, file_hash)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, ('TRI', nemo, fecha_str, pat_uf, periodo, source_file, file_hash))
            insertados += 1
        except Exception:
            pass

    conn.commit()
    return {'insertados': insertados, 'fechas_unicas': len(set(k[0] for k in acum))}


def ingest_ar_pt(excel_path: str) -> Dict:
    """
    Lee hoja 'A&R PT' del CDG extract y persiste en un solo pase:
      - raw_dividendo_line            (Detalle='Dividendo')
      - raw_valor_cuota_line          (Detalle='VR Contable', tipo='contable')
      - raw_precio_cuota_line         (Detalle='VR Bursátil', para gap-fill histórico)
      - raw_cuota_en_circulacion_line (cuotas del período de VR Contable)
      - raw_patrimonio_bursatil_line  (Detalle='VR Bursátil', patrimonio total)

    PT tiene SERIE ÚNICA → nemotecnico = 'CFITRIPT-E' en todas las filas.
    Columnas hoja (0-based): A=año, B=mes, C=id, D=fecha_sf, E=detalle, F=serie,
      G=tipo, H=monto_clp, I=monto_clp_cuota, J=cuotas, K=uf_dia, L=monto_uf, M=monto_uf_cuota.
    Datos desde fila 13 (índice 0 = fila 1). Terminan cuando col A es None.
    """
    path = Path(excel_path)
    if not path.exists():
        return {'error': f'Archivo no encontrado: {excel_path}'}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['A&R PT']
    except Exception as e:
        return {'error': f'Error abriendo Excel: {e}'}

    NEMO = 'CFITRIPT-E'
    FONDO = 'PT'
    START_ROW = 13  # primera fila de datos (1-based)

    rows = list(ws.iter_rows(min_row=START_ROW, values_only=True))
    wb.close()

    file_hash = _hash_file(excel_path)
    source_file = Path(excel_path).name

    # Colectar por tipo
    dividendos: list = []
    vr_contable: list = []
    vr_bursatil: list = []

    for i, row in enumerate(rows):
        if row[0] is None and row[3] is None:
            break
        fecha_val = row[3]
        detalle = str(row[4]).strip() if row[4] else ''
        monto_clp = row[7]
        monto_clp_cuota = row[8]
        cuotas = row[9]
        uf_dia = row[10]
        monto_uf = row[11]
        monto_uf_cuota = row[12]

        if not fecha_val or not detalle:
            continue

        fecha_str = fecha_val.strftime('%Y-%m-%d') if hasattr(fecha_val, 'strftime') else str(fecha_val)
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
        periodo = f"{fecha_obj.year}-{fecha_obj.month:02d}"
        src_row = START_ROW + i

        if 'Dividendo' in detalle:
            dividendos.append((FONDO, NEMO, fecha_str, monto_clp_cuota, monto_uf_cuota, periodo, source_file, file_hash, src_row))
        elif 'VR Contable' in detalle:
            vr_contable.append((FONDO, NEMO, fecha_str, 'contable', monto_uf_cuota, monto_clp_cuota, cuotas, uf_dia, periodo, source_file, file_hash, src_row))
        elif 'VR Burs' in detalle:
            vr_bursatil.append((FONDO, NEMO, fecha_str, monto_clp_cuota, monto_uf, uf_dia, periodo, source_file, file_hash, src_row))

    conn = get_conn()

    # ── dividendos ────────────────────────────────────────────────────────────
    div_ins = 0
    for row in dividendos:
        fk, nm, fecha, clp_c, uf_c, per, sf, fh, sr = row
        try:
            conn.execute("""
                INSERT OR IGNORE INTO raw_dividendo_line
                (fondo_key, nemotecnico, fecha_pago, monto_clp_cuota, monto_uf_cuota,
                 periodo, source_file, file_hash)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (fk, nm, fecha, clp_c, uf_c, per, sf, fh))
            div_ins += conn.execute("SELECT changes()").fetchone()[0]
        except Exception:
            pass

    # ── valor cuota libro (VR Contable) ───────────────────────────────────────
    vc_ins = 0
    cuotas_ins = 0
    for row in vr_contable:
        fk, nm, fecha, tipo, precio_uf, precio_clp, cuotas, uf_dia, per, sf, fh, sr = row

        # EEFF PDF tiene precedencia si ya existe
        exists = conn.execute("""
            SELECT 1 FROM raw_valor_cuota_line
            WHERE fondo_key=? AND nemotecnico=? AND fecha=? AND tipo='contable'
              AND source_file NOT LIKE '%cdg%'
        """, (fk, nm, fecha)).fetchone()
        if not exists and precio_uf:
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO raw_valor_cuota_line
                    (fondo_key, nemotecnico, fecha, tipo, precio_uf, precio_clp, uf_dia,
                     cuotas, periodo, source_file, file_hash)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (fk, nm, fecha, tipo, precio_uf, precio_clp, uf_dia, cuotas, per, sf, fh))
                vc_ins += conn.execute("SELECT changes()").fetchone()[0]
            except Exception:
                pass

        # Cuotas en circulación (siempre del VR Contable, más confiable)
        if cuotas:
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO raw_cuota_en_circulacion_line
                    (fondo_key, nemotecnico, fecha, cuotas, periodo, source_file, file_hash)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (fk, nm, fecha, cuotas, per, sf, fh))
                cuotas_ins += conn.execute("SELECT changes()").fetchone()[0]
            except Exception:
                pass

    # ── precio cuota bursátil (gap-fill histórico) ────────────────────────────
    # raw_precio_cuota_line tiene desde 2024-05; el CDG tiene desde 2017
    precio_ins = 0
    pat_ins = 0
    for row in vr_bursatil:
        fk, nm, fecha, clp_cuota, pat_uf, uf_dia, per, sf, fh, sr = row
        if clp_cuota:
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO raw_precio_cuota_line
                    (nemotecnico, fecha, precio_clp, fuente, loaded_at)
                    VALUES (?, ?, ?, 'cdg_ar_pt', datetime('now'))
                """, (nm, fecha, clp_cuota))
                precio_ins += conn.execute("SELECT changes()").fetchone()[0]
            except Exception:
                pass
        if pat_uf:
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO raw_patrimonio_bursatil_line
                    (fondo_key, nemotecnico, fecha, patrimonio_uf, periodo, source_file, file_hash)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (fk, nm, fecha, pat_uf, per, sf, fh))
                pat_ins += conn.execute("SELECT changes()").fetchone()[0]
            except Exception:
                pass

    conn.commit()
    return {
        'dividendos_insertados': div_ins,
        'valor_cuota_contable_insertados': vc_ins,
        'cuotas_circulacion_insertadas': cuotas_ins,
        'precios_bursatiles_insertados': precio_ins,
        'patrimonio_bursatil_insertados': pat_ins,
    }


def ingest_ar_apo(excel_path: str) -> Dict:
    """
    Lee hoja 'A&R Apoquindo' del CDG extract y persiste:
      - raw_valor_cuota_line          (Detalle='VR Contable', tipo='contable')
      - raw_cuota_en_circulacion_line (cuotas de cada VR Contable)
      - raw_dividendo_line            (Detalle='Dividendo' o 'Disminución')

    Apo no cotiza en bolsa → sin VR Bursátil ni raw_precio_cuota_line.
    Serie única → nemotecnico = 'Apo'.
    Columnas: A=año, B=mes, C=id, D=fecha, E=detalle, F=serie(None),
      G=tipo, H=monto_clp, I=monto_clp_cuota, J=cuotas, K=uf_dia,
      L=monto_uf, M=monto_uf_cuota. Datos desde fila 13.
    """
    path = Path(excel_path)
    if not path.exists():
        return {'error': f'Archivo no encontrado: {excel_path}'}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['A&R Apoquindo']
    except Exception as e:
        return {'error': f'Error abriendo Excel: {e}'}

    NEMO = 'Apo'
    FONDO = 'Apo'
    START_ROW = 13

    rows = list(ws.iter_rows(min_row=START_ROW, values_only=True))
    wb.close()

    file_hash = _hash_file(excel_path)
    source_file = Path(excel_path).name

    vr_contable: list = []
    distribuciones: list = []

    for i, row in enumerate(rows):
        año = row[0]
        if año is None or (isinstance(año, int) and año < 2000):
            break
        fecha_val = row[3]
        detalle = str(row[4]).strip() if row[4] else ''
        monto_clp_cuota = row[8]
        cuotas = row[9]
        uf_dia = row[10]
        monto_uf = row[11]
        monto_uf_cuota = row[12]

        if not fecha_val or not detalle:
            continue

        fecha_str = fecha_val.strftime('%Y-%m-%d') if hasattr(fecha_val, 'strftime') else str(fecha_val)
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
        periodo = f"{fecha_obj.year}-{fecha_obj.month:02d}"

        if 'VR Contable' in detalle:
            vr_contable.append((FONDO, NEMO, fecha_str, 'contable', monto_uf_cuota,
                                monto_clp_cuota, cuotas, uf_dia, periodo, source_file, file_hash))
        elif 'Dividendo' in detalle:
            distribuciones.append((FONDO, NEMO, fecha_str, 'dividendo',
                                   monto_clp_cuota, monto_uf_cuota, periodo, source_file, file_hash))
        elif 'Disminuci' in detalle:  # Disminución (con tilde o sin tilde)
            distribuciones.append((FONDO, NEMO, fecha_str, 'disminucion',
                                   monto_clp_cuota, monto_uf_cuota, periodo, source_file, file_hash))

    conn = get_conn()

    vc_ins = cuotas_ins = 0
    for fk, nm, fecha, tipo, precio_uf, precio_clp, cuotas, uf_dia, per, sf, fh in vr_contable:
        if precio_uf:
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO raw_valor_cuota_line
                    (fondo_key, nemotecnico, fecha, tipo, precio_uf, precio_clp,
                     uf_dia, cuotas, periodo, source_file, file_hash)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (fk, nm, fecha, tipo, precio_uf, precio_clp, uf_dia, cuotas, per, sf, fh))
                vc_ins += conn.execute("SELECT changes()").fetchone()[0]
            except Exception:
                pass
        if cuotas:
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO raw_cuota_en_circulacion_line
                    (fondo_key, nemotecnico, fecha, cuotas, periodo, source_file, file_hash)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (fk, nm, fecha, cuotas, per, sf, fh))
                cuotas_ins += conn.execute("SELECT changes()").fetchone()[0]
            except Exception:
                pass

    # Idempotencia sin UNIQUE en raw_dividendo_line: verificar por file_hash
    existing_keys = {
        r[0] for r in conn.execute(
            "SELECT fecha_pago || '|' || tipo FROM raw_dividendo_line WHERE fondo_key=? AND file_hash=?",
            (FONDO, file_hash)
        ).fetchall()
    }
    div_ins = 0
    for fk, nm, fecha, tipo, clp_c, uf_c, per, sf, fh in distribuciones:
        key = f"{fecha}|{tipo}"
        if key in existing_keys:
            continue
        try:
            conn.execute("""
                INSERT INTO raw_dividendo_line
                (fondo_key, nemotecnico, fecha_pago, tipo, monto_clp_cuota, monto_uf_cuota,
                 periodo, source_file, file_hash)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (fk, nm, fecha, tipo, clp_c, uf_c, per, sf, fh))
            div_ins += 1
            existing_keys.add(key)
        except Exception:
            pass

    conn.commit()
    return {
        'valor_cuota_contable_insertados': vc_ins,
        'cuotas_circulacion_insertadas': cuotas_ins,
        'dividendos_insertados': div_ins,
    }


def _hash_file(path: str) -> str:
    """SHA256 del archivo."""
    sha = hashlib.sha256()
    with open(path, 'rb') as f:
        sha.update(f.read())
    return sha.hexdigest()[:16]
