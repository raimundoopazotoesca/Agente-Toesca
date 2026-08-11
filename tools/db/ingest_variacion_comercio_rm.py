"""Ingesta de la variación real anual de ventas RM (CNC), Total Locales:
Total Comercio + Supermercado Tradicional.

Fuente: 2 XLSX históricos descargados de CNC > Estudios:
  1. "Serie Histórica de Índice de Ventas del Comercio – R. Metropolitana"
     hoja "TL MINORISTA- BASE 2023", sección "VARIACIÓN ANUAL (%)",
     columna "TOTAL COMERCIO".
  2. "Serie Histórica del Índice de Ventas de Supermercados – R. Metropolitana"
     hoja "TL SUPERM. CNC - BASE 2023", bloque "VARIACIÓN REAL ANUAL",
     columna "Supermercados Tradicional".

Cada mes CNC revisa retrospectivamente la serie completa, así que cada
ingesta reemplaza toda la historia vigente (no solo agrega el último mes).

No expone CLI; lo consume scripts/ingesta_server.py (Flask).
"""
from __future__ import annotations

import hashlib
import re
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DB_PATH = ROOT / "memory" / "agente_toesca_v2.db"

sys.path.insert(0, str(ROOT))
from tools.db.connection import get_conn_for  # noqa: E402

_MES_A_NUM = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "SEPT": 9, "OCT": 10, "NOV": 11, "DIC": 12,
}

_MES_RE = re.compile(r"^([A-ZÑ]+)\.?(\d{4})?$")


def _mes_periodo(label: str, año_actual: int | None) -> tuple[str, int] | None:
    """Convierte 'ENE.2019' o 'FEB' (con año arrastrado) a ('YYYY-MM', año).

    Retorna None si el label no es un mes reconocible (p.ej. '2019 Prom.').
    """
    m = _MES_RE.match(label.strip().upper())
    if not m:
        return None
    mes_txt, año_txt = m.groups()
    if mes_txt not in _MES_A_NUM:
        return None
    año = int(año_txt) if año_txt else año_actual
    if año is None:
        return None
    return f"{año:04d}-{_MES_A_NUM[mes_txt]:02d}", año


def _parse_comercio(xlsx_bytes: bytes) -> dict[str, float]:
    """Lee 'TL MINORISTA- BASE 2023' -> VARIACIÓN ANUAL (%) -> TOTAL COMERCIO.

    Retorna {periodo: fracción}. Divide por 100 (el Excel trae % en puntos).
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.upper().startswith("TL ") and "MINORISTA" in s.upper()), None)
    if sheet_name is None:
        raise ValueError(
            "No se encontró la hoja 'TL MINORISTA- BASE 2023' (Total Locales) en el Excel de Comercio."
        )
    ws = wb[sheet_name]

    fila_seccion = next(
        (r for r in range(1, ws.max_row + 1) if isinstance(ws.cell(r, 2).value, str)
         and "VARIAC" in ws.cell(r, 2).value.upper() and "ANUAL" in ws.cell(r, 2).value.upper()
         and "ACUM" not in ws.cell(r, 2).value.upper()),
        None,
    )
    if fila_seccion is None:
        raise ValueError("No se encontró la sección 'VARIACIÓN ANUAL (%)' en la hoja de Comercio.")

    serie: dict[str, float] = {}
    año_actual = None
    for r in range(fila_seccion + 1, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if label is None:
            continue
        if isinstance(label, str) and label.strip().upper().startswith("FUENTE"):
            break
        if not isinstance(label, str):
            continue
        parsed = _mes_periodo(label, año_actual)
        if parsed is None:
            continue
        periodo, año_actual = parsed
        valor = ws.cell(r, 3).value
        if valor is None or valor == "-":
            continue
        serie[periodo] = float(valor) / 100.0
    if not serie:
        raise ValueError("La sección 'VARIACIÓN ANUAL (%)' de Comercio no tiene datos parseables.")
    return serie


def _parse_supermercado(xlsx_bytes: bytes) -> dict[str, float]:
    """Lee 'TL SUPERM. CNC - BASE 2023' -> bloque VARIACIÓN REAL ANUAL ->
    columna 'Supermercados Tradicional' (primera columna de datos del bloque).

    Retorna {periodo: fracción}.
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.upper().startswith("TL ") and "SUPERM" in s.upper()), None)
    if sheet_name is None:
        raise ValueError(
            "No se encontró la hoja 'TL SUPERM. CNC - BASE 2023' (Total Locales) en el Excel de Supermercados."
        )
    ws = wb[sheet_name]

    fila_seccion = None
    col_mes = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "VARIAC" in v.upper() and "ANUAL" in v.upper():
                fila_seccion = r
                col_mes = c
                break
        if fila_seccion is not None:
            break
    if fila_seccion is None:
        raise ValueError("No se encontró el bloque 'VARIACIÓN REAL ANUAL' en la hoja de Supermercados.")

    # Fila de 'MES' + subencabezados está debajo del título; los datos parten
    # un par de filas más abajo. Los valores quedan en la primera columna de
    # datos a la derecha de la columna 'MES' de ese bloque.
    col_valor = col_mes + 1

    serie: dict[str, float] = {}
    año_actual = None
    for r in range(fila_seccion + 1, ws.max_row + 1):
        label = ws.cell(r, col_mes).value
        if label is None:
            continue
        if isinstance(label, str) and label.strip().upper().startswith("FUENTE"):
            break
        if not isinstance(label, str):
            continue
        parsed = _mes_periodo(label, año_actual)
        if parsed is None:
            continue
        periodo, año_actual = parsed
        valor = ws.cell(r, col_valor).value
        if valor is None or valor == "-":
            continue
        serie[periodo] = float(valor) / 100.0
    if not serie:
        raise ValueError("El bloque 'VARIACIÓN REAL ANUAL' de Supermercados no tiene datos parseables.")
    return serie


class ValidationResult:
    def __init__(self):
        self.ok = True
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.data: dict = {}

    def add_error(self, msg: str):
        self.errors.append(msg)
        self.ok = False

    def to_dict(self) -> dict:
        return {"ok": self.ok, "errors": self.errors, "warnings": self.warnings, **self.data}


def _merge(comercio: dict[str, float], supermercado: dict[str, float]) -> dict:
    periodos = sorted(set(comercio) | set(supermercado))
    filas = []
    for periodo in periodos:
        filas.append({
            "periodo": periodo,
            "total_comercio_var_anual": comercio.get(periodo),
            "supermercado_tradicional_var_anual": supermercado.get(periodo),
        })
    return {
        "filas": filas,
        "ultimo_periodo_comercio": max(comercio) if comercio else None,
        "ultimo_periodo_supermercado": max(supermercado) if supermercado else None,
    }


def validate(comercio_bytes: bytes, comercio_filename: str, supermercado_bytes: bytes, supermercado_filename: str) -> ValidationResult:
    result = ValidationResult()
    if not comercio_bytes:
        result.add_error("Sube el XLSX histórico de Índice de Ventas del Comercio RM.")
    if not supermercado_bytes:
        result.add_error("Sube el XLSX histórico de Índice de Ventas de Supermercados RM.")
    if not result.ok:
        return result

    try:
        comercio = _parse_comercio(comercio_bytes)
    except ValueError as exc:
        result.add_error(f"Comercio: {exc}")
        comercio = {}
    try:
        supermercado = _parse_supermercado(supermercado_bytes)
    except ValueError as exc:
        result.add_error(f"Supermercados: {exc}")
        supermercado = {}
    if not result.ok:
        return result

    for periodo, valor in {**comercio, **supermercado}.items():
        if valor is not None and abs(valor) > 1:
            result.add_error(f"{periodo}: valor fuera de rango razonable ({valor * 100:.1f}%)")
    if not result.ok:
        return result

    merged = _merge(comercio, supermercado)
    if merged["ultimo_periodo_comercio"] != merged["ultimo_periodo_supermercado"]:
        result.warnings.append(
            f"El último mes no coincide entre las dos fuentes: Comercio={merged['ultimo_periodo_comercio']}, "
            f"Supermercados={merged['ultimo_periodo_supermercado']}. Verifica que ambos XLSX sean del mismo mes de CNC."
        )

    fhash = hashlib.sha256(comercio_bytes + b"|" + supermercado_bytes).hexdigest()
    con = get_conn_for(str(DB_PATH))
    try:
        n_vigentes = con.execute(
            "SELECT COUNT(*) FROM raw_variacion_comercio_rm WHERE superseded_at IS NULL"
        ).fetchone()[0]
        ya_mismo_hash = con.execute(
            "SELECT COUNT(*) FROM raw_variacion_comercio_rm WHERE file_hash=?", (fhash,)
        ).fetchone()[0]
    finally:
        con.close()

    if n_vigentes:
        result.warnings.append(
            f"Se reemplazará la serie histórica completa: {n_vigentes} fila(s) vigentes actuales quedarán "
            f"superseded y se insertarán {len(merged['filas'])} fila(s) nuevas."
        )

    result.data = {
        "filas": merged["filas"],
        "ultimo_periodo": merged["ultimo_periodo_comercio"],
        "file_hash": fhash,
        "ya_ingestado": bool(ya_mismo_hash),
    }
    return result


def commit(comercio_bytes: bytes, comercio_filename: str, supermercado_bytes: bytes, supermercado_filename: str) -> dict:
    result = validate(comercio_bytes, comercio_filename, supermercado_bytes, supermercado_filename)
    if not result.ok:
        raise ValueError("No se puede ingestar: " + "; ".join(result.errors))

    filas = result.data["filas"]
    fhash = result.data["file_hash"]

    con = get_conn_for(str(DB_PATH))
    try:
        if result.data["ya_ingestado"]:
            return {"status": "skipped_duplicate", "run_id": None, "filas_insertadas": 0, "filas_superseded": 0}

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur = con.execute(
            """INSERT INTO ingest_run (tool, source_file, file_hash, started_at, status, periodo_declarado)
               VALUES (?,?,?,?,?,?)""",
            ("ingest_variacion_comercio_rm", f"{comercio_filename}+{supermercado_filename}", fhash, now, "running", result.data["ultimo_periodo"]),
        )
        run_id = cur.lastrowid

        cur2 = con.execute(
            "UPDATE raw_variacion_comercio_rm SET superseded_at=? WHERE superseded_at IS NULL",
            (now,),
        )
        filas_superseded = cur2.rowcount if cur2.rowcount > 0 else 0

        con.executemany(
            """INSERT INTO raw_variacion_comercio_rm
               (periodo, total_comercio_var_anual, supermercado_tradicional_var_anual, file_hash, ingest_run_id)
               VALUES (?,?,?,?,?)""",
            [(f["periodo"], f["total_comercio_var_anual"], f["supermercado_tradicional_var_anual"], fhash, run_id) for f in filas],
        )

        con.execute(
            "UPDATE ingest_run SET status=?, ended_at=?, rows_in=?, rows_loaded=? WHERE id=?",
            ("ok", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), len(filas), len(filas), run_id),
        )
        con.commit()
        return {
            "status": "ok",
            "run_id": run_id,
            "filas_insertadas": len(filas),
            "filas_superseded": filas_superseded,
            "ultimo_periodo": result.data["ultimo_periodo"],
        }
    finally:
        con.close()
