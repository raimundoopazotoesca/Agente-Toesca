"""Ingesta mensual de ingresos/NOI por activo desde la planilla EEFF fuente
('NUEVO ORDEN/Renta Comercial/Ingresos Activos/<activo>/<año>/MM-YYYY
INFORME ....xlsx'), hoja 'ESTADO DE RESULTADO' (con o sin sufijo de año).

A diferencia de ingest_er_vina.py / ingest_er_curico.py (que leen el
histórico consolidado 'RAW/NOI VIÑA.xlsx' / 'RAW/NOI Curico.xlsx' completo),
este parser lee la planilla EEFF mensual del activo y extrae **un solo mes**
por corrida — el que el usuario elige en la UI (ver [[project_ingesta_er_mensual]]).
El archivo trae 13 columnas de valor (diciembre del año anterior + los 12
meses del año en curso, aunque solo estén llenos hasta el mes de cierre) más
una columna 'TOTAL ACUM.' que se ignora.

Layout confirmado 2026-07-27 sobre archivos de Viña Centro y Mall Curicó
(estructura idéntica entre ambos activos):
  - hoja cuyo nombre empieza con 'ESTADO DE RESULTADO' (a veces con sufijo
    de año, ej. 'ESTADO DE RESULTADO 2026'; Curicó no lo trae)
  - fila ancla: columna B == 'Ingreso de Explotacion' (case-insensitive)
  - fila de encabezado de meses: la misma fila ancla (Curicó) o hasta 4 filas
    arriba (Viña) — se detecta por la fila con más celdas de mes parseables
  - columna B combina código de cuenta + nombre en una sola celda
    (ej. '4-1-01-100  ARRIENDO TIENDAS ANCLAS'); columna A puede traer el
    código repetido o vacío, no se usa
  - secciones (en orden): Ingreso de Explotacion (INGRESOS_OPERACION) →
    Ingreso Fuera De Explotacion (INGRESO_FUERA_EXPLOTACION, no operacional)
    → Gastos de administración y ventas (GASTOS_OPERACION) → terminador
    'Total Operacional'
  - subtotales de control: 'Total Resultado Operación' (valida suma de
    Ingreso de Explotación) y 'Total Gastos de administración y ventas'

monto_clp = pesos reales del mes elegido. monto_uf = monto_clp / UF de fin
de mes (fact_uf de la DB, igual criterio que los parsers históricos).
"""
from __future__ import annotations

import calendar
import hashlib
import re
import sqlite3
from typing import Optional

import openpyxl

_ACTIVOS: dict[str, str] = {
    "Viña Centro": "vina",
    "Mall Curicó": "curico",
}

_ACCOUNT_RE = re.compile(r"^(\d(?:-\d{1,3}){3})\s+(.+)$")
_ANCLA = re.compile(r"^ingreso de explotacion$", re.I)
_HEADER_FUERA_EXPLOTACION = re.compile(r"^ingreso fuera de explotacion$", re.I)
_HEADER_GASTOS_OPERACION = re.compile(r"^gastos de administraci.n y ventas$", re.I)
_SUBTOTAL_RESULTADO_OPERACION = re.compile(r"^total resultado operaci.n$", re.I)
_SUBTOTAL_GASTOS_ADMIN = re.compile(r"^total gastos de administraci.n y ventas$", re.I)
_TERMINATOR = re.compile(r"^total operacional$", re.I)

_ES_OPERACIONAL = {
    "INGRESOS_OPERACION": 1,
    "GASTOS_OPERACION": 1,
    "INGRESO_FUERA_EXPLOTACION": 0,
}

_MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

_TOLERANCIA_CLP = 2000.0


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip())


def _file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def _uf_fin_mes(conn: sqlite3.Connection, periodo: str) -> float:
    from tools.db import repo_fact

    year, month = (int(x) for x in periodo.split("-"))
    last_day = calendar.monthrange(year, month)[1]
    fecha = f"{year:04d}-{month:02d}-{last_day:02d}"
    return repo_fact.get_uf(conn, fecha)


def _pick_sheet(wb) -> "openpyxl.worksheet.worksheet.Worksheet":
    candidatas = [ws for ws in wb.worksheets if _norm(ws.title).lower().startswith("estado de resultado")]
    if not candidatas:
        raise ValueError("No se encontró una hoja 'ESTADO DE RESULTADO' en el archivo.")
    return candidatas[-1]


def _find_title_year(ws, max_row: int = 20) -> int:
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
        for cell in row:
            if isinstance(cell, str):
                m = re.findall(r"\b(20\d{2})\b", cell)
                if m:
                    return int(m[-1])
    raise ValueError(f"No se pudo determinar el año del informe en la hoja '{ws.title}'.")


def _find_ancla(all_rows) -> int:
    for i, row in enumerate(all_rows):
        val = _norm(row[1].value if len(row) > 1 else None)
        if _ANCLA.match(val):
            return i
    raise ValueError("No se encontró la fila ancla 'Ingreso de Explotacion'.")


def _find_header_row(all_rows, ancla_idx: int) -> dict[int, tuple]:
    """Busca en [ancla_idx-4, ancla_idx] la fila con más celdas mes-parseables.

    Devuelve {col_idx (0-based): (mes_num, es_datetime)}.
    """
    best_idx, best_cells = None, {}
    for i in range(max(0, ancla_idx - 4), ancla_idx + 1):
        row = all_rows[i]
        cells = {}
        for col_idx, cell in enumerate(row):
            v = cell.value
            if hasattr(v, "month") and hasattr(v, "year"):
                cells[col_idx] = (v.month, True)
            elif isinstance(v, str) and _norm(v).lower() in _MESES:
                cells[col_idx] = (_MESES[_norm(v).lower()], False)
        if len(cells) > len(best_cells):
            best_idx, best_cells = i, cells
    if best_idx is None or len(best_cells) < 10:
        raise ValueError("No se encontró la fila de encabezado de meses (se esperaban ~13 columnas de mes).")
    return best_cells


def _period_by_col(header_cells: dict[int, tuple], all_rows, header_idx: int, title_year: int) -> dict[int, str]:
    out: dict[int, str] = {}
    running_year: Optional[int] = None
    prev_month: Optional[int] = None
    for col_idx in sorted(header_cells):
        mes_num, es_datetime = header_cells[col_idx]
        if es_datetime:
            cell_val = all_rows[header_idx][col_idx].value
            year = cell_val.year
        else:
            if running_year is None:
                year = title_year - 1 if mes_num == 12 else title_year
            else:
                year = running_year + 1 if (prev_month == 12 and mes_num == 1) else running_year
        running_year = year
        prev_month = mes_num
        out[col_idx] = f"{year:04d}-{mes_num:02d}"
    return out


def parse_mes(xlsx_path: str, activo_key: str, periodo: str,
              conn: "sqlite3.Connection | None" = None) -> list[dict]:
    """Extrae las cuentas de un único mes de la hoja ESTADO DE RESULTADO.

    Lanza ValueError si el periodo no existe en el archivo, si falta la
    columna con datos para ese mes, o si la validación de integridad falla.
    """
    if activo_key not in _ACTIVOS:
        raise ValueError(f"Activo no soportado por ingest_er_mensual: {activo_key!r}")

    owns_conn = conn is None
    if owns_conn:
        from tools.db.connection import get_conn
        conn = get_conn()

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = _pick_sheet(wb)
        sheet_name = ws.title
        all_rows = list(ws.iter_rows(values_only=False))
        title_year = _find_title_year(ws)
        wb.close()

        ancla_idx = _find_ancla(all_rows)
        header_cells = _find_header_row(all_rows, ancla_idx)
        header_idx = ancla_idx  # se recalcula abajo si el header está en otra fila
        for i in range(max(0, ancla_idx - 4), ancla_idx + 1):
            candidato = {}
            for col_idx, cell in enumerate(all_rows[i]):
                v = cell.value
                if hasattr(v, "month") and hasattr(v, "year"):
                    candidato[col_idx] = True
                elif isinstance(v, str) and _norm(v).lower() in _MESES:
                    candidato[col_idx] = True
            if len(candidato) == len(header_cells) and set(candidato) == set(header_cells):
                header_idx = i
                break

        period_by_col = _period_by_col(header_cells, all_rows, header_idx, title_year)
        col_for_periodo = next((c for c, p in period_by_col.items() if p == periodo), None)
        if col_for_periodo is None:
            disponibles = ", ".join(sorted(period_by_col.values()))
            raise ValueError(f"El archivo no tiene columna para el periodo {periodo}. Disponibles: {disponibles}")

        out: list[dict] = []
        current_seccion = "INGRESOS_OPERACION"  # la fila ancla YA es esta sección
        terminador_encontrado = False
        suma_ingreso: float = 0.0
        suma_gastos: float = 0.0
        subtotal_ingreso: Optional[float] = None
        subtotal_gastos: Optional[float] = None

        for i in range(ancla_idx + 1, len(all_rows)):
            row = all_rows[i]
            raw_label = row[1].value if len(row) > 1 else None
            label = _norm(raw_label)
            if not label:
                continue

            if _TERMINATOR.match(label):
                terminador_encontrado = True
                break
            if _HEADER_FUERA_EXPLOTACION.match(label):
                current_seccion = "INGRESO_FUERA_EXPLOTACION"
                continue
            if _HEADER_GASTOS_OPERACION.match(label):
                current_seccion = "GASTOS_OPERACION"
                continue
            if _SUBTOTAL_RESULTADO_OPERACION.match(label):
                cell = row[col_for_periodo] if col_for_periodo < len(row) else None
                if cell is not None and cell.value is not None:
                    subtotal_ingreso = float(cell.value)
                continue
            if _SUBTOTAL_GASTOS_ADMIN.match(label):
                cell = row[col_for_periodo] if col_for_periodo < len(row) else None
                if cell is not None and cell.value is not None:
                    subtotal_gastos = float(cell.value)
                continue

            m = _ACCOUNT_RE.match(label)
            if not m:
                continue  # header de categoría (ej. "SEGURIDAD"), sin código de cuenta

            cuenta_codigo, cuenta_nombre = m.group(1), m.group(2).strip()
            cell = row[col_for_periodo] if col_for_periodo < len(row) else None
            monto_clp = float(cell.value) if cell is not None and cell.value is not None else 0.0

            if current_seccion == "INGRESOS_OPERACION":
                suma_ingreso += monto_clp
            elif current_seccion == "GASTOS_OPERACION":
                suma_gastos += monto_clp

            out.append({
                "activo_key":     activo_key,
                "periodo":        periodo,
                "cuenta_codigo":  cuenta_codigo,
                "cuenta_nombre":  cuenta_nombre,
                "monto_clp":      monto_clp,
                "monto_uf":       monto_clp / _uf_fin_mes(conn, periodo),
                "seccion":        current_seccion,
                "es_operacional": _ES_OPERACIONAL[current_seccion],
                "source_file":    xlsx_path,
                "source_sheet":   sheet_name,
                "source_row":     i + 1,
            })

        if not terminador_encontrado:
            raise ValueError("No se encontró la fila 'Total Operacional'.")
        if not out:
            raise ValueError(f"No se encontraron cuentas con datos para el periodo {periodo}.")

        if subtotal_ingreso is not None and abs(suma_ingreso - subtotal_ingreso) >= _TOLERANCIA_CLP:
            raise ValueError(
                f"Validación de integridad falló en {periodo}: suma Ingreso Explotación={suma_ingreso!r} "
                f"!= Total Resultado Operación={subtotal_ingreso!r}"
            )
        if subtotal_gastos is not None and abs(suma_gastos - subtotal_gastos) >= _TOLERANCIA_CLP:
            raise ValueError(
                f"Validación de integridad falló en {periodo}: suma Gastos Admin y Ventas={suma_gastos!r} "
                f"!= Total Gastos de administración y ventas={subtotal_gastos!r}"
            )

        return out
    finally:
        if owns_conn:
            conn.close()


def ultimo_periodo_disponible(xlsx_path: str, activo_key: str) -> Optional[str]:
    """Best-effort: último periodo del archivo con datos en 'Total Resultado
    Operación', para avisar en la UI si el usuario eligió un mes distinto."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = _pick_sheet(wb)
    all_rows = list(ws.iter_rows(values_only=False))
    title_year = _find_title_year(ws)
    wb.close()

    ancla_idx = _find_ancla(all_rows)
    header_cells = _find_header_row(all_rows, ancla_idx)
    header_idx = ancla_idx
    for i in range(max(0, ancla_idx - 4), ancla_idx + 1):
        candidato = {}
        for col_idx, cell in enumerate(all_rows[i]):
            v = cell.value
            if hasattr(v, "month") and hasattr(v, "year"):
                candidato[col_idx] = True
            elif isinstance(v, str) and _norm(v).lower() in _MESES:
                candidato[col_idx] = True
        if len(candidato) == len(header_cells) and set(candidato) == set(header_cells):
            header_idx = i
            break
    period_by_col = _period_by_col(header_cells, all_rows, header_idx, title_year)

    for i in range(ancla_idx + 1, len(all_rows)):
        row = all_rows[i]
        label = _norm(row[1].value if len(row) > 1 else None)
        if _TERMINATOR.match(label):
            break
        if not _ACCOUNT_RE.match(label):
            continue  # subtotales/headers de categoría: pueden dar 0 por fórmula aunque el mes esté vacío
        filled = [
            period_by_col[c] for c in sorted(period_by_col)
            if c < len(row) and row[c].value is not None
        ]
        if filled:
            return max(filled)
        return None
    return None


# ── Persistencia ─────────────────────────────────────────────────────────

def persist_mes(xlsx_path: str, activo_key: str, periodo: str,
                 conn: "sqlite3.Connection | None" = None) -> dict:
    """Ingesta idempotente de UN mes en raw_er_activo_line.

    A diferencia de los parsers históricos (idempotencia por file_hash sobre
    todo el activo), acá el supersede es por (activo_key, periodo): subir el
    mismo archivo mensual el próximo mes con una columna más no debe tocar
    los meses ya cerrados de otros periodos.
    """
    from tools.db import repo_audit, repo_er_activo

    owns_conn = conn is None
    if owns_conn:
        from tools.db.connection import get_conn
        conn = get_conn()

    try:
        with open(xlsx_path, "rb") as f:
            file_hash = _file_hash(f.read())

        prev = conn.execute(
            """SELECT 1 FROM raw_er_activo_line
                WHERE file_hash=? AND activo_key=? AND periodo=? AND superseded_at IS NULL
                LIMIT 1""",
            (file_hash, activo_key, periodo),
        ).fetchone()
        if prev is not None:
            return {"status": "skipped_idempotent", "rows": 0, "file_hash": file_hash, "ingest_run_id": None}

        lines = parse_mes(xlsx_path, activo_key, periodo, conn=conn)
        for line in lines:
            line["file_hash"] = file_hash

        n_previas = conn.execute(
            """SELECT COUNT(*) FROM raw_er_activo_line
                WHERE activo_key=? AND periodo=? AND superseded_at IS NULL""",
            (activo_key, periodo),
        ).fetchone()[0]
        status = "superseded_and_reinserted" if n_previas else "inserted"
        conn.execute(
            """UPDATE raw_er_activo_line SET superseded_at=datetime('now')
                WHERE activo_key=? AND periodo=? AND superseded_at IS NULL""",
            (activo_key, periodo),
        )

        run_id = repo_audit.start_ingest_run(
            conn, tool="ingest_er_mensual", source_file=xlsx_path, file_hash=file_hash,
        )
        inserted = repo_er_activo.insert_lines(conn, lines, run_id)
        repo_audit.finish_ingest_run(
            conn, run_id, rows_in=len(lines), rows_loaded=inserted, status="ok",
        )
        return {"status": status, "rows": inserted, "file_hash": file_hash, "ingest_run_id": run_id}
    finally:
        if owns_conn:
            conn.close()
