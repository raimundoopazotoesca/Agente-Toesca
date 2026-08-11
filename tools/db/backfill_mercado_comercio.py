"""Carga única del histórico de mercado de comercio minorista RM (CNC)
desde el Excel armado por el usuario (hoja "Histórico publicado").

Uso: python -m tools.db.backfill_mercado_comercio <ruta_al_xlsx>

Idempotente: usa hash del archivo completo como file_hash — reejecutar
con el mismo archivo no duplica filas.
"""
from __future__ import annotations

import hashlib
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DB_PATH = ROOT / "memory" / "agente_toesca_v2.db"

sys.path.insert(0, str(ROOT))
from tools.db.connection import get_conn_for  # noqa: E402
from tools.db.ingest_mercado_comercio import CATEGORIAS_ORDEN  # noqa: E402

SHEET_NAME = "Histórico publicado"
HEADER_ROW = 5
FIRST_DATA_ROW = 6

# Nombre de columna en el Excel -> nombre de categoría en la DB (idéntico
# a CATEGORIAS_ORDEN salvo "Supermercado Tradicional" -> "Supermercados").
_COL_A_CATEGORIA = {
    "Total Comercio": "Total Comercio",
    "Vestuario": "Vestuario",
    "Calzado": "Calzado",
    "Artefactos Eléctricos": "Artefactos Eléctricos",
    "Línea Hogar": "Línea Hogar",
    "Muebles": "Muebles",
    "Supermercado Tradicional": "Supermercados",
}


def backfill(xlsx_path: str, db_path: str | None = None) -> dict:
    db_path = db_path or str(DEFAULT_DB_PATH)
    file_hash = hashlib.sha256(Path(xlsx_path).read_bytes()).hexdigest()

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]
    header = [c.value for c in ws[HEADER_ROW]]
    col_idx = {name: i for i, name in enumerate(header) if name in _COL_A_CATEGORIA}
    mes_idx = header.index("Mes")

    con = get_conn_for(db_path)
    try:
        ya_cargado = con.execute(
            "SELECT COUNT(*) FROM raw_mercado_comercio WHERE file_hash=?", (file_hash,)
        ).fetchone()[0]
        if ya_cargado:
            return {"status": "skipped_duplicate", "filas_insertadas": 0, "periodos": []}

        rows_to_insert = []
        periodos = []
        source_row = 0
        for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
            if row[mes_idx] is None:
                continue
            mes_val = row[mes_idx]
            periodo = mes_val.strftime("%Y-%m") if hasattr(mes_val, "strftime") else str(mes_val)[:7]
            periodos.append(periodo)
            for col_excel, categoria in _COL_A_CATEGORIA.items():
                valor = row[col_idx[col_excel]]
                rows_to_insert.append((periodo, categoria, valor, file_hash, source_row))
                source_row += 1

        con.executemany(
            """INSERT INTO raw_mercado_comercio
               (periodo, categoria, variacion_acumulada_pct, file_hash, source_row)
               VALUES (?,?,?,?,?)""",
            rows_to_insert,
        )
        con.commit()
        return {"status": "ok", "filas_insertadas": len(rows_to_insert), "periodos": periodos}
    finally:
        con.close()


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python -m tools.db.backfill_mercado_comercio <ruta_al_xlsx>")
        sys.exit(1)
    resultado = backfill(sys.argv[1])
    print(resultado)
