"""Ingesta de SharePoint RAW/"Vacancia histórica DB.xlsx" -> raw_vacancia_manual.

Formato fuente: hoja única, fila=activo, columna=mes. Bloque "m2 Útiles
Totales" (GLA) y bloque "m2 Vacantes", con nombres de activo en español que
no siempre calzan con dim_activo.activo_key (ver mapeo abajo).
"""
import argparse
import hashlib
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from tools.db.connection import get_conn

# Etiqueta de fila en el Excel -> (activo_key, tipo_unidad)
# tipo_unidad=None => el valor ya es el total consolidado del activo.
GLA_ROW_MAP = {
    "INMOSA": ("INMOSA", None),
    "Strip Machalí": ("Strip Machalí", None),
    "Chañarcillo": ("Sucden", None),  # mismo sociedad_key que Sucden
    "Parque Titanium Oficinas": ("PT_consolidado", "Oficinas"),
    "Parque Titanium Locales": ("PT_consolidado", "Locales Comerciales"),
    "Parque Titanium Bodegas": ("PT_consolidado", "Bodegas"),
    "Viña Centro": ("Viña Centro", None),
    "Fondo Apoquindo": ("Fondo Apoquindo", None),  # legacy: agregado Apo4501+Apo4700
    "Curicó": ("Mall Curicó", None),
    "Apoquindo 3001": ("Apo3001", None),
}

VACANCIA_ROW_MAP = {
    "INMOSA": ("INMOSA", None),
    "Strip Machalí": ("Strip Machalí", None),
    "SUCDEN": ("Sucden", None),
    "Parque Titanium Oficinas": ("PT_consolidado", "Oficinas"),
    "Parque Titanium Locales": ("PT_consolidado", "Locales Comerciales"),
    "Parque Titanium Bodegas": ("PT_consolidado", "Bodegas"),
    "Viña Centro": ("Viña Centro", None),
    "Apoquindo 4700": ("Apo4700", None),
    "Aporquindo 4501": ("Apo4501", None),
    "Fondo Apoquindo": ("Fondo Apoquindo", None),
    "Curicó": ("Mall Curicó", None),
    "Apoquindo 3001": ("Apo3001", None),
}

GLA_ROWS = range(3, 13)      # filas 3..12 (13 es el total, se ignora)
VACANCIA_ROWS = range(17, 29)  # filas 17..28 (29 es el total, se ignora)


def _file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _read_sheet(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Hoja1"]

    periods = {}
    for col in range(4, ws.max_column + 1):
        v = ws.cell(row=2, column=col).value
        if v is not None:
            periods[col] = v.strftime("%Y-%m")

    gla = {}  # (activo_key, tipo_unidad, periodo) -> m2
    for r in GLA_ROWS:
        label = ws.cell(row=r, column=1).value
        if not label:
            continue
        label = label.strip()
        mapped = GLA_ROW_MAP.get(label)
        if not mapped:
            print(f"[WARN] fila GLA sin mapear: {label!r} (fila {r})")
            continue
        activo_key, tipo_unidad = mapped
        for col, periodo in periods.items():
            val = ws.cell(row=r, column=col).value
            if val is not None:
                gla[(activo_key, tipo_unidad, periodo)] = float(val)

    vac = {}
    for r in VACANCIA_ROWS:
        label = ws.cell(row=r, column=1).value
        if not label:
            continue
        label = label.strip()
        mapped = VACANCIA_ROW_MAP.get(label)
        if not mapped:
            print(f"[WARN] fila vacancia sin mapear: {label!r} (fila {r})")
            continue
        activo_key, tipo_unidad = mapped
        for col, periodo in periods.items():
            val = ws.cell(row=r, column=col).value
            if val is not None:
                vac[(activo_key, tipo_unidad, periodo)] = float(val)

    return gla, vac


def ingest(path: Path, dry_run: bool = False) -> int:
    file_hash = _file_hash(path)
    gla, vac = _read_sheet(path)

    keys = sorted(set(gla) | set(vac))
    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        "INSERT INTO ingest_run (tool, source_file, file_hash) VALUES (?, ?, ?)",
        ("ingest_vacancia_manual", str(path), file_hash),
    )
    ingest_run_id = cur.lastrowid

    cur.execute(
        """
        UPDATE raw_vacancia_manual
           SET superseded_at = datetime('now')
         WHERE source_file = ? AND superseded_at IS NULL
        """,
        (str(path),),
    )

    n = 0
    for activo_key, tipo_unidad, periodo in keys:
        m2_gla = gla.get((activo_key, tipo_unidad, periodo))
        m2_vacantes = vac.get((activo_key, tipo_unidad, periodo))
        if m2_gla is None and m2_vacantes is None:
            continue
        cur.execute(
            """
            INSERT INTO raw_vacancia_manual
                (activo_key, tipo_unidad, periodo, m2_gla, m2_vacantes,
                 source_file, source_sheet, file_hash, ingest_run_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                activo_key,
                tipo_unidad,
                periodo,
                m2_gla,
                m2_vacantes,
                str(path),
                "Hoja1",
                file_hash,
                ingest_run_id,
            ),
        )
        n += 1

    if dry_run:
        conn.rollback()
        print(f"[DRY RUN] {n} filas se habrían insertado")
    else:
        conn.commit()
        print(f"{n} filas insertadas (ingest_run_id={ingest_run_id})")

    conn.close()
    return n


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("path", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    ingest(args.path, dry_run=args.dry_run)
