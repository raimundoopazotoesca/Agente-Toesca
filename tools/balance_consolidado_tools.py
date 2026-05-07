"""
Herramientas para actualizar el Balance Consolidado Rentas PT.

Entidades: Toesca Rentas PT (holding) + Inmobiliaria Boulevard + Torre A S.A.
Planilla: {MM.YYYY}- Balance Consolidado Rentas PT vF.xlsx
Resultado: {MM.YYYY}- Balance Consolidado Rentas PT vAgente.xlsx

Defaults por hoja (solo fallback; manda la regla de periodos pasados):
  Fondo PT    → EEFF PDF (M$ × 1000)
  Inmob Blvd  → balance EEFF PDF (M$ × 1000) + EERR Análisis xlsx (pesos)
  Torre A     → balance + EERR Análisis xlsx (pesos directos)
"""
import glob as glob_module
import os
import re
import shutil
import unicodedata
from calendar import monthrange
from datetime import date, datetime

import openpyxl
from openpyxl.utils import get_column_letter
from config import SHAREPOINT_DIR, WORK_DIR

# ─── Rutas base ────────────────────────────────────────────────────────────────

BALANCES_DIR = os.path.join(
    SHAREPOINT_DIR, "Control de Gestión", "Balances Consolidados"
)
TRI_EEFF_DIR = os.path.join(
    SHAREPOINT_DIR, "Fondos", "Rentas TRI", "EEFF"
)
PT_EEFF_DIR = os.path.join(SHAREPOINT_DIR, "Fondos", "Parque Titanium", "EEFF")
APO_EEFF_DIR = os.path.join(SHAREPOINT_DIR, "Fondos", "Apoquindo", "EEFF")

HOJAS_INPUT = ["Fondo PT", "Inmob Boulevard", "Torre A"]
APO_HOJAS_INPUT = ["Fondo Apoquindo", "Inmobilaria Apoquindo"]

# ─── Helpers de rutas ─────────────────────────────────────────────────────────

def _quarter_end(mes: int, año: int) -> date:
    return date(año, mes, monthrange(año, mes)[1])


def _mes_a_q(mes: int) -> int:
    return {3: 1, 6: 2, 9: 3, 12: 4}[mes]


def _period_status_label(mes: int, año: int) -> str:
    return f"Actualizado {_mes_a_q(mes)}Q{año % 100:02d}"


def _find_quarter_folder(parent: str, mes: int) -> str | None:
    q = _mes_a_q(mes)
    if not os.path.isdir(parent):
        return None
    for e in os.listdir(parent):
        full = os.path.join(parent, e)
        if os.path.isdir(full) and e.lower().strip() in (f"{q}t", f"{q}q"):
            return full
    return None


def _find_latest_vf(año: int, mes: int) -> str | None:
    """Encuentra el archivo vF más reciente para el año dado."""
    año_dir = os.path.join(BALANCES_DIR, str(año))
    q_dir = _find_quarter_folder(año_dir, mes)
    if q_dir:
        for f in os.listdir(q_dir):
            if "Balance Consolidado Rentas PT" in f and "vF" in f and f.endswith(".xlsx"):
                return os.path.join(q_dir, f)
    # búsqueda exhaustiva
    pattern = os.path.join(BALANCES_DIR, "**", "*Balance Consolidado Rentas PT*vF*.xlsx")
    matches = glob_module.glob(pattern, recursive=True)
    return max(matches, key=os.path.getmtime) if matches else None


def _find_eeff_fondo_pt(mes: int, año: int) -> str | None:
    año_dir = os.path.join(PT_EEFF_DIR, str(año))
    q_dir = _find_quarter_folder(año_dir, mes)
    if q_dir:
        for f in os.listdir(q_dir):
            if f.lower().endswith(".pdf") and "rentas pt" in f.lower():
                return os.path.join(q_dir, f)
    pattern = os.path.join(PT_EEFF_DIR, "**", f"*{año}{mes:02d}*PT*.pdf")
    matches = glob_module.glob(pattern, recursive=True)
    return matches[0] if matches else None


def _find_latest_vf_apoquindo(año: int, mes: int) -> str | None:
    """Encuentra el archivo vF mas reciente de Balance Consolidado Rentas Apoquindo."""
    año_dir = os.path.join(BALANCES_DIR, str(año))
    q_dir = _find_quarter_folder(año_dir, mes)
    if q_dir:
        for f in os.listdir(q_dir):
            if "Balance Consolidado Rentas Apoquindo" in f and "vF" in f and f.endswith(".xlsx"):
                return os.path.join(q_dir, f)
    pattern = os.path.join(BALANCES_DIR, "**", "*Balance Consolidado Rentas Apoquindo*vF*.xlsx")
    matches = glob_module.glob(pattern, recursive=True)
    return max(matches, key=os.path.getmtime) if matches else None


def _find_eeff_fondo_apoquindo(mes: int, año: int) -> str | None:
    año_dir = os.path.join(APO_EEFF_DIR, str(año))
    q_dir = _find_quarter_folder(año_dir, mes)
    if q_dir:
        for f in os.listdir(q_dir):
            fl = f.lower()
            if fl.endswith(".pdf") and "apoquindo" in fl:
                return os.path.join(q_dir, f)
    if mes == 3:
        prev_q_dir = _find_quarter_folder(os.path.join(APO_EEFF_DIR, str(año - 1)), 12)
        if prev_q_dir:
            for f in os.listdir(prev_q_dir):
                fl = f.lower()
                if fl.endswith(".pdf") and "apoquindo" in fl:
                    return os.path.join(prev_q_dir, f)
    pattern = os.path.join(APO_EEFF_DIR, "**", f"*Apoquindo*{año}*{mes:02d}*.pdf")
    matches = glob_module.glob(pattern, recursive=True)
    return matches[0] if matches else None


def _find_eeff_inmobiliaria_apoquindo(mes: int, año: int) -> str | None:
    apo_dir = os.path.join(TRI_EEFF_DIR, "Activos", "Inmobiliaria Apoquindo")
    if os.path.isdir(apo_dir):
        for f in os.listdir(apo_dir):
            fl = f.lower()
            if fl.endswith(".pdf") and "apoquindo" in fl:
                return os.path.join(apo_dir, f)
    pattern = os.path.join(TRI_EEFF_DIR, "**", "*Inmobiliaria Apoquindo*.pdf")
    matches = glob_module.glob(pattern, recursive=True)
    return matches[0] if matches else None


def _find_analisis_inmobiliaria_apoquindo(mes: int, año: int) -> str | None:
    apo_dir = os.path.join(TRI_EEFF_DIR, "Activos", "Inmobiliaria Apoquindo")
    if not os.path.isdir(apo_dir):
        return None

    exact_candidates = []
    undated_candidates = []
    period_tokens = (
        f"{mes:02d}-{año}",
        f"{año}-{mes:02d}",
        f"{año} {mes:02d}",
        f"{mes:02d}.{año}",
    )
    for f in os.listdir(apo_dir):
        fl = f.lower()
        fnorm = _norm_label(f)
        if fl.startswith("~$") or not fl.endswith((".xlsx", ".xlsm")):
            continue
        if "apoquindo" not in fnorm:
            continue
        if "anal" not in fnorm and "matriz" not in fnorm:
            continue
        full = os.path.join(apo_dir, f)
        score = int(os.path.getmtime(full) // 1_000_000)
        if any(token in f for token in period_tokens):
            exact_candidates.append((score, full))
        elif not re.search(r"\d{2}[-.]\d{4}|\d{4}[- ]\d{2}", f):
            undated_candidates.append((score, full))

    if exact_candidates:
        return max(exact_candidates)[1]
    if undated_candidates:
        return max(undated_candidates)[1]
    return None


def _find_boulevard_files(mes: int, año: int) -> tuple[str | None, str | None]:
    bvd_dir = os.path.join(TRI_EEFF_DIR, "Activos", "Boulevard")
    eeff_pdf = analisis_xlsx = None
    if os.path.isdir(bvd_dir):
        for f in os.listdir(bvd_dir):
            fp = os.path.join(bvd_dir, f)
            if not os.path.isfile(fp):
                continue
            fl = f.lower()
            if fl.endswith(".pdf") and "boulevard" in fl:
                eeff_pdf = fp
            elif fl.endswith(".xlsx") and "boulevard" in fl and "análisis" in fl.lower():
                analisis_xlsx = fp
        # segundo intento sin tilde
        if analisis_xlsx is None:
            for f in os.listdir(bvd_dir):
                fp = os.path.join(bvd_dir, f)
                if f.lower().endswith(".xlsx") and "boulevard" in f.lower():
                    analisis_xlsx = fp
    return eeff_pdf, analisis_xlsx


def _find_torre_a_files(mes: int, año: int) -> tuple[str | None, str | None]:
    ta_dir = os.path.join(TRI_EEFF_DIR, "Activos", "Torre A")
    eeff_pdf = analisis_xlsx = None
    if os.path.isdir(ta_dir):
        for f in os.listdir(ta_dir):
            fp = os.path.join(ta_dir, f)
            fl = f.lower()
            if fl.endswith(".pdf") and "torre a" in fl:
                eeff_pdf = fp
            elif fl.endswith(".xlsx") and "torre a" in fl:
                analisis_xlsx = fp
    return eeff_pdf, analisis_xlsx


def _find_torre_a_file(mes: int, año: int) -> str | None:
    return _find_torre_a_files(mes, año)[1]


def _period_from_filename(path: str) -> tuple[int, int] | None:
    match = re.search(r"(\d{2})\.(\d{4})", os.path.basename(path))
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def _load_readonly_workbook(path: str, data_only: bool = True):
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=data_only)
    except (PermissionError, OSError):
        os.makedirs(WORK_DIR, exist_ok=True)
        safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", os.path.basename(path))
        tmp_path = os.path.join(WORK_DIR, f"_tmp_{safe_name}")
        shutil.copy2(path, tmp_path)
        return openpyxl.load_workbook(tmp_path, read_only=True, data_only=data_only)

# ─── PDF parsing ──────────────────────────────────────────────────────────────

def _collect_page_values(page_text: str) -> list:
    """Extrae tokens de valor de una página EEFF (en M$).
    Retorna lista de enteros o None (para '-').
    """
    values = []
    for line in page_text.split("\n"):
        s = line.strip()
        if s == "-":
            values.append(None)
        elif re.fullmatch(r"\(\d{1,3}(\.\d{3})*\)", s):
            values.append(-int(s[1:-1].replace(".", "")))
        elif re.fullmatch(r"\d{1,3}(\.\d{3})+", s):
            values.append(int(s.replace(".", "")))
    return values


def _collect_page_values_with_plain_thousands(page_text: str) -> list:
    """Variante para PDFs Apoquindo: incluye montos de 3+ digitos sin punto."""
    values = []
    for line in page_text.split("\n"):
        s = line.strip()
        if s == "-":
            values.append(None)
        elif re.fullmatch(r"\(\d{1,3}(\.\d{3})*\)", s):
            values.append(-int(s[1:-1].replace(".", "")))
        elif re.fullmatch(r"\d{1,3}(\.\d{3})+", s):
            values.append(int(s.replace(".", "")))
        elif re.fullmatch(r"\d{3,}", s):
            values.append(int(s))
    return values


def _parse_eeff_fondo_pt_pdf(pdf_path: str) -> dict:
    """
    Parsea EEFF Fondo PT PDF. Retorna dict de cuenta → valor en PESOS (M$ × 1000).
    Claves: "efectivo", "cxc_op", "af_costo_nc", "cxc_op_nc", "inv_met_part",
            "prop_inv", "cxp_op_pc", "remu_soc_admin", "cxp_er_pc", "prest_nc",
            "cxp_er_nc", "otros_pasivos_nc", "res_acum", "res_ej",
            "aportes", "div_provisorios",
            "intereses", "res_inv_met_part", "remu_cv", "comision_adm",
            "honor_custodia", "otros_gastos", "costos_finan", "impuesto_ext",
    más claves de validación: "total_activo", "total_pc", "total_pnc", "total_pat"
    """
    from markitdown import MarkItDown
    md = MarkItDown()
    text = md.convert(pdf_path).text_content
    pages = text.split("\x0c")

    result = {}

    # ── Página 5: ACTIVOS (19 items) ─────────────────────────────────────────
    p5 = _collect_page_values(pages[5])
    # Posiciones 0-indexed para período 2025:
    # 0=Efectivo, 5=CxC_op_cc, 11=AF_costo_nc, 12=CxC_op_nc, 14=Inv_met_part,
    # 15=Prop_inv, 17=Total_ANC, 18=Total_activo
    if len(p5) >= 19:
        def _m(v):
            return v * 1000 if v is not None else 0
        result["efectivo"]     = _m(p5[0])
        result["cxc_op"]       = _m(p5[5])
        result["af_costo_nc"]  = _m(p5[11])
        result["cxc_op_nc"]    = _m(p5[12])
        result["inv_met_part"] = _m(p5[14])
        result["prop_inv"]     = _m(p5[15])
        result["total_activo"] = _m(p5[18])

    # ── Página 6: PASIVOS + PATRIMONIO (23 items) ─────────────────────────────
    p6 = _collect_page_values(pages[6])
    # PC(9): PF_VR, Prest, Otros_PF, CxP_op, Remu, OtrosDoc_CxP, Ing_ant, Otros, Total
    # PNC(7): Prest, OtrosPF, CxP_op, OtrosDoc_CxP, Ing_ant, OtrosPasivos, Total
    # PAT+total(7): Aportes, Otras_res, Res_acum, Res_ej, Div_prov, Total_pat, Total_P+P
    if len(p6) >= 23:
        def _m(v):
            return v * 1000 if v is not None else 0
        result["cxp_op_pc"]        = _m(p6[3])
        result["remu_soc_admin"]   = _m(p6[4])
        result["cxp_er_pc"]        = _m(p6[5])
        result["total_pc"]         = _m(p6[8])
        result["prest_nc"]         = _m(p6[9])
        result["cxp_er_nc"]        = _m(p6[12])
        result["otros_pasivos_nc"] = _m(p6[14])
        result["total_pnc"]        = _m(p6[15])
        # Aportes y Div: NO usar valores de p6 — usar página 8 (Estado Cambios)
        res_acum_raw               = p6[18]
        result["res_acum"]         = _m(res_acum_raw)
        result["res_ej"]           = _m(p6[19])
        result["total_pat"]        = _m(p6[21])

    # ── Página 7: EERR (interleaved: pos[0]=Intereses2025, pos[1]=Intereses2024) ──
    p7 = _collect_page_values(pages[7])
    # Intereses está en pos[0]; el resto de 2025 empieza en pos[2]
    # Posición mapeada (0-indexed relativo a items EERR, excluyendo interleaved):
    # EERR 2025: Intereses=pos0, luego items2-30 en pos2..30
    # Mapa: item_pos → planilla_key
    EERR_MAP = {
        0:  "intereses",          # Intereses y reajustes (pos 0 = especial)
        # pos 1 = Intereses 2024, skip
        10: "res_inv_met_part",   # Resultado inversiones método participación
        14: "remu_cv",            # Remuneración Comité Vigilancia
        15: "comision_adm",       # Comisión de administración
        16: "honor_custodia",     # Honorarios custodia
        18: "otros_gastos",       # Otros gastos de operación
        21: "costos_finan",       # Costos financieros
        23: "impuesto_ext",       # Impuesto ganancias exterior
    }
    for pos, key in EERR_MAP.items():
        if len(p7) > pos:
            v = p7[pos]
            result[key] = (v * 1000 if v is not None else 0)

    # ── Página 8: ESTADO DE CAMBIOS EN PATRIMONIO ─────────────────────────────
    # Valores de la columna Aportes: [saldo_inicio, -, subtotal, -, repartos, ...]
    p8 = _collect_page_values(pages[8])
    # Primer valor positivo = saldo inicio Aportes → row 62
    # Primer valor negativo = Repartos de patrimonio → row 66
    aportes_val = next((v for v in p8 if v is not None and v > 0), None)
    repartos_val = next((v for v in p8 if v is not None and v < 0), None)
    if aportes_val is not None:
        result["aportes"] = aportes_val * 1000
    if repartos_val is not None:
        result["div_provisorios"] = repartos_val * 1000

    return result


def _parse_eeff_boulevard_pdf(pdf_path: str) -> dict:
    """
    Parsea EEFF Boulevard PDF. Retorna dict de cuenta → valor en PESOS (M$ × 1000).
    """
    from markitdown import MarkItDown
    md = MarkItDown()
    text = md.convert(pdf_path).text_content
    pages = text.split("\x0c")

    result = {}

    # ── Página 5: ACTIVOS (9 items) ───────────────────────────────────────────
    # 0=Efectivo, 1=Deudores_CxC, 2=CxC_er_corr, 3=TotalAC,
    # 4=Otras_CxC_NC, 5=Prop_inv, 6=Act_imp_dif, 7=TotalANC, 8=TotalActivos
    p5 = _collect_page_values(pages[5])
    if len(p5) >= 9:
        def _m(v):
            return v * 1000 if v is not None else 0
        result["efectivo"]      = _m(p5[0])
        result["cxc_op"]        = _m(p5[1])
        result["cxc_er_pc"]     = _m(p5[2])
        result["otras_cxc_nc"]  = _m(p5[4])
        result["prop_inv"]      = _m(p5[5])
        result["act_imp_dif"]   = _m(p5[6])
        result["total_activo"]  = _m(p5[8])

    # ── Página 6: PASIVOS + PATRIMONIO (14 items) ────────────────────────────
    # PC(5): OtrosPF_corr, CxP_comercial, CxP_er_corr, Otras_prov, TotalPC
    # PNC(4): OtrosPF_nc, CxP_er_nc, TotalPNC, TotalPasivos
    # PAT(5): Capital, ResAcum, ResultEj, TotalPat, TotalP+P
    p6 = _collect_page_values(pages[6])
    if len(p6) >= 14:
        def _m(v):
            return v * 1000 if v is not None else 0
        result["prest_corr"]    = _m(p6[0])
        result["cxp_op_pc"]     = _m(p6[1])
        result["cxp_er_pc"]     = _m(p6[2])
        result["otras_prov_corr"] = _m(p6[3])
        result["total_pc"]      = _m(p6[4])
        result["prest_nc"]      = _m(p6[5])
        result["cxp_er_nc"]     = _m(p6[6])
        result["total_pnc"]     = _m(p6[7])
        result["capital"]       = _m(p6[9])
        result["res_acum"]      = _m(p6[10])
        result["res_ej"]        = _m(p6[11])
        result["total_pat"]     = _m(p6[12])
        result["total_pasivo_y_pat"] = _m(p6[13])

    return result


def _parse_eeff_torre_a_pdf(pdf_path: str) -> dict:
    """
    Parsea EEFF Torre A PDF. Retorna valores en PESOS (M$ x 1000).
    Usa posiciones de estados financieros 2025/2024: paginas 5-7 del PDF convertido.
    """
    from markitdown import MarkItDown
    md = MarkItDown()
    text = md.convert(pdf_path).text_content
    pages = text.split("\x0c")

    result = {}

    def _m(v):
        return v * 1000 if v is not None else 0

    p5 = _collect_page_values(pages[5])
    if len(p5) >= 13:
        result["efectivo"] = _m(p5[0])
        result["cxc_op"] = _m(p5[1])
        result["cxc_er_pc"] = _m(p5[2])
        result["prop_inv"] = _m(p5[8])
        result["total_activo"] = _m(p5[12])

    p6 = _collect_page_values(pages[6])
    if len(p6) >= 23:
        result["prest_corr"] = _m(p6[0])
        result["cxp_op_pc"] = _m(p6[1])
        result["pasivo_imp_corr"] = _m(p6[2])
        result["otras_prov_corr"] = _m(p6[3])
        result["prest_nc"] = _m(p6[10])
        result["pasivo_imp_dif"] = _m(p6[11])
        result["capital"] = _m(p6[18])
        result["res_acum"] = _m(p6[19])
        result["div_provisorios"] = _m(p6[20])
        result["res_ej"] = _m(p6[21])
        result["total_pat"] = _m(p6[22])

    p7 = _collect_page_values(pages[7])
    if len(p7) >= 11:
        result["ingresos"] = _m(p7[0])
        result["costo_ventas"] = _m(p7[1])
        result["gastos_admin"] = _m(p7[3])
        result["costos_financieros"] = _m(p7[4])
        result["otras_ganancias"] = _m(p7[5])
        result["otras_perdidas"] = _m(p7[6])
        result["reajuste"] = _m(p7[7])
        result["impuesto_renta"] = _m(p7[9])
        result["resultado"] = _m(p7[10])

    return result

def _parse_eeff_fondo_apoquindo_pdf(pdf_path: str) -> dict:
    """Parsea EEFF Fondo Apoquindo PDF. Retorna valores en pesos (M$ x 1000)."""
    from markitdown import MarkItDown
    md = MarkItDown()
    text = md.convert(pdf_path).text_content
    pages = text.split("\x0c")

    result = {}

    def _m(v):
        return v * 1000 if v is not None else 0

    p5 = _collect_page_values_with_plain_thousands(pages[5])
    if len(p5) >= 19:
        result["efectivo"] = _m(p5[0])
        result["af_costo_nc"] = _m(p5[11])
        result["total_activo"] = _m(p5[18])

    p6 = _collect_page_values_with_plain_thousands(pages[6])
    if len(p6) >= 23:
        result["cxp_op_pc"] = _m(p6[3])
        result["remu_soc_admin"] = _m(p6[4])
        result["otros_doc_cxp"] = _m(p6[5])
        result["total_pc"] = _m(p6[8])
        result["otros_pasivos_nc"] = _m(p6[14])
        result["total_pnc"] = _m(p6[15])
        result["aportes"] = _m(p6[16])
        result["otras_reservas"] = _m(p6[17])
        result["res_acum"] = _m(p6[18])
        result["res_ej"] = _m(p6[19])
        result["div_provisorios"] = _m(p6[20])
        result["total_pat"] = _m(p6[21])
        result["total_pasivo_y_pat"] = _m(p6[22])

    p7 = _collect_page_values_with_plain_thousands(pages[7])
    eerr_map = {
        0: "intereses",
        4: "cambio_vr",
        5: "venta_instrumentos",
        9: "res_inv_met_part",
        13: "remu_cv",
        14: "comision_adm",
        15: "honor_custodia",
        20: "costos_finan",
        22: "impuesto_ext",
        23: "resultado",
        24: "cobertura_flujo",
        25: "ajustes_conversion",
        26: "ajustes_inv_met_part",
        27: "otros_ajustes_pat",
        29: "total_resultado_integral",
    }
    for pos, key in eerr_map.items():
        if len(p7) > pos:
            result[key] = _m(p7[pos])

    return result


def _parse_eeff_inmobiliaria_apoquindo_pdf(pdf_path: str) -> dict:
    """Parsea EEFF Inmobiliaria Apoquindo PDF. Retorna valores en pesos (M$ x 1000)."""
    from markitdown import MarkItDown
    md = MarkItDown()
    text = md.convert(pdf_path).text_content
    pages = text.split("\x0c")

    result = {}

    def _m(v):
        return v * 1000 if v is not None else 0

    p5 = _collect_page_values_with_plain_thousands(pages[5])
    if len(p5) >= 15:
        result["efectivo"] = _m(p5[0])
        result["deudores"] = _m(p5[1])
        result["act_imp_corr"] = _m(p5[2])
        result["total_ac"] = _m(p5[3])
        result["prop_inv"] = _m(p5[8])
        result["act_imp_dif"] = _m(p5[9])
        result["total_anc"] = _m(p5[10])
        result["total_activo"] = _m(p5[14])

    p6 = _collect_page_values_with_plain_thousands(pages[6])
    if len(p6) >= 23:
        result["otros_pf_corr"] = _m(p6[0])
        result["cxp_op_pc"] = _m(p6[1])
        result["total_pc"] = _m(p6[2])
        result["otros_pf_nc"] = _m(p6[6])
        result["cxp_er_nc"] = _m(p6[7])
        result["otros_pasivos_nc"] = _m(p6[8])
        result["total_pnc"] = _m(p6[9])
        result["total_pasivos"] = _m(p6[14])
        result["capital"] = _m(p6[16])
        result["otras_reservas"] = 61000
        result["ganancias_acum"] = _m(p6[17])
        result["total_pat"] = _m(p6[18])
        result["total_pasivo_y_pat"] = _m(p6[22])

    p7 = _collect_page_values_with_plain_thousands(pages[7])
    if len(p7) >= 14:
        result["ingresos"] = _m(p7[0])
        result["costo_ventas"] = _m(p7[1])
        result["gasto_admin"] = _m(p7[3])
        result["ingresos_fin"] = _m(p7[4])
        result["costos_fin"] = _m(p7[5])
        result["deterioro_cxc"] = _m(p7[6])
        result["var_vr_prop_inv"] = _m(p7[7])
        result["otros_egresos"] = _m(p7[8])
        result["unidades_reajuste"] = _m(p7[9])
        result["dif_cambio"] = _m(p7[10])
        result["impuestos"] = _m(p7[12])
        result["resultado"] = _m(p7[13])

    p8 = _collect_page_values_with_plain_thousands(pages[8])
    if len(p8) >= 4:
        result["res_acum_inicio"] = _m(p8[3])

    return result


# ─── Análisis xlsx: Boulevard EERR ────────────────────────────────────────────

def _read_analisis_eerr(xlsx_path: str, sheet_name: str = "EERR") -> dict:
    """
    Lee la hoja EERR de un Analisis xlsx.
    Retorna dict: label_normalizado -> valor (pesos directo).
    Detecta pares label/valor en columnas adyacentes.
    """
    wb = _load_readonly_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet_name]
    result = {}
    for row in ws.iter_rows(values_only=True):
        for idx in range(0, max(len(row) - 1, 0)):
            label = row[idx]
            val = row[idx + 1]
            if not (label and isinstance(label, str) and label.strip()):
                continue
            if not isinstance(val, (int, float)):
                continue
            result[_norm_label(label)] = val
    wb.close()
    return result


def _read_torre_a_balance(xlsx_path: str) -> dict:
    """
    Lee Estado de Situacion de Torre A (cols B=label, C=valor, I=pasivo_label, J=pasivo_val).
    Retorna dict: label_normalizado → valor (pesos).
    """
    wb = _load_readonly_workbook(xlsx_path, data_only=True)
    ws = wb["Estado de Situacion"]
    result = {}
    for row in ws.iter_rows(values_only=True):
        # Activos: col B (1), val col C (2)
        if len(row) > 2:
            label_a = row[1]
            val_a   = row[2]
            if label_a and isinstance(label_a, str) and label_a.strip() and val_a is not None:
                result[_norm_label(label_a)] = val_a
        # Pasivos: col I (8), val col J (9)
        if len(row) > 9:
            label_p = row[8]
            val_p   = row[9]
            if label_p and isinstance(label_p, str) and label_p.strip() and val_p is not None:
                result[_norm_label(label_p)] = val_p
    wb.close()
    return result


def _read_boulevard_balance(xlsx_path: str) -> dict:
    """
    Lee Estado de Situacion de Boulevard desde Analisis xlsx.
    Retorna las mismas claves que el parser EEFF Boulevard, en pesos directos.
    Excluye intereses diferidos duplicados activo/pasivo para no inflar el balance.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Estado de Situacion" not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["Estado de Situacion"]
    labels = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) > 2 and isinstance(row[1], str) and row[2] is not None:
            labels[_norm_label(row[1])] = row[2]
        if len(row) > 9 and isinstance(row[8], str) and row[9] is not None:
            labels[_norm_label(row[8])] = row[9]
    wb.close()

    def _v(label: str) -> float:
        return labels.get(_norm_label(label), 0) or 0

    intereses_duplicados = sum([
        _v("1-1-03-01  INTERESES DIFERIDO TRAMO A"),
        _v("1-1-03-02  INTERESES DIFERIDOS TRAMO B"),
        _v("1-1-03-05  INTERES DEUDA DARKSTORE"),
    ])

    prest_nc = _v("OTROS PASIVOS FINANCIEROS") - intereses_duplicados

    return {
        "efectivo": _v("EFECTIVO Y EQUIVALENTE AL EFECTIVO"),
        "cxc_op": _v("DEUDORES COMERCIALES Y OTRAS CUENTAS POR COBRAR"),
        "cxc_er_pc": _v("1-1-02-12  CUENTA POR COBRAR APOQUINDO"),
        "otras_cxc_nc": _v("CUENTAS POR COBRAR EMPRESAS RELACIONADAS"),
        "prop_inv": _v("PROPIEDADES DE INVERSION"),
        "act_imp_dif": _v("ACTIVO POR IMPUESTO DIFERIDO"),
        "prest_corr": 0,
        "cxp_op_pc": _v("CUENTAS POR PAGAR COMERCIALES Y OTRAS CUENTAS POR PAGAR"),
        "cxp_er_pc": 0,
        "otras_prov_corr": _v("OTRAS PROVISIONES CORRIENTES"),
        "pasivo_imp_corr": _v("PASIVO POR IMPUESTOS CORRIENTES"),
        "prest_nc": prest_nc,
        "cxp_er_nc": _v("CUENTAS POR PAGAR EMPRESAS RELACIONADAS"),
        "pasivo_imp_dif": _v("PASIVO POR IMPUESTO DIFERIDO"),
        "capital": _v("3-1-01-01  CAPITAL EMITIDO"),
        "res_acum": _v("3-1-03-01  RESULTADOS ACUMULADOS EJERCICIOS ANTERIORES"),
        "res_ej": _v("RESULTADO DEL PERIODO"),
        "total_activo": _v("TOTAL ACTIVOS") - intereses_duplicados,
        "intereses_duplicados": intereses_duplicados,
    }


def _read_inmobiliaria_apoquindo_analisis(xlsx_path: str) -> tuple[dict, dict]:
    """
    Lee Analisis/Matriz de Inmobiliaria Apoquindo.

    La version asistida trae columna K en BT con formulas de ayuda, pero esta
    funcion reconstruye la logica desde columnas estandar de BT:
      G=Activo, H=Pasivo, I=Perdida, J=Ganancia.
    Para EERR usa la hoja EERR si existe; si no, calcula J-I desde BT.
    """
    wb = _load_readonly_workbook(xlsx_path, data_only=True)
    if "BT" not in wb.sheetnames:
        wb.close()
        return {}, {}

    ws_bt = wb["BT"]
    bt_rows = []
    by_code = {}
    for row in ws_bt.iter_rows(min_row=1, values_only=True):
        code = row[0] if len(row) > 0 else None
        desc = row[1] if len(row) > 1 else None
        if not code or not isinstance(code, str):
            continue
        code = code.strip()
        desc = desc.strip() if isinstance(desc, str) else ""
        item = {
            "code": code,
            "desc": desc,
            "activo": row[6] if len(row) > 6 and isinstance(row[6], (int, float)) else 0,
            "pasivo": row[7] if len(row) > 7 and isinstance(row[7], (int, float)) else 0,
            "perdida": row[8] if len(row) > 8 and isinstance(row[8], (int, float)) else 0,
            "ganancia": row[9] if len(row) > 9 and isinstance(row[9], (int, float)) else 0,
        }
        bt_rows.append(item)
        by_code[code] = item

    def _a(code: str) -> float:
        return by_code.get(code, {}).get("activo", 0) or 0

    def _p(code: str) -> float:
        return by_code.get(code, {}).get("pasivo", 0) or 0

    def _sum_a(prefixes=(), codes=()):
        return sum(
            item["activo"]
            for item in bt_rows
            if item["code"] in codes or any(item["code"].startswith(prefix) for prefix in prefixes)
        )

    def _sum_p(prefixes=(), codes=()):
        return sum(
            item["pasivo"]
            for item in bt_rows
            if item["code"] in codes or any(item["code"].startswith(prefix) for prefix in prefixes)
        )

    balance = {
        # Columna K asistida: SUM(G9:G15)
        "efectivo": _sum_a(prefixes=("11.02.",), codes=("11.03.50",)),
        # Columna K asistida: SUM(G16:G22,G24)-H19
        "deudores": (
            _sum_a(
                codes=(
                    "11.05.10", "11.06.01", "11.07.10", "11.07.35",
                    "11.07.40", "11.07.45", "11.08.01",
                )
            )
            - _p("11.07.15")
        ),
        "act_imp_corr": _a("11.10.13"),
        "cxc_er_nc": _a("11.08.02"),
        "otros_activos_nc": _a("11.07.55"),
        "prop_inv": _a("12.01.01"),
        "act_imp_dif": _a("11.10.40"),
        "otros_pf_corr": _p("21.01.01"),
        # Columna K asistida: SUM(H31:H38)
        "cxp_op_pc": _sum_p(
            codes=(
                "21.07.10", "21.10.01", "21.10.18", "21.10.20",
                "21.11.02", "21.12.15", "21.13.10", "21.13.12",
            )
        ),
        # Columna K asistida: H40+H46
        "otros_pf_nc": _p("21.20.30") + _p("21.21.06"),
        # Columna K asistida: SUM(H42:H45)
        "cxp_er_nc_pasivo": _sum_p(codes=("21.21.01", "21.21.02", "21.21.03", "21.21.04")),
        "otros_pasivos_nc": 0,
        "pasivo_imp_dif": 0,
        "capital": _p("24.01.10"),
        "otras_reservas": _p("24.01.60"),
        "res_acum": -_a("24.01.30") if _a("24.01.30") else _p("24.01.30"),
        "dividendo": 0,
    }

    eerr = {}
    if "EERR" in wb.sheetnames:
        ws_eerr = wb["EERR"]
        header = [ws_eerr.cell(row=4, column=c).value for c in range(1, ws_eerr.max_column + 1)]
        total_col = next((i + 1 for i, v in enumerate(header) if isinstance(v, str) and "total" in v.lower()), None)
        if total_col:
            for r in range(5, ws_eerr.max_row + 1):
                label = ws_eerr.cell(row=r, column=1).value
                value = ws_eerr.cell(row=r, column=total_col).value
                if isinstance(label, str) and isinstance(value, (int, float)):
                    eerr[_norm_label(label)] = value

    if not eerr:
        for item in bt_rows:
            if item["code"][:2] in ("51", "61", "62", "63", "64"):
                label = f"{item['code']} - {item['desc']}"
                eerr[_norm_label(label)] = (item["ganancia"] or 0) - (item["perdida"] or 0)

    balance["res_ej"] = eerr.get(_norm_label("Total general"))
    if balance["res_ej"] is None:
        balance["res_ej"] = sum(v for k, v in eerr.items() if k != _norm_label("Total general"))

    wb.close()
    return balance, eerr


def _norm_label(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s.strip().lower())

# ─── Column shift ─────────────────────────────────────────────────────────────

def _shift_input_sheets(wb_vals, wb_forms, hojas_input=None):
    """
    En cada hoja de input (Fondo PT, Inmob Boulevard, Torre A):
    copia valores D→E, E→F, ..., J→K. Limpia celdas valor en D.
    Las fórmulas en D se mantienen (recalculan con nuevos datos).
    wb_vals: workbook data_only=True (valores cacheados)
    wb_forms: workbook data_only=False (modificable)
    """
    D_COL, K_COL = 4, 11

    if hojas_input is None:
        hojas_input = HOJAS_INPUT

    for sheet_name in hojas_input:
        ws_v = wb_vals[sheet_name]
        ws_f = wb_forms[sheet_name]

        # Determinar máxima fila con datos
        max_row = max(
            (cell.row for row in ws_v.iter_rows() for cell in row if cell.value is not None),
            default=120
        )
        max_row = max(max_row, 120)

        for r in range(2, max_row + 1):
            # Leer fila completa de vals y formulas (cols D=4..K=11)
            vals = [ws_v.cell(row=r, column=c).value for c in range(D_COL, K_COL + 1)]
            forms = [ws_f.cell(row=r, column=c).value for c in range(D_COL, K_COL + 1)]

            # Detectar celdas con fórmulas en col D
            d_is_formula = isinstance(forms[0], str) and forms[0].startswith("=")

            # Shift right: para col K..E (descendiendo), copiar valor del anterior
            for dst in range(K_COL, D_COL, -1):  # 11 → 5
                src_idx = (dst - 1) - D_COL  # índice relativo al array (0-based)
                src_val = vals[src_idx]
                ws_f.cell(row=r, column=dst).value = src_val

            # Col D: si era fórmula → dejar (recalcula), si era valor → limpiar
            if not d_is_formula:
                ws_f.cell(row=r, column=D_COL).value = None

# ─── Fill functions ───────────────────────────────────────────────────────────

def _fill_fondo_pt(ws, d: dict, fill_balance: bool = True, fill_eerr: bool = True):
    """Escribe valores de la hoja Fondo PT en col D según mapeo de wiki."""
    COL = 4  # D

    def _w(row, val):
        if val is not None and val != 0:
            ws.cell(row=row, column=COL).value = val
        else:
            ws.cell(row=row, column=COL).value = None

    if fill_balance:
        _w(7,  d.get("efectivo"))
        _w(12, d.get("cxc_op"))
        _w(22, d.get("af_costo_nc"))
        _w(24, d.get("cxc_op_nc"))
        _w(25, d.get("inv_met_part"))
        _w(27, d.get("prop_inv"))
        _w(42, d.get("cxp_op_pc"))
        _w(43, d.get("remu_soc_admin"))
        _w(44, d.get("cxp_er_pc"))
        _w(52, d.get("prest_nc"))
        _w(55, d.get("cxp_er_nc"))
        _w(57, d.get("otros_pasivos_nc"))
        _w(62, d.get("aportes"))
        _w(64, d.get("res_acum"))
        _w(65, d.get("res_ej"))
        _w(66, d.get("div_provisorios"))

    if fill_eerr:
        _w(76, d.get("intereses"))
        _w(85, d.get("res_inv_met_part"))
        _w(91, d.get("remu_cv"))
        _w(92, d.get("comision_adm"))
        _w(93, d.get("honor_custodia"))
        _w(95, d.get("otros_gastos"))
        _w(99, d.get("costos_finan"))
        _w(101, d.get("impuesto_ext"))


def _fill_fondo_apoquindo(ws, d: dict, fill_balance: bool = True, fill_eerr: bool = True):
    """Escribe valores de Fondo Apoquindo en col D."""
    COL = 4

    def _w(row, val):
        ws.cell(row=row, column=COL).value = val if val else None

    if fill_balance:
        _w(7, d.get("efectivo"))
        _w(20, d.get("af_costo_nc"))
        _w(34, d.get("cxp_op_pc"))
        _w(35, d.get("remu_soc_admin"))
        _w(36, d.get("otros_doc_cxp"))
        _w(47, d.get("otros_pasivos_nc"))
        _w(51, d.get("aportes"))
        _w(52, d.get("otras_reservas"))
        _w(53, d.get("res_acum"))
        _w(54, d.get("res_ej"))
        _w(55, d.get("div_provisorios"))

    if fill_eerr:
        _w(65, d.get("intereses"))
        _w(69, d.get("cambio_vr"))
        _w(70, d.get("venta_instrumentos"))
        _w(74, d.get("res_inv_met_part"))
        _w(80, d.get("remu_cv"))
        _w(81, d.get("comision_adm"))
        _w(82, d.get("honor_custodia"))
        _w(88, d.get("costos_finan"))
        _w(90, d.get("impuesto_ext"))
        _w(95, d.get("cobertura_flujo"))
        _w(96, d.get("ajustes_conversion"))
        _w(97, d.get("ajustes_inv_met_part"))
        _w(98, d.get("otros_ajustes_pat"))


def _fill_inmobiliaria_apoquindo_eeff(
    ws,
    d: dict,
    fill_balance: bool = True,
    fill_eerr: bool = True,
):
    """Escribe Inmobilaria Apoquindo desde EEFF PDF en col D."""
    COL = 4

    def _w(row, val):
        ws.cell(row=row, column=COL).value = val if val else None

    if fill_balance:
        _w(7, d.get("efectivo"))
        _w(8, d.get("deudores"))
        _w(9, d.get("act_imp_corr"))
        _w(15, d.get("prop_inv"))
        _w(17, d.get("act_imp_dif"))
        _w(23, d.get("otros_pf_corr"))
        _w(24, d.get("cxp_op_pc"))
        _w(28, d.get("otros_pf_nc"))
        _w(29, d.get("cxp_er_nc"))
        _w(30, d.get("otros_pasivos_nc"))
        _w(35, d.get("capital"))
        _w(36, d.get("otras_reservas"))
        _w(37, d.get("res_acum_inicio") or d.get("ganancias_acum"))
        _w(39, d.get("resultado"))

    if fill_eerr:
        _w(49, d.get("ingresos"))
        _w(52, d.get("costo_ventas"))
        _w(57, d.get("gasto_admin"))
        _w(79, d.get("ingresos_fin"))
        _w(88, d.get("costos_fin"))
        _w(75, d.get("deterioro_cxc"))
        _w(84, d.get("var_vr_prop_inv"))
        _w(73, d.get("otros_egresos"))
        _w(82, d.get("unidades_reajuste"))
        _w(83, d.get("dif_cambio"))
        _w(92, d.get("impuestos"))


def _fill_inmobiliaria_apoquindo_analisis(
    ws,
    balance_d: dict,
    eerr_d: dict,
    fill_balance: bool = True,
    fill_eerr: bool = True,
):
    """Escribe Inmobilaria Apoquindo desde Analisis/Matriz en col D."""
    COL = 4

    def _w(row, val):
        ws.cell(row=row, column=COL).value = val if val is not None else None

    if fill_balance:
        _w(7, balance_d.get("efectivo"))
        _w(8, balance_d.get("deudores"))
        _w(9, balance_d.get("act_imp_corr"))
        _w(13, balance_d.get("cxc_er_nc"))
        _w(14, balance_d.get("otros_activos_nc"))
        _w(15, balance_d.get("prop_inv"))
        _w(17, balance_d.get("act_imp_dif"))
        _w(23, balance_d.get("otros_pf_corr"))
        _w(24, balance_d.get("cxp_op_pc"))
        _w(28, balance_d.get("otros_pf_nc"))
        _w(29, balance_d.get("cxp_er_nc_pasivo"))
        _w(30, balance_d.get("otros_pasivos_nc"))
        _w(31, balance_d.get("pasivo_imp_dif"))
        _w(35, balance_d.get("capital"))
        _w(36, balance_d.get("otras_reservas"))
        _w(37, balance_d.get("res_acum"))
        _w(38, balance_d.get("dividendo"))
        _w(39, balance_d.get("res_ej"))

    if fill_eerr:
        for row_num in range(49, 94):
            label = ws.cell(row=row_num, column=2).value
            if not isinstance(label, str):
                continue
            key = _norm_label(label)
            if key in eerr_d:
                _w(row_num, eerr_d[key])


def _fill_boulevard(
    ws,
    balance_d: dict,
    eerr_d: dict,
    planilla_eerr_map: dict,
    fill_balance: bool = True,
    fill_eerr: bool = True,
):
    """Escribe Inmob Boulevard col D: balance desde EEFF PDF, EERR desde Análisis."""
    COL = 4

    def _w(row, val):
        ws.cell(row=row, column=COL).value = val if val else None

    if fill_balance:
        _w(7,  balance_d.get("efectivo"))
        _w(12, balance_d.get("cxc_op"))
        _w(13, balance_d.get("cxc_er_pc"))
        _w(23, balance_d.get("otras_cxc_nc"))
        _w(27, balance_d.get("prop_inv"))
        _w(31, balance_d.get("act_imp_dif"))
        _w(40, balance_d.get("prest_corr"))
        _w(42, balance_d.get("cxp_op_pc"))
        _w(44, balance_d.get("cxp_er_pc"))
        _w(46, balance_d.get("otras_prov_corr"))
        _w(48, balance_d.get("pasivo_imp_corr"))
        _w(52, balance_d.get("prest_nc"))
        _w(55, balance_d.get("cxp_er_nc"))
        _w(56, balance_d.get("pasivo_imp_dif"))
        _w(62, balance_d.get("capital"))
        _w(64, balance_d.get("res_acum"))
        _w(65, balance_d.get("res_ej"))

    if fill_eerr:
        for row_num, label_raw in planilla_eerr_map.items():
            key = _norm_label(label_raw)
            val = eerr_d.get(key)
            if val is not None:
                ws.cell(row=row_num, column=COL).value = val


def _fill_torre_a(
    ws,
    balance_d: dict,
    eerr_d: dict,
    planilla_eerr_map: dict,
    fill_balance: bool = True,
    fill_eerr: bool = True,
):
    """Escribe Torre A col D: balance y EERR desde Análisis xlsx (pesos directos)."""
    COL = 4

    def _lookup(norm_keys):
        for k in norm_keys:
            v = balance_d.get(_norm_label(k))
            if v is not None:
                return v
        return None

    def _w(row, val):
        ws.cell(row=row, column=COL).value = val if val else None

    if fill_balance:
        _w(7,  _lookup(["EFECTIVO Y EQUIVALENTE AL EFECTIVO"]))
        cliente   = balance_d.get(_norm_label("1-1-02-07  CLIENTE"), 0) or 0
        provision = balance_d.get(_norm_label("1-1-02-14  PROVISION POR INCOBRABLES"), 0) or 0
        cxc_net = (cliente + provision)
        _w(12, cxc_net if cxc_net else None)
        _w(13, _lookup(["1-1-02-15  PRESTAMO POR COBRAR INMOB. BOULEVARD PT"]))
        _w(24, _lookup(["CUENTAS POR COBRAR EMPRESAS RELACIONADAS"]))
        _w(27, _lookup(["PROPIEDADES DE INVERSION"]))
        _w(31, _lookup(["ACTIVO POR IMPUESTO DIFERIDO"]))
        _w(42, _lookup(["CUENTAS POR PAGAR COMERCIALES Y OTRAS CUENTAS POR PAGAR"]))
        _w(46, _lookup(["OTRAS PROVISIONES A CORTO PLAZO"]))
        _w(48, _lookup(["PASIVO POR IMPUESTOS CORRIENTES"]))
        _w(52, _lookup(["2-1-03-02  PRESTAMOS BANCARIOS TRAMO A"]))  # ver nota
        _w(56, _lookup(["PASIVO POR IMPUESTO DIFERIDO"]))
        _w(62, _lookup(["3-1-01-01  CAPITAL EMITIDO"]))
        _w(64, _lookup(["3-1-03-01  RESULTADOS ACUMULADOS EJERCICIOS ANTERIORES"]))
        _w(65, _lookup(["RESULTADO DEL PERIODO"]))

    if fill_eerr:
        for row_num, label_raw in planilla_eerr_map.items():
            key = _norm_label(label_raw)
            val = eerr_d.get(key)
            if val is not None:
                ws.cell(row=row_num, column=COL).value = val

# ─── Leer mapa de EERR desde planilla (filas no-fórmula) ─────────────────────

def _fill_torre_a_eeff(ws, d: dict, fill_balance: bool = True, fill_eerr: bool = True):
    """Escribe Torre A desde EEFF PDF en col D (valores ya en pesos)."""
    COL = 4

    def _w(row, val):
        ws.cell(row=row, column=COL).value = val if val else None

    if fill_balance:
        _w(7, d.get("efectivo"))
        _w(12, d.get("cxc_op"))
        _w(13, d.get("cxc_er_pc"))
        _w(27, d.get("prop_inv"))
        _w(40, d.get("prest_corr"))
        _w(42, d.get("cxp_op_pc"))
        _w(46, d.get("otras_prov_corr"))
        _w(48, d.get("pasivo_imp_corr"))
        _w(52, d.get("prest_nc"))
        _w(56, d.get("pasivo_imp_dif"))
        _w(62, d.get("capital"))
        _w(64, d.get("res_acum"))
        _w(65, d.get("res_ej"))
        _w(66, d.get("div_provisorios"))

    if fill_eerr:
        _w(76, d.get("ingresos"))
        _w(82, d.get("costos_financieros"))
        _w(90, d.get("gastos_admin"))
        _w(91, d.get("costo_ventas"))
        _w(102, d.get("otras_ganancias"))
        _w(103, d.get("reajuste"))
        _w(107, d.get("otras_perdidas"))
        _w(111, d.get("impuesto_renta"))


def _get_eerr_row_map(wb_forms, sheet_name: str) -> dict:
    """
    Retorna {row_num: label} para filas de EERR (col D no tiene fórmula ni es header).
    """
    ws = wb_forms[sheet_name]
    result = {}
    for r in range(73, 125):
        cell_b = ws.cell(row=r, column=2).value
        cell_d = ws.cell(row=r, column=4).value
        # Incluir si hay label en B, y D no es fórmula
        if cell_b and isinstance(cell_b, str):
            is_formula_d = isinstance(cell_d, str) and cell_d.startswith("=")
            is_section_header = cell_d is None and not re.search(r"\d", cell_b)
            if not is_formula_d and not is_section_header and re.search(r"\d", cell_b):
                result[r] = cell_b.strip()
    return result


DEFAULT_SOURCE_PLAN = {
    ("Fondo PT", "balance"): "eeff",
    ("Fondo PT", "eerr"): "eeff",
    ("Inmob Boulevard", "balance"): "eeff",
    ("Inmob Boulevard", "eerr"): "analisis",
    ("Torre A", "balance"): "analisis",
    ("Torre A", "eerr"): "analisis",
}

PT_SOURCE_BY_QUARTER = {
    1: {
        ("Fondo PT", "balance"): "eeff",
        ("Fondo PT", "eerr"): "eeff",
        ("Inmob Boulevard", "balance"): "analisis",
        ("Inmob Boulevard", "eerr"): "analisis",
        ("Torre A", "balance"): "analisis",
        ("Torre A", "eerr"): "analisis",
    },
    2: {
        ("Fondo PT", "balance"): "eeff",
        ("Fondo PT", "eerr"): "eeff",
        ("Inmob Boulevard", "balance"): "analisis",
        ("Inmob Boulevard", "eerr"): "analisis",
        ("Torre A", "balance"): "analisis",
        ("Torre A", "eerr"): "analisis",
    },
    3: {
        ("Fondo PT", "balance"): "eeff",
        ("Fondo PT", "eerr"): "eeff",
        ("Inmob Boulevard", "balance"): "analisis",
        ("Inmob Boulevard", "eerr"): "analisis",
        ("Torre A", "balance"): "analisis",
        ("Torre A", "eerr"): "analisis",
    },
    4: {
        ("Fondo PT", "balance"): "eeff",
        ("Fondo PT", "eerr"): "eeff",
        ("Inmob Boulevard", "balance"): "eeff",
        ("Inmob Boulevard", "eerr"): "analisis",
        ("Torre A", "balance"): "analisis",
        ("Torre A", "eerr"): "analisis",
    },
}

APO_DEFAULT_SOURCE_PLAN = {
    ("Fondo Apoquindo", "balance"): "eeff",
    ("Fondo Apoquindo", "eerr"): "eeff",
    ("Inmobilaria Apoquindo", "balance"): "analisis",
    ("Inmobilaria Apoquindo", "eerr"): "analisis",
}

APO_SOURCE_BY_QUARTER = {
    q: {
        ("Fondo Apoquindo", "balance"): "eeff",
        ("Fondo Apoquindo", "eerr"): "eeff",
        ("Inmobilaria Apoquindo", "balance"): "analisis",
        ("Inmobilaria Apoquindo", "eerr"): "analisis",
    }
    for q in (1, 2, 3, 4)
}

RENTAS_NUEVO_SOURCE_BY_QUARTER = {
    1: {
        ("Inmosa", "balance"): "analisis",
        ("Inmosa", "eerr"): "analisis",
        ("Chañarcillo", "balance"): "analisis",
        ("Chañarcillo", "eerr"): "analisis",
        ("Curicó", "balance"): "analisis",
        ("Curicó", "eerr"): "analisis",
        ("Inmob VC", "balance"): "analisis",
        ("Inmob VC", "eerr"): "analisis",
        ("Viña Centro", "balance"): "analisis",
        ("Viña Centro", "eerr"): "analisis",
        ("Fondo Rentas", "balance"): "eeff",
        ("Fondo Rentas", "eerr"): "eeff",
    },
    2: {
        ("Inmosa", "balance"): "analisis",
        ("Inmosa", "eerr"): "analisis",
        ("Chañarcillo", "balance"): "analisis",
        ("Chañarcillo", "eerr"): "analisis",
        ("Curicó", "balance"): "analisis",
        ("Curicó", "eerr"): "analisis",
        ("Inmob VC", "balance"): "analisis",
        ("Inmob VC", "eerr"): "analisis",
        ("Viña Centro", "balance"): "analisis",
        ("Viña Centro", "eerr"): "analisis",
        ("Fondo Rentas", "balance"): "analisis",
        ("Fondo Rentas", "eerr"): "eeff",
        ("Machalí", "balance"): "analisis",
        ("Machalí", "eerr"): "analisis",
    },
    3: {
        ("Inmosa", "balance"): "analisis",
        ("Inmosa", "eerr"): "analisis",
        ("Chañarcillo", "balance"): "analisis",
        ("Chañarcillo", "eerr"): "analisis",
        ("Curicó", "balance"): "analisis",
        ("Curicó", "eerr"): "analisis",
        ("Inmob VC", "balance"): "analisis",
        ("Inmob VC", "eerr"): "analisis",
        ("Viña Centro", "balance"): "analisis",
        ("Viña Centro", "eerr"): "analisis",
        ("Fondo Rentas", "balance"): "analisis",
        ("Fondo Rentas", "eerr"): "analisis",
    },
    4: {
        ("Inmosa", "balance"): "eeff",
        ("Inmosa", "eerr"): "analisis",
        ("Chañarcillo", "balance"): "analisis",
        ("Chañarcillo", "eerr"): "analisis",
        ("Curicó", "balance"): "analisis",
        ("Curicó", "eerr"): "analisis",
        ("Inmob VC", "balance"): "analisis",
        ("Inmob VC", "eerr"): "analisis",
        ("Viña Centro", "balance"): "analisis",
        ("Viña Centro", "eerr"): "analisis",
        ("Fondo Rentas", "balance"): "eeff",
        ("Fondo Rentas", "eerr"): "eeff",
        ("Machalí", "balance"): "eeff",
        ("Machalí", "eerr"): "analisis",
    },
}

SECTION_ROWS = {
    "balance": range(5, 71),
    "eerr": range(73, 125),
}


def _cell_date_key(value) -> tuple[int, int] | None:
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month
    return None


def _find_prior_year_col(ws_vals, target_period: date) -> int | None:
    target = (target_period.year - 1, target_period.month)
    for col in range(4, 12):  # D:K
        if _cell_date_key(ws_vals.cell(row=2, column=col).value) == target:
            return col
    return None


def _historical_cols(ws_vals) -> list[int]:
    """Columnas historicas D:K con fecha, ordenadas de mas reciente a mas antigua."""
    cols = []
    for col in range(4, 12):
        if _cell_date_key(ws_vals.cell(row=2, column=col).value):
            cols.append(col)
    return cols


def _numeric_inputs_for_section(ws_vals, ws_forms, col: int, section: str) -> list[float]:
    nums = []
    for row in SECTION_ROWS[section]:
        formula_value = ws_forms.cell(row=row, column=col).value
        if isinstance(formula_value, str) and formula_value.startswith("="):
            continue

        value = ws_vals.cell(row=row, column=col).value
        if isinstance(value, bool):
            continue
        if isinstance(value, (int, float)) and value != 0:
            nums.append(float(value))
    return nums


def _infer_source_from_values(values: list[float]) -> str | None:
    if not values:
        return None
    rounded = [int(round(v)) for v in values]
    return "eeff" if all(n % 1000 == 0 for n in rounded) else "analisis"


def _infer_source_from_history(ws_vals, ws_forms, target_period: date, section: str):
    """
    La regla manda por periodos pasados:
    1) mismo periodo del año anterior, si existe;
    2) si no, cualquier periodo historico D:K con inputs suficientes.
    """
    prior_col = _find_prior_year_col(ws_vals, target_period)
    if prior_col is not None:
        values = _numeric_inputs_for_section(ws_vals, ws_forms, prior_col, section)
        inferred = _infer_source_from_values(values)
        if inferred is not None:
            return inferred, prior_col, len(values), "mismo periodo ano anterior"

    for col in _historical_cols(ws_vals):
        values = _numeric_inputs_for_section(ws_vals, ws_forms, col, section)
        inferred = _infer_source_from_values(values)
        if inferred is not None:
            return inferred, col, len(values), "historico disponible"

    return None, None, 0, "sin historico con inputs"


def _build_source_plan(
    wb_vals,
    wb_forms,
    target_period: date,
    hojas_input=None,
    default_source_plan=None,
    source_by_quarter=None,
) -> tuple[dict, list[str]]:
    """Wiki rule wins: infer EEFF vs Analisis from historical periods."""
    if hojas_input is None:
        hojas_input = HOJAS_INPUT
    if default_source_plan is None:
        default_source_plan = DEFAULT_SOURCE_PLAN
    plan = dict(default_source_plan)
    notes = []
    quarter = _mes_a_q(target_period.month)

    quarter_plan = (source_by_quarter or {}).get(quarter, {})
    if quarter_plan:
        for key, source in quarter_plan.items():
            if key[0] in hojas_input:
                plan[key] = source
        notes.append(f"  Fuente definida por tabla fija Q{quarter} (derivada de historico 2025)")

    for sheet_name in hojas_input:
        for section in ("balance", "eerr"):
            if (sheet_name, section) in quarter_plan:
                notes.append(f"  {sheet_name} {section}: {plan[(sheet_name, section)]} (tabla fija Q{quarter})")
                continue
            inferred, col, n_values, basis = _infer_source_from_history(
                wb_vals[sheet_name], wb_forms[sheet_name], target_period, section
            )
            if inferred is None:
                notes.append(f"  {sheet_name} {section}: {basis}; uso default documentado")
                continue
            plan[(sheet_name, section)] = inferred
            col_letter = get_column_letter(col)
            notes.append(
                f"  {sheet_name} {section}: {inferred} "
                f"({basis}, col {col_letter}, {n_values} inputs)"
            )

    return plan, notes


def _source(plan: dict, sheet_name: str, section: str) -> str:
    return plan[(sheet_name, section)]


def _warn_unsupported_sources(plan: dict) -> list[str]:
    unsupported = []
    if _source(plan, "Fondo PT", "balance") != "eeff":
        unsupported.append("  Fondo PT balance: wiki indica Analisis, pero no hay lector implementado")
    if _source(plan, "Fondo PT", "eerr") != "eeff":
        unsupported.append("  Fondo PT EERR: wiki indica Analisis, pero no hay lector implementado")
    if _source(plan, "Inmob Boulevard", "eerr") != "analisis":
        unsupported.append("  Inmob Boulevard EERR: wiki indica EEFF, pero no hay parser EERR PDF implementado")
    return unsupported


def _warn_unsupported_sources_apoquindo(plan: dict) -> list[str]:
    unsupported = []
    if _source(plan, "Fondo Apoquindo", "balance") != "eeff":
        unsupported.append("  Fondo Apoquindo balance: wiki indica Analisis, pero no hay lector implementado")
    if _source(plan, "Fondo Apoquindo", "eerr") != "eeff":
        unsupported.append("  Fondo Apoquindo EERR: wiki indica Analisis, pero no hay lector implementado")
    return unsupported

def _find_vf_with_fallback(finder, año: int, mes: int) -> str | None:
    vf_path = finder(año, mes)
    if vf_path:
        return vf_path
    if not os.path.isdir(BALANCES_DIR):
        return None
    for y in sorted(os.listdir(BALANCES_DIR), reverse=True):
        ydir = os.path.join(BALANCES_DIR, y)
        if os.path.isdir(ydir):
            vf_path = finder(int(y) if y.isdigit() else año, mes)
            if vf_path:
                return vf_path
    return None


def _sp_path(path: str | None) -> str:
    if not path:
        return ""
    if SHAREPOINT_DIR and path.startswith(SHAREPOINT_DIR):
        rel = path[len(SHAREPOINT_DIR):].lstrip(os.sep)
        return f"SP: {rel}"
    return path


def _check_required(encontrados: list[tuple[str, str]], faltantes: list[str], nombre: str, ruta: str | None):
    if ruta and os.path.isfile(ruta):
        encontrados.append((nombre, ruta))
    else:
        faltantes.append(nombre)


def _format_balance_check(
    title: str,
    año: int,
    mes: int,
    encontrados: list[tuple[str, str]],
    faltantes: list[str],
    bloqueos: list[str] | None = None,
    notas: list[str] | None = None,
) -> str:
    total = len(encontrados) + len(faltantes)
    lines = [f"=== {title} {mes:02d}.{año} ===", ""]
    lines.append(f"Archivos encontrados ({len(encontrados)}/{total})")
    lines.extend(
        f"  - {nombre}: {_sp_path(ruta)}" for nombre, ruta in encontrados
    )
    if not encontrados:
        lines.append("  - Ninguno")
    lines.append("")
    lines.append(f"Archivos faltantes ({len(faltantes)}/{total})")
    lines.extend(f"  - {nombre}" for nombre in faltantes)
    if not faltantes:
        lines.append("  - Ninguno")
    if bloqueos:
        lines.append("")
        lines.append("Bloqueos")
        lines.extend(bloqueos)
    if notas:
        lines.append("")
        lines.append("Regla wiki EEFF/Analisis")
        lines.extend(notas)
    return "\n".join(lines)


def verificar_archivos_balance_consolidado_pt(mes: int, año: int) -> str:
    """Verifica los inputs necesarios antes de actualizar el Balance Consolidado PT."""
    if mes not in (3, 6, 9, 12):
        return f"Error: mes={mes} no es fin de trimestre (usar 3, 6, 9 o 12)"

    encontrados: list[tuple[str, str]] = []
    faltantes: list[str] = []
    bloqueos: list[str] = []
    notas: list[str] = []

    vf_path = _find_vf_with_fallback(_find_latest_vf, año, mes)
    _check_required(encontrados, faltantes, "Balance Consolidado Rentas PT vF base", vf_path)
    if not vf_path:
        return _format_balance_check("Balance Consolidado PT", año, mes, encontrados, faltantes, bloqueos, notas)
    if _period_from_filename(vf_path) == (mes, año):
        bloqueos.append(
            f"  - El periodo ya existe como vF: {os.path.basename(vf_path)}. No se crea vAgente para un periodo cerrado."
        )

    wb_vals = _load_readonly_workbook(vf_path, data_only=True)
    wb_forms = _load_readonly_workbook(vf_path, data_only=False)
    try:
        source_plan, notas = _build_source_plan(
            wb_vals,
            wb_forms,
            _quarter_end(mes, año),
            source_by_quarter=PT_SOURCE_BY_QUARTER,
        )
        bloqueos.extend(_warn_unsupported_sources(source_plan))
    finally:
        wb_forms.close()
        wb_vals.close()

    if _source(source_plan, "Fondo PT", "balance") == "eeff" or _source(source_plan, "Fondo PT", "eerr") == "eeff":
        _check_required(encontrados, faltantes, "EEFF PDF Fondo PT", _find_eeff_fondo_pt(mes, año))

    eeff_bvd, analisis_bvd = _find_boulevard_files(mes, año)
    if _source(source_plan, "Inmob Boulevard", "balance") == "eeff":
        _check_required(encontrados, faltantes, "EEFF PDF Boulevard", eeff_bvd)
    if _source(source_plan, "Inmob Boulevard", "balance") == "analisis" or _source(source_plan, "Inmob Boulevard", "eerr") == "analisis":
        _check_required(encontrados, faltantes, "Analisis Boulevard", analisis_bvd)

    eeff_ta, analisis_ta = _find_torre_a_files(mes, año)
    if _source(source_plan, "Torre A", "balance") == "eeff" or _source(source_plan, "Torre A", "eerr") == "eeff":
        _check_required(encontrados, faltantes, "EEFF PDF Torre A", eeff_ta)
    if _source(source_plan, "Torre A", "balance") == "analisis" or _source(source_plan, "Torre A", "eerr") == "analisis":
        _check_required(encontrados, faltantes, "Analisis Torre A", analisis_ta)

    return _format_balance_check("Balance Consolidado PT", año, mes, encontrados, faltantes, bloqueos, notas)


def verificar_archivos_balance_consolidado_apoquindo(mes: int, año: int) -> str:
    """Verifica los inputs necesarios antes de actualizar el Balance Consolidado Apoquindo."""
    if mes not in (3, 6, 9, 12):
        return f"Error: mes={mes} no es fin de trimestre (usar 3, 6, 9 o 12)"

    encontrados: list[tuple[str, str]] = []
    faltantes: list[str] = []
    bloqueos: list[str] = []
    notas: list[str] = []

    vf_path = _find_vf_with_fallback(_find_latest_vf_apoquindo, año, mes)
    _check_required(encontrados, faltantes, "Balance Consolidado Rentas Apoquindo vF base", vf_path)
    if not vf_path:
        return _format_balance_check("Balance Consolidado Apoquindo", año, mes, encontrados, faltantes, bloqueos, notas)
    if _period_from_filename(vf_path) == (mes, año):
        bloqueos.append(
            f"  - El periodo ya existe como vF: {os.path.basename(vf_path)}. No se crea vAgente para un periodo cerrado."
        )

    wb_vals = _load_readonly_workbook(vf_path, data_only=True)
    wb_forms = _load_readonly_workbook(vf_path, data_only=False)
    try:
        missing_sheets = [s for s in APO_HOJAS_INPUT if s not in wb_forms.sheetnames]
        if missing_sheets:
            bloqueos.append("  - Faltan hojas esperadas en la planilla: " + ", ".join(missing_sheets))
            source_plan = dict(APO_DEFAULT_SOURCE_PLAN)
        else:
            source_plan, notas = _build_source_plan(
                wb_vals,
                wb_forms,
                _quarter_end(mes, año),
                hojas_input=APO_HOJAS_INPUT,
                default_source_plan=APO_DEFAULT_SOURCE_PLAN,
                source_by_quarter=APO_SOURCE_BY_QUARTER,
            )
            bloqueos.extend(_warn_unsupported_sources_apoquindo(source_plan))
    finally:
        wb_forms.close()
        wb_vals.close()

    if _source(source_plan, "Fondo Apoquindo", "balance") == "eeff" or _source(source_plan, "Fondo Apoquindo", "eerr") == "eeff":
        _check_required(encontrados, faltantes, "EEFF PDF Fondo Apoquindo", _find_eeff_fondo_apoquindo(mes, año))

    pdf_inmob = _find_eeff_inmobiliaria_apoquindo(mes, año)
    analisis_inmob = _find_analisis_inmobiliaria_apoquindo(mes, año)
    if _source(source_plan, "Inmobilaria Apoquindo", "balance") == "eeff" or _source(source_plan, "Inmobilaria Apoquindo", "eerr") == "eeff":
        _check_required(encontrados, faltantes, "EEFF PDF Inmobiliaria Apoquindo", pdf_inmob)
    if _source(source_plan, "Inmobilaria Apoquindo", "balance") == "analisis" or _source(source_plan, "Inmobilaria Apoquindo", "eerr") == "analisis":
        _check_required(encontrados, faltantes, "Analisis Inmobiliaria Apoquindo", analisis_inmob)

    return _format_balance_check("Balance Consolidado Apoquindo", año, mes, encontrados, faltantes, bloqueos, notas)


def _balance_check_is_complete(check_result: str) -> bool:
    return "Archivos faltantes (0/" in check_result and "\nBloqueos\n" not in check_result


def actualizar_balance_consolidado_pt_si_completo(mes: int, año: int) -> str:
    """Verifica archivos y actualiza PT solo si no hay faltantes ni bloqueos."""
    check_result = verificar_archivos_balance_consolidado_pt(mes, año)
    if not _balance_check_is_complete(check_result):
        return check_result + "\n\nNo actualice el balance porque hay archivos faltantes o bloqueos."
    return check_result + "\n\n" + actualizar_balance_consolidado_pt(mes, año)


def actualizar_balance_consolidado_apoquindo_si_completo(mes: int, año: int) -> str:
    """Verifica archivos y actualiza Apoquindo solo si no hay faltantes ni bloqueos."""
    check_result = verificar_archivos_balance_consolidado_apoquindo(mes, año)
    if not _balance_check_is_complete(check_result):
        return check_result + "\n\nNo actualice el balance porque hay archivos faltantes o bloqueos."
    return check_result + "\n\n" + actualizar_balance_consolidado_apoquindo(mes, año)


def actualizar_balance_consolidado_rentas_si_completo(mes: int, año: int) -> str:
    """Actualiza Balance Consolidado Rentas Nuevo para el trimestre mes/año."""
    return actualizar_balance_consolidado_rentas_nuevo(mes, año)


def actualizar_balances_consolidados_si_completos(mes: int, año: int) -> str:
    """Actualiza PT, Apoquindo y Rentas, verificando cada fondo antes de ejecutar."""
    resultados = [
        actualizar_balance_consolidado_pt_si_completo(mes, año),
        actualizar_balance_consolidado_apoquindo_si_completo(mes, año),
        actualizar_balance_consolidado_rentas_si_completo(mes, año),
    ]
    return ("\n\n" + "=" * 72 + "\n\n").join(resultados)


# ─── Validación ───────────────────────────────────────────────────────────────

def _validate_sheet(ws_vals, sheet_name: str) -> list[str]:
    """
    Verifica cuadre: Total Activos == Pasivos + Patrimonio.
    Retorna lista de mensajes.
    """
    COL = 4

    def _v(row):
        return ws_vals.cell(row=row, column=COL).value or 0

    msgs = []
    if sheet_name == "Fondo PT":
        total_a  = _v(35)
        total_pp = _v(70)
        res_ej_balance = _v(65)
        res_ej_eerr    = _v(103)
        if total_a and abs(total_a - total_pp) > 1000:
            msgs.append(f"  ⚠ Activos ({total_a:,.0f}) ≠ Pasivos+Pat ({total_pp:,.0f})")
        if res_ej_balance and res_ej_eerr and abs(res_ej_balance - res_ej_eerr) > 1000:
            msgs.append(f"  ⚠ Resultado balance ({res_ej_balance:,.0f}) ≠ EERR ({res_ej_eerr:,.0f})")
    elif sheet_name == "Inmob Boulevard":
        total_a  = _v(35)
        total_pp = _v(70)
        if total_a and abs(total_a - total_pp) > 1000:
            msgs.append(f"  ⚠ Activos ({total_a:,.0f}) ≠ Pasivos+Pat ({total_pp:,.0f})")
    elif sheet_name == "Torre A":
        total_a  = _v(35)
        total_pp = _v(70)
        if total_a and abs(total_a - total_pp) > 1000:
            msgs.append(f"  ⚠ Activos ({total_a:,.0f}) ≠ Pasivos+Pat ({total_pp:,.0f})")
    return msgs

# ─── Función principal ────────────────────────────────────────────────────────

def actualizar_balance_consolidado_pt(mes: int, año: int) -> str:
    """
    Actualiza el Balance Consolidado Rentas PT para el período dado.

    Pasos:
      1. Busca el último archivo vF y crea copia vAgente
      2. Desplaza columnas D-K (shift right) en las 3 hojas input
      3. Escribe la fecha del período en row 2 col D
      4. Decide EEFF vs Analisis con la regla general de la wiki
      5. Rellena Fondo PT, Inmob Boulevard y Torre A con la fuente inferida
         para cada seccion; usa defaults solo si no hay periodo comparable
      7. Valida cuadres y retorna informe
    """
    if mes not in (3, 6, 9, 12):
        return f"Error: mes={mes} no es fin de trimestre (usar 3, 6, 9 o 12)"

    lines = [f"=== Balance Consolidado PT {mes:02d}.{año} ===", ""]

    # 1. Encontrar vF más reciente
    vf_path = _find_latest_vf(año, mes)
    if not vf_path:
        # Buscar en cualquier año disponible
        for y in sorted(os.listdir(BALANCES_DIR), reverse=True):
            ydir = os.path.join(BALANCES_DIR, y)
            if os.path.isdir(ydir):
                vf_path = _find_latest_vf(int(y) if y.isdigit() else año, mes)
                if vf_path:
                    break
    if not vf_path:
        return "Error: no se encontró ningún archivo vF de Balance Consolidado Rentas PT"

    if _period_from_filename(vf_path) == (mes, año):
        return (
            f"Balance Consolidado Rentas PT {mes:02d}.{año} ya existe como vF: "
            f"{os.path.basename(vf_path)}. No se crea vAgente para un periodo ya cerrado."
        )

    lines.append(f"Fuente vF: {os.path.basename(vf_path)}")

    # 2. Crear archivo vAgente
    mm_yyyy = f"{mes:02d}.{año}"
    dest_name = f"{mm_yyyy}- Balance Consolidado Rentas PT vAgente.xlsx"
    dest_dir = os.path.dirname(vf_path)
    dest_path = os.path.join(dest_dir, dest_name)
    shutil.copy2(vf_path, dest_path)
    lines.append(f"Archivo destino: {dest_name}")
    lines.append("")

    # 3. Cargar workbooks (data_only=True para valores, data_only=False para fórmulas)
    wb_vals  = openpyxl.load_workbook(dest_path, data_only=True)
    wb_forms = openpyxl.load_workbook(dest_path, data_only=False)

    fecha_periodo = _quarter_end(mes, año)
    source_plan, source_notes = _build_source_plan(
        wb_vals,
        wb_forms,
        fecha_periodo,
        source_by_quarter=PT_SOURCE_BY_QUARTER,
    )
    lines.append("Regla wiki EEFF/Analisis:")
    lines.extend(source_notes)
    unsupported_sources = _warn_unsupported_sources(source_plan)
    if unsupported_sources:
        lines.append("  ⚠ Fuentes inferidas no soportadas por la herramienta:")
        lines.extend(unsupported_sources)
    lines.append("")

    # 4. Shift D-K en hojas input
    _shift_input_sheets(wb_vals, wb_forms)

    # 5. Fecha del período (último día del trimestre)
    fecha_dt = datetime(fecha_periodo.year, fecha_periodo.month, fecha_periodo.day)
    status_label = _period_status_label(mes, año)
    for hn in HOJAS_INPUT:
        wb_forms[hn].cell(row=2, column=2).value = status_label
        wb_forms[hn].cell(row=2, column=4).value = fecha_dt
    lines.append(f"Fecha período: {fecha_periodo}")
    lines.append(f"Estado B2: {status_label}")
    lines.append("")

    # ── EEFF Fondo PT ─────────────────────────────────────────────────────────
    fondo_balance_from_eeff = _source(source_plan, "Fondo PT", "balance") == "eeff"
    fondo_eerr_from_eeff = _source(source_plan, "Fondo PT", "eerr") == "eeff"
    pdf_pt = _find_eeff_fondo_pt(mes, año)
    if pdf_pt and (fondo_balance_from_eeff or fondo_eerr_from_eeff):
        lines.append(f"EEFF Fondo PT: {os.path.basename(pdf_pt)}")
        try:
            fondo_pt_d = _parse_eeff_fondo_pt_pdf(pdf_pt)
            _fill_fondo_pt(
                wb_forms["Fondo PT"],
                fondo_pt_d,
                fill_balance=fondo_balance_from_eeff,
                fill_eerr=fondo_eerr_from_eeff,
            )
            lines.append(f"  Efectivo: {fondo_pt_d.get('efectivo', 0):,.0f}")
            lines.append(f"  Inv. método participación: {fondo_pt_d.get('inv_met_part', 0):,.0f}")
            lines.append(f"  Total Activo: {fondo_pt_d.get('total_activo', 0):,.0f}")
            lines.append(f"  Aportes (bruto): {fondo_pt_d.get('aportes', 0):,.0f}")
            lines.append(f"  Resultado ejercicio: {fondo_pt_d.get('res_ej', 0):,.0f}")
        except Exception as e:
            lines.append(f"  ⚠ Error parseo PDF PT: {e}")
    elif not (fondo_balance_from_eeff or fondo_eerr_from_eeff):
        lines.append("⚠ Fondo PT no actualizado: regla wiki no pide EEFF y no hay lector de Analisis")
    else:
        lines.append("⚠ No se encontró EEFF PDF Fondo PT — hoja Fondo PT no actualizada")
    lines.append("")

    # ── Boulevard EEFF PDF (balance) + Análisis xlsx (EERR) ───────────────────
    eeff_bvd, analisis_bvd = _find_boulevard_files(mes, año)
    eerr_bvd_map = _get_eerr_row_map(wb_forms, "Inmob Boulevard")
    bvd_balance_from_eeff = _source(source_plan, "Inmob Boulevard", "balance") == "eeff"
    bvd_balance_from_analisis = _source(source_plan, "Inmob Boulevard", "balance") == "analisis"
    bvd_eerr_from_analisis = _source(source_plan, "Inmob Boulevard", "eerr") == "analisis"

    if eeff_bvd and bvd_balance_from_eeff:
        lines.append(f"EEFF Boulevard: {os.path.basename(eeff_bvd)}")
        try:
            bvd_balance = _parse_eeff_boulevard_pdf(eeff_bvd)
            lines.append(f"  Efectivo: {bvd_balance.get('efectivo', 0):,.0f}")
            lines.append(f"  Prop. Inversión: {bvd_balance.get('prop_inv', 0):,.0f}")
            lines.append(f"  Total Activo: {bvd_balance.get('total_activo', 0):,.0f}")
        except Exception as e:
            bvd_balance = {}
            lines.append(f"  ⚠ Error parseo PDF Boulevard: {e}")
    elif analisis_bvd and bvd_balance_from_analisis:
        lines.append(f"Análisis Boulevard balance: {os.path.basename(analisis_bvd)}")
        try:
            bvd_balance = _read_boulevard_balance(analisis_bvd)
            lines.append(f"  Efectivo: {bvd_balance.get('efectivo', 0):,.0f}")
            lines.append(f"  Prop. Inversión: {bvd_balance.get('prop_inv', 0):,.0f}")
            lines.append(f"  Total Activo ajustado: {bvd_balance.get('total_activo', 0):,.0f}")
            lines.append(f"  Intereses duplicados excluidos: {bvd_balance.get('intereses_duplicados', 0):,.0f}")
        except Exception as e:
            bvd_balance = {}
            lines.append(f"  ⚠ Error Análisis Boulevard balance: {e}")
    elif bvd_balance_from_analisis:
        bvd_balance = {}
        lines.append("⚠ No se encontró Análisis Boulevard — balance no actualizado")
    else:
        bvd_balance = {}
        lines.append("⚠ No se encontró EEFF PDF Boulevard — balance no actualizado")

    if analisis_bvd and bvd_eerr_from_analisis:
        lines.append(f"Análisis Boulevard: {os.path.basename(analisis_bvd)}")
        try:
            eerr_bvd = _read_analisis_eerr(analisis_bvd, "EERR")
            lines.append(f"  Cuentas EERR leídas: {len(eerr_bvd)}")
        except Exception as e:
            eerr_bvd = {}
            lines.append(f"  ⚠ Error Análisis Boulevard: {e}")
    elif not bvd_eerr_from_analisis:
        eerr_bvd = {}
        lines.append("⚠ EERR Boulevard no actualizado: regla wiki pide EEFF y no hay parser EERR PDF")
    else:
        eerr_bvd = {}
        lines.append("⚠ No se encontró Análisis Boulevard — EERR no actualizado")

    if bvd_balance or eerr_bvd:
        _fill_boulevard(
            wb_forms["Inmob Boulevard"],
            bvd_balance,
            eerr_bvd,
            eerr_bvd_map,
            fill_balance=bvd_balance_from_eeff or bvd_balance_from_analisis,
            fill_eerr=bvd_eerr_from_analisis,
        )
    lines.append("")

    # ── Torre A: Análisis xlsx ─────────────────────────────────────────────────
    eeff_ta, analisis_ta = _find_torre_a_files(mes, año)
    eerr_ta_map = _get_eerr_row_map(wb_forms, "Torre A")
    ta_balance_from_analisis = _source(source_plan, "Torre A", "balance") == "analisis"
    ta_eerr_from_analisis = _source(source_plan, "Torre A", "eerr") == "analisis"
    ta_balance_from_eeff = _source(source_plan, "Torre A", "balance") == "eeff"
    ta_eerr_from_eeff = _source(source_plan, "Torre A", "eerr") == "eeff"
    ta_updated = False

    if analisis_ta and (ta_balance_from_analisis or ta_eerr_from_analisis):
        lines.append(f"Análisis Torre A: {os.path.basename(analisis_ta)}")
        try:
            ta_balance = _read_torre_a_balance(analisis_ta) if ta_balance_from_analisis else {}
            eerr_ta = _read_analisis_eerr(analisis_ta, "EERR") if ta_eerr_from_analisis else {}
            _fill_torre_a(
                wb_forms["Torre A"],
                ta_balance,
                eerr_ta,
                eerr_ta_map,
                fill_balance=ta_balance_from_analisis,
                fill_eerr=ta_eerr_from_analisis,
            )
            # Buscar valores claves para verificar
            efectivo_ta = ta_balance.get(_norm_label("EFECTIVO Y EQUIVALENTE AL EFECTIVO"))
            lines.append(f"  Efectivo: {efectivo_ta:,.0f}" if efectivo_ta else "  Efectivo: no encontrado")
            prop_ta = ta_balance.get(_norm_label("PROPIEDADES DE INVERSION"))
            lines.append(f"  Prop. Inversión: {prop_ta:,.0f}" if prop_ta else "  Prop. Inversión: no encontrado")
            ta_updated = True
        except Exception as e:
            lines.append(f"  ⚠ Error Análisis Torre A: {e}")
    if eeff_ta and (ta_balance_from_eeff or ta_eerr_from_eeff):
        lines.append(f"EEFF Torre A: {os.path.basename(eeff_ta)}")
        try:
            ta_eeff = _parse_eeff_torre_a_pdf(eeff_ta)
            _fill_torre_a_eeff(
                wb_forms["Torre A"],
                ta_eeff,
                fill_balance=ta_balance_from_eeff,
                fill_eerr=ta_eerr_from_eeff,
            )
            lines.append(f"  Efectivo: {ta_eeff.get('efectivo', 0):,.0f}")
            lines.append(f"  Prop. Inversión: {ta_eeff.get('prop_inv', 0):,.0f}")
            lines.append(f"  Total Activo: {ta_eeff.get('total_activo', 0):,.0f}")
            lines.append(f"  Resultado ejercicio: {ta_eeff.get('res_ej', 0):,.0f}")
            ta_updated = True
        except Exception as e:
            lines.append(f"  ⚠ Error EEFF Torre A: {e}")
    if not eeff_ta and (ta_balance_from_eeff or ta_eerr_from_eeff):
        lines.append("⚠ No se encontró EEFF Torre A — hoja no actualizada")
    if not analisis_ta and (ta_balance_from_analisis or ta_eerr_from_analisis):
        lines.append("⚠ No se encontró Análisis Torre A — hoja no actualizada")
    if not ta_updated:
        lines.append("⚠ Torre A no fue actualizada")
    lines.append("")

    # ── Validar en memoria (antes de guardar) ─────────────────────────────────
    lines.append("=== Validaciones (en memoria) ===")
    # Para validar, leer los valores escritos directo de wb_forms (formulas no calculadas)
    # Solo verificar que hay valores en celdas clave
    for hn in HOJAS_INPUT:
        ws_f = wb_forms[hn]
        fecha = ws_f.cell(row=2, column=4).value
        val_key = ws_f.cell(row=7, column=4).value  # Efectivo
        if fecha and val_key is not None:
            lines.append(f"  {hn}: fecha={fecha.strftime('%d/%m/%Y') if hasattr(fecha, 'strftime') else fecha}, efectivo={val_key:,.0f}")
        else:
            lines.append(f"  {hn}: {'' if fecha else 'sin fecha'}{'' if val_key is not None else ', sin efectivo'}")

    # ── Guardar ────────────────────────────────────────────────────────────────
    wb_forms.save(dest_path)
    wb_forms.close()
    wb_vals.close()
    lines.append("")
    lines.append(f"Archivo guardado en: {dest_path}")
    lines.append("Abrir en Excel y verificar que las formulas recalculen correctamente.")

    return "\n".join(lines)


def actualizar_balance_consolidado_apoquindo(mes: int, año: int) -> str:
    """
    Actualiza el Balance Consolidado Rentas Apoquindo para el periodo dado.

    Usa la misma logica de PT: copia vF, desplaza D:K, escribe fecha nueva,
    infiere EEFF vs Analisis desde periodos pasados y rellena solo las fuentes
    que la herramienta sabe parsear.
    """
    if mes not in (3, 6, 9, 12):
        return f"Error: mes={mes} no es fin de trimestre (usar 3, 6, 9 o 12)"

    lines = [f"=== Balance Consolidado Apoquindo {mes:02d}.{año} ===", ""]

    vf_path = _find_latest_vf_apoquindo(año, mes)
    if not vf_path:
        for y in sorted(os.listdir(BALANCES_DIR), reverse=True):
            ydir = os.path.join(BALANCES_DIR, y)
            if os.path.isdir(ydir):
                vf_path = _find_latest_vf_apoquindo(int(y) if y.isdigit() else año, mes)
                if vf_path:
                    break
    if not vf_path:
        return "Error: no se encontro ningun archivo vF de Balance Consolidado Rentas Apoquindo"

    if _period_from_filename(vf_path) == (mes, año):
        return (
            f"Balance Consolidado Rentas Apoquindo {mes:02d}.{año} ya existe como vF: "
            f"{os.path.basename(vf_path)}. No se crea vAgente para un periodo ya cerrado."
        )

    lines.append(f"Fuente vF: {os.path.basename(vf_path)}")

    mm_yyyy = f"{mes:02d}.{año}"
    dest_name = f"{mm_yyyy}- Balance Consolidado Rentas Apoquindo vAgente.xlsx"
    dest_dir = os.path.dirname(vf_path)
    dest_path = os.path.join(dest_dir, dest_name)
    shutil.copy2(vf_path, dest_path)
    lines.append(f"Archivo destino: {dest_name}")
    lines.append("")

    wb_vals = openpyxl.load_workbook(dest_path, data_only=True)
    wb_forms = openpyxl.load_workbook(dest_path, data_only=False)

    missing_sheets = [s for s in APO_HOJAS_INPUT if s not in wb_forms.sheetnames]
    if missing_sheets:
        wb_forms.close()
        wb_vals.close()
        return "Error: faltan hojas esperadas en la planilla: " + ", ".join(missing_sheets)

    fecha_periodo = _quarter_end(mes, año)
    source_plan, source_notes = _build_source_plan(
        wb_vals,
        wb_forms,
        fecha_periodo,
        hojas_input=APO_HOJAS_INPUT,
        default_source_plan=APO_DEFAULT_SOURCE_PLAN,
        source_by_quarter=APO_SOURCE_BY_QUARTER,
    )
    lines.append("Regla wiki EEFF/Analisis:")
    lines.extend(source_notes)
    unsupported_sources = _warn_unsupported_sources_apoquindo(source_plan)
    if unsupported_sources:
        lines.append("  Avisos de fuente no soportada:")
        lines.extend(unsupported_sources)
    lines.append("")

    _shift_input_sheets(wb_vals, wb_forms, APO_HOJAS_INPUT)

    fecha_dt = datetime(fecha_periodo.year, fecha_periodo.month, fecha_periodo.day)
    status_label = _period_status_label(mes, año)
    for hn in APO_HOJAS_INPUT:
        wb_forms[hn].cell(row=2, column=2).value = status_label
        wb_forms[hn].cell(row=2, column=4).value = fecha_dt
    lines.append(f"Fecha periodo: {fecha_periodo}")
    lines.append(f"Estado B2: {status_label}")
    lines.append("")

    fondo_balance_from_eeff = _source(source_plan, "Fondo Apoquindo", "balance") == "eeff"
    fondo_eerr_from_eeff = _source(source_plan, "Fondo Apoquindo", "eerr") == "eeff"
    pdf_fondo = _find_eeff_fondo_apoquindo(mes, año)
    if pdf_fondo and (fondo_balance_from_eeff or fondo_eerr_from_eeff):
        lines.append(f"EEFF Fondo Apoquindo: {os.path.basename(pdf_fondo)}")
        try:
            fondo_d = _parse_eeff_fondo_apoquindo_pdf(pdf_fondo)
            _fill_fondo_apoquindo(
                wb_forms["Fondo Apoquindo"],
                fondo_d,
                fill_balance=fondo_balance_from_eeff,
                fill_eerr=fondo_eerr_from_eeff,
            )
            lines.append(f"  Efectivo: {fondo_d.get('efectivo', 0):,.0f}")
            lines.append(f"  Act. fin. costo amortizado NC: {fondo_d.get('af_costo_nc', 0):,.0f}")
            lines.append(f"  Total Activo: {fondo_d.get('total_activo', 0):,.0f}")
            lines.append(f"  Resultado ejercicio: {fondo_d.get('res_ej', 0):,.0f}")
        except Exception as e:
            lines.append(f"  Error parseo PDF Fondo Apoquindo: {e}")
    elif not (fondo_balance_from_eeff or fondo_eerr_from_eeff):
        lines.append("Fondo Apoquindo no actualizado: regla wiki no pide EEFF y no hay lector de Analisis")
    else:
        lines.append("No se encontro EEFF PDF Fondo Apoquindo; hoja Fondo Apoquindo no actualizada")
    lines.append("")

    inmob_balance_from_eeff = _source(source_plan, "Inmobilaria Apoquindo", "balance") == "eeff"
    inmob_eerr_from_eeff = _source(source_plan, "Inmobilaria Apoquindo", "eerr") == "eeff"
    inmob_balance_from_analisis = _source(source_plan, "Inmobilaria Apoquindo", "balance") == "analisis"
    inmob_eerr_from_analisis = _source(source_plan, "Inmobilaria Apoquindo", "eerr") == "analisis"
    pdf_inmob = _find_eeff_inmobiliaria_apoquindo(mes, año)
    analisis_inmob = _find_analisis_inmobiliaria_apoquindo(mes, año)
    if analisis_inmob and (inmob_balance_from_analisis or inmob_eerr_from_analisis):
        lines.append(f"Analisis Inmobiliaria Apoquindo: {os.path.basename(analisis_inmob)}")
        try:
            inmob_balance, inmob_eerr = _read_inmobiliaria_apoquindo_analisis(analisis_inmob)
            _fill_inmobiliaria_apoquindo_analisis(
                wb_forms["Inmobilaria Apoquindo"],
                inmob_balance,
                inmob_eerr,
                fill_balance=inmob_balance_from_analisis,
                fill_eerr=inmob_eerr_from_analisis,
            )
            lines.append(f"  Efectivo: {inmob_balance.get('efectivo', 0):,.0f}")
            lines.append(f"  Deudores netos: {inmob_balance.get('deudores', 0):,.0f}")
            lines.append(f"  Prop. inversion: {inmob_balance.get('prop_inv', 0):,.0f}")
            lines.append(f"  Resultado ejercicio: {inmob_balance.get('res_ej', 0):,.0f}")
        except Exception as e:
            lines.append(f"  Error Analisis Inmobiliaria Apoquindo: {e}")
    elif pdf_inmob and (inmob_balance_from_eeff or inmob_eerr_from_eeff):
        lines.append(f"EEFF Inmobiliaria Apoquindo: {os.path.basename(pdf_inmob)}")
        try:
            inmob_d = _parse_eeff_inmobiliaria_apoquindo_pdf(pdf_inmob)
            _fill_inmobiliaria_apoquindo_eeff(
                wb_forms["Inmobilaria Apoquindo"],
                inmob_d,
                fill_balance=inmob_balance_from_eeff,
                fill_eerr=inmob_eerr_from_eeff,
            )
            lines.append(f"  Efectivo: {inmob_d.get('efectivo', 0):,.0f}")
            lines.append(f"  Prop. inversion: {inmob_d.get('prop_inv', 0):,.0f}")
            lines.append(f"  Total Activo: {inmob_d.get('total_activo', 0):,.0f}")
            lines.append(f"  Resultado ejercicio: {inmob_d.get('resultado', 0):,.0f}")
        except Exception as e:
            lines.append(f"  Error parseo PDF Inmobiliaria Apoquindo: {e}")
    elif inmob_balance_from_analisis or inmob_eerr_from_analisis:
        lines.append("Inmobilaria Apoquindo no actualizada: historico indica Analisis/Matriz y no se encontro archivo")
    else:
        lines.append("No se encontro EEFF PDF Inmobiliaria Apoquindo; hoja no actualizada")
    lines.append("")

    lines.append("=== Validaciones (en memoria) ===")
    for hn in APO_HOJAS_INPUT:
        ws_f = wb_forms[hn]
        fecha = ws_f.cell(row=2, column=4).value
        val_key = ws_f.cell(row=7, column=4).value
        if fecha and val_key is not None:
            fecha_txt = fecha.strftime("%d/%m/%Y") if hasattr(fecha, "strftime") else fecha
            lines.append(f"  {hn}: fecha={fecha_txt}, efectivo={val_key:,.0f}")
        else:
            lines.append(f"  {hn}: {'' if fecha else 'sin fecha'}{'' if val_key is not None else ', sin efectivo'}")

    wb_forms.save(dest_path)
    wb_forms.close()
    wb_vals.close()
    lines.append("")
    lines.append(f"Archivo guardado en: {dest_path}")
    lines.append("Abrir en Excel y verificar que las formulas recalculen correctamente.")

    return "\n".join(lines)


# ─── Balance Consolidado Rentas Nuevo ─────────────────────────────────────────

RAW_DIR = os.path.join(SHAREPOINT_DIR, "RAW")
RENTAS_TRI_ACTIVOS_DIR = os.path.join(SHAREPOINT_DIR, "Fondos", "Rentas TRI", "Activos")

RENTAS_NUEVO_HOJAS_INPUT = [
    "Inmosa",
    "Chañarcillo",
    "Curicó",
    "Inmob VC",
    "Viña Centro",
    "Fondo Rentas",
]

# Balance maps verified against Dec 2025 planilla.
# type='A' → A - P; type='P' → P - A

CHANAR_BALANCE_MAP = {
    7:  ("A", ["1-1-01-"]),
    12: ("A", ["1-1-02-02", "1-1-02-07", "1-1-02-14", "1-1-02-15", "1-1-02-16"]),
    27: ("A", ["1-2-01-", "1-2-03-01"]),
    31: ("A", ["1-1-04-03"]),
    32: ("A", ["1-1-03-"]),
    42: ("P", ["2-1-01-08", "2-1-01-09", "2-1-01-11", "2-1-01-12"]),
    44: ("P", ["2-1-01-04"]),
    46: ("P", ["2-1-05-"]),
    48: ("P", ["2-1-06-05"]),
    52: ("P", ["2-1-03-01", "2-1-03-13"]),
    55: ("P", ["2-1-03-07", "2-1-03-08", "2-1-03-09", "2-1-03-10", "2-1-03-11", "2-1-03-12"]),
    56: ("P", ["2-1-06-04"]),
    57: ("P", ["2-1-04-04", "2-1-04-06", "2-1-04-07"]),
    62: ("P", ["3-1-01-01"]),
    63: ("P", ["3-1-05-"]),
    64: ("P", ["3-1-03-"]),
}

CURICO_BALANCE_MAP = {
    7:  ("A", ["1-1-01-", "1-1-02-012", "1-1-05-062", "1-1-05-066"]),
    12: ("A", ["1-1-04-010", "1-1-04-011", "1-1-05-064", "1-1-10-010"]),
    15: ("A", ["1-1-06-020", "1-1-09-012"]),
    26: ("A", ["1-2-01-002", "1-2-01-005", "1-2-03-010", "1-2-03-011"]),
    # R31/R56 handled separately via _apply_curico_impdif (A(1-2-04-001) - P(2-2-06-010))
    32: ("A", ["1-3-06-010", "1-3-06-011", "1-1-05-013", "1-1-05-012"]),
    40: ("P", ["2-1-05-010"]),
    41: ("P", ["2-1-05-014"]),
    42: ("P", ["2-1-04-010", "2-1-04-020", "2-1-04-030", "2-1-04-042"]),
    46: ("P", ["2-1-08-020", "2-1-08-040", "2-1-08-041"]),
    47: ("P", ["2-1-09-010"]),
    52: ("P", ["2-1-05-011"]),
    53: ("P", ["2-1-05-031", "2-1-05-032", "2-1-05-041", "2-1-05-051"]),
    55: ("P", ["2-1-06-"]),
    62: ("P", ["2-3-01-001"]),
    63: ("P", ["2-3-02-002"]),
    64: ("P", ["2-3-02-001"]),
}

INMOB_VC_BALANCE_MAP = {
    7:  ("A", ["1-1-01-01", "1-1-01-11"]),
    25: ("A", ["1-2-03-02", "1-2-03-03"]),
    31: ("A", ["1-1-04-03"]),
    42: ("P", ["2-1-01-08", "2-1-01-10"]),
    44: ("P", ["2-1-02-01", "1-1-02-06"]),  # 1-1-02-06 holds a P balance despite 1-x class
    46: ("P", ["2-1-05-02"]),
    52: ("P", ["2-1-01-05"]),
    55: ("P", ["2-2-01-02"]),
    56: ("P", ["2-1-06-04"]),
    62: ("P", ["3-1-01-01"]),
    64: ("P", ["3-1-03-01"]),
}

# EERR map Inmob VC — codes extracted from col B labels in planilla, verified Dec 2025.
INMOB_VC_EERR_MAP = {
    76:  ["4-1-01-01"],   # INGRESOS POR ARRIENDO
    77:  ["4-3-01-01"],   # OTROS INGRESOS (operacionales)
    81:  ["4-2-01-02"],   # INTERESES PAGARE
    82:  ["4-2-01-03"],   # INTERESES POR PRESTAMOS BANCARIOS
    83:  ["5-1-01-01"],   # REPARACION Y MANTENCION
    84:  ["5-1-01-02"],   # SEGUROS
    85:  ["5-1-01-03"],   # ASESORIAS LEGALES
    86:  ["5-1-01-05"],   # HONORARIOS AUDITORIA
    87:  ["5-1-01-06"],   # IMPUESTO TIMBRE
    88:  ["5-1-01-07"],   # PATENTES
    89:  ["5-1-01-08"],   # ASESORIAS EXTERNAS
    90:  ["5-1-01-10"],   # OTROS GASTOS
    91:  ["5-1-01-11"],   # CONTRIBUCIONES
    92:  ["5-1-01-13"],   # COMISIONES BANCARIAS
    93:  ["5-1-01-14"],   # ASESORIAS CONTABLES
    94:  ["5-1-01-15"],   # GASTOS NOTARIALES
    95:  ["5-1-01-18"],   # GASTOS DEUDORES INCOBRABLES
    96:  ["5-1-01-19"],   # PROPORCIONALIDAD IVA
    97:  ["5-1-01-20"],   # GASTOS COMUNES
    104: ["5-2-01-02"],   # RESULTADOS METODO DE LA PARTICIPACION
    105: ["5-2-01-03"],   # FLUCTUACION VALOR CUOTA FONDOS MUTUOS
    106: ["5-2-01-04"],   # REAJUSTE UF
    107: ["5-2-01-10"],   # REAJUSTES PAGARE
    108: ["5-3-01-01"],   # CUADRE PESOS
    109: ["5-2-01-05"],   # REAJUSTE IMPUESTOS
    112: ["5-1-01-16"],   # IMPUESTO RENTA
}

VINA_BALANCE_MAP = {
    7:  ("A", ["1-1-01-020", "1-1-01-022", "1-1-01-023", "1-1-03-030"]),
    9:  ("A", ["1-1-10-031"]),
    11: ("A", ["1-1-10-032"]),
    12: ("A", ["1-1-04-", "1-1-05-010", "1-1-05-012", "1-1-05-013", "1-1-06-020"]),
    15: ("A", ["1-1-09-020", "1-1-10-010", "1-1-10-030", "1-1-06-012"]),
    22: ("A", ["1-1-10-033"]),
    26: ("A", ["1-2-01-001", "1-2-01-010", "1-2-01-040"]),
    31: ("A", ["1-1-20-010"]),
    32: ("A", ["1-1-06-093", "1-1-06-094"]),
    41: ("P", ["2-1-40-025", "2-1-06-011"]),
    42: ("P", ["2-1-04-010", "2-1-04-042", "2-1-08-070"]),
    46: ("P", ["2-1-08-071"]),
    48: ("P", ["2-1-09-037", "2-1-09-020", "2-1-09-010", "2-1-09-060"]),
    52: ("P", ["2-1-40-024"]),
    53: ("P", ["2-1-40-026"]),
    54: ("P", ["2-2-03-111", "2-2-03-112", "2-2-03-113", "2-2-03-114"]),
    55: ("P", ["2-1-04-050", "2-2-03-020"]),
    56: ("P", ["2-2-06-010"]),
    62: ("P", ["2-3-01-001"]),
    64: ("P", ["2-3-02-001"]),
    66: ("P", ["2-3-01-071"]),
}

# INMOSA EERR via Senior Assist xlsx (dot-notation codes).
# Value = G - Pd per code; positive = income, negative = expense.
INMOSA_SA_EERR_MAP = {
    76:  ["3.1.1010.20.01"],
    77:  ["3.1.1010.20.06"],
    78:  ["3.1.1010.20.02"],
    82:  ["4.5.1030.10.02"],
    83:  ["4.5.1030.10.03"],
    84:  ["4.5.1030.10.05"],
    85:  ["4.5.1030.10.06"],
    86:  ["4.5.1030.10.07"],
    87:  ["4.5.1030.20.03"],
    88:  ["4.5.1030.20.04"],
    89:  ["4.5.1030.20.06"],
    90:  ["4.5.1030.20.09"],
    91:  ["4.5.1030.20.11"],
    92:  ["4.5.1030.20.12"],
    94:  ["4.5.1030.30.01"],
    95:  ["4.5.1030.30.02"],
    96:  ["4.5.1030.30.04"],
    97:  ["4.5.1030.30.07"],
    98:  ["4.5.1030.50.03"],
    105: ["4.5.1030.50.05"],
    106: ["4.5.1070.10.02"],
    107: ["4.5.1050.10.01"],
    108: ["4.5.1070.10.04"],
    109: ["4.5.1070.10.05"],
    110: ["4.5.1070.10.06"],
    111: ["3.5.1050.10.02"],
    112: ["4.5.1090.10.01"],
    113: ["4.5.1090.10.02"],
    114: ["4.5.2110.10.02"],
    115: ["3.5.1090.10.01"],
    120: ["4.5.1070.10.08"],
}


# ─── Source finders ───────────────────────────────────────────────────────────

def _find_vf_rentas_nuevo(año: int, mes: int) -> str | None:
    año_dir = os.path.join(BALANCES_DIR, str(año))
    q_dir = _find_quarter_folder(año_dir, mes)
    if q_dir:
        for f in os.listdir(q_dir):
            if "Rentas Nuevo" in f and "vF" in f and f.endswith(".xlsx"):
                return os.path.join(q_dir, f)
    pattern = os.path.join(BALANCES_DIR, "**", "*Rentas Nuevo*vF*.xlsx")
    hits = glob_module.glob(pattern, recursive=True)
    return max(hits, key=os.path.getmtime) if hits else None


def _find_analisis_chanar_rn(mes: int, año: int) -> str | None:
    for pat in [
        os.path.join(RAW_DIR, f"{mes:02d}-{año}*Ch*arcillo*.xlsx"),
        os.path.join(RAW_DIR, f"{mes:02d}-{año}*Cha*.xlsx"),
    ]:
        hits = glob_module.glob(pat)
        if hits:
            return sorted(hits)[-1]
    return None


def _find_analisis_inmob_vc_rn(mes: int, año: int) -> str | None:
    for pat in [
        os.path.join(RAW_DIR, f"{mes:02d}-{año}*Inmobiliaria*VC*.xlsx"),
        os.path.join(RAW_DIR, f"{mes:02d}-{año}*VC*.xlsx"),
    ]:
        hits = glob_module.glob(pat)
        if hits:
            return sorted(hits)[-1]
    return None


def _find_curico_informe_rn(mes: int, año: int) -> str | None:
    base = os.path.join(RENTAS_TRI_ACTIVOS_DIR, "Curicó", "EEFF")
    for y in (año, año - 1):
        ydir = os.path.join(base, str(y))
        if not os.path.isdir(ydir):
            continue
        for pat in [
            os.path.join(ydir, f"{mes:02d}-{año}*INFORME*CURIC*.xlsx"),
            os.path.join(ydir, f"*{año}*CURIC*.xlsx"),
        ]:
            hits = glob_module.glob(pat)
            if hits:
                return sorted(hits)[-1]
    return None


def _find_vina_trial_balance_rn(mes: int, año: int) -> str | None:
    for pat in [
        os.path.join(RAW_DIR, f"{mes:02d}-{año}*INFORME*EFF*VI*A*CENTRO*.xlsx"),
        os.path.join(RAW_DIR, f"*{mes:02d}*{año}*VI*A*CENTRO*.xlsx"),
        os.path.join(RAW_DIR, f"*VI*A*CENTRO*{año}*.xlsx"),
    ]:
        hits = glob_module.glob(pat)
        if hits:
            return sorted(hits)[-1]
    return None


def _find_senior_assist_rn(mes: int, año: int) -> str | None:
    for pat in [
        os.path.join(RAW_DIR, f"*{año}*Senior*Assist*.xlsx"),
        os.path.join(RAW_DIR, f"Balance*{año}*Senior*.xlsx"),
        os.path.join(RAW_DIR, f"Balance*General*{año}*.xlsx"),
    ]:
        hits = glob_module.glob(pat)
        if hits:
            return sorted(hits)[-1]
    return None


# ─── Trial balance reader ─────────────────────────────────────────────────────

def _read_trial_balance_rn(xlsx_path: str, sheet_name: str | None = None) -> dict:
    """
    Read a 9-column trial balance (Cuenta|Deb|Cred|Deudor|Acreedor|Activo|Pasivo|Perdida|Ganancia).
    Returns: {code: {'A': activo, 'P': pasivo, 'Pd': perdida, 'G': ganancia}}
    Col indices (0-based): 0=code+name, 5=Activo, 6=Pasivo, 7=Perdida, 8=Ganancia.
    """
    wb = _load_readonly_workbook(xlsx_path, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif sheet_name:
            norm = _norm_label(sheet_name)
            matched = next((sn for sn in wb.sheetnames if _norm_label(sn) == norm), None)
            ws = wb[matched] if matched else wb.worksheets[0]
        else:
            ws = wb.worksheets[0]

        result = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            raw = str(row[0]).strip()
            m = re.match(r"^(\d[\d\.\-]*\d)", raw)
            if not m:
                continue
            code = m.group(1)

            def _n(v):
                if v is None:
                    return 0.0
                if isinstance(v, (int, float)):
                    return float(v)
                try:
                    return float(str(v).replace(".", "").replace(",", "."))
                except ValueError:
                    return 0.0

            a  = _n(row[5]) if len(row) > 5 else 0.0
            p  = _n(row[6]) if len(row) > 6 else 0.0
            pd = _n(row[7]) if len(row) > 7 else 0.0
            g  = _n(row[8]) if len(row) > 8 else 0.0

            if a or p or pd or g:
                result[code] = {"A": a, "P": p, "Pd": pd, "G": g}
        return result
    finally:
        wb.close()


def _tb_sum(tb: dict, prefixes: list, field: str) -> float:
    return sum(
        vals.get(field, 0.0)
        for code, vals in tb.items()
        if any(code.startswith(pfx) for pfx in prefixes)
    )


def _apply_balance_map_rn(ws, tb: dict, bmap: dict, col: int):
    for row, (typ, prefixes) in bmap.items():
        if typ == "A":
            val = _tb_sum(tb, prefixes, "A") - _tb_sum(tb, prefixes, "P")
        else:
            val = _tb_sum(tb, prefixes, "P") - _tb_sum(tb, prefixes, "A")
        ws.cell(row=row, column=col).value = val if val != 0 else None


def _apply_curico_impdif(ws, tb: dict, col: int):
    """R31/R56: net deferred tax = A(1-2-04-001) - P(2-2-06-010). R31 if >0, R56 if <0."""
    net = _tb_sum(tb, ["1-2-04-001"], "A") - _tb_sum(tb, ["2-2-06-010"], "P")
    ws.cell(row=31, column=col).value = net if net > 0 else None
    ws.cell(row=56, column=col).value = -net if net < 0 else None


def _apply_eerr_sa_map_rn(ws, tb: dict, col: int, eerr_map: dict | None = None):
    """Fill EERR rows using {row: [codes]} map. Value = G - Pd per code."""
    if eerr_map is None:
        eerr_map = INMOSA_SA_EERR_MAP
    for row, codes in eerr_map.items():
        total = sum(tb[c].get("G", 0.0) - tb[c].get("Pd", 0.0) for c in codes if c in tb)
        ws.cell(row=row, column=col).value = total if total != 0 else None


# ─── Sheet lookup helper ──────────────────────────────────────────────────────

def _find_ws_rn(wb, target: str):
    """Return worksheet by normalized name (exact then partial match)."""
    norm = _norm_label(target)
    for sn in wb.sheetnames:
        if _norm_label(sn) == norm:
            return wb[sn]
    for sn in wb.sheetnames:
        sn_n = _norm_label(sn)
        if norm in sn_n or sn_n in norm:
            return wb[sn]
    return None


# ─── Column shift (single-workbook variant for Rentas Nuevo) ──────────────────

def _shift_one_sheet_rn(ws_vals, ws_forms):
    """Shift cols D:K right using cached values from ws_vals, write into ws_forms."""
    D, K = 4, 11
    max_row = max(
        (c.row for row in ws_vals.iter_rows() for c in row if c.value is not None),
        default=130,
    )
    max_row = max(max_row, 130)
    for r in range(2, max_row + 1):
        vals = [ws_vals.cell(row=r, column=c).value for c in range(D, K + 1)]
        forms = [ws_forms.cell(row=r, column=c).value for c in range(D, K + 1)]
        d_is_formula = isinstance(forms[0], str) and forms[0].startswith("=")
        for dst in range(K, D, -1):
            ws_forms.cell(row=r, column=dst).value = vals[dst - D - 1]
        if not d_is_formula:
            ws_forms.cell(row=r, column=D).value = None


# ─── PT / Apoquindo sheet copy ────────────────────────────────────────────────

def _copy_vals_sheet_rn(src_path: str, src_sheet: str, wb_dst, dst_sheet: str) -> bool:
    """Copy all non-formula values from src_sheet to dst_sheet."""
    try:
        wb_src = openpyxl.load_workbook(src_path, data_only=True)
    except Exception:
        return False
    ws_src = _find_ws_rn(wb_src, src_sheet)
    ws_dst = _find_ws_rn(wb_dst, dst_sheet)
    if ws_src is None or ws_dst is None:
        wb_src.close()
        return False
    for row in ws_src.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                continue
            ws_dst.cell(row=cell.row, column=cell.column).value = val
    wb_src.close()
    return True


# ─── Main function ────────────────────────────────────────────────────────────

def actualizar_balance_consolidado_rentas_nuevo(mes: int, año: int) -> str:
    """
    Actualiza Balance Consolidado Rentas Nuevo para el trimestre mes/año.

    Input sheets (trial balance): Inmosa, Chañarcillo, Curicó, Inmob VC, Viña Centro, Fondo Rentas
    Copy sheets: Resumen PT, Consolidado Fondo PT, Resumen  Apoquindo, Consolidado Apoquindo
    Output sheets (no tocar): Resumen, Consolidado Fondo Rentas , Resumen Viña, Consolidado Viña
    """
    if mes not in (3, 6, 9, 12):
        return f"Error: mes={mes} no es fin de trimestre (usar 3, 6, 9 o 12)"

    lines = [f"=== Balance Consolidado Rentas Nuevo {mes:02d}.{año} ===", ""]

    vf_path = _find_vf_rentas_nuevo(año, mes)
    if not vf_path:
        return f"Error: no se encontro archivo vF de Balance Consolidado Rentas Nuevo para {mes:02d}.{año}"

    lines.append(f"Fuente vF: {_sp_path(vf_path)}")

    mm_yyyy = f"{mes:02d}.{año}"
    dest_name = f"{mm_yyyy}- Balance Consolidado Rentas Nuevo vAgente.xlsx"
    dest_dir = os.path.dirname(vf_path)
    dest_path = os.path.join(dest_dir, dest_name)
    shutil.copy2(vf_path, dest_path)
    lines.append(f"Destino: {dest_name}")
    lines.append("")

    wb_vals  = openpyxl.load_workbook(dest_path, data_only=True)
    wb_forms = openpyxl.load_workbook(dest_path, data_only=False)

    fecha_periodo = _quarter_end(mes, año)
    fecha_dt = datetime(fecha_periodo.year, fecha_periodo.month, fecha_periodo.day)
    status_label = _period_status_label(mes, año)
    col = 4  # D

    # Shift D:K on all input sheets + PT/Apo copy targets
    all_shift = RENTAS_NUEVO_HOJAS_INPUT + [
        "Resumen PT", "Consolidado Fondo PT",
        "Resumen  Apoquindo", "Consolidado Apoquindo",
    ]
    for sn_target in all_shift:
        ws_v = _find_ws_rn(wb_vals, sn_target)
        ws_f = _find_ws_rn(wb_forms, sn_target)
        if ws_v and ws_f:
            _shift_one_sheet_rn(ws_v, ws_f)
            ws_f.cell(row=2, column=col).value = fecha_dt
            ws_f.cell(row=2, column=2).value = status_label

    lines.append(f"Columnas D:K desplazadas. Fecha: {fecha_periodo}  Estado: {status_label}")
    lines.append("")

    quarter = _mes_a_q(mes)
    qplan = RENTAS_NUEVO_SOURCE_BY_QUARTER.get(quarter, {})

    # ── Chañarcillo ──────────────────────────────────────────────────────────────
    chanar_path = _find_analisis_chanar_rn(mes, año)
    lines.append(f"Chanarcillo: {_sp_path(chanar_path) if chanar_path else 'NO ENCONTRADO'}")
    ws_chanar = _find_ws_rn(wb_forms, "Chañarcillo")
    if chanar_path and ws_chanar:
        try:
            tb = _read_trial_balance_rn(chanar_path, "Bce Tributario")
            if qplan.get(("Chañarcillo", "balance")) == "analisis":
                _apply_balance_map_rn(ws_chanar, tb, CHANAR_BALANCE_MAP, col)
                lines.append(f"  Balance: {len(tb)} cuentas OK")
            lines.append("  EERR: TODO (mapa filas pendiente)")
        except Exception as e:
            lines.append(f"  Error: {e}")
    elif not chanar_path:
        lines.append("  No actualizado: archivo no encontrado")
    lines.append("")

    # ── Curicó ───────────────────────────────────────────────────────────────────
    curico_path = _find_curico_informe_rn(mes, año)
    lines.append(f"Curico: {_sp_path(curico_path) if curico_path else 'NO ENCONTRADO'}")
    ws_curico = _find_ws_rn(wb_forms, "Curicó")
    if curico_path and ws_curico:
        try:
            curico_sheet = f"Acum {mes:02d}-{año}"
            tb = _read_trial_balance_rn(curico_path, curico_sheet)
            if qplan.get(("Curicó", "balance")) == "analisis":
                _apply_balance_map_rn(ws_curico, tb, CURICO_BALANCE_MAP, col)
                _apply_curico_impdif(ws_curico, tb, col)
                lines.append(f"  Balance: {len(tb)} cuentas OK (ImpDif neto aplicado)")
            lines.append("  EERR: TODO (mapa filas pendiente)")
        except Exception as e:
            lines.append(f"  Error: {e}")
    elif not curico_path:
        lines.append("  No actualizado: archivo no encontrado")
    lines.append("")

    # ── Inmob VC ─────────────────────────────────────────────────────────────────
    inmob_vc_path = _find_analisis_inmob_vc_rn(mes, año)
    lines.append(f"Inmob VC: {_sp_path(inmob_vc_path) if inmob_vc_path else 'NO ENCONTRADO'}")
    ws_inmob_vc = _find_ws_rn(wb_forms, "Inmob VC")
    if inmob_vc_path and ws_inmob_vc:
        try:
            tb = _read_trial_balance_rn(inmob_vc_path, "Bce Tributario")
            if qplan.get(("Inmob VC", "balance")) == "analisis":
                _apply_balance_map_rn(ws_inmob_vc, tb, INMOB_VC_BALANCE_MAP, col)
                lines.append(f"  Balance: {len(tb)} cuentas OK")
            if qplan.get(("Inmob VC", "eerr")) == "analisis":
                _apply_eerr_sa_map_rn(ws_inmob_vc, tb, col, INMOB_VC_EERR_MAP)
                lines.append(f"  EERR: {len(INMOB_VC_EERR_MAP)} filas escritas")
        except Exception as e:
            lines.append(f"  Error: {e}")
    elif not inmob_vc_path:
        lines.append("  No actualizado: archivo no encontrado")
    lines.append("")

    # ── Viña Centro ──────────────────────────────────────────────────────────────
    vina_path = _find_vina_trial_balance_rn(mes, año)
    lines.append(f"Vina Centro: {_sp_path(vina_path) if vina_path else 'NO ENCONTRADO'}")
    ws_vina = _find_ws_rn(wb_forms, "Viña Centro")
    if vina_path and ws_vina:
        try:
            tb = _read_trial_balance_rn(vina_path, "BALANCE ACUMULADO")
            if qplan.get(("Viña Centro", "balance")) == "analisis":
                _apply_balance_map_rn(ws_vina, tb, VINA_BALANCE_MAP, col)
                efectivo = (
                    _tb_sum(tb, ["1-1-01-020", "1-1-01-022", "1-1-01-023", "1-1-03-030"], "A")
                    - _tb_sum(tb, ["1-1-01-020", "1-1-01-022", "1-1-01-023", "1-1-03-030"], "P")
                )
                lines.append(f"  Balance: {len(tb)} cuentas OK (efectivo={efectivo:,.0f})")
            lines.append("  EERR: TODO (mapa filas pendiente)")
        except Exception as e:
            lines.append(f"  Error: {e}")
    elif not vina_path:
        lines.append("  No actualizado: archivo no encontrado")
    lines.append("")

    # ── Inmosa ───────────────────────────────────────────────────────────────────
    sa_path = _find_senior_assist_rn(mes, año)
    lines.append(f"Inmosa: {_sp_path(sa_path) if sa_path else 'NO ENCONTRADO'}")
    ws_inmosa = _find_ws_rn(wb_forms, "Inmosa")
    if sa_path and ws_inmosa:
        try:
            tb = _read_trial_balance_rn(sa_path)  # only sheet
            inmosa_bal_src = qplan.get(("Inmosa", "balance"), "eeff")
            if inmosa_bal_src == "analisis":
                lines.append("  Balance: TODO (mapa SA dot-notation pendiente)")
            else:
                lines.append(f"  Balance Q{quarter}: eeff (PDF, pendiente implementacion)")
            if qplan.get(("Inmosa", "eerr")) == "analisis":
                _apply_eerr_sa_map_rn(ws_inmosa, tb, col)
                lines.append(f"  EERR: {len(INMOSA_SA_EERR_MAP)} filas escritas desde Senior Assist")
        except Exception as e:
            lines.append(f"  Error: {e}")
    elif not sa_path:
        lines.append("  No actualizado: Senior Assist no encontrado")
    lines.append("")

    # ── Fondo Rentas ─────────────────────────────────────────────────────────────
    fondo_src = qplan.get(("Fondo Rentas", "balance"), "eeff")
    lines.append(f"Fondo Rentas: TODO (fuente={fondo_src}, parser PDF pendiente)")
    lines.append("")

    # ── Copiar hojas PT y Apoquindo ───────────────────────────────────────────────
    lines.append("=== Copiar hojas PT y Apoquindo ===")
    for src_kw, copy_pairs in (
        ("Rentas PT", [("Resumen", "Resumen PT"), ("Consolidado Fondo PT", "Consolidado Fondo PT")]),
        ("Apoquindo", [("Resumen", "Resumen  Apoquindo"), ("Consolidado Apoquindo", "Consolidado Apoquindo")]),
    ):
        src_hits = glob_module.glob(os.path.join(dest_dir, f"*{src_kw}*vAgente*.xlsx"))
        if not src_hits:
            src_hits = glob_module.glob(os.path.join(WORK_DIR, f"*{src_kw}*vAgente*.xlsx"))
        if src_hits:
            src_file = max(src_hits, key=os.path.getmtime)
            results = []
            for src_sn, dst_sn in copy_pairs:
                ok = _copy_vals_sheet_rn(src_file, src_sn, wb_forms, dst_sn)
                results.append(f"{src_sn}:{'OK' if ok else 'err'}")
            lines.append(f"  {src_kw}: {os.path.basename(src_file)} — {', '.join(results)}")
        else:
            lines.append(f"  {src_kw} vAgente no encontrado en {_sp_path(dest_dir)} ni en WORK_DIR")
    lines.append("")

    wb_vals.close()
    wb_forms.save(dest_path)
    wb_forms.close()
    lines.append(f"Guardado: {_sp_path(dest_path)}")
    lines.append("Abrir en Excel y verificar que las formulas recalculen correctamente.")

    return "\n".join(lines)
