from __future__ import annotations

import io

import openpyxl
import pytest

from tools.db import connection
from tools.db.connection import apply_migrations, get_conn_for


def _block(ws, row: int, name: str, saldo: float) -> int:
    ws.cell(row=row, column=3).value = name
    row += 2
    ws.cell(row=row, column=3).value = "Banco"
    ws.cell(row=row, column=5).value = "Saldo"
    row += 1
    ws.cell(row=row, column=3).value = "Chile"
    ws.cell(row=row, column=5).value = saldo
    return row + 2


_COMPANIES = [
    ("Toesca Rentas Inmobiliarias Fondo Inversión", 1000.0),
    ("Toesca Rentas Inmobiliarias PT Fondo Inversión", 300.0),
    ("Inmobiliaria VC SpA", 100.0),
    ("Inmobiliaria Chañarcillo Limitada", 100.0),
    ("Torre A S.A.", 900.0),
    ("Inmobiliaria Boulevard PT SPA", 300.0),
    ("Toesca Rentas Inmobiliarias Apoquindo Fondo Inversión", 200.0),
    ("Inmobiliaria Apoquindo S.A", 800.0),
    ("Viña Centro SpA", 500.0),
    ("POWER CENTER CURICÓ LTDA.", 100.0),
]


def _fixture_xlsx_bytes(sheet_name: str = "30-04-2026") -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet_name)
    row = 3
    for name, saldo in _COMPANIES:
        row = _block(ws, row, name, saldo)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def client(tmp_db_path, monkeypatch):
    apply_migrations(tmp_db_path)
    monkeypatch.setattr(connection, "DEFAULT_DB_PATH", tmp_db_path)
    from scripts import ingesta_server
    ingesta_server.app.config["TESTING"] = True
    with ingesta_server.app.test_client() as c:
        c.environ_base["HTTP_X_INGESTA_TOKEN"] = ingesta_server.API_TOKEN
        yield c


def test_caja_validate_no_escribe_en_db(client, tmp_db_path):
    data = {"file": (io.BytesIO(_fixture_xlsx_bytes()), "saldo_caja.xlsx"), "periodo": "2026-04"}
    res = client.post("/api/caja/validate", data=data, content_type="multipart/form-data")
    assert res.status_code == 200
    body = res.get_json()
    assert body["ok"] is True
    assert body["periodo"] == "2026-04"
    resumen = body["resumen"]
    assert resumen["tri_clp"] == pytest.approx(1000 + 300 / 3 + 100 + 100 + 900 / 3 + 300 / 3 + 200 * 0.3 + 800 * 0.3 + 500 + 100 * 0.8)

    con = get_conn_for(tmp_db_path)
    n = con.execute("SELECT COUNT(*) AS n FROM raw_caja").fetchone()["n"]
    con.close()
    assert n == 0  # validate es dry-run, no persiste


def test_caja_validate_sin_periodo(client):
    data = {"file": (io.BytesIO(_fixture_xlsx_bytes()), "saldo_caja.xlsx")}
    res = client.post("/api/caja/validate", data=data, content_type="multipart/form-data")
    body = res.get_json()
    assert body["ok"] is False
    assert "período" in body["errors"][0]


def test_caja_validate_periodo_sin_hoja_cercana(client):
    data = {"file": (io.BytesIO(_fixture_xlsx_bytes()), "saldo_caja.xlsx"), "periodo": "2019-01"}
    res = client.post("/api/caja/validate", data=data, content_type="multipart/form-data")
    body = res.get_json()
    assert body["ok"] is False
    assert body["errors"]


def test_caja_commit_persiste_y_es_idempotente(client, tmp_db_path):
    data = {"file": (io.BytesIO(_fixture_xlsx_bytes()), "saldo_caja.xlsx"), "periodo": "2026-04"}
    res = client.post("/api/caja/commit", data=data, content_type="multipart/form-data")
    assert res.status_code == 200
    body = res.get_json()
    assert body["ok"] is True
    assert body["filas_upsert"] == 3

    con = get_conn_for(tmp_db_path)
    rows = con.execute("SELECT fondo_key, saldo_clp FROM raw_caja WHERE fecha='2026-04-30'").fetchall()
    con.close()
    assert {r["fondo_key"] for r in rows} == {"Apo", "PT", "TRI"}

    data2 = {"file": (io.BytesIO(_fixture_xlsx_bytes()), "saldo_caja.xlsx"), "periodo": "2026-04"}
    res2 = client.post("/api/caja/commit", data=data2, content_type="multipart/form-data")
    assert res2.get_json()["filas_upsert"] == 3
    con = get_conn_for(tmp_db_path)
    n = con.execute("SELECT COUNT(*) AS n FROM raw_caja").fetchone()["n"]
    con.close()
    assert n == 3  # no duplica


def test_caja_periodo_check_sin_datos(client):
    res = client.get("/api/caja/periodo_check")
    assert res.status_code == 200
    assert res.get_json() == {"fondos": []}


def test_caja_validate_requiere_token(tmp_db_path, monkeypatch):
    apply_migrations(tmp_db_path)
    monkeypatch.setattr(connection, "DEFAULT_DB_PATH", tmp_db_path)
    from scripts import ingesta_server
    ingesta_server.app.config["TESTING"] = True
    with ingesta_server.app.test_client() as c:
        res = c.post(
            "/api/caja/validate",
            data={"file": (io.BytesIO(_fixture_xlsx_bytes()), "saldo_caja.xlsx")},
            content_type="multipart/form-data",
        )
    assert res.status_code == 401


def test_caja_validate_sin_archivo(client):
    res = client.post("/api/caja/validate", data={}, content_type="multipart/form-data")
    assert res.status_code == 200
    body = res.get_json()
    assert body["ok"] is False
    assert body["errors"]
