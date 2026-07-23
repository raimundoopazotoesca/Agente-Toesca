from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from tools.db.connection import apply_migrations
from tools.db import ingest_balance_consolidado


def _xlsx_balance() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "03-2026"
    blocks = [
        ("Fondo Toesca Rentas Inmobiliarias", 3),
        ("Fondo Toesca Rentas Inmobiliarias Apoquindo", 13),
        ("Fondo Toesca Rentas Inmobiliarias PT", 23),
    ]
    for title, row in blocks:
        ws.cell(row, 2, title)
        ws.cell(row + 1, 2, "BALANCE CONSOLIDADO (en miles de pesos)")
        is_apo = "Apoquindo" in title
        pasivo_diferido_label = None if is_apo else "Pasivos por Impuestos Diferidos"
        patrimonio = 2_300 if is_apo else 2_000
        rows = [
            ("Efectivo y Efectivo Equivalente", 100, "Préstamos Bancarios", 1_500),
            ("Otros Activos Corrientes", 200, pasivo_diferido_label, 300),
            ("Propiedades de Inversión", 3_000, "Otros Pasivos", 200),
            ("Activos Por Impuestos Diferidos", 700, "Patrimonio", patrimonio),
            ("Total Activos", 4_000, "Total Pasivos + Patrimonio", 4_000),
        ]
        for offset, values in enumerate(rows, start=2):
            ws.cell(row + offset, 2, values[0])
            ws.cell(row + offset, 3, values[1])
            if values[2] is not None:
                ws.cell(row + offset, 5, values[2])
                ws.cell(row + offset, 6, values[3])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


@pytest.fixture
def client(tmp_db_path, monkeypatch):
    apply_migrations(tmp_db_path)
    monkeypatch.setattr(ingest_balance_consolidado, "DB_PATH", tmp_db_path)
    from scripts import ingesta_server

    ingesta_server.app.config["TESTING"] = True
    with ingesta_server.app.test_client() as c:
        yield c


def test_balance_validate_endpoint_ok(client):
    data = {
        "periodo": "2026-03",
        "unidad": "M$",
        "file": (BytesIO(_xlsx_balance()), "balance.xlsx"),
    }
    res = client.post("/api/balance/validate", data=data, content_type="multipart/form-data")
    payload = res.get_json()
    assert res.status_code == 200
    assert payload["ok"] is True
    assert payload["n_lineas"] == 30


def test_balance_commit_endpoint_periodo_check(client):
    data = {
        "periodo": "2026-03",
        "unidad": "M$",
        "file": (BytesIO(_xlsx_balance()), "balance.xlsx"),
    }
    res = client.post("/api/balance/commit", data=data, content_type="multipart/form-data")
    payload = res.get_json()
    assert res.status_code == 200
    assert payload["ok"] is True
    assert payload["filas_insertadas"] == 30

    check = client.get("/api/balance/periodo_check?periodo=2026-03").get_json()
    assert check["ya_ingestado"] is True
    assert check["fondos"] == {"Apo": 10, "PT": 10, "TRI": 10}
