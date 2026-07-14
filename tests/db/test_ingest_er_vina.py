"""Tests para tools.db.ingest_er_vina."""
from __future__ import annotations

import datetime
import os
import sqlite3

import openpyxl
import pytest

from tools.db import ingest_er_vina as mod


# ── Fixture xlsx replicando la estructura real de RAW/NOI VIÑA.xlsx ────────
# Fila 6: fechas en col C..L (>=10 celdas fecha para pasar el umbral de
# detección de header). Solo las 2 primeras columnas (C, D) llevan datos
# reales; el resto son fechas dummy para completar el umbral.
# Col C = columna de label (código de cuenta), a partir de la fila ancla.
#
# Layout de columnas de datos: mismas columnas que la fila de fechas.
# Usamos columnas D y E como las 2 periodos reales (jan/feb-2025).

import calendar

_DATA_COL_1 = 4  # D
_DATA_COL_2 = 5  # E
_PERIODOS = ["2025-01", "2025-02"]
_FECHAS = {_DATA_COL_1: datetime.datetime(2025, 1, 31), _DATA_COL_2: datetime.datetime(2025, 2, 28)}

# El parser detecta la fila de fechas exigiendo >=10 celdas tipo fecha (para
# no confundirla con celdas sueltas). Rellenamos 8 columnas "dummy" con
# meses distintos a los 2 periodos reales, sin datos en las cuentas — solo
# para pasar el umbral. Cada mes dummy necesita su propia UF sembrada.
_DUMMY_MESES = [(2024, m) for m in range(5, 13)]  # 2024-05 .. 2024-12


def _fin_mes(year: int, month: int) -> datetime.date:
    return datetime.date(year, month, calendar.monthrange(year, month)[1])


_UF = {"2025-01-31": 37000.0, "2025-02-28": 37200.0}
_UF.update({_fin_mes(y, m).isoformat(): 36000.0 + i * 10 for i, (y, m) in enumerate(_DUMMY_MESES)})

# Ingreso Explotación: 2 cuentas
_ING_1 = {_DATA_COL_1: 100_000_000.0, _DATA_COL_2: 110_000_000.0}
_ING_2 = {_DATA_COL_1: 20_000_000.0, _DATA_COL_2: 21_000_000.0}

# Ingreso Fuera De Explotación: 1 cuenta (debe quedar excluida del NOI)
_ING_FUERA = {_DATA_COL_1: 5_000_000.0, _DATA_COL_2: 5_500_000.0}

# Gastos de Administración y Ventas: 2 cuentas (ya negativas en la fuente)
_GASTO_1 = {_DATA_COL_1: -15_000_000.0, _DATA_COL_2: -16_000_000.0}
_GASTO_2 = {_DATA_COL_1: -8_000_000.0, _DATA_COL_2: -8_500_000.0}

_TOTAL_RESULTADO_OPERACION = {c: _ING_1[c] + _ING_2[c] for c in (_DATA_COL_1, _DATA_COL_2)}
_TOTAL_GASTOS_ADMIN = {c: _GASTO_1[c] + _GASTO_2[c] for c in (_DATA_COL_1, _DATA_COL_2)}


def _build_fixture_xlsx(tmp_path, corrupt_total_gastos=False) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # Fila 6: fechas. 2 columnas reales + 8 dummy para pasar el umbral (>=10).
    for col, fecha in _FECHAS.items():
        ws.cell(row=6, column=col).value = fecha
    for col, (y, m) in zip(range(6, 14), _DUMMY_MESES):
        d = _fin_mes(y, m)
        ws.cell(row=6, column=col).value = datetime.datetime(d.year, d.month, d.day)

    # Bloque de input manual, arranca fila 124 (col C = label/código cuenta).
    r = 124
    ws.cell(row=r, column=3).value = "Ingreso de Explotacion"
    r += 1
    ws.cell(row=r, column=3).value = "4-1-01-100  ARRIENDO TIENDAS ANCLAS"
    for col, v in _ING_1.items():
        ws.cell(row=r, column=col).value = v
    r += 1
    ws.cell(row=r, column=3).value = "4-1-01-101  ARRIENDO TIENDAS MENORES"
    for col, v in _ING_2.items():
        ws.cell(row=r, column=col).value = v
    r += 1
    ws.cell(row=r, column=3).value = "Total Resultado Operación"
    for col, v in _TOTAL_RESULTADO_OPERACION.items():
        ws.cell(row=r, column=col).value = v
    r += 1
    ws.cell(row=r, column=3).value = "Ingreso Fuera De Explotacion"
    r += 1
    ws.cell(row=r, column=3).value = "4-2-01-002  OTROS INGRESOS"
    for col, v in _ING_FUERA.items():
        ws.cell(row=r, column=col).value = v
    r += 1
    ws.cell(row=r, column=3).value = "Total Ingreso Fuera De Explotacion"
    r += 1
    ws.cell(row=r, column=3).value = "GASTOS DE ADMINISTRACIÓN Y VENTAS"
    r += 1
    ws.cell(row=r, column=3).value = "3-1-10-102  FEE ADMINISTRATIVO (REMUNERACIONES)"
    for col, v in _GASTO_1.items():
        ws.cell(row=r, column=col).value = v
    r += 1
    ws.cell(row=r, column=3).value = "3-1-40-102  CONTRIBUCIONES"
    for col, v in _GASTO_2.items():
        ws.cell(row=r, column=col).value = v
    r += 1
    ws.cell(row=r, column=3).value = "Total Gastos de administración y ventas"
    gastos_vals = dict(_TOTAL_GASTOS_ADMIN)
    if corrupt_total_gastos:
        gastos_vals = {c: v + 999_999 for c, v in gastos_vals.items()}
    for col, v in gastos_vals.items():
        ws.cell(row=r, column=col).value = v
    r += 1
    ws.cell(row=r, column=3).value = "Total Operacional"

    os.makedirs(tmp_path, exist_ok=True)
    path = os.path.join(tmp_path, "vina_fixture.xlsx")
    wb.save(path)
    return path


@pytest.fixture
def conn(tmp_path):
    """DB en disco (tmp) con schema mínimo necesario (mismo patrón que
    tests/db/test_ingest_er_sucden.py: sqlite3.connect plano, sin FK
    enforcement ni migraciones completas)."""
    db_path = os.path.join(str(tmp_path), "test.db")
    c = sqlite3.connect(db_path)
    c.row_factory = sqlite3.Row
    c.executescript("""
        CREATE TABLE ingest_run (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tool TEXT, source_file TEXT, file_hash TEXT,
            rows_in INTEGER, rows_loaded INTEGER,
            started_at TEXT, ended_at TEXT, status TEXT, error TEXT
        );
        CREATE TABLE dim_activo (
            activo_key TEXT PRIMARY KEY, fondo_key TEXT, nombre TEXT,
            tipo TEXT, participacion_fondo_activo REAL, categoria TEXT, sociedad TEXT
        );
        INSERT INTO dim_activo (activo_key, fondo_key, nombre, participacion_fondo_activo)
             VALUES ('Viña Centro','TRI','Viña Centro',1.0);
        CREATE TABLE raw_er_activo_line (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            activo_key TEXT NOT NULL REFERENCES dim_activo(activo_key),
            periodo TEXT NOT NULL,
            cuenta_codigo TEXT, cuenta_nombre TEXT,
            monto_clp REAL, monto_uf REAL,
            seccion TEXT, es_operacional INTEGER,
            source_file TEXT, source_sheet TEXT, source_row INTEGER,
            file_hash TEXT, ingest_run_id INTEGER REFERENCES ingest_run(id),
            loaded_at TEXT DEFAULT (datetime('now')),
            superseded_at TEXT
        );
        CREATE TABLE raw_uf_diaria (
            fecha TEXT PRIMARY KEY, valor REAL, fuente TEXT,
            loaded_at TEXT DEFAULT (datetime('now'))
        );
        CREATE VIEW fact_uf AS SELECT fecha, valor FROM raw_uf_diaria ORDER BY fecha;
    """)
    for fecha, valor in _UF.items():
        c.execute(
            "INSERT INTO raw_uf_diaria (fecha, valor, fuente) VALUES (?, ?, 'test')",
            (fecha, valor),
        )
    c.commit()
    yield c
    c.close()


@pytest.fixture
def fixture_xlsx(tmp_path):
    return _build_fixture_xlsx(str(tmp_path))


# ── Tests de parsing ─────────────────────────────────────────────────────

def test_parse_devuelve_5_cuentas_x_10_periodos(fixture_xlsx, conn):
    # 5 cuentas (2 ingreso explotación + 1 fuera explotación + 2 gastos) x
    # 10 columnas de fecha detectadas (2 con datos reales + 8 dummy en 0).
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    assert len(rows) == 50


def test_parse_activo_key_fijo_vina(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    assert all(r["activo_key"] == "Viña Centro" for r in rows)


def test_parse_periodos_yyyy_mm(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    assert set(_PERIODOS) <= {r["periodo"] for r in rows}


def test_parse_codigos_de_cuenta_extraidos_por_regex(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    codigos = {r["cuenta_codigo"] for r in rows}
    assert codigos == {"4-1-01-100", "4-1-01-101", "4-2-01-002", "3-1-10-102", "3-1-40-102"}


def test_parse_seccion_y_es_operacional(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    by_codigo = {r["cuenta_codigo"]: r for r in rows if r["periodo"] == "2025-01"}
    assert by_codigo["4-1-01-100"]["seccion"] == "INGRESOS_OPERACION"
    assert by_codigo["4-1-01-100"]["es_operacional"] == 1
    assert by_codigo["4-2-01-002"]["seccion"] == "INGRESO_FUERA_EXPLOTACION"
    assert by_codigo["4-2-01-002"]["es_operacional"] == 0
    assert by_codigo["3-1-10-102"]["seccion"] == "GASTOS_OPERACION"
    assert by_codigo["3-1-10-102"]["es_operacional"] == 1


def test_parse_monto_uf_usa_fact_uf(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    r = next(r for r in rows if r["cuenta_codigo"] == "4-1-01-100" and r["periodo"] == "2025-01")
    assert r["monto_clp"] == _ING_1[_DATA_COL_1]
    assert abs(r["monto_uf"] - _ING_1[_DATA_COL_1] / _UF["2025-01-31"]) < 1e-6


def test_parse_noi_excluye_ingreso_fuera_de_explotacion(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    noi_2025_01 = sum(r["monto_uf"] for r in rows if r["periodo"] == "2025-01" and r["es_operacional"] == 1)
    esperado = (_ING_1[_DATA_COL_1] + _ING_2[_DATA_COL_1] + _GASTO_1[_DATA_COL_1] + _GASTO_2[_DATA_COL_1]) / _UF["2025-01-31"]
    assert abs(noi_2025_01 - esperado) < 1e-6
    # confirmar que sin excluir fuera-explotación el NOI sería distinto (mayor)
    assert abs(noi_2025_01 - (esperado + _ING_FUERA[_DATA_COL_1] / _UF["2025-01-31"])) > 1e-6


def test_parse_valida_integridad_ingreso_explotacion(fixture_xlsx, conn):
    # No debe lanzar: el fixture ya cuadra por construcción.
    mod.parse_planilla(fixture_xlsx, conn=conn)


def test_parse_falla_si_total_gastos_no_cuadra(tmp_path, conn):
    path = _build_fixture_xlsx(str(tmp_path), corrupt_total_gastos=True)
    with pytest.raises(ValueError, match="Validación de integridad"):
        mod.parse_planilla(path, conn=conn)


# ── Tests de override de datos faltantes ────────────────────────────────

def test_override_seguridad_parking_aplica_valor_fijo(tmp_path, conn):
    """Cuenta 3-1-10-120 en periodo 2025-07 debe usar el override, no la
    fuente (que en este fixture está vacía/None, como en el archivo real)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    col = 4
    fecha = datetime.datetime(2025, 7, 31)
    ws.cell(row=6, column=col).value = fecha
    for c, (y, m) in zip(range(5, 14), _DUMMY_MESES + [(2024, 4)]):
        d = _fin_mes(y, m)
        ws.cell(row=6, column=c).value = datetime.datetime(d.year, d.month, d.day)

    r = 124
    ws.cell(row=r, column=3).value = "Ingreso de Explotacion"
    r += 1
    ws.cell(row=r, column=3).value = "4-1-01-100  ARRIENDO TIENDAS ANCLAS"
    ws.cell(row=r, column=col).value = 100_000_000.0
    r += 1
    ws.cell(row=r, column=3).value = "Total Resultado Operación"
    ws.cell(row=r, column=col).value = 100_000_000.0
    r += 1
    ws.cell(row=r, column=3).value = "Ingreso Fuera De Explotacion"
    r += 1
    ws.cell(row=r, column=3).value = "Total Ingreso Fuera De Explotacion"
    ws.cell(row=r, column=col).value = 0.0
    r += 1
    ws.cell(row=r, column=3).value = "GASTOS DE ADMINISTRACIÓN Y VENTAS"
    r += 1
    ws.cell(row=r, column=3).value = "3-1-10-120  SEGURIDAD PARKING"
    # celda vacía a propósito (None), como en el archivo real jul-2025
    r += 1
    ws.cell(row=r, column=3).value = "Total Gastos de administración y ventas"
    ws.cell(row=r, column=col).value = -57_551_335.0
    r += 1
    ws.cell(row=r, column=3).value = "Total Operacional"

    os.makedirs(str(tmp_path), exist_ok=True)
    path = os.path.join(str(tmp_path), "override_fixture.xlsx")
    wb.save(path)

    conn.execute(
        "INSERT INTO raw_uf_diaria (fecha, valor, fuente) VALUES (?, ?, ?)",
        ("2025-07-31", 39000.0, "test"),
    )
    conn.execute(
        "INSERT INTO raw_uf_diaria (fecha, valor, fuente) VALUES (?, ?, ?)",
        ("2024-04-30", 36000.0, "test"),
    )
    conn.commit()

    rows = mod.parse_planilla(path, conn=conn)
    r_seguridad = next(r for r in rows if r["cuenta_codigo"] == "3-1-10-120")
    assert r_seguridad["monto_clp"] == -57_551_335.0


# ── Tests de persistencia (idempotencia) ────────────────────────────────

def test_persist_inserta_filas(fixture_xlsx, conn):
    res = mod.persist(fixture_xlsx, conn=conn)
    assert res["status"] == "inserted"
    assert res["rows"] == 50


def test_persist_es_idempotente(fixture_xlsx, conn):
    mod.persist(fixture_xlsx, conn=conn)
    res2 = mod.persist(fixture_xlsx, conn=conn)
    assert res2["status"] == "skipped_idempotent"
    assert res2["rows"] == 0


def test_persist_supersede_en_reingesta_con_cambios(tmp_path, conn):
    path1 = _build_fixture_xlsx(str(tmp_path) + "_1")
    res1 = mod.persist(path1, conn=conn)
    assert res1["status"] == "inserted"

    # Segunda versión del archivo con un valor distinto -> hash distinto.
    wb = openpyxl.load_workbook(path1)
    ws = wb["Hoja1"]
    ws.cell(row=125, column=_DATA_COL_1).value = 999_000_000.0
    path2 = os.path.join(str(tmp_path), "vina_fixture_v2.xlsx")
    os.makedirs(str(tmp_path), exist_ok=True)
    # ajustar el Total Resultado Operación para que siga cuadrando
    ws.cell(row=127, column=_DATA_COL_1).value = 999_000_000.0 + _ING_2[_DATA_COL_1]
    wb.save(path2)

    res2 = mod.persist(path2, conn=conn)
    assert res2["status"] == "superseded_and_reinserted"

    activos = conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE activo_key='Viña Centro' AND superseded_at IS NULL"
    ).fetchone()[0]
    assert activos == 50
