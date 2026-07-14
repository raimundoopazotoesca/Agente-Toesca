"""Tests para tools.db.ingest_er_apo3001."""
from __future__ import annotations

import os
import sqlite3

import openpyxl
import pytest

from tools.db import ingest_er_apo3001 as mod


# ── Fixture xlsx replicando la estructura real de RAW/NOI 3001.xlsx ────────
# Fila 4: ancla "Apoquindo 3001" + header de fechas en la MISMA fila.
# Fila 5: agregado "(+) Ingresos por Arriendos" (se descarta).
# Filas 6-7: Taipei / Otros (sub-detalle, se usan).
# Filas 8-13: gastos. Fila 14: NOI Mensual.
# Valores reales de ene/feb/mar-2020 tomados del archivo fuente.

_PERIODOS = ["2020-01", "2020-02", "2020-03"]
_FECHAS = [
    __import__("datetime").datetime(2020, 1, 31),
    __import__("datetime").datetime(2020, 2, 29),
    __import__("datetime").datetime(2020, 3, 31),
]
_INGRESOS_AGREGADO = [2650.8714760577295, 2650.8714760577295, 2660.9636914047614]
_TAIPEI = [359.0314760577296, 359.0314760577296, 369.12369140476113]
_OTROS = [2291.84, 2291.84, 2291.84]
_GASTOS_COMUNES = [-79.83642250315387, -79.48463778564043, -78.95208175831]
_ADMIN = [-21.206971808461837, -21.206971808461837, -21.287709531238093]
_CONTRIB_SOBRETASA = [-204.41666666666666, -204.41666666666666, -204.41666666666666]
_SEGUROS = [None, -29.405833333333334, -29.405833333333334]
_NOI_ESPERADO = [2345.4114150794476, 2316.3573664636274, 2326.901400115213]


def _build_fixture_xlsx(tmp_path, corrupt_noi: bool = False,
                         unknown_label: bool = False,
                         desalinear_agregado: bool = False) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # Fila 4: ancla + header de fechas en la misma fila (col B en adelante)
    ws.cell(row=4, column=1).value = "Apoquindo 3001"
    for j, f in enumerate(_FECHAS, start=2):
        ws.cell(row=4, column=j).value = f

    # Fila 5: agregado (se descarta siempre, incluso si desalineado)
    ws.cell(row=5, column=1).value = "(+) Ingresos por Arriendos"
    agregado = list(_INGRESOS_AGREGADO)
    if desalinear_agregado:
        agregado[0] += 0.5  # simula el redondeo obsoleto real (2026-03/04)
    for j, v in enumerate(agregado, start=2):
        ws.cell(row=5, column=j).value = v

    # Fila 6-7: Taipei / Otros
    ws.cell(row=6, column=1).value = "unknown label xyz" if unknown_label else "Taipei"
    for j, v in enumerate(_TAIPEI, start=2):
        ws.cell(row=6, column=j).value = v
    ws.cell(row=7, column=1).value = "Otros"
    for j, v in enumerate(_OTROS, start=2):
        ws.cell(row=7, column=j).value = v

    # Fila 8: Gastos Comunes
    ws.cell(row=8, column=1).value = "(-) Gastos Comunes"
    for j, v in enumerate(_GASTOS_COMUNES, start=2):
        ws.cell(row=8, column=j).value = v
    # Fila 9: Administración
    ws.cell(row=9, column=1).value = "(-) Administración"
    for j, v in enumerate(_ADMIN, start=2):
        ws.cell(row=9, column=j).value = v
    # Fila 10: Comisión Corredor (vacía en este fixture, como en el ejemplo real)
    ws.cell(row=10, column=1).value = "(-) Comisión Corredor"
    # Fila 11: Provision Incobrables (vacía)
    ws.cell(row=11, column=1).value = "(-) Provision Incobrables"
    # Fila 12: Contribuciones + Sobretasa
    ws.cell(row=12, column=1).value = "(-) Contribuciones + Sobretasa"
    for j, v in enumerate(_CONTRIB_SOBRETASA, start=2):
        ws.cell(row=12, column=j).value = v
    # Fila 13: Seguros (None en enero-2020, como en el archivo real)
    ws.cell(row=13, column=1).value = "(-) Seguros"
    for j, v in enumerate(_SEGUROS, start=2):
        if v is not None:
            ws.cell(row=13, column=j).value = v

    # Fila 14: NOI Mensual (control) — calculado desde Taipei+Otros+gastos,
    # NO desde el agregado (replica el hallazgo real).
    ws.cell(row=14, column=1).value = "NOI Mensual"
    noi_vals = list(_NOI_ESPERADO)
    if corrupt_noi:
        noi_vals[0] += 1000  # rompe la validación de integridad a propósito
    for j, v in enumerate(noi_vals, start=2):
        ws.cell(row=14, column=j).value = v

    os.makedirs(tmp_path, exist_ok=True)
    path = os.path.join(tmp_path, "apo3001_fixture.xlsx")
    wb.save(path)
    return path


@pytest.fixture
def fixture_xlsx(tmp_path):
    return _build_fixture_xlsx(str(tmp_path))


# ── Tests de parsing ─────────────────────────────────────────────────────

def test_parse_devuelve_24_filas(fixture_xlsx):
    # 8 categorías usadas (Taipei, Otros, Gastos Comunes, Administración,
    # Comisión Corredor, Provision Incobrables, Contribuciones+Sobretasa,
    # Seguros) x 3 meses. El agregado "(+) Ingresos por Arriendos" no cuenta.
    rows = mod.parse_planilla(fixture_xlsx)
    assert len(rows) == 24


def test_parse_activo_key_fijo_apo3001(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    assert all(r["activo_key"] == "Apo3001" for r in rows)


def test_parse_periodos_yyyy_mm(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    assert {r["periodo"] for r in rows} == set(_PERIODOS)


def test_parse_pseudo_codigos_completos(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    codigos = {r["cuenta_codigo"] for r in rows}
    esperados = {
        "APO3001_ING_TAIPEI", "APO3001_ING_OTROS", "APO3001_GC", "APO3001_ADM",
        "APO3001_COM_CORR", "APO3001_PROV_INCOB", "APO3001_CONTRIB_SOBRETASA",
        "APO3001_SEG",
    }
    assert codigos == esperados


def test_parse_agregado_descartado(fixture_xlsx):
    """La fila '(+) Ingresos por Arriendos' (agregada) nunca se persiste —
    solo sus sub-detalles Taipei y Otros."""
    rows = mod.parse_planilla(fixture_xlsx)
    assert not any(r["cuenta_nombre"] == "(+) Ingresos por Arriendos" for r in rows)
    montos_agregado = {round(v, 4) for v in _INGRESOS_AGREGADO}
    montos_persistidos = {round(r["monto_clp"], 4) for r in rows}
    assert not montos_agregado & montos_persistidos


def test_parse_agregado_desalineado_no_rompe_integridad(tmp_path):
    """Replica el hallazgo real: el agregado puede diferir de Taipei+Otros
    (redondeo obsoleto en la fuente) sin que la validación de integridad
    falle, porque el parser ignora el agregado y usa el sub-detalle."""
    path = _build_fixture_xlsx(str(tmp_path), desalinear_agregado=True)
    rows = mod.parse_planilla(path)  # no debe lanzar
    assert len(rows) == 24


def test_parse_todas_operacionales(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    assert all(r["es_operacional"] == 1 for r in rows)


def test_parse_valores_reales_taipei_otros(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    taipei_by_periodo = {r["periodo"]: r["monto_clp"] for r in rows if r["cuenta_codigo"] == "APO3001_ING_TAIPEI"}
    otros_by_periodo = {r["periodo"]: r["monto_clp"] for r in rows if r["cuenta_codigo"] == "APO3001_ING_OTROS"}
    for periodo, esperado in zip(_PERIODOS, _TAIPEI):
        assert abs(taipei_by_periodo[periodo] - esperado) < 1e-6
    for periodo, esperado in zip(_PERIODOS, _OTROS):
        assert abs(otros_by_periodo[periodo] - esperado) < 1e-6


def test_parse_categoria_desconocida_falla(tmp_path):
    path = _build_fixture_xlsx(str(tmp_path), unknown_label=True)
    with pytest.raises(ValueError, match=r"(?i)categor[ií]a"):
        mod.parse_planilla(path)


def test_parse_validacion_integridad_falla_si_no_cuadra(tmp_path):
    path = _build_fixture_xlsx(str(tmp_path), corrupt_noi=True)
    with pytest.raises(ValueError, match=r"(?i)noi"):
        mod.parse_planilla(path)


def test_parse_source_metadata(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    assert all(r["source_file"] == fixture_xlsx for r in rows)
    assert all(r["source_sheet"] == "Hoja1" for r in rows)
    assert all(isinstance(r["source_row"], int) and r["source_row"] > 0 for r in rows)


def test_file_hash_estable(fixture_xlsx):
    h1 = mod._file_hash(fixture_xlsx)
    h2 = mod._file_hash(fixture_xlsx)
    assert h1 == h2
    assert len(h1) == 64


# ── Tests de persistencia ────────────────────────────────────────────────

@pytest.fixture
def db_conn(tmp_path):
    db_path = os.path.join(str(tmp_path), "test.db")
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        CREATE TABLE ingest_run (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tool TEXT, source_file TEXT, file_hash TEXT,
            rows_in INTEGER, rows_loaded INTEGER,
            started_at TEXT, ended_at TEXT, status TEXT, error TEXT
        );
        CREATE TABLE dim_activo (
            activo_key TEXT PRIMARY KEY, fondo_key TEXT, nombre TEXT,
            tipo TEXT, participacion_fondo_activo REAL, categoria TEXT, sociedad TEXT
        );
        INSERT INTO dim_activo (activo_key, fondo_key, nombre, participacion_fondo_activo)
             VALUES ('Apo3001','TRI','Apoquindo 3001',0.685);
        CREATE TABLE raw_er_activo_line (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            activo_key TEXT NOT NULL REFERENCES dim_activo(activo_key),
            periodo TEXT NOT NULL,
            cuenta_codigo TEXT, cuenta_nombre TEXT,
            monto_clp REAL, monto_uf REAL,
            seccion TEXT, es_operacional INTEGER,
            source_file TEXT, source_sheet TEXT, source_row INTEGER,
            file_hash TEXT, ingest_run_id INTEGER REFERENCES ingest_run(id),
            loaded_at TEXT DEFAULT (datetime('now')),
            superseded_at TEXT
        );
    """)
    conn.commit()
    yield conn
    conn.close()


def test_persist_inserta_24_filas(fixture_xlsx, db_conn):
    res = mod.persist(fixture_xlsx, conn=db_conn)
    assert res["status"] == "inserted"
    assert res["rows"] == 24
    n = db_conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE superseded_at IS NULL"
    ).fetchone()[0]
    assert n == 24


def test_persist_idempotente_mismo_hash(fixture_xlsx, db_conn):
    mod.persist(fixture_xlsx, conn=db_conn)
    res2 = mod.persist(fixture_xlsx, conn=db_conn)
    assert res2["status"] == "skipped_idempotent"
    n = db_conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE superseded_at IS NULL"
    ).fetchone()[0]
    assert n == 24


def test_persist_reingesta_supersede_previas(fixture_xlsx, tmp_path, db_conn):
    mod.persist(fixture_xlsx, conn=db_conn)
    fixture_xlsx_2 = _build_fixture_xlsx(str(tmp_path / "sub"))
    wb = openpyxl.load_workbook(fixture_xlsx_2)
    wb["Hoja1"].cell(row=4, column=1).value = "Apoquindo 3001 "  # cambia hash
    wb.save(fixture_xlsx_2)

    res = mod.persist(fixture_xlsx_2, conn=db_conn)
    assert res["status"] == "superseded_and_reinserted"
    total = db_conn.execute("SELECT COUNT(*) FROM raw_er_activo_line").fetchone()[0]
    activas = db_conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE superseded_at IS NULL"
    ).fetchone()[0]
    assert activas == 24
    assert total - activas == 24


def test_noi_derivado_matchea_suma_esperada(fixture_xlsx, db_conn):
    mod.persist(fixture_xlsx, conn=db_conn)
    for periodo, esperado in zip(_PERIODOS, _NOI_ESPERADO):
        calc = db_conn.execute("""
            SELECT SUM(monto_clp) FROM raw_er_activo_line
             WHERE activo_key='Apo3001' AND periodo=?
               AND es_operacional=1 AND superseded_at IS NULL
        """, (periodo,)).fetchone()[0]
        assert abs(calc - esperado) < 0.01, f"{periodo}: {calc} != {esperado}"


def test_persist_falla_no_escribe_nada_si_integridad_no_cuadra(tmp_path, db_conn):
    path = _build_fixture_xlsx(str(tmp_path), corrupt_noi=True)
    with pytest.raises(ValueError, match=r"(?i)noi"):
        mod.persist(path, conn=db_conn)
    n = db_conn.execute("SELECT COUNT(*) FROM raw_er_activo_line").fetchone()[0]
    assert n == 0


# ── Test de integración de solo-lectura contra el archivo real ──────────

_REAL_XLSX = (
    r"C:\Users\raimundo.opazo\OneDrive - Toesca\Inmobiliario Toesca - Documentos"
    r"\RAW\NOI 3001.xlsx"
)


def _copy_real_file(dst: str) -> bool:
    import subprocess
    if not os.path.exists(_REAL_XLSX):
        return False
    try:
        subprocess.run(["cp", _REAL_XLSX, dst], check=True, capture_output=True)
        return os.path.exists(dst) and os.path.getsize(dst) > 0
    except (subprocess.CalledProcessError, OSError, FileNotFoundError):
        return False


def _real_file_accessible(tmp_path_str: str) -> bool:
    probe = os.path.join(tmp_path_str, "_probe.xlsx")
    ok = _copy_real_file(probe)
    if ok and os.path.exists(probe):
        os.remove(probe)
    return ok


@pytest.mark.skipif(
    not _real_file_accessible(str(__import__("tempfile").mkdtemp())),
    reason="archivo real no disponible o locked en este entorno",
)
def test_parse_archivo_real_no_lanza_y_cuadra_integridad(tmp_path):
    local_copy = os.path.join(str(tmp_path), "real.xlsx")
    assert _copy_real_file(local_copy), "no se pudo copiar el archivo real pese a pasar el skipif"

    rows = mod.parse_planilla(local_copy)
    assert len(rows) > 0
    periodos = {r["periodo"] for r in rows}
    assert "2020-01" in periodos
    assert "2026-05" in periodos
    assert len(periodos) == 77
