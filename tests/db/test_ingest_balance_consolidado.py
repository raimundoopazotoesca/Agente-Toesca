from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook

from tools.db.connection import apply_migrations
from tools.db import ingest_balance_consolidado as balance


def _xlsx_balance() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "03-2026"
    blocks = [
        ("Fondo Toesca Rentas Inmobiliarias", 3),
        ("Fondo Toesca Rentas Inmobiliarias Apoquindo", 13),
        ("Fondo Toesca Rentas Inmobiliarias PT", 23),
    ]
    for title, row in blocks:
        ws.cell(row, 2, title)
        ws.cell(row + 1, 2, "BALANCE CONSOLIDADO (en miles de pesos)")
        is_apo = "Apoquindo" in title
        pasivo_diferido_label = None if is_apo else "Pasivos por Impuestos Diferidos"
        patrimonio = 2_300 if is_apo else 2_000
        rows = [
            ("Efectivo y Efectivo Equivalente", 100, "Préstamos Bancarios", 1_500),
            ("Otros Activos Corrientes", 200, pasivo_diferido_label, 300),
            ("Propiedades de Inversión", 3_000, "Otros Pasivos", 200),
            ("Activos Por Impuestos Diferidos", 700, "Patrimonio", patrimonio),
            ("Total Activos", 4_000, "Total Pasivos + Patrimonio", 4_000),
        ]
        for offset, values in enumerate(rows, start=2):
            ws.cell(row + offset, 2, values[0])
            ws.cell(row + offset, 3, values[1])
            if values[2] is not None:
                ws.cell(row + offset, 5, values[2])
                ws.cell(row + offset, 6, values[3])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _xlsx_balance_descuadrado() -> bytes:
    wb = load_workbook(BytesIO(_xlsx_balance()))
    ws = wb["03-2026"]
    ws["F9"] = 3_900
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def test_validate_balance_consolidado_ok(tmp_db_path, monkeypatch):
    apply_migrations(tmp_db_path)
    monkeypatch.setattr(balance, "DB_PATH", tmp_db_path)

    result = balance.validate(_xlsx_balance(), "balance.xlsx", "2026-03", "M$")

    assert result.ok is True
    assert result.data["n_lineas"] == 30
    tri = next(f for f in result.data["fondos"] if f["fondo_key"] == "TRI")
    total = next(r for r in tri["rows"] if r["cuenta_codigo"] == "ESF.total_activo")
    assert total["monto_clp"] == 4_000_000
    assert tri["balance_check"]["ok"] is True


def test_validate_balance_consolidado_rechaza_descuadre(tmp_db_path, monkeypatch):
    apply_migrations(tmp_db_path)
    monkeypatch.setattr(balance, "DB_PATH", tmp_db_path)

    result = balance.validate(_xlsx_balance_descuadrado(), "balance.xlsx", "2026-03", "M$")

    assert result.ok is False
    assert any("descuadrado" in e.lower() for e in result.errors)


def test_validate_balance_consolidado_periodo_sin_hoja_mensaje_simple(tmp_db_path, monkeypatch):
    apply_migrations(tmp_db_path)
    monkeypatch.setattr(balance, "DB_PATH", tmp_db_path)

    result = balance.validate(_xlsx_balance(), "balance.xlsx", "2026-06", "M$")

    assert result.ok is False
    assert result.errors == ["No existe hoja '06-2026' en la planilla."]


def test_validate_balance_consolidado_delta_vs_periodo_anterior(tmp_db_path, monkeypatch):
    apply_migrations(tmp_db_path)
    monkeypatch.setattr(balance, "DB_PATH", tmp_db_path)

    import sqlite3

    con = sqlite3.connect(tmp_db_path)
    try:
        con.execute(
            "INSERT INTO raw_balance_consolidado_line "
            "(fondo_key, periodo, cuenta_codigo, monto_clp, source_file) "
            "VALUES ('TRI', '2025-12', 'ESF.total_activo', 2000000, 'test')"
        )
        con.commit()
    finally:
        con.close()

    result = balance.validate(_xlsx_balance(), "balance.xlsx", "2026-03", "M$")

    tri = next(f for f in result.data["fondos"] if f["fondo_key"] == "TRI")
    total = next(r for r in tri["rows"] if r["cuenta_codigo"] == "ESF.total_activo")
    assert tri["periodo_anterior"] == "2025-12"
    assert total["delta_pct"] == 100.0


def test_commit_balance_consolidado_supersede_periodo(tmp_db_path, monkeypatch):
    apply_migrations(tmp_db_path)
    monkeypatch.setattr(balance, "DB_PATH", tmp_db_path)
    xlsx = _xlsx_balance()

    first = balance.commit(xlsx, "balance.xlsx", "2026-03", "M$")
    second = balance.commit(xlsx, "balance.xlsx", "2026-03", "M$")

    assert first["filas_insertadas"] == 30
    assert second["filas_insertadas"] == 30
    assert second["filas_superseded"] == 30

    import sqlite3

    con = sqlite3.connect(tmp_db_path)
    try:
        rows = con.execute(
            "SELECT fondo_key, COUNT(*) FROM raw_balance_consolidado_line "
            "WHERE periodo='2026-03' AND superseded_at IS NULL GROUP BY fondo_key"
        ).fetchall()
    finally:
        con.close()
    assert sorted(rows) == [("Apo", 10), ("PT", 10), ("TRI", 10)]
