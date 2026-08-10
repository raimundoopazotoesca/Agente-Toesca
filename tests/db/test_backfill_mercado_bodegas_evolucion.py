"""Tests para tools.db.backfill_mercado_bodegas_evolucion."""
from __future__ import annotations

import openpyxl
import pytest

from tools.db import backfill_mercado_bodegas_evolucion as mod
from tools.db.connection import apply_migrations, get_conn_for


@pytest.fixture
def xlsx_fixture(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws["D3"] = "UF/m2"
    ws["E3"] = "Vacancia"
    filas = [
        ("2S-2015", 0.119, 0.0995),
        ("1S-2016", 0.118, 0.1228),
        ("2S-2016", 0.109, 0.0749),
    ]
    for i, (semestre, uf, vac) in enumerate(filas, start=4):
        ws[f"C{i}"] = semestre
        ws[f"D{i}"] = uf
        ws[f"E{i}"] = vac
    path = tmp_path / "mercado_bodegas.xlsx"
    wb.save(path)
    return str(path)


def test_ingest_inserta_filas(tmp_path, xlsx_fixture):
    db_path = str(tmp_path / "test.db")
    n = mod.ingest(xlsx_fixture, db_path=db_path)
    assert n == 3

    con = get_conn_for(db_path)
    rows = [tuple(row) for row in con.execute(
        "SELECT semestre, anio, periodo_num, uf_m2, vacancia_pct FROM raw_mercado_bodegas_evolucion "
        "WHERE superseded_at IS NULL ORDER BY anio, periodo_num"
    ).fetchall()]
    con.close()
    assert rows == [
        ("2S-2015", 2015, 2, 0.119, 0.0995),
        ("1S-2016", 2016, 1, 0.118, 0.1228),
        ("2S-2016", 2016, 2, 0.109, 0.0749),
    ]


def test_ingest_idempotente(tmp_path, xlsx_fixture):
    db_path = str(tmp_path / "test.db")
    mod.ingest(xlsx_fixture, db_path=db_path)
    n2 = mod.ingest(xlsx_fixture, db_path=db_path)
    assert n2 == 0

    con = get_conn_for(db_path)
    total = con.execute(
        "SELECT COUNT(*) FROM raw_mercado_bodegas_evolucion WHERE superseded_at IS NULL"
    ).fetchone()[0]
    con.close()
    assert total == 3
