"""Tests para tools.db.ingest_caja."""
from __future__ import annotations

import os
from datetime import date

import openpyxl
import pytest

from tools.db import ingest_caja as mod


def _write_block(ws, row: int, company: str, banks: list[tuple], ffmm: list[tuple]) -> int:
    """Escribe un bloque de sociedad a partir de `row`. banks=[(nombre, saldo, moneda_o_None)],
    ffmm=[(nombre, monto)]. Devuelve la fila siguiente libre tras el bloque + separador."""
    ws.cell(row=row, column=3).value = company
    row += 2
    ws.cell(row=row, column=3).value = "Banco"
    ws.cell(row=row, column=4).value = "N° Cuenta"
    ws.cell(row=row, column=5).value = "Saldo"
    ws.cell(row=row, column=7).value = "Fondo"
    ws.cell(row=row, column=8).value = "FFMM"
    row += 1
    n = max(len(banks), len(ffmm))
    for i in range(n):
        if i < len(banks):
            name, saldo, moneda = banks[i]
            ws.cell(row=row, column=3).value = name
            ws.cell(row=row, column=4).value = "00-1234"
            ws.cell(row=row, column=5).value = saldo
            if moneda:
                ws.cell(row=row, column=6).value = moneda
        if i < len(ffmm):
            fname, fval = ffmm[i]
            ws.cell(row=row, column=7).value = fname
            ws.cell(row=row, column=8).value = fval
        row += 1
    row += 1  # fila en blanco separadora
    return row


def _build_workbook(tmp_path, sheets: dict[str, dict]) -> str:
    """sheets: {nombre_hoja: {company: {'banks': [...], 'ffmm': [...]}}}"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, companies in sheets.items():
        ws = wb.create_sheet(sheet_name)
        row = 3
        for company, data in companies.items():
            row = _write_block(ws, row, company, data.get("banks", []), data.get("ffmm", []))
    path = os.path.join(tmp_path, "saldo_caja_fixture.xlsx")
    wb.save(path)
    return path


# ── parse_sheet_companies ───────────────────────────────────────────────────

def test_parse_sheet_companies_suma_clp_y_separa_usd(tmp_path):
    path = _build_workbook(tmp_path, {
        "27-04-2026": {
            "Inmobiliaria VC SpA": {
                "banks": [("Chile", 100.0, None), ("Security US$", 10.0, None)],
                "ffmm": [("FFMM BTG", 500.0), ("FFMM US$", 5.0)],
            },
        },
    })
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    companies = mod.parse_sheet_companies(wb["27-04-2026"])
    key = mod._normalize_name("Inmobiliaria VC SpA")
    assert companies[key]["clp"] == pytest.approx(600.0)
    assert companies[key]["usd"] == pytest.approx(15.0)


def test_parse_sheet_companies_usa_marcador_columna_f(tmp_path):
    path = _build_workbook(tmp_path, {
        "27-04-2026": {
            "Inmobiliaria Chañarcillo Limitada": {
                "banks": [("Scotiabank", 1000.0, "CLP"), ("Scotiabank", 20.0, "USD")],
                "ffmm": [],
            },
        },
    })
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    companies = mod.parse_sheet_companies(wb["27-04-2026"])
    key = mod._normalize_name("Inmobiliaria Chañarcillo Limitada")
    assert companies[key]["clp"] == pytest.approx(1000.0)
    assert companies[key]["usd"] == pytest.approx(20.0)


def test_parse_sheet_companies_multiples_bloques(tmp_path):
    path = _build_workbook(tmp_path, {
        "27-04-2026": {
            "Torre A S.A.": {"banks": [("Security", 111.0, None)], "ffmm": []},
            "Inmobiliaria Boulevard PT SPA": {"banks": [("Security", 222.0, None)], "ffmm": []},
        },
    })
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    companies = mod.parse_sheet_companies(wb["27-04-2026"])
    assert companies[mod._normalize_name("Torre A S.A.")]["clp"] == pytest.approx(111.0)
    assert companies[mod._normalize_name("Inmobiliaria Boulevard PT SPA")]["clp"] == pytest.approx(222.0)


# ── compute_fund_totals ─────────────────────────────────────────────────────

def _companies_completas_tri():
    keys = [
        mod._TRI_FIP, mod._PT_FIP, mod._VC, mod._CHANARCILLO, mod._TORRE_A,
        mod._BOULEVARD, mod._APO_FIP, mod._APO_SA, mod._VINA, mod._CURICO,
    ]
    return {k: {"clp": 1000.0, "usd": 0.0, "raw_name": k} for k in keys}


def test_compute_fund_totals_aplica_ponderadores_tri():
    companies = _companies_completas_tri()
    totales, faltantes, usd_pend = mod.compute_fund_totals(companies, date(2026, 4, 27), None)
    assert faltantes == {}
    peso_total = sum(mod.FUND_WEIGHTS["TRI"].values())
    assert totales["TRI"] == pytest.approx(1000.0 * peso_total)


def test_compute_fund_totals_usa_dolar_real_si_existe(tmp_db):
    from tools.db import repo_fact
    repo_fact.upsert_dolar(tmp_db, "2026-04-20", 950.0)
    companies = {
        mod._APO_FIP: {"clp": 0.0, "usd": 0.0, "raw_name": "x"},
        mod._APO_SA: {"clp": 0.0, "usd": 100.0, "raw_name": "y"},
    }
    totales, faltantes, usd_fallback = mod.compute_fund_totals(companies, date(2026, 4, 27), tmp_db)
    assert totales["Apo"] == pytest.approx(100.0 * 950.0)
    assert usd_fallback == {}


def test_compute_fund_totals_omite_fondo_con_sociedad_faltante():
    companies = {mod._TRI_FIP: {"clp": 1000.0, "usd": 0.0, "raw_name": "x"}}
    totales, faltantes, usd_pend = mod.compute_fund_totals(companies, date(2026, 4, 27), None)
    assert "TRI" not in totales
    assert "TRI" in faltantes
    assert "Apo" in faltantes and "PT" in faltantes


def test_compute_fund_totals_usd_sin_dolar_usa_fallback_900():
    companies = {
        mod._APO_FIP: {"clp": 1000.0, "usd": 0.0, "raw_name": "x"},
        mod._APO_SA: {"clp": 500.0, "usd": 100.0, "raw_name": "y"},
    }
    totales, faltantes, usd_fallback = mod.compute_fund_totals(companies, date(2026, 4, 27), None)
    assert totales["Apo"] == pytest.approx(1500.0 + 100.0 * mod.FALLBACK_USD_CLP)
    assert usd_fallback["Apo"] == pytest.approx(100.0)


# ── pick_month_end_sheets ───────────────────────────────────────────────────

def test_pick_month_end_sheets_elige_la_mas_cercana_al_cierre():
    sheet_dates = [
        (date(2026, 3, 23), "23-03-2026"),
        (date(2026, 4, 6), "06-04-2026"),
        (date(2026, 4, 27), "27-04-2026"),
    ]
    out = mod.pick_month_end_sheets(sheet_dates)
    # cierre marzo = 31-03: 06-04 (6 días) más cerca que 23-03 (8 días)
    assert out["2026-03"] == (date(2026, 4, 6), "06-04-2026")
    assert out["2026-04"] == (date(2026, 4, 27), "27-04-2026")


def test_pick_month_end_sheets_no_reutiliza_hoja():
    sheet_dates = [(date(2026, 4, 15), "15-04-2026")]
    out = mod.pick_month_end_sheets(sheet_dates)
    assert set(out.keys()) == {"2026-04"}


def test_pick_month_end_sheets_omite_mes_con_gap_grande():
    sheet_dates = [(date(2026, 1, 1), "01-01-2026"), (date(2026, 6, 1), "01-06-2026")]
    out = mod.pick_month_end_sheets(sheet_dates, max_gap_days=20)
    assert "2026-03" not in out


# ── ingestar (end-to-end sobre DB temporal) ─────────────────────────────────

@pytest.fixture
def tmp_db_con_fondos(tmp_db):
    """dim_fondo (TRI/PT/Apo) ya viene seedeada por baseline.sql."""
    return tmp_db


def test_ingestar_es_idempotente(tmp_path, tmp_db_con_fondos):
    companies = {
        "Toesca Rentas Inmobiliarias Fondo Inversión": {"banks": [("Chile", 1000.0, None)], "ffmm": []},
        "Toesca Rentas Inmobiliarias PT Fondo Inversión": {"banks": [("Chile", 300.0, None)], "ffmm": []},
        "Inmobiliaria VC SpA": {"banks": [("Chile", 100.0, None)], "ffmm": []},
        "Inmobiliaria Chañarcillo Limitada": {"banks": [("Chile", 100.0, None)], "ffmm": []},
        "Torre A S.A.": {"banks": [("Security", 900.0, None)], "ffmm": []},
        "Inmobiliaria Boulevard PT SPA": {"banks": [("Security", 300.0, None)], "ffmm": []},
        "Toesca Rentas Inmobiliarias Apoquindo Fondo Inversión": {"banks": [("Chile", 200.0, None)], "ffmm": []},
        "Inmobiliaria Apoquindo S.A": {"banks": [("Chile", 800.0, None)], "ffmm": []},
        "Viña Centro SpA": {"banks": [("Chile", 500.0, None)], "ffmm": []},
        "POWER CENTER CURICÓ LTDA.": {"banks": [("BCI", 100.0, None)], "ffmm": []},
    }
    path = _build_workbook(tmp_path, {"30-04-2026": companies})

    r1 = mod.ingestar(path=path, conn=tmp_db_con_fondos)
    assert r1["meses_procesados"] == ["2026-04"]
    assert r1["filas_upsert"] == 3
    rows = tmp_db_con_fondos.execute(
        "SELECT fondo_key, saldo_clp FROM raw_caja WHERE fecha='2026-04-30' ORDER BY fondo_key"
    ).fetchall()
    assert {r["fondo_key"] for r in rows} == {"Apo", "PT", "TRI"}

    r2 = mod.ingestar(path=path, conn=tmp_db_con_fondos)
    assert r2["filas_upsert"] == 3
    rows2 = tmp_db_con_fondos.execute("SELECT COUNT(*) AS n FROM raw_caja").fetchone()
    assert rows2["n"] == 3  # re-ingestar no duplica (ON CONFLICT upsert)


def test_ingestar_reporta_fondo_con_sociedad_faltante(tmp_path, tmp_db_con_fondos):
    companies = {
        "Toesca Rentas Inmobiliarias Fondo Inversión": {"banks": [("Chile", 1000.0, None)], "ffmm": []},
    }
    path = _build_workbook(tmp_path, {"30-04-2026": companies})
    r = mod.ingestar(path=path, conn=tmp_db_con_fondos)
    assert r["filas_upsert"] == 0
    assert "2026-04" in r["sociedades_faltantes"]
    n = tmp_db_con_fondos.execute("SELECT COUNT(*) AS n FROM raw_caja").fetchone()["n"]
    assert n == 0
