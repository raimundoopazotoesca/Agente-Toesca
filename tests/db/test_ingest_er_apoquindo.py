"""Tests para tools.db.ingest_er_apoquindo."""
from __future__ import annotations

import os
import tempfile
from datetime import date

import openpyxl
import pytest

from tools.db import ingest_er_apoquindo as mod


# ── Fixture xlsx con 3 meses × 2 activos × 10 categorías ────────────────────

_MESES_HEADER = ["dic-24", "ene-25", "feb-25"]
_PERIODOS = ["2024-12", "2025-01", "2025-02"]

_CATS_ORDER = [
    ("(-) Ingresos por Arriendos",       "APO_ING_ARR",   +1),
    ("(-) Gastos Comunes/Vacancia",      "APO_GC_VAC",    -1),
    ("(-) Comisión Corredor",            "APO_COM_CORR",  -1),
    ("(-) Administración",               "APO_ADM",       -1),
    ("Provisión Reparaciones",           "APO_PROV_REP",  -1),
    ("(-) Gastos Bono + Legales + Otros","APO_BONOS_LEG", -1),
    ("(-) Gastos Constructores Asociados (Contabilidad)", "APO_CONSTRUCT", -1),
    ("(-) Gastos IVA no recuperado/Otros Gastos", "APO_IVA_NR", -1),
    ("(-) Contribuciones",               "APO_CONTRIB",   -1),
    ("(-) Seguros",                      "APO_SEG",       -1),
]


def _build_fixture_xlsx(tmp_path) -> str:
    """Construye un xlsx con el layout esperado y valores previsibles.

    monto_activo = mes_index * 1000 + cat_index * 100  (signo aplicado).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apo"
    # Fila 1: A vacío, luego meses
    ws.cell(row=1, column=1).value = "Fondo Apoquindo"
    for j, m in enumerate(_MESES_HEADER, start=2):
        ws.cell(row=1, column=j).value = m
    r = 2
    for cat_idx, (label, _, sign) in enumerate(_CATS_ORDER):
        ws.cell(row=r, column=1).value = label
        r += 1
        # Sub-fila 4700
        ws.cell(row=r, column=1).value = "Apoquindo 4700"
        for mi, _ in enumerate(_MESES_HEADER):
            ws.cell(row=r, column=2 + mi).value = sign * (mi * 1000 + cat_idx * 100 + 47)
        r += 1
        # Sub-fila 4501
        ws.cell(row=r, column=1).value = "Apoquindo 4501"
        for mi, _ in enumerate(_MESES_HEADER):
            ws.cell(row=r, column=2 + mi).value = sign * (mi * 1000 + cat_idx * 100 + 45)
        r += 1
    # Fila NOI Mensual (debe ignorarse)
    ws.cell(row=r, column=1).value = "NOI Mensual"
    for mi in range(len(_MESES_HEADER)):
        ws.cell(row=r, column=2 + mi).value = 99999
    os.makedirs(tmp_path, exist_ok=True)
    path = os.path.join(tmp_path, "apo_fixture.xlsx")
    wb.save(path)
    return path


@pytest.fixture
def fixture_xlsx(tmp_path):
    return _build_fixture_xlsx(str(tmp_path))


# ── Tests ────────────────────────────────────────────────────────────────────

def test_parse_devuelve_60_filas(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    # 3 meses × 2 activos × 10 categorías
    assert len(rows) == 60


def test_parse_activo_keys(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    activos = {r["activo_key"] for r in rows}
    assert activos == {"Apo4501", "Apo4700"}


def test_parse_periodos_yyyy_mm(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    periodos = {r["periodo"] for r in rows}
    assert periodos == {"2024-12", "2025-01", "2025-02"}


def test_parse_pseudo_codigos_completos(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    codigos = {r["cuenta_codigo"] for r in rows}
    esperados = {
        "APO_ING_ARR", "APO_GC_VAC", "APO_COM_CORR", "APO_ADM", "APO_PROV_REP",
        "APO_BONOS_LEG", "APO_CONSTRUCT", "APO_IVA_NR", "APO_CONTRIB", "APO_SEG",
    }
    assert codigos == esperados


def test_parse_signos(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    # Ingresos por arriendos siempre >0
    ings = [r for r in rows if r["cuenta_codigo"] == "APO_ING_ARR"]
    assert all(r["monto_clp"] > 0 for r in ings)
    # Contribuciones (gasto) siempre <0
    contrib = [r for r in rows if r["cuenta_codigo"] == "APO_CONTRIB"]
    assert all(r["monto_clp"] < 0 for r in contrib)


def test_parse_todas_operacionales(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    assert all(r["es_operacional"] == 1 for r in rows)


def test_parse_noi_row_ignorada(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    # 99999 es el valor de la fila NOI del fixture; no debe aparecer
    assert not any(r["monto_clp"] == 99999 for r in rows)


def test_parse_source_metadata(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    assert all(r["source_file"] == fixture_xlsx for r in rows)
    assert all(r["source_sheet"] == "Apo" for r in rows)
    assert all(isinstance(r["source_row"], int) and r["source_row"] > 0 for r in rows)


def test_file_hash_estable(fixture_xlsx):
    h1 = mod._file_hash(fixture_xlsx)
    h2 = mod._file_hash(fixture_xlsx)
    assert h1 == h2
    assert len(h1) == 64


# ── Tests de persistencia ────────────────────────────────────────────────────

import sqlite3


@pytest.fixture
def db_conn(tmp_path):
    """DB en disco (tmp) con schema mínimo necesario."""
    db_path = os.path.join(tmp_path, "test.db")
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        CREATE TABLE ingest_run (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tool TEXT, source_file TEXT, file_hash TEXT,
            rows_in INTEGER, rows_loaded INTEGER,
            started_at TEXT, ended_at TEXT, status TEXT, error TEXT
        );
        CREATE TABLE dim_activo (
            activo_key TEXT PRIMARY KEY, fondo_key TEXT, nombre TEXT,
            tipo TEXT, participacion_fondo_activo REAL, categoria TEXT, sociedad TEXT
        );
        INSERT INTO dim_activo (activo_key, fondo_key, nombre, participacion_fondo_activo)
             VALUES ('Apo4501','Apo','Apoquindo 4501',1.0),
                    ('Apo4700','Apo','Apoquindo 4700',1.0);
        CREATE TABLE raw_er_activo_line (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            activo_key TEXT NOT NULL REFERENCES dim_activo(activo_key),
            periodo TEXT NOT NULL,
            cuenta_codigo TEXT, cuenta_nombre TEXT,
            monto_clp REAL, monto_uf REAL,
            seccion TEXT, es_operacional INTEGER,
            source_file TEXT, source_sheet TEXT, source_row INTEGER,
            file_hash TEXT, ingest_run_id INTEGER REFERENCES ingest_run(id),
            loaded_at TEXT DEFAULT (datetime('now')),
            superseded_at TEXT
        );
    """)
    conn.commit()
    yield conn
    conn.close()


def test_persist_inserta_60_filas(fixture_xlsx, db_conn):
    res = mod.persist(fixture_xlsx, conn=db_conn)
    assert res["status"] == "inserted"
    assert res["rows"] == 60
    n = db_conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE superseded_at IS NULL"
    ).fetchone()[0]
    assert n == 60


def test_persist_idempotente_mismo_hash(fixture_xlsx, db_conn):
    mod.persist(fixture_xlsx, conn=db_conn)
    res2 = mod.persist(fixture_xlsx, conn=db_conn)
    assert res2["status"] == "skipped_idempotent"
    n = db_conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE superseded_at IS NULL"
    ).fetchone()[0]
    assert n == 60  # no duplica


def test_persist_reingesta_supersede_previas(fixture_xlsx, tmp_path, db_conn):
    mod.persist(fixture_xlsx, conn=db_conn)
    # Crear un xlsx con contenido distinto → hash distinto
    fixture_xlsx_2 = _build_fixture_xlsx(str(tmp_path / "sub"))
    # Modificar un valor para cambiar el hash
    import openpyxl as _ox
    wb = _ox.load_workbook(fixture_xlsx_2)
    wb.active.cell(row=1, column=1).value = "Fondo Apoquindo v2"
    wb.save(fixture_xlsx_2)

    res = mod.persist(fixture_xlsx_2, conn=db_conn)
    assert res["status"] == "superseded_and_reinserted"
    activos_total = db_conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line"
    ).fetchone()[0]
    activas = db_conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE superseded_at IS NULL"
    ).fetchone()[0]
    superseded = activos_total - activas
    assert activas == 60
    assert superseded == 60


def test_noi_calculado_matchea_suma(fixture_xlsx, db_conn):
    mod.persist(fixture_xlsx, conn=db_conn)
    # NOI 2025-01 Apo4501 según fixture: sum sobre 10 categorías, mi=1, base=45
    # Ingresos (cat_idx=0, sign=+1): 1*1000 + 0*100 + 45 = 1045
    # Gastos (cat_idx=1..9, sign=-1): sum_i -(1000 + i*100 + 45)
    esperado = 1045 + sum(-(1000 + i*100 + 45) for i in range(1, 10))
    calc = db_conn.execute("""
        SELECT SUM(monto_clp) FROM raw_er_activo_line
         WHERE activo_key='Apo4501' AND periodo='2025-01'
           AND es_operacional=1 AND superseded_at IS NULL
    """).fetchone()[0]
    assert abs(calc - esperado) < 0.01
