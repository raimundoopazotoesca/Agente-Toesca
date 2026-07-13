from __future__ import annotations

from datetime import date
import sqlite3

import openpyxl

from tools.db import ingest_er_pt as mod


def _build_fixture_xlsx(tmp_path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NOI PT"
    ws.cell(row=1, column=1).value = "Cuenta"
    ws.cell(row=1, column=2).value = date(2026, 7, 31)

    labels = {
        3: "(+) Ingresos Torre A S.A",
        11: "(+) Ingresos Inmobiliaria Centro de Convenciones",
        27: "Pago Derecho Uso / Fee Asesor",
        29: "Torre A S.A",
        30: "Inmobiliaria Centro de Convenciones",
        32: "Torre A S.A",
        33: "Inmobiliaria Centro de Convenciones",
        38: "Torre A S.A",
        39: "Inmobiliaria Centro de Convenciones",
        41: "Torre A S.A",
        42: "Inmobiliaria Centro de Convenciones",
        44: "Torre A S.A",
        45: "Inmobiliaria Centro de Convenciones",
    }
    for row, label in labels.items():
        ws.cell(row=row, column=1).value = label

    ws.cell(row=3, column=2).value = 10000.0
    ws.cell(row=11, column=2).value = 20000.0
    ws.cell(row=27, column=2).value = 3000.0
    ws.cell(row=29, column=2).value = 100.0
    ws.cell(row=30, column=2).value = 200.0

    # Valores fuente positivos a proposito: el parser debe convertir gastos a signo negativo
    # y debe sobreescribir los supuestos fijos/derivados definidos para PT.
    for row in (32, 33, 38, 39, 41, 42, 44, 45):
        ws.cell(row=row, column=2).value = 999.0

    path = tmp_path / "noi_pt_fixture.xlsx"
    wb.save(path)
    return str(path)


def _amount(rows: list[dict], activo: str, codigo: str) -> float:
    matches = [r["monto_clp"] for r in rows if r["activo_key"] == activo and r["cuenta_codigo"] == codigo]
    assert len(matches) == 1
    return matches[0]


def test_parse_pt_overrides_gastos_usuario(tmp_path):
    rows = mod.parse_planilla(_build_fixture_xlsx(tmp_path))

    assert _amount(rows, "Torre A", "PT_ADM") == -20.2
    assert _amount(rows, "Boulevard", "PT_ADM") == -46.4

    assert _amount(rows, "Boulevard", "PT_GC_VAC") == -531.0
    assert _amount(rows, "Torre A", "PT_CONTRIB") == -1257.0
    assert _amount(rows, "Boulevard", "PT_CONTRIB") == -621.0
    assert _amount(rows, "Torre A", "PT_SEG") == -173.464166666667
    assert _amount(rows, "Boulevard", "PT_SEG") == -63.46


def test_parse_pt_todos_los_gastos_son_negativos(tmp_path):
    rows = mod.parse_planilla(_build_fixture_xlsx(tmp_path))
    gastos = [r for r in rows if r["seccion"] == "GASTOS_OPERACION"]
    assert gastos
    assert all(r["monto_clp"] < 0 for r in gastos)


def test_parse_pt_no_aplica_reglas_antes_de_vigencia(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NOI PT"
    ws.cell(row=1, column=1).value = "Cuenta"
    ws.cell(row=1, column=2).value = date(2026, 6, 30)
    ws.cell(row=3, column=1).value = "(+) Ingresos Torre A S.A"
    ws.cell(row=3, column=2).value = 10000.0
    ws.cell(row=32, column=1).value = "Torre A S.A"
    ws.cell(row=32, column=2).value = -999.0
    path = tmp_path / "noi_pt_pre_vigencia.xlsx"
    wb.save(path)

    rows = mod.parse_planilla(str(path))

    assert _amount(rows, "Torre A", "PT_ADM") == -999.0


def test_ingest_no_supersede_historia_antes_de_vigencia(tmp_path):
    conn = sqlite3.connect(tmp_path / "test.db")
    conn.executescript("""
        CREATE TABLE ingest_run (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tool TEXT, source_file TEXT, file_hash TEXT,
            rows_in INTEGER, rows_loaded INTEGER,
            started_at TEXT, ended_at TEXT, status TEXT, error TEXT
        );
        CREATE TABLE raw_er_activo_line (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            activo_key TEXT NOT NULL,
            periodo TEXT NOT NULL,
            cuenta_codigo TEXT, cuenta_nombre TEXT,
            monto_clp REAL, monto_uf REAL,
            seccion TEXT, es_operacional INTEGER,
            source_file TEXT, source_sheet TEXT, source_row INTEGER,
            file_hash TEXT, ingest_run_id INTEGER,
            loaded_at TEXT DEFAULT (datetime('now')),
            superseded_at TEXT
        );
        INSERT INTO raw_er_activo_line (
            activo_key, periodo, cuenta_codigo, monto_clp, source_row, file_hash
        ) VALUES
            ('Torre A', '2026-06', 'PT_ADM', -999, 32, 'old_history'),
            ('Torre A', '2026-07', 'PT_ADM', -999, 32, 'old_future');
    """)
    conn.commit()

    res = mod.ingest(_build_fixture_xlsx(tmp_path), conn)

    assert res["status"] == "superseded_and_reinserted"
    june = conn.execute(
        "SELECT superseded_at FROM raw_er_activo_line WHERE file_hash='old_history'"
    ).fetchone()[0]
    july = conn.execute(
        "SELECT superseded_at FROM raw_er_activo_line WHERE file_hash='old_future'"
    ).fetchone()[0]
    assert june is None
    assert july is not None
    conn.close()
