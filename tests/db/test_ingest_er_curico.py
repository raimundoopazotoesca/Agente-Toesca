"""Tests para tools.db.ingest_er_curico."""
from __future__ import annotations

import datetime
import os
import sqlite3

import openpyxl
import pytest

from tools.db import ingest_er_curico as mod


# ── Fixture xlsx replicando la estructura real de RAW/NOI Curico.xlsx ──────
# Fila 4: fechas en columnas D, E (2 periodos reales).
# Columna C = label/código de cuenta a partir de la fila 6 (justo después de
# la fila de fechas, SIN header de texto "Ingreso de Explotacion" — a
# diferencia de Viña, acá la sección INGRESOS_OPERACION es la de arranque
# por defecto).
#
# Estructura del fixture (filas):
#   4  fechas
#   6  leaf ingreso 1 (INGRESOS_OPERACION, sin header previo)
#   7  leaf ingreso 2
#   9  "Total Resultado Operación" (subtotal, cuadra exacto)
#  11  "Ingreso Fuera De Explotacion" (header)
#  12  leaf fuera de explotación
#  13  "Total Ingreso Fuera De Explotación"
#  15  "Gastos de administración y ventas" (header)
#  16  "MANTENCIÓN" (subcategoría, sin código -> se salta)
#  17  leaf gasto 1
#  18  leaf gasto huérfano (3-1-10-115, con valor real pero el subtotal de
#      la fila 19 NO lo incluye, replicando el bug real de la fuente)
#  19  "Total Gastos de administración y ventas" (= solo leaf gasto 1,
#      subestimado a propósito, igual que en el archivo real)
#  20  "Total Operacional" (terminador — PRIMERA ocurrencia)
#  25  "Ingreso Fuera De Explotacion" (repetida, simula la Sección 2 de
#      espejo en UF que está más abajo en el archivo real)
#  26  leaf fantasma (NO debe capturarse — está después del terminador)
#  27  "Total Operacional" (segunda ocurrencia, tampoco debe alcanzarse)

_DATA_COL_1 = 4  # D
_DATA_COL_2 = 5  # E
_PERIODOS = ["2025-01", "2025-02"]
_FECHAS = {_DATA_COL_1: datetime.datetime(2025, 1, 31), _DATA_COL_2: datetime.datetime(2025, 2, 28)}
_UF = {"2025-01-31": 37000.0, "2025-02-28": 37200.0}

_ING_1 = {_DATA_COL_1: 42_000_000.0, _DATA_COL_2: 43_000_000.0}
_ING_2 = {_DATA_COL_1: 10_000_000.0, _DATA_COL_2: 10_500_000.0}
_FUERA = {_DATA_COL_1: 1_000_000.0, _DATA_COL_2: 1_100_000.0}
_GASTO_1 = {_DATA_COL_1: -15_000_000.0, _DATA_COL_2: -16_000_000.0}
_GASTO_HUERFANO = {_DATA_COL_1: -2_000_000.0, _DATA_COL_2: -2_200_000.0}

_TOTAL_RESULTADO_OPERACION = {c: _ING_1[c] + _ING_2[c] for c in (_DATA_COL_1, _DATA_COL_2)}
_TOTAL_FUERA = dict(_FUERA)
# Subestimado a propósito: NO incluye _GASTO_HUERFANO (replica el bug real).
_TOTAL_GASTOS_ADMIN = dict(_GASTO_1)


def _build_fixture_xlsx(tmp_path,
                         corrupt_ingreso_explotacion=False,
                         corrupt_gastos_admin_por_debajo=False) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    for col, fecha in _FECHAS.items():
        ws.cell(row=4, column=col).value = fecha

    ws.cell(row=6, column=3).value = "4-1-01-100  ARRIENDO TIENDAS ANCLAS"
    for col, v in _ING_1.items():
        ws.cell(row=6, column=col).value = v
    ws.cell(row=7, column=3).value = "4-1-01-101  ARRIENDO TIENDAS MENORES"
    for col, v in _ING_2.items():
        ws.cell(row=7, column=col).value = v

    ws.cell(row=9, column=3).value = "Total Resultado Operación"
    tro_vals = dict(_TOTAL_RESULTADO_OPERACION)
    if corrupt_ingreso_explotacion:
        tro_vals = {c: v + 999_999 for c, v in tro_vals.items()}
    for col, v in tro_vals.items():
        ws.cell(row=9, column=col).value = v

    ws.cell(row=11, column=3).value = "Ingreso Fuera De Explotacion"
    ws.cell(row=12, column=3).value = "4-2-01-002  OTROS INGRESOS"
    for col, v in _FUERA.items():
        ws.cell(row=12, column=col).value = v
    ws.cell(row=13, column=3).value = "Total Ingreso Fuera De Explotación"
    for col, v in _TOTAL_FUERA.items():
        ws.cell(row=13, column=col).value = v

    ws.cell(row=15, column=3).value = "Gastos de administración y ventas"
    ws.cell(row=16, column=3).value = "MANTENCIÓN"  # subcategoría, sin código
    ws.cell(row=17, column=3).value = "3-1-10-102  FEE ADMINISTRATIVO (REMUNERACIONES)"
    for col, v in _GASTO_1.items():
        ws.cell(row=17, column=col).value = v
    ws.cell(row=18, column=3).value = "3-1-10-115  MANTENCION COBRO DIRECTO"
    for col, v in _GASTO_HUERFANO.items():
        ws.cell(row=18, column=col).value = v

    ws.cell(row=19, column=3).value = "Total Gastos de administración y ventas"
    tga_vals = dict(_TOTAL_GASTOS_ADMIN)
    if corrupt_gastos_admin_por_debajo:
        # La suma calculada por el parser (gasto1 + huérfano = -17M) debe ser
        # menor en magnitud que el subtotal de la fuente. Hacemos el subtotal
        # más negativo (25% más) para asegurar que la validación falle.
        tga_vals = {c: v * 1.25 for c, v in _GASTO_1.items()}
    for col, v in tga_vals.items():
        ws.cell(row=19, column=col).value = v

    ws.cell(row=20, column=3).value = "Total Operacional"

    # Sección 2 (espejo, más abajo) — nunca debe procesarse.
    ws.cell(row=25, column=3).value = "Ingreso Fuera De Explotacion"
    ws.cell(row=26, column=3).value = "9-9-99-999  CUENTA FANTASMA SECCION 2"
    for col in (_DATA_COL_1, _DATA_COL_2):
        ws.cell(row=26, column=col).value = 999_999_999.0
    ws.cell(row=27, column=3).value = "Total Operacional"

    os.makedirs(tmp_path, exist_ok=True)
    path = os.path.join(tmp_path, "curico_fixture.xlsx")
    wb.save(path)
    return path


@pytest.fixture
def conn(tmp_path):
    """DB en disco (tmp) con schema mínimo necesario (mismo patrón que
    tests/db/test_ingest_er_vina.py)."""
    db_path = os.path.join(str(tmp_path), "test.db")
    c = sqlite3.connect(db_path)
    c.row_factory = sqlite3.Row
    c.executescript("""
        CREATE TABLE ingest_run (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tool TEXT, source_file TEXT, file_hash TEXT,
            rows_in INTEGER, rows_loaded INTEGER,
            started_at TEXT, ended_at TEXT, status TEXT, error TEXT
        );
        CREATE TABLE dim_activo (
            activo_key TEXT PRIMARY KEY, fondo_key TEXT, nombre TEXT,
            tipo TEXT, participacion_fondo_activo REAL, categoria TEXT, sociedad TEXT
        );
        INSERT INTO dim_activo (activo_key, fondo_key, nombre, participacion_fondo_activo)
             VALUES ('Mall Curicó','TRI','Mall Curicó',1.0);
        CREATE TABLE raw_er_activo_line (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            activo_key TEXT NOT NULL REFERENCES dim_activo(activo_key),
            periodo TEXT NOT NULL,
            cuenta_codigo TEXT, cuenta_nombre TEXT,
            monto_clp REAL, monto_uf REAL,
            seccion TEXT, es_operacional INTEGER,
            source_file TEXT, source_sheet TEXT, source_row INTEGER,
            file_hash TEXT, ingest_run_id INTEGER REFERENCES ingest_run(id),
            loaded_at TEXT DEFAULT (datetime('now')),
            superseded_at TEXT
        );
        CREATE TABLE raw_uf_diaria (
            fecha TEXT PRIMARY KEY, valor REAL, fuente TEXT,
            loaded_at TEXT DEFAULT (datetime('now'))
        );
        CREATE VIEW fact_uf AS SELECT fecha, valor FROM raw_uf_diaria ORDER BY fecha;
    """)
    for fecha, valor in _UF.items():
        c.execute(
            "INSERT INTO raw_uf_diaria (fecha, valor, fuente) VALUES (?, ?, 'test')",
            (fecha, valor),
        )
    c.commit()
    yield c
    c.close()


@pytest.fixture
def fixture_xlsx(tmp_path):
    return _build_fixture_xlsx(str(tmp_path))


# ── Tests de parsing ─────────────────────────────────────────────────────

def test_parse_devuelve_5_cuentas_x_2_periodos(fixture_xlsx, conn):
    # 2 ingreso explotación + 1 fuera explotación + 2 gastos (incl. huérfana) = 5 cuentas x 2 periodos.
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    assert len(rows) == 10


def test_parse_activo_key_fijo_curico(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    assert all(r["activo_key"] == "Mall Curicó" for r in rows)


def test_parse_periodos_yyyy_mm(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    assert {r["periodo"] for r in rows} == set(_PERIODOS)


def test_parse_codigos_de_cuenta_extraidos_por_regex(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    codigos = {r["cuenta_codigo"] for r in rows}
    assert codigos == {"4-1-01-100", "4-1-01-101", "4-2-01-002", "3-1-10-102", "3-1-10-115"}


def test_parse_ingresos_operacion_sin_header_explicito(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    by_codigo = {r["cuenta_codigo"]: r for r in rows if r["periodo"] == "2025-01"}
    assert by_codigo["4-1-01-100"]["seccion"] == "INGRESOS_OPERACION"
    assert by_codigo["4-1-01-100"]["es_operacional"] == 1


def test_parse_seccion_y_es_operacional(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    by_codigo = {r["cuenta_codigo"]: r for r in rows if r["periodo"] == "2025-01"}
    assert by_codigo["4-2-01-002"]["seccion"] == "INGRESO_FUERA_EXPLOTACION"
    assert by_codigo["4-2-01-002"]["es_operacional"] == 0
    assert by_codigo["3-1-10-102"]["seccion"] == "GASTOS_OPERACION"
    assert by_codigo["3-1-10-102"]["es_operacional"] == 1


def test_parse_monto_uf_usa_fact_uf(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    r = next(r for r in rows if r["cuenta_codigo"] == "4-1-01-100" and r["periodo"] == "2025-01")
    assert r["monto_clp"] == _ING_1[_DATA_COL_1]
    assert abs(r["monto_uf"] - _ING_1[_DATA_COL_1] / _UF["2025-01-31"]) < 1e-6


def test_parse_noi_incluye_cuenta_huerfana(fixture_xlsx, conn):
    """La cuenta 3-1-10-115 (huérfana, excluida del subtotal de la fuente)
    debe seguir sumando al NOI recalculado."""
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    codigos_gastos = {r["cuenta_codigo"] for r in rows if r["seccion"] == "GASTOS_OPERACION"}
    assert "3-1-10-115" in codigos_gastos
    noi_2025_01 = sum(r["monto_uf"] for r in rows if r["periodo"] == "2025-01" and r["es_operacional"] == 1)
    esperado = (_ING_1[_DATA_COL_1] + _ING_2[_DATA_COL_1] + _GASTO_1[_DATA_COL_1] + _GASTO_HUERFANO[_DATA_COL_1]) / _UF["2025-01-31"]
    assert abs(noi_2025_01 - esperado) < 1e-6


def test_parse_noi_excluye_ingreso_fuera_de_explotacion(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    noi_2025_01 = sum(r["monto_uf"] for r in rows if r["periodo"] == "2025-01" and r["es_operacional"] == 1)
    esperado = (_ING_1[_DATA_COL_1] + _ING_2[_DATA_COL_1] + _GASTO_1[_DATA_COL_1] + _GASTO_HUERFANO[_DATA_COL_1]) / _UF["2025-01-31"]
    assert abs(noi_2025_01 - esperado) < 1e-6
    assert abs(noi_2025_01 - (esperado + _FUERA[_DATA_COL_1] / _UF["2025-01-31"])) > 1e-6


def test_parse_corta_en_primera_ocurrencia_de_total_operacional(fixture_xlsx, conn):
    rows = mod.parse_planilla(fixture_xlsx, conn=conn)
    codigos = {r["cuenta_codigo"] for r in rows}
    assert "9-9-99-999" not in codigos
    assert len({r["seccion"] for r in rows if r["cuenta_codigo"] == "4-2-01-002"}) == 1


def test_parse_valida_integridad_ingreso_explotacion_ok(fixture_xlsx, conn):
    # No debe lanzar: el fixture cuadra por construcción.
    mod.parse_planilla(fixture_xlsx, conn=conn)


def test_parse_falla_si_ingreso_explotacion_no_cuadra(tmp_path, conn):
    path = _build_fixture_xlsx(str(tmp_path), corrupt_ingreso_explotacion=True)
    with pytest.raises(ValueError, match="Validación de integridad"):
        mod.parse_planilla(path, conn=conn)


def test_parse_gastos_admin_permite_gap_de_huerfana_sin_fallar(fixture_xlsx, conn):
    # El fixture ya tiene el subtotal de fuente subestimado (no incluye la
    # huérfana) — la validación blanda debe pasar igual, no lanzar.
    mod.parse_planilla(fixture_xlsx, conn=conn)


def test_parse_falla_si_gastos_admin_cae_por_debajo_del_subtotal_fuente(tmp_path, conn):
    path = _build_fixture_xlsx(str(tmp_path), corrupt_gastos_admin_por_debajo=True)
    with pytest.raises(ValueError, match="Validación de integridad"):
        mod.parse_planilla(path, conn=conn)


# ── Tests de persistencia (idempotencia) ────────────────────────────────

def test_persist_inserta_filas(fixture_xlsx, conn):
    res = mod.persist(fixture_xlsx, conn=conn)
    assert res["status"] == "inserted"
    assert res["rows"] == 10


def test_persist_es_idempotente(fixture_xlsx, conn):
    mod.persist(fixture_xlsx, conn=conn)
    res2 = mod.persist(fixture_xlsx, conn=conn)
    assert res2["status"] == "skipped_idempotent"
    assert res2["rows"] == 0


def test_persist_supersede_en_reingesta_con_cambios(tmp_path, conn):
    path1 = _build_fixture_xlsx(str(tmp_path) + "_1")
    res1 = mod.persist(path1, conn=conn)
    assert res1["status"] == "inserted"

    wb = openpyxl.load_workbook(path1)
    ws = wb["Hoja1"]
    ws.cell(row=6, column=_DATA_COL_1).value = 50_000_000.0
    ws.cell(row=9, column=_DATA_COL_1).value = 50_000_000.0 + _ING_2[_DATA_COL_1]
    path2 = os.path.join(str(tmp_path), "curico_fixture_v2.xlsx")
    os.makedirs(str(tmp_path), exist_ok=True)
    wb.save(path2)

    res2 = mod.persist(path2, conn=conn)
    assert res2["status"] == "superseded_and_reinserted"

    activos = conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE activo_key='Mall Curicó' AND superseded_at IS NULL"
    ).fetchone()[0]
    assert activos == 10


# ── Test de integración de solo-lectura contra el archivo real ──────────
# Se salta automáticamente si el archivo no está disponible en este entorno.

_REAL_XLSX = (
    r"C:\Users\raimundo.opazo\OneDrive - Toesca\Inmobiliario Toesca - Documentos"
    r"\RAW\NOI Curico.xlsx"
)


@pytest.mark.skipif(not os.path.exists(_REAL_XLSX), reason="archivo real no disponible en este entorno")
def test_parse_archivo_real_no_lanza_y_cuadra_integridad(tmp_path):
    """Copia el archivo real a tmp_path (evita locks de OneDrive) y confirma
    que parse_planilla no lanza ValueError sobre el histórico completo
    (34 periodos, 44 cuentas, validación estricta de Ingreso Explotación y
    blanda de Gastos Admin y Ventas)."""
    import shutil
    local_copy = os.path.join(str(tmp_path), "real.xlsx")
    try:
        shutil.copy(_REAL_XLSX, local_copy)
    except (PermissionError, OSError) as e:
        pytest.skip(f"archivo real bloqueado o inaccesible en este entorno: {e}")

    rows = mod.parse_planilla(local_copy)
    assert len(rows) == 1496
    periodos = {r["periodo"] for r in rows}
    assert periodos == {f"2023-{m:02d}" for m in range(8, 13)} | \
                        {f"2024-{m:02d}" for m in range(1, 13)} | \
                        {f"2025-{m:02d}" for m in range(1, 13)} | \
                        {f"2026-{m:02d}" for m in range(1, 6)}
    codigos = {r["cuenta_codigo"] for r in rows}
    assert {"3-1-10-115", "3-1-10-116", "3-1-10-117"} <= codigos
