"""Tests para tools.db.ingest_er_inmosa."""
from __future__ import annotations

import os
import sqlite3

import openpyxl
import pytest

from tools.db import ingest_er_inmosa as mod


# ── Fixture xlsx replicando la estructura real de RAW/NOI INMOSA.xlsx ──────
# Fila 3: header de fechas. Fila 4: vacía. Fila 5: ancla "INMOSA".
# Filas 6-14: 9 filas de categoría (fila 7 duplica fila 6). Fila 15: NOI Mensual.
# Valores reales de ene/feb/mar-2018 tomados del archivo fuente.

_PERIODOS = ["2018-01", "2018-02", "2018-03"]
_FECHAS = [
    __import__("datetime").datetime(2018, 1, 31),
    __import__("datetime").datetime(2018, 2, 28),
    __import__("datetime").datetime(2018, 3, 31),
]
_INGRESOS = [6440.0915337339, 6434.459445817043, 6437.583242252402]
_ADMIN = [-175, -175, -175]
_PROV_REP = [-35.82156799923614, -19.76057110904785, -69.61203164324844]
_NOI_ESPERADO = [6229.2699657346675, 6239.698874707995, 6192.971210609153]


def _build_fixture_xlsx(tmp_path, corrupt_noi: bool = False,
                         unknown_label: bool = False) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # Fila 3: header de fechas (col B=2 en adelante)
    for j, f in enumerate(_FECHAS, start=2):
        ws.cell(row=3, column=j).value = f
    # Fila 4: vacía (implícito)
    # Fila 5: ancla
    ws.cell(row=5, column=1).value = "INMOSA"
    # Fila 6-7: Ingresos (duplicado)
    label_ingresos = "(+) Ingresos por Arriendos"
    ws.cell(row=6, column=1).value = label_ingresos
    ws.cell(row=7, column=1).value = label_ingresos
    for j, v in enumerate(_INGRESOS, start=2):
        ws.cell(row=6, column=j).value = v
        ws.cell(row=7, column=j).value = v
    # Fila 8: Contribuciones (vacía en este fixture, como en el ejemplo real)
    ws.cell(row=8, column=1).value = "(+) Contribuciones"
    # Fila 9: Administración (mojibake real)
    ws.cell(row=9, column=1).value = "unknown label xyz" if unknown_label else "(-) Administraci�n"
    for j, v in enumerate(_ADMIN, start=2):
        ws.cell(row=9, column=j).value = v
    # Fila 10: Provision Reparaciones (con espacio final real)
    ws.cell(row=10, column=1).value = "(-) Provision Reparaciones "
    for j, v in enumerate(_PROV_REP, start=2):
        ws.cell(row=10, column=j).value = v
    # Fila 11: Aseo, Mantención y Otros (mojibake, valores 0)
    ws.cell(row=11, column=1).value = "(-) Aseo, Mantenci�n y Otros"
    for j in range(2, 2 + len(_PERIODOS)):
        ws.cell(row=11, column=j).value = 0
    # Fila 12: Otros Gastos Operacionales (vacía)
    ws.cell(row=12, column=1).value = "(-) Otros Gastos Operacionales"
    # Fila 13: IVA (vacía)
    ws.cell(row=13, column=1).value = "(-) IVA"
    # Fila 14: Seguros (valores 0)
    ws.cell(row=14, column=1).value = "(-) Seguros"
    for j in range(2, 2 + len(_PERIODOS)):
        ws.cell(row=14, column=j).value = 0
    # Fila 15: NOI Mensual (control)
    ws.cell(row=15, column=1).value = "NOI Mensual"
    noi_vals = list(_NOI_ESPERADO)
    if corrupt_noi:
        noi_vals[0] += 1000  # rompe la validación de integridad a propósito
    for j, v in enumerate(noi_vals, start=2):
        ws.cell(row=15, column=j).value = v

    os.makedirs(tmp_path, exist_ok=True)
    path = os.path.join(tmp_path, "inmosa_fixture.xlsx")
    wb.save(path)
    return path


@pytest.fixture
def fixture_xlsx(tmp_path):
    return _build_fixture_xlsx(str(tmp_path))


# ── Tests de parsing ─────────────────────────────────────────────────────

def test_parse_devuelve_24_filas(fixture_xlsx):
    # 8 categorías (sin la duplicada) × 3 meses = 24
    # (fixture has 9 category rows: Ingresos duplicated in rows 6-7, then
    # Contribuciones, Administración, Provision Reparaciones, Aseo, Otros Gastos,
    # IVA, Seguros = 8 unique categories)
    rows = mod.parse_planilla(fixture_xlsx)
    assert len(rows) == 24


def test_parse_activo_key_fijo_inmosa(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    assert all(r["activo_key"] == "INMOSA" for r in rows)


def test_parse_periodos_yyyy_mm(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    assert {r["periodo"] for r in rows} == set(_PERIODOS)


def test_parse_pseudo_codigos_completos(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    codigos = {r["cuenta_codigo"] for r in rows}
    esperados = {
        "INMOSA_ING_ARR", "INMOSA_CONTRIB", "INMOSA_ADM", "INMOSA_PROV_REP",
        "INMOSA_ASEO", "INMOSA_OTROS_GASTOS", "INMOSA_IVA", "INMOSA_SEG",
    }
    assert codigos == esperados


def test_parse_ingresos_no_duplicados(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    ing = [r for r in rows if r["cuenta_codigo"] == "INMOSA_ING_ARR"]
    assert len(ing) == 3  # una fila por periodo, no dos


def test_parse_todas_operacionales(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    assert all(r["es_operacional"] == 1 for r in rows)


def test_parse_noi_row_ignorada(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    # La fila "NOI Mensual" nunca debe generar una fila con cuenta_codigo
    # (no hay categoría mapeada para ese label — si apareciera, cat_meta
    # sería None y el parser fallaría en vez de emitir una fila silenciosa).
    assert all(r["cuenta_codigo"] is not None for r in rows)
    # Ningún monto individual persistido iguala el NOI total del periodo
    # (confirma que la fila de control no se coló como si fuera una categoría).
    montos = {round(r["monto_clp"], 4) for r in rows}
    assert not montos & {round(v, 4) for v in _NOI_ESPERADO}


def test_parse_valores_reales_ingresos(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    ing_by_periodo = {r["periodo"]: r["monto_clp"] for r in rows if r["cuenta_codigo"] == "INMOSA_ING_ARR"}
    for periodo, esperado in zip(_PERIODOS, _INGRESOS):
        assert abs(ing_by_periodo[periodo] - esperado) < 1e-6


def test_parse_contribuciones_puede_ser_negativa(tmp_path):
    """Contribuciones con valor negativo se clasifica igual como INGRESOS_OPERACION
    (la seccion no depende del signo del valor, solo del prefijo de la fuente)."""
    path = _build_fixture_xlsx(str(tmp_path))
    wb = openpyxl.load_workbook(path)
    ws = wb["Hoja1"]
    ws.cell(row=8, column=2).value = -1381.0  # Contribuciones ene-2018 negativa
    # Ajustar NOI de ese periodo para que la suma siga cuadrando
    ws.cell(row=15, column=2).value = _NOI_ESPERADO[0] + (-1381.0)
    wb.save(path)

    rows = mod.parse_planilla(path)
    contrib = [r for r in rows if r["cuenta_codigo"] == "INMOSA_CONTRIB" and r["periodo"] == "2018-01"]
    assert len(contrib) == 1
    assert contrib[0]["monto_clp"] == -1381.0
    assert contrib[0]["seccion"] == "INGRESOS_OPERACION"


def test_parse_categoria_desconocida_falla(tmp_path):
    path = _build_fixture_xlsx(str(tmp_path), unknown_label=True)
    with pytest.raises(ValueError, match=r"(?i)categor[ií]a"):
        mod.parse_planilla(path)


def test_parse_validacion_integridad_falla_si_no_cuadra(tmp_path):
    path = _build_fixture_xlsx(str(tmp_path), corrupt_noi=True)
    with pytest.raises(ValueError, match=r"(?i)noi"):
        mod.parse_planilla(path)


def test_parse_source_metadata(fixture_xlsx):
    rows = mod.parse_planilla(fixture_xlsx)
    assert all(r["source_file"] == fixture_xlsx for r in rows)
    assert all(r["source_sheet"] == "Hoja1" for r in rows)
    assert all(isinstance(r["source_row"], int) and r["source_row"] > 0 for r in rows)


def test_file_hash_estable(fixture_xlsx):
    h1 = mod._file_hash(fixture_xlsx)
    h2 = mod._file_hash(fixture_xlsx)
    assert h1 == h2
    assert len(h1) == 64


# ── Tests de persistencia ────────────────────────────────────────────────

@pytest.fixture
def db_conn(tmp_path):
    """DB en disco (tmp) con schema mínimo necesario."""
    db_path = os.path.join(str(tmp_path), "test.db")
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        CREATE TABLE ingest_run (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tool TEXT, source_file TEXT, file_hash TEXT,
            rows_in INTEGER, rows_loaded INTEGER,
            started_at TEXT, ended_at TEXT, status TEXT, error TEXT
        );
        CREATE TABLE dim_activo (
            activo_key TEXT PRIMARY KEY, fondo_key TEXT, nombre TEXT,
            tipo TEXT, participacion_fondo_activo REAL, categoria TEXT, sociedad TEXT
        );
        INSERT INTO dim_activo (activo_key, fondo_key, nombre, participacion_fondo_activo)
             VALUES ('INMOSA','TRI','INMOSA',0.43);
        CREATE TABLE raw_er_activo_line (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            activo_key TEXT NOT NULL REFERENCES dim_activo(activo_key),
            periodo TEXT NOT NULL,
            cuenta_codigo TEXT, cuenta_nombre TEXT,
            monto_clp REAL, monto_uf REAL,
            seccion TEXT, es_operacional INTEGER,
            source_file TEXT, source_sheet TEXT, source_row INTEGER,
            file_hash TEXT, ingest_run_id INTEGER REFERENCES ingest_run(id),
            loaded_at TEXT DEFAULT (datetime('now')),
            superseded_at TEXT
        );
    """)
    conn.commit()
    yield conn
    conn.close()


def test_persist_inserta_24_filas(fixture_xlsx, db_conn):
    res = mod.persist(fixture_xlsx, conn=db_conn)
    assert res["status"] == "inserted"
    assert res["rows"] == 24
    n = db_conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE superseded_at IS NULL"
    ).fetchone()[0]
    assert n == 24


def test_persist_idempotente_mismo_hash(fixture_xlsx, db_conn):
    mod.persist(fixture_xlsx, conn=db_conn)
    res2 = mod.persist(fixture_xlsx, conn=db_conn)
    assert res2["status"] == "skipped_idempotent"
    n = db_conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE superseded_at IS NULL"
    ).fetchone()[0]
    assert n == 24  # no duplica


def test_persist_reingesta_supersede_previas(fixture_xlsx, tmp_path, db_conn):
    mod.persist(fixture_xlsx, conn=db_conn)
    fixture_xlsx_2 = _build_fixture_xlsx(str(tmp_path / "sub"))
    wb = openpyxl.load_workbook(fixture_xlsx_2)
    wb["Hoja1"].cell(row=5, column=1).value = "INMOSA "  # cambia contenido → cambia hash
    wb.save(fixture_xlsx_2)

    res = mod.persist(fixture_xlsx_2, conn=db_conn)
    assert res["status"] == "superseded_and_reinserted"
    total = db_conn.execute("SELECT COUNT(*) FROM raw_er_activo_line").fetchone()[0]
    activas = db_conn.execute(
        "SELECT COUNT(*) FROM raw_er_activo_line WHERE superseded_at IS NULL"
    ).fetchone()[0]
    assert activas == 24
    assert total - activas == 24


def test_noi_derivado_matchea_suma_esperada(fixture_xlsx, db_conn):
    mod.persist(fixture_xlsx, conn=db_conn)
    for periodo, esperado in zip(_PERIODOS, _NOI_ESPERADO):
        calc = db_conn.execute("""
            SELECT SUM(monto_clp) FROM raw_er_activo_line
             WHERE activo_key='INMOSA' AND periodo=?
               AND es_operacional=1 AND superseded_at IS NULL
        """, (periodo,)).fetchone()[0]
        assert abs(calc - esperado) < 0.01, f"{periodo}: {calc} != {esperado}"


def test_persist_falla_no_escribe_nada_si_integridad_no_cuadra(tmp_path, db_conn):
    """Si la validación de integridad falla, no debe quedar ninguna fila
    persistida (falla atómica antes de tocar la DB)."""
    path = _build_fixture_xlsx(str(tmp_path), corrupt_noi=True)
    with pytest.raises(ValueError, match=r"(?i)noi"):
        mod.persist(path, conn=db_conn)
    n = db_conn.execute("SELECT COUNT(*) FROM raw_er_activo_line").fetchone()[0]
    assert n == 0


# ── Test de integración de solo-lectura contra el archivo real ──────────
# Se salta automáticamente si el archivo no está disponible en este entorno
# (por ejemplo, en CI sin acceso a SharePoint local).

_REAL_XLSX = (
    r"C:\Users\raimundo.opazo\OneDrive - Toesca\Inmobiliario Toesca - Documentos"
    r"\RAW\NOI INMOSA.xlsx"
)


def _copy_real_file(dst: str) -> bool:
    """Copia el archivo real a `dst`. Devuelve True si la copia funcionó.

    En este entorno, el archivo vive en una carpeta sincronizada por OneDrive
    y puede quedar bloqueado para `open()`/`shutil.copy` de Python (WinError
    32/PermissionError) incluso cuando el comando `cp` de shell sí puede
    leerlo (distinto modo de apertura a nivel Win32). Por eso se usa
    `subprocess` con `cp` en vez de `shutil.copy` — verificado necesario en
    este entorno real, no una preferencia arbitraria.
    """
    import subprocess
    if not os.path.exists(_REAL_XLSX):
        return False
    try:
        subprocess.run(["cp", _REAL_XLSX, dst], check=True, capture_output=True)
        return os.path.exists(dst) and os.path.getsize(dst) > 0
    except (subprocess.CalledProcessError, OSError, FileNotFoundError):
        return False


def _real_file_accessible(tmp_path_str: str) -> bool:
    probe = os.path.join(tmp_path_str, "_probe.xlsx")
    ok = _copy_real_file(probe)
    if ok and os.path.exists(probe):
        os.remove(probe)
    return ok


@pytest.mark.skipif(
    not _real_file_accessible(str(__import__("tempfile").mkdtemp())),
    reason="archivo real no disponible o locked en este entorno",
)
def test_parse_archivo_real_no_lanza_y_cuadra_integridad(tmp_path):
    """Copia el archivo real a tmp_path (evita locks de OneDrive) y confirma
    que parse_planilla no lanza ValueError — es decir, la validación de
    integridad (SUM(componentes)==NOI Mensual) cuadra en las 99 columnas
    reales, no solo en el fixture sintético."""
    local_copy = os.path.join(str(tmp_path), "real.xlsx")
    assert _copy_real_file(local_copy), "no se pudo copiar el archivo real pese a pasar el skipif"

    rows = mod.parse_planilla(local_copy)
    assert len(rows) > 0
    periodos = {r["periodo"] for r in rows}
    assert "2018-01" in periodos
    assert "2026-03" in periodos
    assert len(periodos) == 99
