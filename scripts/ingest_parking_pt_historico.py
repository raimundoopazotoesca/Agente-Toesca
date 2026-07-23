"""Ingesta one-shot del histórico de parking Parque Titanium (SABA).

Lee 'RAW/Parking PT DB.xlsx' (hojas Ingresos + Tickets) y consolida en
raw_parking_ingreso_line, raw_parking_gasto_line, raw_parking_ticket_line,
raw_parking_facturacion_line.

No reusable: los rangos de filas están hardcodeados contra esta planilla
puntual. Ver docs/superpowers/specs/2026-07-23-parking-pt-consolidacion-design.md

Uso: python scripts/ingest_parking_pt_historico.py <ruta_a_Parking_PT_DB.xlsx>
"""
import hashlib
import shutil
import sqlite3
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl

DB = Path(__file__).parent.parent / "memory" / "agente_toesca_v2.db"
ACTIVO_KEY = "Parking PT"

VENTAS_ROWS = range(5, 12)      # filas 5-11
GASTOS_ROWS = range(14, 27)     # filas 14-26 (algunas sin código -> se saltan)
TOTAL_INGRESOS_ROW = 13
TOTAL_GASTOS_ROW = 27

FACTURACION_ROWS = {
    29: "saba_neto",
    30: "saba_iva",
    31: "saba_bruto",
    33: "liquidacion_neto",
    34: "liquidacion_iva",
    35: "liquidacion_bruto",
    37: "pago_a_pt",
}

PERIOD_HEADER_ROW = 3
FIRST_PERIOD_COL = 4  # columna D


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    h.update(path.read_bytes())
    return h.hexdigest()


def periodo_cols(ws):
    """columna -> 'YYYY-MM' mientras haya fecha en la fila de encabezado."""
    cols = {}
    c = FIRST_PERIOD_COL
    while True:
        val = ws.cell(row=PERIOD_HEADER_ROW, column=c).value
        if val is None:
            break
        cols[c] = f"{val.year:04d}-{val.month:02d}"
        c += 1
    return cols


def get_or_create_concepto(cur, codigo, nombre, tipo, signo):
    row = cur.execute(
        "SELECT id FROM dim_concepto_parking WHERE codigo IS ? AND nombre=? AND signo=?",
        (str(codigo) if codigo is not None else None, nombre, signo),
    ).fetchone()
    if row:
        return row[0]
    cur.execute(
        "INSERT INTO dim_concepto_parking (codigo, nombre, tipo, signo) VALUES (?,?,?,?)",
        (str(codigo) if codigo is not None else None, nombre, tipo, signo),
    )
    return cur.lastrowid


def already_ingested(cur, file_hash) -> bool:
    for table in (
        "raw_parking_ingreso_line",
        "raw_parking_gasto_line",
        "raw_parking_ticket_line",
        "raw_parking_facturacion_line",
    ):
        row = cur.execute(f"SELECT 1 FROM {table} WHERE file_hash=? LIMIT 1", (file_hash,)).fetchone()
        if row:
            return True
    return False


def ingest_ingresos(cur, ws, periodos, source_file, file_hash, run_id, now):
    cols = periodos
    for r in list(VENTAS_ROWS) + list(GASTOS_ROWS):
        codigo = ws.cell(row=r, column=1).value
        signo_txt = ws.cell(row=r, column=2).value
        nombre = ws.cell(row=r, column=3).value
        if nombre is None:
            continue
        tipo = "venta" if r in VENTAS_ROWS else "gasto"
        if signo_txt == "+":
            signo = 1
        elif signo_txt == "-":
            signo = -1
        else:
            # Filas sin columna de signo explícita (ej. 'Ticket', 'Otras
            # mantenciones'): siguen la convención del resto de su tipo.
            signo = 1 if tipo == "venta" else -1
        concepto_id = get_or_create_concepto(cur, codigo, nombre.strip(), tipo, signo)
        table = "raw_parking_ingreso_line" if tipo == "venta" else "raw_parking_gasto_line"
        for c, periodo in cols.items():
            monto = ws.cell(row=r, column=c).value
            if monto is None:
                continue
            cur.execute(
                f"""INSERT INTO {table}
                    (activo_key, periodo, concepto_id, monto_clp, source_file, file_hash, ingest_run_id, loaded_at)
                    VALUES (?,?,?,?,?,?,?,?)""",
                (ACTIVO_KEY, periodo, concepto_id, float(monto), source_file, file_hash, run_id, now),
            )


def ingest_facturacion(cur, ws, periodos, source_file, file_hash, run_id, now):
    for r, concepto in FACTURACION_ROWS.items():
        for c, periodo in periodos.items():
            monto = ws.cell(row=r, column=c).value
            if monto is None:
                continue
            cur.execute(
                """INSERT INTO raw_parking_facturacion_line
                    (activo_key, periodo, concepto, monto_clp, source_file, file_hash, ingest_run_id, loaded_at)
                    VALUES (?,?,?,?,?,?,?,?)""",
                (ACTIVO_KEY, periodo, concepto, float(monto), source_file, file_hash, run_id, now),
            )


def ingest_tickets(cur, ws, source_file, file_hash, run_id, now):
    r = 2
    n = 0
    while True:
        fecha = ws.cell(row=r, column=1).value
        if fecha is None:
            break
        tickets = ws.cell(row=r, column=6).value
        feriado = ws.cell(row=r, column=7).value
        bruto = ws.cell(row=r, column=8).value
        if tickets is not None:
            cur.execute(
                """INSERT INTO raw_parking_ticket_line
                    (activo_key, fecha, tickets, feriado, monto_bruto_clp,
                     source_file, file_hash, ingest_run_id, loaded_at)
                    VALUES (?,?,?,?,?,?,?,?,?)""",
                (ACTIVO_KEY, fecha.strftime("%Y-%m-%d"), int(tickets), 1 if feriado else 0,
                 float(bruto) if bruto is not None else None,
                 source_file, file_hash, run_id, now),
            )
            n += 1
        r += 1
    return n


def verify_parking_ingest(con: sqlite3.Connection) -> bool:
    cur = con.cursor()
    ok = True

    wb_path = getattr(con, "_verify_source_path", None)

    for label, table in (("ingresos", "raw_parking_ingreso_line"), ("gastos", "raw_parking_gasto_line")):
        rows = cur.execute(
            f"SELECT periodo, SUM(monto_clp) FROM {table} WHERE superseded_at IS NULL GROUP BY periodo"
        ).fetchall()
        print(f"  {label}: {len(rows)} periodos con datos")

    n_null = cur.execute(
        "SELECT COUNT(*) FROM raw_parking_ingreso_line WHERE concepto_id IS NULL OR periodo IS NULL"
    ).fetchone()[0]
    n_null += cur.execute(
        "SELECT COUNT(*) FROM raw_parking_gasto_line WHERE concepto_id IS NULL OR periodo IS NULL"
    ).fetchone()[0]
    if n_null:
        print(f"  MISMATCH nulls en concepto/periodo: {n_null}")
        ok = False
    else:
        print("  OK nulls concepto/periodo = 0")

    return ok


def verify_against_sheet(con, ws, periodos):
    cur = con.cursor()
    ok = True
    for c, periodo in periodos.items():
        esperado_ing = ws.cell(row=TOTAL_INGRESOS_ROW, column=c).value
        if esperado_ing is not None:
            obtenido = cur.execute(
                """SELECT SUM(i.monto_clp) FROM raw_parking_ingreso_line i
                   WHERE i.periodo=? AND i.superseded_at IS NULL""",
                (periodo,),
            ).fetchone()[0] or 0.0
            if abs(obtenido - float(esperado_ing)) > 1:
                print(f"  MISMATCH ingresos periodo={periodo} esperado={esperado_ing} obtenido={obtenido}")
                ok = False

        esperado_gas = ws.cell(row=TOTAL_GASTOS_ROW, column=c).value
        if esperado_gas is not None:
            obtenido = cur.execute(
                """SELECT SUM(g.monto_clp) FROM raw_parking_gasto_line g
                   WHERE g.periodo=? AND g.superseded_at IS NULL""",
                (periodo,),
            ).fetchone()[0] or 0.0
            if abs(obtenido - float(esperado_gas)) > 1:
                print(f"  MISMATCH gastos periodo={periodo} esperado={esperado_gas} obtenido={obtenido}")
                ok = False
    if ok:
        print("  OK ingresos/gastos vs totales de planilla (tolerancia 1 CLP)")
    return ok


def verify_tickets(con, ws):
    cur = con.cursor()
    ok = True
    conteo_planilla = {}
    r = 2
    while True:
        fecha = ws.cell(row=r, column=1).value
        if fecha is None:
            break
        tickets = ws.cell(row=r, column=6).value
        if tickets is not None:
            conteo_planilla[fecha.year] = conteo_planilla.get(fecha.year, 0) + 1
        r += 1
    for anio, esperado in conteo_planilla.items():
        obtenido = cur.execute(
            "SELECT COUNT(*) FROM raw_parking_ticket_line WHERE fecha LIKE ? AND superseded_at IS NULL",
            (f"{anio}-%",),
        ).fetchone()[0]
        if obtenido != esperado:
            print(f"  MISMATCH tickets anio={anio} esperado={esperado} obtenido={obtenido}")
            ok = False
    if ok:
        print("  OK tickets por año")
    return ok


def main(original_path_str):
    original_path = Path(original_path_str)
    with tempfile.TemporaryDirectory() as tmp:
        scratch = Path(tmp) / original_path.name
        shutil.copy2(original_path, scratch)
        file_hash = sha256_file(scratch)

        con = sqlite3.connect(str(DB))
        con.execute("PRAGMA foreign_keys = ON")
        cur = con.cursor()

        if already_ingested(cur, file_hash):
            print(f"ABORT: file_hash {file_hash} ya fue ingestado previamente.")
            con.close()
            return

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        source_file = str(original_path)

        cur.execute(
            """INSERT INTO ingest_run (tool, source_file, file_hash, started_at, status)
               VALUES (?,?,?,?,?)""",
            ("ingest_parking_pt_historico", source_file, file_hash, now, "running"),
        )
        run_id = cur.lastrowid

        wb = openpyxl.load_workbook(scratch, data_only=True)
        ws_ing = wb["Ingresos"]
        ws_tk = wb["Tickets"]
        periodos = periodo_cols(ws_ing)

        try:
            ingest_ingresos(cur, ws_ing, periodos, source_file, file_hash, run_id, now)
            ingest_facturacion(cur, ws_ing, periodos, source_file, file_hash, run_id, now)
            n_tickets = ingest_tickets(cur, ws_tk, source_file, file_hash, run_id, now)
            cur.execute(
                "UPDATE ingest_run SET ended_at=?, status='ok', rows_loaded=? WHERE id=?",
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), n_tickets, run_id),
            )
            con.commit()
        except Exception:
            con.rollback()
            raise

        print(f"Ingest run {run_id} OK. Periodos: {len(periodos)}. Tickets: {n_tickets}.")
        print("\nVerificación:")
        ok1 = verify_parking_ingest(con)
        ok2 = verify_against_sheet(con, ws_ing, periodos)
        ok3 = verify_tickets(con, ws_tk)
        print("\nRESULTADO:", "OK" if (ok1 and ok2 and ok3) else "MISMATCH — revisar arriba")

        con.close()


if __name__ == "__main__":
    main(sys.argv[1])
