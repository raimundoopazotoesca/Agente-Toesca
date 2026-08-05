"""Ingesta las filas "Cuota Financiamiento (UF)" (fila 28) y "RCSD"
(fila 30) de RAW/Cuotas Leasing TRI.xlsx como series oficiales HISTÓRICAS
del gráfico "Evolución NOI y RCSD" del fact sheet TRI.

Son las filas que alimentan directamente el gráfico de referencia del
usuario (bloque filas 24-33: Ingresos/NOI/Cuota Financiamiento/RCSD).

Persiste en derived_kpi: entidad_tipo='fondo', entidad_key='TRI',
kpi='cuota_financiamiento_uf' | 'rcsd_oficial'. Solo cubre los períodos con
dato en la planilla (histórico) — los períodos posteriores al último dato
se calculan en build_factsheet._fetch_noi_rcsd desde raw_amortizacion
(look-through TRI/PT/Apo) y noi_mes/cuota (RCSD derivado).

Re-ejecutar si el usuario actualiza las filas 28/30 de la planilla.
"""
import datetime

import openpyxl

from tools.db.connection import get_conn

SOURCE = (
    r"C:\Users\raimundo.opazo\OneDrive - Toesca\Inmobiliario Toesca - Documentos"
    r"\RAW\Cuotas Leasing TRI.xlsx"
)
LABEL_ROW = 24
LABEL_COL = 6
ROWS = {
    28: ("cuota_financiamiento_uf", "cuota financiamiento",
         "raw_cuotas_leasing_tri_row28_cuota_financiamiento"),
    31: ("rcsd_oficial", "rcsd media m",
         "raw_cuotas_leasing_tri_row31_rcsd_media_movil"),
}

_MESES = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}


def _parse_periodo(label: str) -> str | None:
    # Formato "ene.-18" -> "2018-01"
    if not label or "-" not in label:
        return None
    mes_txt, anio_txt = label.replace(".", "").split("-")
    mes = _MESES.get(mes_txt.strip().lower())
    if mes is None:
        return None
    anio = 2000 + int(anio_txt.strip())
    return f"{anio:04d}-{mes:02d}"


def _read_row(ws, data_row: int, expect_label: str) -> dict:
    label = (ws.cell(row=data_row, column=LABEL_COL).value or "").lower()
    if expect_label not in label:
        raise SystemExit(f"Fila {data_row} no es la esperada: {label!r}")

    maxcol = ws.max_column
    data = {}
    for c in range(7, maxcol + 1):
        raw_label = ws.cell(row=LABEL_ROW, column=c).value
        v = ws.cell(row=data_row, column=c).value
        if v is None:
            continue
        if isinstance(raw_label, datetime.datetime):
            p = raw_label.strftime("%Y-%m")
        elif isinstance(raw_label, str):
            p = _parse_periodo(raw_label)
        else:
            p = None
        if p:
            data[p] = float(v)
    return data


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    ws = wb["Hoja1"]
    conn = get_conn()
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for data_row, (kpi, expect_label, formula) in ROWS.items():
        data = _read_row(ws, data_row, expect_label)
        conn.execute(
            "DELETE FROM derived_kpi WHERE entidad_tipo='fondo' AND entidad_key='TRI' AND kpi=?",
            (kpi,),
        )
        rows = [
            ("fondo", "TRI", p, kpi, None, v, None, formula, None, now)
            for p, v in data.items()
        ]
        conn.executemany(
            "INSERT INTO derived_kpi (entidad_tipo, entidad_key, periodo, kpi, variante, "
            "valor, unidad, formula, ingest_run_id, computed_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        print(f"OK -> {len(rows)} períodos en derived_kpi (kpi={kpi}, TRI), rango {min(data)}..{max(data)}")
    conn.commit()


if __name__ == "__main__":
    main()
